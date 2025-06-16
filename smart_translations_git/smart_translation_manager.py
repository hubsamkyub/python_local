import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from tkinterdnd2 import DND_FILES, TkinterDnD
import os
from dotenv import load_dotenv
import sqlite3
import json
import time
from datetime import datetime
import pandas as pd
import re
from difflib import SequenceMatcher
import threading
from collections import defaultdict
import deepl
import requests
import uuid
from collections import defaultdict, Counter
import openpyxl
from datetime import datetime
from typing import List, Dict, Optional, Tuple

# 분리된 모듈들 import
from utils import TextProtector, TranslationMetrics, ScrollableCheckList
from dialogs.edit_dialogs import InlineEditDialog, GlossaryEditDialog
from dialogs.preview_dialogs import UpdatePreviewDialog, TranslationReportDialog
from dialogs.selection_dialogs import LanguageSelectionDialog
from dialogs.speaker_dialog import SpeakerEditDialog
from config import DEEPL_API_KEY, AZURE_API_KEY, AZURE_REGION, OPENAI_API_KEY, LANG_CODES
from translation_helpers import TranslationHelper
from scenario_manager import ScenarioTranslationManager
from data_quality_manager import DataQualityManager, SmartTranslationManagerIntegration



# # main.spec 파일에 추가
# a = Analysis(
#     ['smart_translation_manager.py'],
#     # ... 기타 설정 ...
#     datas=[('.env.template', '.'), ('credentials.template.json', '.')],  # 템플릿 파일 포함
# )
# === 용어집 매칭 시스템 Import (수정/추가) ===
try:
    from glossary_matcher import SmartGlossaryMatcher
    from text_preprocessor import TextPreprocessor
    from translation_pipeline import ImprovedTranslationPipeline, safe_translate_dict
    PIPELINE_AVAILABLE = True
    print("✅ 용어집 및 번역 파이프라인 모듈 로드 완료")
except ImportError as e:
    print(f"ℹ️ 고급 번역 시스템을 사용할 수 없습니다: {e}")
    print("   기본 번역 방식을 사용합니다.")
    PIPELINE_AVAILABLE = False

class SmartTranslationManager:
    def __init__(self, root):
        self.root = root
        self.root.title("스마트 번역 자동화 시스템")
        self.root.geometry("1600x800")  # 1900x900 → 1400x800로 축소
        
        # 최소 크기 설정
        self.root.minsize(1200, 700)
        
        # === 설정 파일 확인 및 생성 ===
        if not self.check_and_create_config_files():
            return  # 설정 파일 생성이 취소되면 프로그램 종료
        
        # === 기본 속성 초기화 (UI 설정 전에 먼저) ===
        self.translation_db_path = "smart_translations.db"
        self.unique_texts_db_path = "unique_texts.db"
        
        self.translation_memory = {}
        self.glossary = {}
        self.pending_translations = []
        self.exact_matches = {}

        self.excel_import_folder_var = tk.StringVar()
        self.excel_import_files = [] # (파일명, 파일경로) 튜플 저장
        self.excel_import_lang_vars = {} # 언어 선택 체크박스 변수 저장
        self.db_build_mode_var = tk.StringVar(value="conflict") # DB 빌드 모드 선택 변수

        self.MULTI_LANG_GROUP = ["TH", "PT", "FR", "DE", "ES"]
        self.VISIBLE_LANGS = ["EN", "CN", "TW", "TH", "PT", "ES", "FR", "DE"]
        
        # === 중요: 시나리오 매니저 초기화 (UI 설정 전에) ===
        self.scenario_manager = None
        self.metrics = None
        self.text_protector = None
        self.tm_view_search_var = None
        self.tm_search_exact_var = None
        self.improved_pipeline = None
        self.translation_cache_file = "translation_cache.json"
        
        # === UI 설정 ===
        self.setup_ui()
        
        # === 키보드 단축키 설정 ===
        self.setup_keyboard_shortcuts()
        
        # === 데이터베이스 초기화 ===
        self.init_database()
        
        # === 번역 메모리 로드 ===
        self.load_translation_memory()
        
        # === 기타 초기화 ===
        try:
            from utils import TranslationMetrics, TextProtector
            self.metrics = TranslationMetrics()
            self.text_protector = TextProtector()
        except ImportError as e:
            print(f"경고: 유틸리티 모듈 로드 실패: {e}")
            self.metrics = None
            self.text_protector = None
        
        # GPT-4o-mini 최적화를 위한 새로운 속성들
        self.translation_cache = {}  # 번역 결과 캐싱
        self.token_usage_tracker = {
            'total_input_tokens': 0,
            'total_output_tokens': 0,
            'total_requests': 0,
            'cache_hits': 0
        }
        self.batch_size = 5  # 배치 처리 크기

        # 용어집 매칭 시스템 초기화 (선택사항)
        self.glossary_matcher = None
        self.text_preprocessor = None
        self.reassembly_pipeline = None # <--- 새로 추가

        if PIPELINE_AVAILABLE:
            try:
                self.text_preprocessor = TextPreprocessor() # <--- 새로 추가
                self.glossary_matcher = SmartGlossaryMatcher(self.translation_db_path)
                # 파이프라인 인스턴스 생성
                self.improved_pipeline = ImprovedTranslationPipeline(
                    self.glossary_matcher, 
                    self.text_preprocessor
                )
                print("✅ 고급 번역 파이프라인 초기화 완료")
            except Exception as e:
                print(f"⚠️ 고급 번역 시스템 초기화 실패: {e}")
                self.glossary_matcher = None
                self.text_preprocessor = None
                self.reassembly_pipeline = None

        # === 시나리오 매니저 초기화 ===
        self.ensure_scenario_manager()

        # 🆕 용어집을 가장 먼저 로드 (exact_matches 초기화)
        self.load_glossary()
        print(f"용어집 로드 완료: {len(getattr(self, 'exact_matches', {}))}개 항목")
        
        self.load_exclusion_rules()
        self.load_tm_view()
        self.load_conflicts_to_view()
        
        # === 레퍼런스 데이터 영구 저장 시스템 초기화 ===
        self.setup_reference_persistence()
 
    def setup_ui(self):
        """UI 구성 (컴팩트 버전)"""
        # 메인 프레임
        canvas = tk.Canvas(self.root)
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        main_frame = ttk.Frame(scrollable_frame, padding="8")  # 10 → 8
        main_frame.pack(fill="both", expand=True)
        
        main_frame.drop_target_register(DND_FILES)
        main_frame.dnd_bind('<<Drop>>', self.handle_drop)
        
        # 1. 파일 선택 영역 (기존과 동일하지만 패딩 축소)
        file_frame = ttk.LabelFrame(main_frame, text="📁 파일 선택")
        file_frame.pack(fill="x", pady=3)  # 5 → 3
        
        file_inner = ttk.Frame(file_frame, padding="3")  # 5 → 3
        file_inner.pack(fill="x")
        
        ttk.Label(file_inner, text="대상 파일:").grid(row=0, column=0, sticky="w")
        self.file_path_var = tk.StringVar()
        ttk.Entry(file_inner, textvariable=self.file_path_var, width=45).grid(row=0, column=1, sticky="ew", padx=3)  # width 50→45, padx 5→3
        ttk.Button(file_inner, text="찾기", command=self.select_file).grid(row=0, column=2, padx=1)  # padx 2→1
        ttk.Button(file_inner, text="로드", command=self.load_data).grid(row=0, column=3, padx=1)
        
        file_inner.grid_columnconfigure(1, weight=1)
        
        # 2. 번역 설정 & LLM 프롬프트 영역 (너비 조정)
        settings_container = ttk.Frame(main_frame)
        settings_container.pack(fill="x", pady=3)
        
        # 좌측: 번역 설정 (더 컴팩트하게)
        left_settings_frame = ttk.LabelFrame(settings_container, text="⚙️ 번역 설정")
        left_settings_frame.pack(side="left", fill="both", expand=False, padx=(0, 3))  # 5 → 3
        
        settings_inner = ttk.Frame(left_settings_frame, padding="6")  # 10 → 6
        settings_inner.pack(fill="both", expand=True)
        
        # 번역 엔진 선택 (세로 배치를 가로 배치로 변경)
        engine_frame = ttk.LabelFrame(settings_inner, text="번역 엔진")
        engine_frame.pack(fill="x", pady=1)
        
        self.api_engine_var = tk.StringVar(value="deepl")
        engine_inner = ttk.Frame(engine_frame, padding="3")
        engine_inner.pack(fill="x")
        
        ttk.Radiobutton(engine_inner, text="DeepL", variable=self.api_engine_var, 
                    value="deepl", command=self.on_engine_changed).pack(side="left")
        ttk.Radiobutton(engine_inner, text="Azure", variable=self.api_engine_var, 
                    value="azure", command=self.on_engine_changed).pack(side="left", padx=5)
        ttk.Radiobutton(engine_inner, text="LLM", variable=self.api_engine_var, 
                    value="llm", command=self.on_engine_changed).pack(side="left")
        
        # 번역 옵션 (가로 배치)
        options_frame = ttk.LabelFrame(settings_inner, text="번역 옵션")
        options_frame.pack(fill="x", pady=1)
        
        options_inner = ttk.Frame(options_frame, padding="3")
        options_inner.pack(fill="x")
        
        self.translate_en_var = tk.BooleanVar(value=True)
        self.translate_multi_var = tk.BooleanVar(value=False)
        self.translate_cn_tw_var = tk.BooleanVar(value=False)
        
        ttk.Checkbutton(options_inner, text="EN", variable=self.translate_en_var).pack(side="left")
        ttk.Checkbutton(options_inner, text="다국어", variable=self.translate_multi_var).pack(side="left", padx=5)
        ttk.Checkbutton(options_inner, text="CN/TW", variable=self.translate_cn_tw_var).pack(side="left")
        
        # 고급 옵션 (가로 배치)
        advanced_frame = ttk.LabelFrame(settings_inner, text="고급 옵션")
        advanced_frame.pack(fill="x", pady=1)
        
        advanced_inner = ttk.Frame(advanced_frame, padding="3")
        advanced_inner.pack(fill="x")
        
        self.protect_tags_var = tk.BooleanVar(value=True)
        self.complex_markup_var = tk.BooleanVar(value=True)
        self.scenario_translation_var = tk.BooleanVar(value=False)
        self.use_glossary_reassembly_var = tk.BooleanVar(value=True)
        
        ttk.Checkbutton(advanced_inner, text="태그보호", variable=self.protect_tags_var).pack(side="left")
        ttk.Checkbutton(advanced_inner, text="마크업", variable=self.complex_markup_var).pack(side="left", padx=5)
        ttk.Checkbutton(advanced_inner, text="🎭시나리오", variable=self.scenario_translation_var, 
                    command=self.on_scenario_option_changed).pack(side="left")

        # '용어 재조립' 체크박스 추가
        reassembly_cb = ttk.Checkbutton(advanced_inner, text="🧩용어 재조립", variable=self.use_glossary_reassembly_var) # <--- 새로 추가
        reassembly_cb.pack(side="left", padx=5) # <--- 새로 추가
        if not PIPELINE_AVAILABLE: # <--- 새로 추가
             reassembly_cb.config(state="disabled") # <--- 새로 추가

        # 우측: LLM 프롬프트 설정 영역 (높이 축소)
        self.llm_settings_frame = ttk.LabelFrame(settings_container, text="🤖 LLM 프롬프트 설정")
        self.llm_settings_frame.pack(side="right", fill="both", expand=True, padx=(3, 0))
        
        prompt_inner = ttk.Frame(self.llm_settings_frame, padding="6")
        prompt_inner.pack(fill="both", expand=True)
        
        # 프롬프트 템플릿 버튼들 (더 작게)
        template_frame = ttk.Frame(prompt_inner)
        template_frame.pack(fill="x", pady=(0, 3))
        
        ttk.Label(template_frame, text="템플릿:", font=("맑은 고딕", 8)).pack(side="left")
        
        template_buttons = ttk.Frame(template_frame)
        template_buttons.pack(side="right")
        
        button_configs = [("게임", "game"), ("일반", "natural"), ("기술", "technical"), ("초기화", "default")]
        for text, template in button_configs:
            ttk.Button(template_buttons, text=text, 
                    command=lambda t=template: self.set_prompt_template(t), 
                    width=6).pack(side="left", padx=1)
        
        # 프롬프트 입력 영역 (높이 축소)
        ttk.Label(prompt_inner, text="프롬프트:", font=("맑은 고딕", 8)).pack(anchor="w", pady=(3, 1))
        
        prompt_text_frame = ttk.Frame(prompt_inner)
        prompt_text_frame.pack(fill="both", expand=True)
        
        self.llm_prompt_entry = tk.Text(prompt_text_frame, height=6, wrap="word", font=("맑은 고딕", 9))  # height 8→6
        prompt_scrollbar = ttk.Scrollbar(prompt_text_frame, orient="vertical", command=self.llm_prompt_entry.yview)
        self.llm_prompt_entry.configure(yscrollcommand=prompt_scrollbar.set)
        
        self.llm_prompt_entry.pack(side="left", fill="both", expand=True)
        prompt_scrollbar.pack(side="right", fill="y")
        
        # 기본 프롬프트 설정
        default_prompt = """한국어를 영어로 번역해주세요. 10-20대 영미권 사용자가 이해하기 쉽게, 간결하고 자연스럽게 번역하며 특수 태그는 그대로 유지하세요."""
        self.llm_prompt_entry.insert("1.0", default_prompt)
        
        # 프롬프트 하단 도구
        prompt_tools_frame = ttk.Frame(prompt_inner)
        prompt_tools_frame.pack(fill="x", pady=(5, 0))
        
        ttk.Button(prompt_tools_frame, text="📊 최적화 통계", 
                command=self.show_optimization_stats).pack(side="left", padx=2)
        
        # === 3. 중앙: 탭 컨트롤 ===
        tab_control = ttk.Notebook(main_frame)
        tab_control.pack(fill="both", expand=True, pady=5)
        
        # 탭들 설정 (기존과 동일)
        self.translation_tab = ttk.Frame(tab_control)
        tab_control.add(self.translation_tab, text="📝 번역 대상")
        self.setup_translation_tab()
        
        self.scenario_tab = ttk.Frame(tab_control)
        tab_control.add(self.scenario_tab, text="🎭 시나리오 번역")
        self.setup_scenario_tab()
        
        self.tm_management_tab = ttk.Frame(tab_control)
        tab_control.add(self.tm_management_tab, text="💾 TM 관리")
        self.setup_tm_management_tab()
        
        self.conflict_tab = ttk.Frame(tab_control)
        tab_control.add(self.conflict_tab, text="⚠️ 충돌 해결")
        self.setup_conflict_tab()
        
        self.glossary_tab = ttk.Frame(tab_control)
        tab_control.add(self.glossary_tab, text="📚 용어집")
        self.setup_glossary_tab()
        
        self.exclusion_tab = ttk.Frame(tab_control)
        tab_control.add(self.exclusion_tab, text="🚫 제외 목록")
        self.setup_exclusion_tab()
        
        self.history_tab = ttk.Frame(tab_control)
        tab_control.add(self.history_tab, text="📈 번역 이력")
        self.setup_history_tab()
        
        def on_tab_changed(event):
            selected_tab = event.widget.tab('current')['text']
            if selected_tab == "🎭 시나리오 번역":
                self.root.after(100, self.refresh_speaker_list)
        
        tab_control.bind("<<NotebookTabChanged>>", on_tab_changed)
            
        # 4. 하단: 실행 버튼 및 상태 (더 컴팩트)
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill="x", pady=3)
        
        # 버튼들을 2줄로 배치
        button_row1 = ttk.Frame(bottom_frame)
        button_row1.pack(fill="x")
        
        # 첫 번째 줄: 주요 기능
        ttk.Button(button_row1, text="🔍 분석", command=self.analyze_translations, width=8).pack(side="left", padx=1)
        ttk.Button(button_row1, text="🚀 번역", command=self.execute_translation, width=8).pack(side="left", padx=1)
        ttk.Button(button_row1, text="💾 저장", command=self.save_results, width=8).pack(side="left", padx=1)
        
        # 상태 표시
        self.status_label = ttk.Label(button_row1, text="준비됨", foreground="blue", font=("맑은 고딕", 9))
        self.status_label.pack(side="right", padx=5)
        
        # 두 번째 줄: 보조 기능
        button_row2 = ttk.Frame(bottom_frame)
        button_row2.pack(fill="x", pady=1)
        
        ttk.Button(button_row2, text="🔄 재번역", command=self.force_retranslate_selected, width=8).pack(side="left", padx=1)
        ttk.Button(button_row2, text="🗑️ TM삭제", command=self.remove_from_tm, width=8).pack(side="left", padx=1)
        
        # 5. 진행 상황 (높이 축소)
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var)  # height=15 제거
        self.progress_bar.pack(fill="x", pady=1)
            
        # 스크롤바 설정
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 마우스 휠 스크롤 지원
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

    def on_engine_changed(self):
        """번역 엔진 선택이 변경되었을 때 호출 (레이아웃 개선 반영)"""
        selected_engine = self.api_engine_var.get()
        
        if selected_engine == "llm":
            # LLM 선택 시 프롬프트 설정 프레임 표시 (이미 항상 표시됨)
            self.update_status("🤖 LLM 엔진 선택됨 - 우측 프롬프트 영역을 활용하세요")
        else:
            if selected_engine == "deepl":
                self.update_status("🔷 DeepL 엔진 선택됨")
            elif selected_engine == "azure":
                self.update_status("🔶 Azure 엔진 선택됨")

    def set_window_proportions(self):
        """창 크기에 따른 프레임 비율 조정 (선택적 개선)"""
        def adjust_layout():
            window_width = self.root.winfo_width()
            if window_width > 1600:  # 큰 화면
                # 프롬프트 영역을 더 넓게
                left_width = int(window_width * 0.3)   # 30%
                right_width = int(window_width * 0.7)  # 70%
            else:  # 일반 화면
                left_width = int(window_width * 0.4)   # 40%
                right_width = int(window_width * 0.6)  # 60%
        
        # 창 크기 변경 시 자동 조정 (선택사항)
        self.root.bind('<Configure>', lambda e: adjust_layout() if e.widget == self.root else None)

    def setup_scenario_tab(self):
        """시나리오 번역 전용 탭 설정 (컴팩트 버전)"""
        main_frame = ttk.Frame(self.scenario_tab, padding="6")  # 10→6
        main_frame.pack(fill="both", expand=True)
        
        self.ensure_scenario_manager()
        
        # 시나리오 번역 안내 (더 짧게)
        info_frame = ttk.LabelFrame(main_frame, text="ℹ️ 시나리오 번역")
        info_frame.pack(fill="x", pady=3)
        
        info_text = """화자별 맞춤 번역: 캐릭터 말투에 맞는 번역 제공
    사용법: 레퍼런스 설정 → 화자 분석 → 시나리오모드 체크 → 번역 실행"""
        
        ttk.Label(info_frame, text=info_text, font=("맑은 고딕", 8), foreground="navy").pack(padx=5, pady=3)
        
        # 1. 레퍼런스 데이터 설정 (버튼 크기 축소)
        ref_frame = ttk.LabelFrame(main_frame, text="1️⃣ 레퍼런스 데이터")
        ref_frame.pack(fill="x", padx=3, pady=3)
        
        ref_inner = ttk.Frame(ref_frame, padding="3")
        ref_inner.pack(fill="x")
        
        ttk.Label(ref_inner, text="파일:", font=("맑은 고딕", 8)).grid(row=0, column=0, sticky="w")
        self.ref_file_var = tk.StringVar()
        ttk.Entry(ref_inner, textvariable=self.ref_file_var, width=30).grid(row=0, column=1, sticky="ew", padx=3)  # 40→30
        ttk.Button(ref_inner, text="엑셀", command=self.select_reference_file, width=6).grid(row=0, column=2, padx=1)
        ttk.Button(ref_inner, text="구글", command=self.load_reference_from_gsheet, width=6).grid(row=0, column=3, padx=1)
        
        ref_inner.grid_columnconfigure(1, weight=1)
        
        # 분석 버튼들 (더 작게)
        analysis_frame = ttk.Frame(ref_inner)
        analysis_frame.grid(row=1, column=0, columnspan=4, sticky="ew", pady=3)
        
        buttons = [
            ("🔍 구조확인", self.debug_file_structure_detailed),
            ("🚀 자동분석", self.analyze_reference_data_smart),
            ("🔧 수동매핑", self.manual_column_mapping_dialog),
            ("📜 로그", self.show_latest_debug_log),
            ("💾 관리", self.show_reference_dataset_manager)
        ]
        
        for text, command in buttons:
            ttk.Button(analysis_frame, text=text, command=command, width=8).pack(side="left", padx=1)
        
        # 2. 화자 관리 (높이 축소)
        speaker_frame = ttk.LabelFrame(main_frame, text="2️⃣ 화자 관리")
        speaker_frame.pack(fill="both", expand=True, padx=3, pady=3)
        
        speaker_inner = ttk.Frame(speaker_frame, padding="3")
        speaker_inner.pack(fill="both", expand=True)
        
        # 화자 상태 및 새로고침
        status_frame = ttk.Frame(speaker_inner)
        status_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 3))
        
        self.speaker_status_label = ttk.Label(status_frame, text="화자 정보 로딩 중...", foreground="blue", font=("맑은 고딕", 8))
        self.speaker_status_label.pack(side="left")
        
        ttk.Button(status_frame, text="🔄", command=self.refresh_speaker_list, width=4).pack(side="right")
        
        # 화자 리스트 (높이 축소)
        columns = ("화자", "성별", "말투", "번역 스타일", "수")  # "레퍼런스 수" → "수"
        self.speaker_tree = ttk.Treeview(speaker_inner, columns=columns, show="headings", height=6)  # 8→6
        
        # 컬럼 너비 축소
        widths = {"화자": 60, "성별": 40, "말투": 50, "번역 스타일": 150, "수": 40}  # 전체적으로 축소
        for col in columns:
            self.speaker_tree.heading(col, text=col)
            self.speaker_tree.column(col, width=widths.get(col, 80))
        
        speaker_scroll = ttk.Scrollbar(speaker_inner, orient="vertical", command=self.speaker_tree.yview)
        self.speaker_tree.configure(yscrollcommand=speaker_scroll.set)
        
        self.speaker_tree.grid(row=1, column=0, sticky="nsew")
        speaker_scroll.grid(row=1, column=1, sticky="ns")
        
        speaker_inner.grid_rowconfigure(1, weight=1)
        speaker_inner.grid_columnconfigure(0, weight=1)
        
        # 화자 관리 버튼들 (더 작게)
        speaker_btn_frame = ttk.Frame(speaker_inner)
        speaker_btn_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=3)
        
        ttk.Button(speaker_btn_frame, text="➕", command=self.add_speaker, width=4).pack(side="left", padx=1)
        ttk.Button(speaker_btn_frame, text="✏️", command=self.edit_speaker, width=4).pack(side="left", padx=1)
        ttk.Button(speaker_btn_frame, text="🗑️", command=self.delete_speaker, width=4).pack(side="left", padx=1)
        
        # 3. 사용 안내 (더 짧게)
        usage_frame = ttk.LabelFrame(main_frame, text="3️⃣ 사용법")
        usage_frame.pack(fill="x", padx=3, pady=3)
        
        usage_text = """✅ 완료 후: 메인에서 '시나리오모드' 체크 → LLM 선택 → 번역 실행
    💡 레퍼런스는 한 번 분석하면 자동 저장되어 재사용 가능"""
        
        ttk.Label(usage_frame, text=usage_text, font=("맑은 고딕", 8), 
                foreground="darkgreen").pack(padx=5, pady=3)
        
        # 초기 화자 리스트 로드
        self.refresh_speaker_list()
        
    def ensure_scenario_manager(self):
        """시나리오 매니저가 초기화되어 있는지 확인하고, 없으면 생성"""
        try:
            # scenario_manager 속성이 없으면 먼저 None으로 초기화
            if not hasattr(self, 'scenario_manager'):
                self.scenario_manager = None
                
            if not self.scenario_manager:
                try:
                    from scenario_manager import ScenarioTranslationManager
                    self.scenario_manager = ScenarioTranslationManager(self.translation_db_path)
                    print("✅ 시나리오 매니저 초기화 완료")
                except Exception as e:
                    print(f"❌ 시나리오 매니저 초기화 실패: {e}")
                    self.scenario_manager = None
                    
        except Exception as e:
            print(f"ensure_scenario_manager 오류: {e}")
            if not hasattr(self, 'scenario_manager'):
                self.scenario_manager = None

    def refresh_speaker_list(self):
        """화자 리스트 새로고침 (안전한 초기화 포함)"""
        try:
            # 시나리오 매니저 확인
            self.ensure_scenario_manager()
            
            if not self.scenario_manager:
                if hasattr(self, 'speaker_status_label'):
                    self.speaker_status_label.config(text="❌ 시나리오 매니저 초기화 실패", foreground="red")
                return
            
            # 기존 항목 삭제
            if hasattr(self, 'speaker_tree') and self.speaker_tree.winfo_exists():
                for item in self.speaker_tree.get_children():
                    self.speaker_tree.delete(item)
            
            # DB에서 화자 정보 다시 로드
            self.scenario_manager.load_speakers()
            
            # 화자 리스트 업데이트
            speaker_count = 0
            if self.scenario_manager.speakers and hasattr(self, 'speaker_tree'):
                for speaker in self.scenario_manager.speakers.values():
                    self.speaker_tree.insert("", "end", values=(
                        speaker.name,
                        speaker.gender,
                        speaker.tone,
                        speaker.style[:30] + "..." if len(speaker.style) > 30 else speaker.style,
                        speaker.reference_count
                    ))
                    speaker_count += 1
            
            # 상태 라벨 업데이트
            if hasattr(self, 'speaker_status_label'):
                if speaker_count > 0:
                    self.speaker_status_label.config(
                        text=f"✅ {speaker_count}명의 화자 로드 완료", 
                        foreground="green"
                    )
                else:
                    self.speaker_status_label.config(
                        text="ℹ️ 저장된 화자가 없습니다. 레퍼런스 분석을 먼저 진행하세요.", 
                        foreground="orange"
                    )
            
            print(f"화자 리스트 새로고침 완료: {speaker_count}명")
            
        except Exception as e:
            print(f"화자 리스트 새로고침 오류: {e}")
            if hasattr(self, 'speaker_status_label'):
                self.speaker_status_label.config(text=f"❌ 오류: {str(e)[:50]}...", foreground="red")
                
    def update_speaker_list(self):
        """화자 목록 업데이트 (안전한 초기화 포함)"""
        try:
            if not hasattr(self, 'scenario_manager') or not self.scenario_manager:
                self.ensure_scenario_manager()
                
            if not self.scenario_manager:
                return
                
            # 기존 항목 삭제
            if hasattr(self, 'speaker_tree') and self.speaker_tree.winfo_exists():
                for item in self.speaker_tree.get_children():
                    self.speaker_tree.delete(item)
                
                # 화자들 추가
                speaker_count = 0
                for speaker in self.scenario_manager.speakers.values():
                    self.speaker_tree.insert("", "end", values=(
                        speaker.name,
                        speaker.gender,
                        speaker.tone,
                        speaker.style[:30] + "..." if len(speaker.style) > 30 else speaker.style,
                        speaker.reference_count
                    ))
                    speaker_count += 1
                
                # 상태 업데이트
                if hasattr(self, 'speaker_status_label'):
                    self.speaker_status_label.config(
                        text=f"✅ {speaker_count}명의 화자 정보", 
                        foreground="green"
                    )
                    
        except Exception as e:
            print(f"화자 목록 업데이트 오류: {e}")

    def on_scenario_option_changed(self):
        """시나리오 번역 옵션 변경 시 호출 (프롬프트 창 업데이트 포함)"""
        if self.scenario_translation_var.get():
            # LLM 엔진 강제 선택
            self.api_engine_var.set("llm")
            self.on_engine_changed()
            
            # 시나리오용 프롬프트로 변경
            self.set_scenario_mode_prompt()
            
            self.update_status("✅ 시나리오 번역 모드 활성화 (화자별 맞춤 번역)")
            
            # 시나리오 탭으로 안내
            messagebox.showinfo("시나리오 번역 모드", 
                            "🎭 시나리오 번역이 활성화되었습니다!\n\n"
                            "✨ 번역 시 화자별 맞춤 프롬프트가 자동 생성됩니다.\n"
                            "📝 프롬프트 창은 기본 템플릿이며, 실제로는 화자별로 최적화된 프롬프트를 사용합니다.\n\n"
                            "'🎭 시나리오 번역' 탭에서 레퍼런스 데이터와 화자를 설정하세요.")
        else:
            # 일반 모드로 복원
            self.set_prompt_template("default")
            self.update_status("❌ 시나리오 번역 모드 비활성화")

    def setup_keyboard_shortcuts(self):
        """키보드 단축키 설정"""
        self.root.bind('<Control-o>', lambda e: self.select_file())          # Ctrl+O: 파일 열기
        self.root.bind('<Control-l>', lambda e: self.load_data())            # Ctrl+L: 데이터 로드  
        self.root.bind('<Control-t>', lambda e: self.analyze_translations()) # Ctrl+T: 번역 분석
        self.root.bind('<Control-r>', lambda e: self.execute_translation())  # Ctrl+R: 번역 실행
        self.root.bind('<Control-s>', lambda e: self.save_results())         # Ctrl+S: 결과 저장
        self.root.bind('<F5>', lambda e: self.update_translation_table())    # F5: 새로고침

    def setup_compact_tabs(self):
        """탭 제목을 더 컴팩트하게 수정"""
        # 이미 위의 setup_ui에서 적용됨
        pass

    # 진행바 높이 조정 및 스타일 개선
    def setup_progress_bar(self):
        """진행바 설정 개선"""
        style = ttk.Style()
        style.configure("Compact.Horizontal.TProgressbar", thickness=15)
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            self.main_frame, 
            variable=self.progress_var, 
            style="Compact.Horizontal.TProgressbar"
        )

    # 상태바에 도움말 추가
    def setup_status_help(self):
        """상태바 하단에 도움말 텍스트 추가"""
        help_frame = ttk.Frame(self.main_frame)
        help_frame.pack(fill="x", pady=2)
        
        help_text = "💡 단축키: Ctrl+O(파일열기) | Ctrl+L(로드) | Ctrl+T(분석) | Ctrl+R(번역) | Ctrl+S(저장) | F5(새로고침)"
        help_label = ttk.Label(help_frame, text=help_text, font=("맑은 고딕", 8), foreground="gray")
        help_label.pack()

    # 반응형 윈도우 크기 조정
    def setup_responsive_layout(self):
        """반응형 레이아웃 설정"""
        # 최소 크기 설정
        self.root.minsize(1200, 700)
        
        # 윈도우 크기 변경 시 컴포넌트 크기 자동 조정
        def on_window_resize(event):
            if event.widget == self.root:
                # 윈도우가 너무 작아지면 스크롤바 표시
                if self.root.winfo_width() < 1200:
                    # 컴포넌트들을 더 작게 조정
                    pass
        
        self.root.bind('<Configure>', on_window_resize)

    # 다크모드 지원 (선택사항)
    def setup_theme_support(self):
        """테마 지원 설정"""
        style = ttk.Style()
        
        # 사용 가능한 테마 확인
        available_themes = style.theme_names()
        
        # 시스템에 따른 기본 테마 선택
        if 'vista' in available_themes:
            style.theme_use('vista')  # Windows 10/11
        elif 'aqua' in available_themes:
            style.theme_use('aqua')   # macOS
        else:
            style.theme_use('clam')   # 기본

    # 메모리 사용량 표시 (디버깅용, 선택사항)
    def setup_memory_monitor(self):
        """메모리 사용량 모니터링 (개발자용)"""
        import psutil
        import threading
        
        def update_memory_info():
            while True:
                try:
                    process = psutil.Process()
                    memory_mb = process.memory_info().rss / 1024 / 1024
                    
                    if hasattr(self, 'memory_label'):
                        self.root.after(0, lambda: self.memory_label.config(
                            text=f"메모리: {memory_mb:.1f}MB"
                        ))
                    
                    time.sleep(5)  # 5초마다 업데이트
                except:
                    break
        
        # 메모리 표시 라벨 (상태바에 추가)
        self.memory_label = ttk.Label(self.status_frame, text="메모리: 0MB", font=("맑은 고딕", 8))
        self.memory_label.pack(side="left", padx=10)
        
        # 백그라운드 스레드로 모니터링 시작
        memory_thread = threading.Thread(target=update_memory_info, daemon=True)
        memory_thread.start()

    # 통합 초기화 메서드
    def setup_enhanced_ui(self):
        """향상된 UI 설정 통합 메서드"""
        self.setup_ui()                    # 메인 UI
        self.setup_keyboard_shortcuts()    # 키보드 단축키
        self.setup_status_help()          # 도움말
        self.setup_responsive_layout()     # 반응형 레이아웃
        self.setup_theme_support()        # 테마 지원
        # self.setup_memory_monitor()     # 메모리 모니터 (선택사항)

    def load_data(self):
        """데이터 로드 (제외 규칙 적용 및 #번역요청 컬럼 처리)"""
        file_path = self.file_path_var.get()
        if not file_path or not os.path.exists(file_path):
            messagebox.showwarning("경고", "파일을 선택하세요.")
            return
            
        try:
            self.update_status("최신 번역 메모리 로딩 중...")
            self.load_translation_memory()
            
            # DB에서 활성화된 제외 규칙들만 미리 불러옴
            conn = sqlite3.connect(self.translation_db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT id, rule_type, field, value FROM exclusion_rules WHERE is_enabled = 1")
            active_rules = [{'id': r[0], 'type': r[1], 'field': r[2], 'value': r[3]} for r in cursor.fetchall()]
            conn.close()

            df = pd.read_excel(file_path, skiprows=3)
            if "STRING_ID" not in df.columns or "KR" not in df.columns:
                messagebox.showerror("오류", "STRING_ID와 KR 컬럼이 필요합니다.")
                return
            
            # #번역요청 컬럼 체크
            has_translation_request = "#번역요청" in df.columns
            if has_translation_request:
                self.update_status("'#번역요청' 컬럼이 발견되었습니다.")
                
            self.pending_translations = []
            excluded_count = 0
            
            for _, row in df.iterrows():
                if pd.isna(row["KR"]) or str(row["KR"]).strip() == "":
                    continue

                # 제외 규칙 검사
                entry_dict = row.to_dict()
                if self._is_entry_excluded(entry_dict, active_rules):
                    excluded_count += 1
                    continue
                    
                item = {
                    "STRING_ID": str(row["STRING_ID"]), 
                    "KR": str(row["KR"]).strip(),
                    "status": self.determine_status(str(row["KR"]).strip()),
                    "translations": {}, 
                    "method": "",
                    "translation_request": ""  # 새로운 필드 추가
                }
                
                # #번역요청 컬럼 처리
                if has_translation_request and not pd.isna(row["#번역요청"]):
                    item["translation_request"] = str(row["#번역요청"]).strip()
                
                for lang in self.VISIBLE_LANGS:
                    if lang in df.columns and not pd.isna(row[lang]):
                        item["translations"][lang] = str(row[lang])
                        
                self.pending_translations.append(item)
                
            # 번역요청 필터 옵션 업데이트
            if has_translation_request:
                self.update_translation_request_filter_options()
                
            self.update_translation_table()
            self.update_status(f"데이터 로드 완료: {len(self.pending_translations)}개 항목 (제외: {excluded_count}개)")
            
        except Exception as e:
            messagebox.showerror("오류", f"파일 로드 실패: {str(e)}")

    def setup_translation_tab(self):
        """번역 대상 탭 설정 (초컴팩트 버전)"""
        # 검색 및 필터 프레임
        filter_frame = ttk.Frame(self.translation_tab)
        filter_frame.pack(fill="x", padx=3, pady=3)
        
        # 첫 번째 행: 검색 및 상태 필터 (더 컴팩트)
        filter_row1 = ttk.Frame(filter_frame)
        filter_row1.pack(fill="x")
        
        ttk.Label(filter_row1, text="검색:", font=("맑은 고딕", 8)).pack(side="left")
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(filter_row1, textvariable=self.search_var, width=20)  # 30→20
        search_entry.pack(side="left", padx=3)
        search_entry.bind("<KeyRelease>", lambda e: self.filter_translations())
        
        ttk.Label(filter_row1, text="상태:", font=("맑은 고딕", 8)).pack(side="left", padx=(10, 3))
        self.filter_vars = {
            "신규": tk.BooleanVar(value=True), 
            "변경": tk.BooleanVar(value=True),
            "확인필요": tk.BooleanVar(value=True), 
            "확정": tk.BooleanVar(value=True),
            "완료": tk.BooleanVar(value=True),
            "재번역완료": tk.BooleanVar(value=True)
        }
        for status, var in self.filter_vars.items():
            ttk.Checkbutton(filter_row1, text=status, variable=var,
                        command=self.filter_translations).pack(side="left", padx=2)
        
        self.stats_frame = ttk.Frame(filter_row1)
        self.stats_frame.pack(side="right", padx=5)
        
        # 두 번째 행: 방법 필터 및 번역요청 필터 (더 컴팩트)
        filter_row2 = ttk.Frame(filter_frame)
        filter_row2.pack(fill="x", pady=1)
        
        ttk.Label(filter_row2, text="방법:", font=("맑은 고딕", 8)).pack(side="left")
        self.method_filter_var = tk.StringVar(value="전체")
        self.method_filter_combo = ttk.Combobox(filter_row2, textvariable=self.method_filter_var, state="readonly", width=10)  # 15→10
        self.method_filter_combo.pack(side="left", padx=3)
        self.method_filter_combo.bind("<<ComboboxSelected>>", lambda e: self.filter_translations())
        
        ttk.Label(filter_row2, text="번역요청:", font=("맑은 고딕", 8)).pack(side="left", padx=(10, 3))
        self.translation_request_filter_var = tk.StringVar(value="전체")
        self.translation_request_filter_combo = ttk.Combobox(filter_row2, textvariable=self.translation_request_filter_var, state="readonly", width=10)  # 15→10
        self.translation_request_filter_combo.pack(side="left", padx=3)
        self.translation_request_filter_combo.bind("<<ComboboxSelected>>", lambda e: self.filter_translations())
        
        self.update_stats_label()
            
        # 전체 선택/해제 체크박스
        select_all_frame = ttk.Frame(self.translation_tab)
        select_all_frame.pack(fill="x", padx=3, pady=(3,0))
        self.select_all_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(select_all_frame, text="전체 선택/해제", variable=self.select_all_var, command=self.toggle_all_selections).pack(side="left")
        
        # 동적 컬럼 리스트 생성 (#번역요청 컬럼 추가)
        base_columns = ["선택", "STRING_ID", "KR", "상태"]
        lang_columns = self.VISIBLE_LANGS
        end_columns = ["번역요청", "번역방법"]
        columns = base_columns + lang_columns + end_columns

        tree_frame = ttk.Frame(self.translation_tab)
        tree_frame.pack(fill="both", expand=True, padx=3, pady=3)
        
        self.translation_tree = ttk.Treeview(tree_frame, columns=columns, show="tree headings")
        
        # 컬럼 설정 (대폭 축소)
        self.translation_tree.column("#0", width=0, stretch=False)
        self.translation_tree.heading("#0", text="")

        # 고정 컬럼 설정 (너비 대폭 축소)
        self.translation_tree.column("선택", width=30, anchor="center")  # 40→30
        self.translation_tree.column("STRING_ID", width=100)  # 150→100
        self.translation_tree.column("KR", width=180)  # 250→180
        self.translation_tree.column("상태", width=60)  # 80→60
        self.translation_tree.column("번역요청", width=70)  # 100→70
        self.translation_tree.column("번역방법", width=80)  # 100→80

        # 동적 언어 컬럼 설정 (대폭 축소)
        for lang_col in lang_columns:
            self.translation_tree.column(lang_col, width=100)  # 150→100

        # 헤더 텍스트도 짧게
        for col in columns:
            header_text = col
            if col in ["번역요청", "번역방법"]:
                header_text = col[:4]  # "번역요청" → "번역요", "번역방법" → "번역방"
            self.translation_tree.heading(col, text=header_text)
        
        # 스크롤바 및 이벤트 바인딩 (기존과 동일)
        v_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.translation_tree.yview)
        h_scroll = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.translation_tree.xview)
        self.translation_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        
        self.translation_tree.grid(row=0, column=0, sticky="nsew")
        v_scroll.grid(row=0, column=1, sticky="ns")
        h_scroll.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # 나머지는 기존과 동일...
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="🔄 이 항목 재번역 (API 강제)", command=self.force_retranslate_selected)
        self.context_menu.add_command(label="✏️ 직접 편집", command=self.edit_translation_inline)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="🗑️ TM에서 삭제", command=self.remove_from_tm)
        self.context_menu.add_command(label="📋 TM 항목 보기", command=self.view_tm_entry)
        
        self.translation_tree.bind("<Button-3>", self.show_context_menu)
        self.translation_tree.bind("<Control-Button-1>", self.show_context_menu)
        self.translation_tree.bind("<Button-1>", self.on_tree_click)
        self.translation_tree.bind("<KeyPress-space>", self.toggle_selected_checkboxes)
        self.translation_tree.bind("<Return>", self.toggle_selected_checkboxes)
        self.translation_tree.bind("<Control-a>", self.select_all_items)
        self.translation_tree.bind("<Control-d>", self.deselect_all_items)
        self.translation_tree.bind("<Control-i>", self.invert_selection)
        self.translation_tree.bind("<Delete>", self.clear_selected_translations)
        self.translation_tree.bind("<F2>", self.edit_selected_item)
        
        self.translation_tree.focus_set()
        
        # 키보드 단축키 안내 (더 짧게)
        help_frame = ttk.Frame(self.translation_tab)
        help_frame.pack(fill="x", padx=3, pady=1)
        
        help_text = "💡 Space=토글 | Ctrl+A=전체선택 | Ctrl+D=전체해제 | F2=편집 | Del=삭제"
        help_label = ttk.Label(help_frame, text=help_text, font=("맑은 고딕", 7), foreground="gray")  # 8→7
        help_label.pack(side="right")
        
        self.check_states = {}

    def update_translation_request_filter_options(self):
        """번역요청 필터 옵션 업데이트"""
        if not hasattr(self, 'translation_request_filter_combo'):
            return
            
        # 고유한 번역요청 값들 수집
        request_values = set()
        for trans in self.pending_translations:
            request = trans.get("translation_request", "").strip()
            if request:
                request_values.add(request)
        
        # 필터 옵션 업데이트
        options = ["전체", "요청없음"] + sorted(list(request_values))
        self.translation_request_filter_combo['values'] = options

    def update_translation_table(self):
        """번역 테이블 업데이트 (필터 개선 및 #번역요청 컬럼 포함)"""
        self.translation_tree.delete(*self.translation_tree.get_children())
            
        search_text = self.search_var.get().lower()
        active_filters = {status for status, var in self.filter_vars.items() if var.get()}
        selected_method = self.method_filter_var.get()
        selected_request = self.translation_request_filter_var.get()
        
        # 번역 방법 필터 값 업데이트
        all_methods = sorted(list(set(t["method"] for t in self.pending_translations if t["method"])))
        self.method_filter_combo['values'] = ["전체"] + all_methods
        
        for trans in self.pending_translations:
            # 검색 필터
            if search_text and search_text not in trans["KR"].lower() and search_text not in trans["STRING_ID"].lower():
                continue
                
            # 상태 필터
            status_text = trans["status"].strip("[]")
            if status_text not in active_filters:
                continue

            # 방법 필터
            if selected_method != "전체" and trans["method"] != selected_method:
                continue
                
            # 번역요청 필터
            request_text = trans.get("translation_request", "").strip()
            if selected_request != "전체":
                if selected_request == "요청없음" and request_text:
                    continue
                elif selected_request != "요청없음" and request_text != selected_request:
                    continue
                        
            # 값 리스트 생성
            values = [
                "☑",
                trans["STRING_ID"],
                trans["KR"],
                trans["status"],
            ]
            
            # 언어별 번역문 추가
            for lang in self.VISIBLE_LANGS:
                values.append(trans["translations"].get(lang, ""))
                
            values.append(trans.get("translation_request", ""))  # 번역요청 컬럼 추가
            values.append(trans["method"])
            
            item_id = self.translation_tree.insert("", "end", values=values)
            self.check_states[item_id] = True
            
        self.update_stats_label()


    #GPT 설정
    def calculate_token_estimate(self, text):
        """토큰 수 간단 추정 (GPT-4o-mini 기준)"""
        # 간단한 토큰 추정: 영어는 ~4글자당 1토큰, 한국어는 ~2글자당 1토큰
        return len(text) // 3

    def should_use_batch_translation(self, items_to_translate):
        """배치 번역 사용 여부 결정"""
        if len(items_to_translate) < 3:
            return False
        
        # 총 토큰 수 추정
        total_tokens = sum(self.calculate_token_estimate(item["KR"]) for item in items_to_translate)
        
        # 3000 토큰 이하이고 같은 화자면 배치 처리 효율적
        return total_tokens < 3000

    def get_cache_key(self, text, prompt_hash, speaker=None):
        """캐싱을 위한 키 생성"""
        import hashlib
        cache_string = f"{text}|{prompt_hash}|{speaker or 'default'}"
        return hashlib.md5(cache_string.encode()).hexdigest()

    def get_cached_translation(self, cache_key):
        """캐시된 번역 결과 확인"""
        if cache_key in self.translation_cache:
            self.token_usage_tracker['cache_hits'] += 1
            return self.translation_cache[cache_key]
        return None

    def cache_translation(self, cache_key, result):
        """번역 결과 캐싱 (메모리 관리 포함)"""
        # 캐시 크기 제한 (최대 1000개)
        if len(self.translation_cache) > 1000:
            # 오래된 항목부터 제거 (FIFO)
            oldest_key = list(self.translation_cache.keys())[0]
            del self.translation_cache[oldest_key]
        
        self.translation_cache[cache_key] = result

    def translate_with_llm_optimized(self, text, prompt, speaker=None, max_retries=3):
        """GPT-4o-mini 최적화된 LLM 번역"""
        # 1. 캐시 확인
        import hashlib
        prompt_hash = hashlib.md5(prompt.encode()).hexdigest()[:8]
        cache_key = self.get_cache_key(text, prompt_hash, speaker)
        
        cached_result = self.get_cached_translation(cache_key)
        if cached_result:
            return cached_result
        
        # 2. 토큰 수 사전 체크
        estimated_tokens = self.calculate_token_estimate(text + prompt)
        if estimated_tokens > 3500:  # GPT-4o-mini 컨텍스트 고려
            print(f"⚠️ 토큰 수 초과 예상: {estimated_tokens}")
            # 긴 텍스트는 분할 처리
            return self.translate_long_text(text, prompt, speaker)
        
        # 3. 최적화된 API 호출
        result = self.call_gpt4o_mini_api(text, prompt, max_retries)
        
        # 4. 결과 캐싱
        if result:
            self.cache_translation(cache_key, result)
        
        return result

    def call_gpt4o_mini_api(self, text, prompt, max_retries=3):
        """GPT-4o-mini API 호출 (최적화된 설정)"""
        for attempt in range(max_retries):
            try:
                import openai
                
                client = openai.OpenAI(api_key=OPENAI_API_KEY)
                
                # GPT-4o-mini 최적화 프롬프트
                full_prompt = f"{prompt}\n\n{text}"
                
                response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {
                            "role": "system", 
                            "content": "You are a professional translator. Return only the translation, no explanations."
                        },
                        {
                            "role": "user", 
                            "content": full_prompt
                        }
                    ],
                    max_tokens=min(800, len(text) * 3),  # 동적 토큰 제한
                    temperature=0.2,  # 일관성 우선
                    top_p=0.9,  # 품질 향상
                    frequency_penalty=0.1,  # 반복 방지
                    presence_penalty=0.1,   # 다양성 소폭 증가
                    timeout=25  # 타임아웃 단축
                )
                
                # 토큰 사용량 추적
                if hasattr(response, 'usage'):
                    self.token_usage_tracker['total_input_tokens'] += response.usage.prompt_tokens
                    self.token_usage_tracker['total_output_tokens'] += response.usage.completion_tokens
                self.token_usage_tracker['total_requests'] += 1
                
                if response.choices and len(response.choices) > 0:
                    result = response.choices[0].message.content.strip()
                    result = self.post_process_llm_result(result, text)
                    return result
                        
            except Exception as e:
                print(f"GPT-4o-mini API 호출 실패 (시도 {attempt + 1}): {e}")
                if attempt == max_retries - 1:
                    return None
                    
                # 지수적 백오프 (토큰 제한 오류 시 더 긴 대기)
                import time
                wait_time = (2 ** attempt) * (2 if "rate_limit" in str(e).lower() else 1)
                time.sleep(min(wait_time, 30))
        
        return None

    def translate_long_text(self, text, prompt, speaker=None):
        """긴 텍스트 분할 번역 처리"""
        # 간단한 분할 전략: 문장 단위로 분할
        sentences = text.split('. ')
        if len(sentences) == 1:
            # 단일 문장이 너무 긴 경우는 그대로 시도
            return self.call_gpt4o_mini_api(text[:2000], prompt, 2)  # 길이 제한
        
        # 여러 문장을 적절히 그룹화하여 번역
        grouped_results = []
        current_group = []
        current_length = 0
        
        for sentence in sentences:
            if current_length + len(sentence) > 1500:  # 그룹 크기 제한
                if current_group:
                    group_text = '. '.join(current_group) + '.'
                    result = self.call_gpt4o_mini_api(group_text, prompt, 2)
                    if result:
                        grouped_results.append(result)
                    current_group = [sentence]
                    current_length = len(sentence)
                else:
                    # 단일 문장이 너무 긴 경우
                    result = self.call_gpt4o_mini_api(sentence[:1500], prompt, 2)
                    if result:
                        grouped_results.append(result)
            else:
                current_group.append(sentence)
                current_length += len(sentence)
        
        # 마지막 그룹 처리
        if current_group:
            group_text = '. '.join(current_group) + '.'
            result = self.call_gpt4o_mini_api(group_text, prompt, 2)
            if result:
                grouped_results.append(result)
        
        # 결과 합치기
        return ' '.join(grouped_results) if grouped_results else None

    def batch_translate_similar_items(self, items_group, prompt):
        """유사한 항목들의 배치 번역 (토큰 효율성 향상)"""
        if len(items_group) < 2:
            return {}
        
        # 배치 프롬프트 생성
        batch_prompt = f"{prompt}\n\nTranslate the following Korean texts to English. Return only the translations in the same order, separated by '|||':\n\n"
        
        for i, item in enumerate(items_group, 1):
            batch_prompt += f"{i}. {item['KR']}\n"
        
        batch_prompt += "\nTranslations:"
        
        # 배치 번역 실행
        try:
            batch_result = self.call_gpt4o_mini_api("", batch_prompt, 2)
            if not batch_result:
                return {}
            
            # 결과 파싱
            translations = [t.strip() for t in batch_result.split('|||')]
            
            # 결과 매핑
            results = {}
            for i, item in enumerate(items_group):
                if i < len(translations) and translations[i]:
                    results[item['STRING_ID']] = translations[i]
            
            return results
            
        except Exception as e:
            print(f"배치 번역 오류: {e}")
            return {}

    def get_optimization_stats(self):
        """최적화 통계 정보 반환"""
        total_tokens = self.token_usage_tracker['total_input_tokens'] + self.token_usage_tracker['total_output_tokens']
        
        # GPT-4o-mini 가격 기준 (2024년 기준)
        input_cost = self.token_usage_tracker['total_input_tokens'] * 0.00015 / 1000  # $0.15 per 1K tokens
        output_cost = self.token_usage_tracker['total_output_tokens'] * 0.0006 / 1000  # $0.60 per 1K tokens
        total_cost = input_cost + output_cost
        
        return {
            'total_requests': self.token_usage_tracker['total_requests'],
            'total_tokens': total_tokens,
            'input_tokens': self.token_usage_tracker['total_input_tokens'],
            'output_tokens': self.token_usage_tracker['total_output_tokens'],
            'cache_hits': self.token_usage_tracker['cache_hits'],
            'cache_hit_rate': self.token_usage_tracker['cache_hits'] / max(1, self.token_usage_tracker['total_requests']),
            'estimated_cost': total_cost,
            'avg_tokens_per_request': total_tokens / max(1, self.token_usage_tracker['total_requests'])
        }

    def show_optimization_stats(self):
        """최적화 통계를 사용자에게 표시"""
        if hasattr(self, 'get_optimization_stats'):        
            stats = self.get_optimization_stats()
        
        stats_message = f"""🔧 GPT-4o-mini 최적화 통계

    📊 사용량:
    • 총 요청 수: {stats['total_requests']:,}
    • 총 토큰 수: {stats['total_tokens']:,}
    • 입력 토큰: {stats['input_tokens']:,}
    • 출력 토큰: {stats['output_tokens']:,}

    ⚡ 효율성:
    • 캐시 적중률: {stats['cache_hit_rate']:.1%}
    • 캐시 절약: {stats['cache_hits']}회
    • 평균 토큰/요청: {stats['avg_tokens_per_request']:.1f}

    💰 예상 비용:
    • 총 비용: ${stats['estimated_cost']:.4f}
    • 요청당 평균: ${stats['estimated_cost']/max(1,stats['total_requests']):.4f}"""

        messagebox.showinfo("최적화 통계", stats_message)

    
    #파일 세팅
    def check_and_create_config_files(self):
        """설정 파일들이 존재하는지 확인하고, 없으면 생성 안내"""
        missing_files = []
        
        # .env 파일 확인
        if not os.path.exists('.env'):
            missing_files.append('.env (API 키 설정 파일)')
        
        # credentials.json 파일 확인
        if not os.path.exists('credentials.json'):
            missing_files.append('credentials.json (구글 시트 인증 파일)')
        
        if missing_files:
            return self.show_config_setup_dialog(missing_files)
        
        return True


    def select_reference_file(self):
        """레퍼런스 엑셀 파일 선택 (간소화)"""
        file_path = filedialog.askopenfilename(
            title="레퍼런스 데이터 파일 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.ref_file_var.set(file_path)
            self.update_status(f"📁 레퍼런스 파일: {os.path.basename(file_path)}")

            
    def set_prompt_template(self, template_type):
        """프롬프트 템플릿 설정 (GPT-4o-mini 최적화 및 자연스러움 강화)"""
        templates = {
            "game": """You are a professional game localizer. Translate Korean to natural English for English-speaking gamers aged 10-20.

    RULES:
    • Natural flow > literal accuracy
    • Use standard gaming terminology
    • Keep it concise and impactful
    • Preserve special tags: {}, [#color#], etc.
    • Make it feel native, not translated

    Examples:
    KR: "적을 물리쳤습니다!" → EN: "Enemy defeated!"
    KR: "아이템을 획득했습니다." → EN: "Item acquired!"

    Text to translate:""",
            
            "natural": """Translate Korean to natural, fluent English that sounds like it was originally written in English.

    FOCUS ON:
    • Natural word choice and flow
    • Meaning and nuance over literal translation
    • What native speakers would actually say
    • Appropriate formality level

    Think: "How would an English speaker express this idea?"

    Text:""",
            
            "technical": """Translate Korean technical content to precise, professional English.

    REQUIREMENTS:
    • Use standard technical terminology
    • Maintain accuracy and clarity
    • Professional tone
    • Preserve technical concepts exactly

    Text:""",
            
            "casual": """Translate Korean to casual, conversational English.

    STYLE:
    • Relaxed, friendly tone
    • Natural contractions (it's, don't, etc.)
    • Everyday vocabulary
    • How friends would talk

    Text:""",
            
            "story": """Translate Korean narrative text to engaging English.

    GOALS:
    • Maintain storytelling flow
    • Preserve emotional impact
    • Natural dialogue and descriptions
    • Immersive reading experience

    Text:""",
            
            "default": """Translate Korean to natural English for young English speakers.

    KEY POINTS:
    • Sound natural, not translated
    • Age-appropriate language (teens/young adults)
    • Preserve special formatting
    • Clear and engaging

    Text:"""
        }
        
        prompt = templates.get(template_type, templates["default"])
        
        # 텍스트 위젯 내용 교체
        self.llm_prompt_entry.delete("1.0", "end")
        self.llm_prompt_entry.insert("1.0", prompt)
        
        self.update_status(f"🎯 {template_type.title()} 프롬프트 템플릿 적용됨 (GPT-4o-mini 최적화)")

    def set_scenario_mode_prompt(self):
        """시나리오 모드용 프롬프트 설정"""
        scenario_prompt = """🎭 시나리오 번역 모드 활성화됨

    이 프롬프트는 기본 템플릿입니다.
    실제 번역 시에는 각 화자별로 다음과 같이 최적화된 프롬프트가 자동 생성됩니다:

    1️⃣ 화자 프로필 정보 (성별, 말투, 성격)
    2️⃣ 기존 번역 예시 (일관성 유지)
    3️⃣ 문맥 기반 스타일 지침
    4️⃣ 자연스러운 번역 규칙

    💡 화자가 등록되지 않은 경우에만 이 기본 프롬프트를 사용합니다.

    === 기본 시나리오 번역 프롬프트 ===

    You are localizing game dialogue. Translate Korean to natural English maintaining character consistency.

    NATURAL TRANSLATION RULES:
    • Preserve Korean specificity: "that kid" not "the kid" when 그 아이
    • Time expressions: "a little while ago" for 조금 전에  
    • Emotional nuance: keep hesitation, uncertainty, excitement
    • Character voice: maintain personality in speech patterns

    PRESERVE:
    • Special tags: {}, [#color#], etc.
    • Character personality and speaking style
    • Emotional undertones and atmosphere

    Text to translate:"""
        
        # 프롬프트 창 업데이트
        self.llm_prompt_entry.delete("1.0", "end")
        self.llm_prompt_entry.insert("1.0", scenario_prompt)


    def show_current_translation_info(self, speaker_name=None, kr_text=None):
        """현재 번역에 사용될 프롬프트 정보 표시 (디버깅/확인용)"""
        if not self.scenario_translation_var.get():
            messagebox.showinfo("번역 정보", "일반 번역 모드\n프롬프트 창의 내용을 그대로 사용합니다.")
            return
        
        if speaker_name and self.scenario_manager:
            if speaker_name in self.scenario_manager.speakers:
                speaker = self.scenario_manager.speakers[speaker_name]
                info = f"""🎭 시나리오 번역 정보

    👤 화자: {speaker_name}
    • 성별: {speaker.gender}
    • 말투: {speaker.tone}  
    • 특징: {speaker.style}
    • 레퍼런스: {speaker.reference_count}개 문장

    📝 사용될 프롬프트:
    화자 맞춤형 프롬프트가 자동 생성됩니다.
    - 화자 특성 반영
    - 기존 번역 예시 포함
    - 일관성 검증 기능

    💡 프롬프트 창의 내용은 기본 템플릿이며,
    실제로는 훨씬 더 상세한 화자별 프롬프트를 사용합니다."""
            else:
                info = f"""🎭 시나리오 번역 정보

    ⚠️ 화자 '{speaker_name}'가 등록되지 않았습니다.

    📝 사용될 프롬프트:
    프롬프트 창의 기본 시나리오 프롬프트를 사용합니다.

    💡 더 나은 번역을 위해:
    '🎭 시나리오 번역' 탭에서 화자를 등록하거나
    레퍼런스 데이터를 분석하여 자동 생성하세요."""
        else:
            info = """🎭 시나리오 번역 모드

    📝 프롬프트 생성 방식:
    • 화자 등록됨 → 맞춤형 프롬프트 자동 생성
    • 화자 미등록 → 기본 시나리오 프롬프트 사용

    🔍 현재 상태 확인:
    번역을 실행하면 각 문장마다 적절한 프롬프트가 선택됩니다."""
        
        messagebox.showinfo("번역 프롬프트 정보", info)
        

    def get_llm_prompt(self):
        """현재 설정된 LLM 프롬프트 가져오기 (개선된 버전)"""
        prompt = self.llm_prompt_entry.get("1.0", "end-1c").strip()
        
        # 빈 프롬프트인 경우 기본값 사용
        if not prompt:
            self.set_prompt_template("default")
            prompt = self.llm_prompt_entry.get("1.0", "end-1c").strip()
        
        return prompt

    def translate_with_llm(self, text, prompt, max_retries=3):
        """LLM으로 번역 (개선된 버전 - GPT-4o-mini 최적화)"""
        # 용어집 기능이 활성화되어 있으면 프롬프트 향상
        enhanced_prompt = prompt
        glossary_applied = False
        
        if self.glossary_matcher:
            try:
                enhanced_prompt = self.glossary_matcher.create_enhanced_prompt(text, prompt)
                glossary_applied = len(enhanced_prompt) > len(prompt) + len(text) + 10  # 용어집이 실제로 추가되었는지 확인
                
                if glossary_applied:
                    # 관련 용어 정보 로깅
                    relevant_terms = self.glossary_matcher.find_relevant_terms(text, max_terms=5)
                    print(f"🔍 용어집 적용: {len(relevant_terms)}개 용어 참조")
                
            except Exception as e:
                print(f"용어집 적용 실패, 기본 프롬프트 사용: {e}")
                enhanced_prompt = prompt

        for attempt in range(max_retries):
            try:
                import openai
                
                client = openai.OpenAI(api_key=OPENAI_API_KEY)
                
                response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {
                            "role": "system", 
                            "content": "You are a professional translator. Return only the translation, no explanations."
                        },
                        {
                            "role": "user", 
                            "content": enhanced_prompt  # 향상된 프롬프트 사용
                        }
                    ],
                    max_tokens=min(800, len(text) * 3),
                    temperature=0.2,
                    top_p=0.9,
                    frequency_penalty=0.1,
                    presence_penalty=0.1,
                    timeout=25
                )
                
                if response.choices and len(response.choices) > 0:
                    result = response.choices[0].message.content.strip()
                    result = self.post_process_llm_result(result, text)
                    
                    # 용어집이 적용되었으면 성공 로그
                    if glossary_applied:
                        print(f"✅ 용어집 번역 성공: {text[:30]}...")
                    
                    return result
                        
            except Exception as e:
                print(f"LLM 번역 시도 {attempt + 1} 실패: {e}")
                if attempt == max_retries - 1:
                    return None
                    
                import time
                wait_time = (2 ** attempt) * (2 if "rate_limit" in str(e).lower() else 1)
                time.sleep(min(wait_time, 30))
        
        return None

    def post_process_llm_result(self, result, original_text):
        """LLM 결과 후처리 (개선된 버전)"""
        if not result:
            return result
        
        # 1. 불필요한 설명문 제거
        unwanted_phrases = [
            "Here's the translation:",
            "Translation:",
            "영어 번역:",
            "번역:",
            "The translation is:",
            "Translated text:"
        ]
        
        for phrase in unwanted_phrases:
            if result.lower().startswith(phrase.lower()):
                result = result[len(phrase):].strip()
                break
        
        # 2. 따옴표 처리
        if ((result.startswith('"') and result.endswith('"')) or 
            (result.startswith("'") and result.endswith("'"))):
            result = result[1:-1].strip()
        
        # 3. 특수 태그 검증 및 복구
        original_tags = self.extract_special_tags(original_text)
        result_tags = self.extract_special_tags(result)
        
        # 태그가 누락되었으면 경고 (하지만 결과는 그대로 반환)
        if len(original_tags) != len(result_tags):
            print(f"⚠️ 특수 태그 불일치: 원본 {original_tags} → 번역 {result_tags}")
        
        # 4. 앞뒤 공백 정리
        result = result.strip()
        
        return result


# 선택사항: 용어집 성능 통계를 보여주는 메서드 추가
    def show_glossary_stats(self):
        """용어집 매칭 시스템 성능 통계 표시"""
        if not self.glossary_matcher:
            messagebox.showinfo("용어집 통계", "용어집 매칭 시스템이 활성화되지 않았습니다.")
            return
        
        try:
            stats = self.glossary_matcher.get_performance_stats()
            
            stats_message = f"""📊 용어집 매칭 시스템 통계

📚 용어집 정보:
• 로드된 용어 수: {stats['glossary_terms']:,}개
• Trie 노드 수: {stats['trie_nodes']:,}개
• 메모리 효율성: {stats['memory_efficiency']:.2f}

⚡ 캐시 성능:
• 캐시 적중률: {stats['cache_hit_rate']:.1%}
• 캐시 적중: {stats['cache_hits']}회
• 캐시 미스: {stats['cache_misses']}회
• 캐시 크기: {stats['cache_size']}개

💡 효과:
용어집 적중률이 높을수록 번역 일관성이 향상됩니다."""

            messagebox.showinfo("용어집 시스템 통계", stats_message)
            
        except Exception as e:
            messagebox.showerror("통계 오류", f"통계를 가져오는 중 오류 발생: {e}")

    # 선택사항: 디버깅용 메서드 추가
    def debug_glossary_matching(self, text: str):
        """특정 텍스트의 용어집 매칭 디버깅"""
        if not self.glossary_matcher:
            print("용어집 매칭 시스템이 비활성화되어 있습니다.")
            return
        
        try:
            debug_info = self.glossary_matcher.get_debug_info(text)
            
            print(f"\n=== 용어집 매칭 디버깅 ===")
            print(f"원본 텍스트: {debug_info['original_text']}")
            print(f"검색 대상: '{debug_info.get('searchable', 'N/A')}'")
            print(f"제외된 태그: {debug_info.get('excluded_tags', [])}")
            print(f"포함된 태그: {debug_info.get('included_tags', [])}")
            print(f"매칭된 용어: {debug_info.get('final_terms', [])}")
            print(f"매칭 개수: {debug_info.get('final_count', 0)}개")
            
        except Exception as e:
            print(f"디버깅 오류: {e}")


    def extract_special_tags(self, text):
        """특수 태그 추출 (디버깅 및 검증용)"""
        import re
        tags = []
        
        # {} 형태 태그
        tags.extend(re.findall(r'\{[^}]*\}', text))
        
        # [#색상#] 형태 태그  
        tags.extend(re.findall(r'\[#[^#]*#\]', text))
        
        # [@변수] 형태 태그
        tags.extend(re.findall(r'\[@[^\]]*\]', text))
        
        return tags

    def show_config_setup_dialog(self, missing_files):
        """설정 파일 생성 안내 다이얼로그"""
        missing_text = '\n• '.join([''] + missing_files)
        
        message = f"""번역 프로그램 실행에 필요한 설정 파일이 없습니다:{missing_text}

    설정 파일을 자동으로 생성하시겠습니까?

    - 예: 템플릿 파일을 생성하고 설정 방법을 안내합니다
    - 아니오: 프로그램을 종료합니다"""

        result = messagebox.askyesno(
            "설정 파일 없음", 
            message,
            icon="question"
        )
        
        if result:
            self.create_config_templates()
            self.show_setup_instructions()
            return True
        else:
            self.root.quit()
            return False

    def create_config_templates(self):
        """설정 파일 템플릿 생성"""
        created_files = []
        
        # .env 템플릿 생성
        if not os.path.exists('.env'):
            env_template = """# API 키 설정 파일
    # 각 서비스에서 발급받은 API 키를 입력하세요

    # DeepL API 키 (필수)
    DEEPL_API_KEY=여기에_DeepL_API_키를_입력하세요

    # Azure Translator API 키 (선택)
    AZURE_API_KEY=여기에_Azure_API_키를_입력하세요
    AZURE_REGION=koreacentral

    # OpenAI API 키 (AI 후편집용, 선택)
    OPENAI_API_KEY=여기에_OpenAI_API_키를_입력하세요
    """
            with open('.env', 'w', encoding='utf-8') as f:
                f.write(env_template)
            created_files.append('.env')
        
        # credentials.json 템플릿 생성
        if not os.path.exists('credentials.json'):
            cred_template = {
                "type": "service_account",
                "project_id": "여기에_프로젝트_ID_입력",
                "private_key_id": "여기에_private_key_id_입력",
                "private_key": "-----BEGIN PRIVATE KEY-----\n여기에_private_key_입력\n-----END PRIVATE KEY-----\n",
                "client_email": "여기에_client_email_입력",
                "client_id": "여기에_client_id_입력",
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
                "client_x509_cert_url": "여기에_client_x509_cert_url_입력"
            }
            with open('credentials.json', 'w', encoding='utf-8') as f:
                json.dump(cred_template, f, indent=2, ensure_ascii=False)
            created_files.append('credentials.json')
        
        if created_files:
            files_text = ', '.join(created_files)
            messagebox.showinfo("파일 생성 완료", f"템플릿 파일이 생성되었습니다: {files_text}")

    def show_setup_instructions(self):
        """설정 방법 안내 창"""
        instructions = """📋 설정 파일 작성 방법

    🔑 .env 파일 설정:
    1. DeepL 계정에서 API 키 발급 (https://www.deepl.com/pro-api)
    2. .env 파일을 메모장으로 열기
    3. 'DEEPL_API_KEY=' 뒤에 실제 API 키 입력
    4. Azure Translate도 동일하게 API 키 발급
    5. GPT-API 키도 발급 (하단에 상세)
    4. 파일 저장

    ------------------------------------------------------------------
    API 키 발급 페이지(로그인 필요): platform.openai.com/account/api-keys
    위 링크로 이동해 OpenAI 계정에 로그인하세요.
    “Create new secret key” 버튼을 눌러 새 키를 생성합니다.
    생성된 키는 한 번만 표시되므로, 안전한 곳에 복사해 보관하세요. 
    ------------------------------------------------------------------

    📊 credentials.json 파일 설정:
    1. Google Cloud Console에서 서비스 계정 생성
    2. JSON 키 파일 다운로드
    3. 다운로드한 파일 내용을 credentials.json에 복사
    4. 파일 저장

    ⚠️ 주의사항:
    - API 키는 절대 다른 사람과 공유하지 마세요
    - 설정 완료 후 프로그램을 다시 시작하세요

    설정이 완료되면 프로그램을 다시 실행해주세요."""

        # 별도 창으로 안내 표시
        instruction_window = tk.Toplevel(self.root)
        instruction_window.title("설정 방법 안내")
        instruction_window.geometry("600x500")
        instruction_window.transient(self.root)
        instruction_window.grab_set()
        
        main_frame = ttk.Frame(instruction_window, padding="20")
        main_frame.pack(fill="both", expand=True)
        
        text_widget = tk.Text(main_frame, wrap="word", font=("맑은 고딕", 10))
        text_widget.insert("1.0", instructions)
        text_widget.config(state="disabled")
        
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        button_frame = ttk.Frame(instruction_window)
        button_frame.pack(fill="x", pady=10)
        
        ttk.Button(button_frame, text="프로그램 종료", 
                command=lambda: [instruction_window.destroy(), self.root.quit()]).pack(side="right", padx=5)
        ttk.Button(button_frame, text="설정 폴더 열기", 
                command=lambda: os.startfile(os.getcwd())).pack(side="right", padx=5)

    def handle_drop(self, event):
        """파일을 드래그 앤 드랍했을 때 실행되는 이벤트 핸들러"""
        try:
            # event.data에는 파일 경로가 문자열로 들어옵니다.
            # 여러 파일을 드랍하면 공백으로 구분되고, 경로에 공백이 있으면 { }로 묶입니다.
            # 여기서는 첫 번째 파일만 처리하도록 간단하게 구현합니다.
            file_path = event.data.strip()
            if file_path.startswith('{') and file_path.endswith('}'):
                file_path = file_path[1:-1]

            # 엑셀 파일인지 확인
            if file_path.lower().endswith((".xlsx", ".xls")):
                self.update_status(f"파일 드랍: {os.path.basename(file_path)}")
                
                # 1. 파일 경로 변수를 업데이트합니다.
                self.file_path_var.set(file_path)
                
                # 2. 즉시 데이터 로드를 실행합니다.
                self.load_data()
            else:
                self.update_status("엑셀 파일(.xlsx, .xls)만 드랍할 수 있습니다.")
                messagebox.showwarning("파일 형식 오류", "엑셀 파일만 드랍할 수 있습니다.")
                
        except Exception as e:
            self.update_status(f"파일 드랍 처리 중 오류 발생: {e}")
            messagebox.showerror("오류", f"파일을 처리하는 중 오류가 발생했습니다:\n{e}")


    def toggle_all_selections(self):
        """'전체 선택/해제' 체크박스의 상태에 따라 모든 항목의 체크 상태를 변경합니다."""
        is_checked = self.select_all_var.get()
        new_symbol = "☑" if is_checked else "☐"

        for item_id in self.translation_tree.get_children():
            self.check_states[item_id] = is_checked
            
            current_values = list(self.translation_tree.item(item_id, "values"))
            if len(current_values) > 0:
                current_values[0] = new_symbol
                self.translation_tree.item(item_id, values=current_values)
        
        action = "선택" if is_checked else "해제"
        total_count = len(self.translation_tree.get_children())
        self.update_status(f"전체 {action}: {total_count}개 항목")


    def toggle_selected_checkboxes(self, event):
        """선택된 항목들의 체크박스 토글 (Spacebar/Enter 이벤트)"""
        selected_items = self.translation_tree.selection()
        
        if not selected_items:
            # 아무것도 선택되지 않았으면 현재 포커스된 항목을 토글
            focused_item = self.translation_tree.focus()
            if focused_item:
                selected_items = [focused_item]
            else:
                return "break"  # 이벤트 전파 중단
        
        # 선택된 항목들의 현재 체크 상태 확인
        current_states = [self.check_states.get(item, True) for item in selected_items]
        
        # 토글 방식 결정: 하나라도 체크되어 있으면 모두 해제, 모두 해제되어 있으면 모두 체크
        if any(current_states):
            new_state = False  # 하나라도 체크되어 있으면 모두 해제
            symbol = "☐"
        else:
            new_state = True   # 모두 해제되어 있으면 모두 체크
            symbol = "☑"
        
        # 선택된 모든 항목의 체크 상태 변경
        for item in selected_items:
            self.check_states[item] = new_state
            
            # UI에서 체크박스 심볼 업데이트
            current_values = list(self.translation_tree.item(item, "values"))
            if len(current_values) > 0:
                current_values[0] = symbol
                self.translation_tree.item(item, values=current_values)
        
        # 상태 메시지 업데이트
        action = "선택" if new_state else "해제"
        self.update_status(f"{len(selected_items)}개 항목 {action}됨")
        
        # 전체 선택/해제 체크박스 상태도 업데이트
        self.update_select_all_checkbox()
        
        return "break"  # 기본 이벤트 처리 중단

    def select_all_items(self, event):
        """Ctrl+A: 모든 항목 선택"""
        all_items = self.translation_tree.get_children()
        
        # 모든 항목 체크
        for item in all_items:
            self.check_states[item] = True
            current_values = list(self.translation_tree.item(item, "values"))
            if len(current_values) > 0:
                current_values[0] = "☑"
                self.translation_tree.item(item, values=current_values)
        
        # 전체 선택/해제 체크박스도 업데이트
        self.select_all_var.set(True)
        
        self.update_status(f"모든 항목 선택됨 ({len(all_items)}개)")
        return "break"

    def deselect_all_items(self, event):
        """Ctrl+D: 모든 항목 선택 해제"""
        all_items = self.translation_tree.get_children()
        
        # 모든 항목 해제
        for item in all_items:
            self.check_states[item] = False
            current_values = list(self.translation_tree.item(item, "values"))
            if len(current_values) > 0:
                current_values[0] = "☐"
                self.translation_tree.item(item, values=current_values)
        
        # 전체 선택/해제 체크박스도 업데이트
        self.select_all_var.set(False)
        
        self.update_status(f"모든 항목 선택 해제됨 ({len(all_items)}개)")
        return "break"

    def update_select_all_checkbox(self):
        """전체 선택/해제 체크박스 상태를 현재 선택 상태에 맞게 업데이트"""
        all_items = self.translation_tree.get_children()
        if not all_items:
            return
        
        # 모든 항목의 체크 상태 확인
        checked_count = sum(1 for item in all_items if self.check_states.get(item, True))
        
        if checked_count == len(all_items):
            self.select_all_var.set(True)   # 모두 체크됨
        elif checked_count == 0:
            self.select_all_var.set(False)  # 모두 해제됨
        else:
            # 일부만 체크된 경우는 현재 상태 유지 (또는 특별한 상태 표시 가능)
            pass


    def invert_selection(self, event):
        """Ctrl+I: 선택 상태 반전"""
        all_items = self.translation_tree.get_children()
        
        for item in all_items:
            current_state = self.check_states.get(item, True)
            new_state = not current_state
            self.check_states[item] = new_state
            
            symbol = "☑" if new_state else "☐"
            current_values = list(self.translation_tree.item(item, "values"))
            if len(current_values) > 0:
                current_values[0] = symbol
                self.translation_tree.item(item, values=current_values)
        
        self.update_select_all_checkbox()
        self.update_status("선택 상태가 반전되었습니다.")
        return "break"

    def clear_selected_translations(self, event):
        """Delete: 선택된 항목들의 번역 내용 삭제"""
        selected_items = self.translation_tree.selection()
        if not selected_items:
            return "break"
        
        if not messagebox.askyesno("번역 삭제 확인", 
                                f"선택된 {len(selected_items)}개 항목의 번역 내용을 삭제하시겠습니까?"):
            return "break"
        
        # 선택된 항목들의 번역 내용 삭제
        for item in selected_items:
            values = self.translation_tree.item(item, "values")
            string_id = values[1]
            
            # pending_translations에서 해당 항목 찾아서 번역 내용 삭제
            for trans in self.pending_translations:
                if trans["STRING_ID"] == string_id:
                    trans["translations"].clear()
                    trans["method"] = ""
                    trans["status"] = "[신규]"
                    break
        
        self.update_translation_table()
        self.update_status(f"{len(selected_items)}개 항목의 번역 내용이 삭제되었습니다.")
        return "break"

    def edit_selected_item(self, event):
        """F2: 선택된 항목 편집"""
        selected_items = self.translation_tree.selection()
        if len(selected_items) == 1:
            self.edit_translation_inline()
        return "break"

    def setup_glossary_tab(self):
        """용어집 관리 탭 UI 구성 (STRING_ID 제거)"""
        main_frame = ttk.Frame(self.glossary_tab, padding="10")
        main_frame.pack(fill="both", expand=True)

        # 상단 컨트롤 프레임
        control_frame = ttk.LabelFrame(main_frame, text="용어집 관리 도구")
        control_frame.pack(fill="x", pady=5)
        
        # 좌측: 핵심 기능
        left_control = ttk.Frame(control_frame)
        left_control.pack(side="left", padx=5, pady=5)
        
        ttk.Button(left_control, text="🔄 구글 시트와 동기화", command=self.sync_glossary_from_gsheet, style="Accent.TButton").pack(side="left")
        
        # 우측: 검색 기능
        search_frame = ttk.Frame(control_frame)
        search_frame.pack(side="right", padx=10, pady=5)
        
        ttk.Label(search_frame, text="KR 검색:").pack(side="left")
        self.glossary_search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.glossary_search_var, width=30)
        search_entry.pack(side="left", padx=5)
        ttk.Button(search_frame, text="🔍 필터", command=self.filter_glossary).pack(side="left", padx=2)
        ttk.Button(search_frame, text="🔄 전체", command=self.clear_glossary_filter).pack(side="left", padx=2)
        
        # 사용자 안내
        info_label = ttk.Label(control_frame, text="ℹ️ 용어 추가/수정/삭제는 마스터 구글 시트에서 진행 후, 동기화 버튼을 눌러주세요.", foreground="blue")
        info_label.pack(side="bottom", padx=10, pady=2)

        # 용어집 목록 Treeview (STRING_ID 제거)
        list_frame = ttk.LabelFrame(main_frame, text="용어집 목록 (마스터 구글 시트의 로컬 사본)")
        list_frame.pack(fill="both", expand=True, pady=10)

        # 🆕 STRING_ID 제거된 컬럼 목록
        self.glossary_cols = ["kr", "en", "cn", "tw", "contributor", "verified"]
        self.glossary_tree = ttk.Treeview(list_frame, columns=self.glossary_cols, show="headings")
        
        # 컬럼 너비 조정
        col_widths = {"kr": 200, "en": 200, "cn": 120, "tw": 120, "contributor": 100, "verified": 70}
        for col in self.glossary_cols:
            self.glossary_tree.heading(col, text=col.upper())
            self.glossary_tree.column(col, width=col_widths.get(col, 100), anchor="w")

        # 스크롤바
        v_scroll = ttk.Scrollbar(list_frame, orient="vertical", command=self.glossary_tree.yview)
        h_scroll = ttk.Scrollbar(list_frame, orient="horizontal", command=self.glossary_tree.xview)
        self.glossary_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        
        self.glossary_tree.pack(side="left", fill="both", expand=True)
        v_scroll.pack(side="right", fill="y")
        h_scroll.pack(side="bottom", fill="x")

    def load_glossary(self, kr_filter=None):
        """용어집 로드 (STRING_ID 제거 버전, 필터링 지원)"""
        self.glossary_tree.delete(*self.glossary_tree.get_children())
        
        conn = sqlite3.connect(self.translation_db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # 필터링 쿼리 (STRING_ID 제거)
        if kr_filter:
            cursor.execute("SELECT * FROM glossary WHERE kr LIKE ?", (f"%{kr_filter}%",))
        else:
            cursor.execute("SELECT * FROM glossary")
        
        # self.glossary와 self.exact_matches에 전체 데이터 로드
        self.glossary = {}
        self.exact_matches = {}  # 🆕 exact_matches도 함께 로드
        
        for row in cursor.fetchall():
            row_dict = dict(row)
            kr_text = row_dict.get("kr")
            
            if kr_text:
                self.glossary[kr_text] = row_dict
                
                # exact_matches에도 저장 (빠른 매칭용)
                self.exact_matches[kr_text] = {
                    'korean': kr_text,
                    'english': row_dict.get("en", ""),
                    'category': 'general',  # 기본 카테고리
                    'string_id': None
                }

        # Treeview에 표시 (STRING_ID 제외)
        for record in self.glossary.values():
            values = []
            for col in self.glossary_cols:
                val = record.get(col)
                if col == 'verified':
                    val = "Y" if val == 1 else "N"
                values.append(val or "")
            self.glossary_tree.insert("", "end", values=values)
                
        conn.close()
        
        # 상태 표시
        if kr_filter:
            self.update_status(f"용어집 필터 적용: '{kr_filter}' 포함 항목")
        else:
            self.update_status("용어집 전체 로드 완료")

    def filter_glossary(self):
        """용어집 KR 필터링"""
        search_term = self.glossary_search_var.get().strip()
        self.load_glossary(kr_filter=search_term)

    def clear_glossary_filter(self):
        """용어집 필터 초기화"""
        self.glossary_search_var.set("")
        self.load_glossary()

    def load_glossary(self, kr_filter=None):
        """새로운 스키마의 용어집을 DB에서 로드하여 Treeview에 표시 (KR 필터 추가)"""
        self.glossary_tree.delete(*self.glossary_tree.get_children())
        
        conn = sqlite3.connect(self.translation_db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # 필터링 쿼리
        if kr_filter:
            cursor.execute("SELECT * FROM glossary WHERE kr LIKE ?", (f"%{kr_filter}%",))
        else:
            cursor.execute("SELECT * FROM glossary")
        
        # self.glossary에 전체 데이터 로드 (용어집 우선 적용 기능에 사용)
        self.glossary = {}
        for row in cursor.fetchall():
            self.glossary[row["kr"]] = dict(row)

        # Treeview에는 지정된 컬럼만 표시
        for record in self.glossary.values():
            values = []
            for col in self.glossary_cols:
                val = record.get(col)
                if col == 'verified':
                    val = "Y" if val == 1 else "N"
                values.append(val or "")
            self.glossary_tree.insert("", "end", values=values)
                
        conn.close()
        
        # 필터 상태 표시
        if kr_filter:
            self.update_status(f"용어집 필터 적용: '{kr_filter}' 포함 항목")
        else:
            self.update_status("용어집 전체 로드 완료")

            
    def setup_history_tab(self):
        """번역 이력 탭 설정"""
        # 검색
        search_frame = ttk.Frame(self.history_tab)
        search_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(search_frame, text="검색:").pack(side="left")
        self.history_search_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=self.history_search_var, width=40).pack(side="left", padx=5)
        ttk.Button(search_frame, text="검색", command=self.search_history).pack(side="left")
        
        # 이력 테이블
        columns = ["시간", "STRING_ID", "KR", "번역방법", "상태"]
        self.history_tree = ttk.Treeview(self.history_tab, columns=columns, show="headings")
        
        for col in columns:
            self.history_tree.heading(col, text=col)
        
        self.history_tree.column("시간", width=150)
        self.history_tree.column("STRING_ID", width=150)
        self.history_tree.column("KR", width=300)
        self.history_tree.column("번역방법", width=100)
        self.history_tree.column("상태", width=100)
        
        history_scroll = ttk.Scrollbar(self.history_tab, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=history_scroll.set)
        
        self.history_tree.pack(side="left", fill="both", expand=True, padx=10, pady=5)
        history_scroll.pack(side="right", fill="y")


    def setup_conflict_tab(self):
        """'충돌 해결' 탭의 UI를 구성합니다."""
        main_frame = ttk.Frame(self.conflict_tab, padding="10")
        main_frame.pack(fill="both", expand=True)

        # 1. 상단 액션 프레임
        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill="x", pady=5)
        ttk.Button(action_frame, text="🚀 가장 많이 쓰인 번역으로 전체 자동 해결", command=self.auto_resolve_all_conflicts).pack(side="left")
        ttk.Button(action_frame, text="🔄 목록 새로고침", command=self.load_conflicts_to_view).pack(side="left", padx=10)

        # 2. 콘텐츠 프레임 (충돌 목록 + 해결 패널)
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill="both", expand=True, pady=5)

        # 2-1. 충돌 목록 Treeview
        list_frame = ttk.LabelFrame(content_frame, text="충돌 항목 목록")
        list_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))

        columns = ["KR"] + self.VISIBLE_LANGS
        self.conflict_tree = ttk.Treeview(list_frame, columns=columns, show="headings")
        self.conflict_tree.heading("KR", text="KR (충돌 항목)")
        self.conflict_tree.column("KR", width=200)
        for lang in self.VISIBLE_LANGS:
            self.conflict_tree.heading(lang, text=lang)
            self.conflict_tree.column(lang, width=120)
        
        self.conflict_tree.tag_configure('conflict', foreground='red')
        self.conflict_tree.bind("<<TreeviewSelect>>", self.on_conflict_row_selected)
        self.conflict_tree.pack(fill="both", expand=True)

        # 2-2. 충돌 해결 패널
        resolve_panel = ttk.LabelFrame(content_frame, text="충돌 해결")
        resolve_panel.pack(side="right", fill="y")

        self.conflict_kr_var = tk.StringVar()
        ttk.Label(resolve_panel, text="KR:").pack(anchor="w", padx=5, pady=5)
        ttk.Entry(resolve_panel, textvariable=self.conflict_kr_var, state="readonly", width=40).pack(anchor="w", padx=5)
        
        self.conflict_combos = {}
        for lang in self.VISIBLE_LANGS:
            ttk.Label(resolve_panel, text=f"{lang} 후보:").pack(anchor="w", padx=5, pady=(10, 0))
            combo = ttk.Combobox(resolve_panel, state="disabled", width=38)
            combo.pack(anchor="w", padx=5)
            self.conflict_combos[lang] = combo
        
        ttk.Button(resolve_panel, text="✅ 선택된 값으로 충돌 해결", command=self.resolve_selected_conflict).pack(pady=20, padx=5)


    def load_conflicts_to_view(self):
        """DB에서 충돌 상태인 항목을 불러와 충돌 해결 탭에 표시"""
        self.conflict_tree.delete(*self.conflict_tree.get_children())
        
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT kr_text, translations, conflict_info FROM translation_memory WHERE status = 'conflict'")
        
        conflicts = cursor.fetchall()
        conn.close()

        for kr, trans_json, conflict_json in conflicts:
            translations = json.loads(trans_json)
            conflict_info = json.loads(conflict_json) if conflict_json else {}
            
            display_values = [kr]
            tags = [''] * (len(self.VISIBLE_LANGS) + 1)
            
            for i, lang in enumerate(self.VISIBLE_LANGS):
                display_values.append(translations.get(lang, ""))
                if lang in conflict_info:
                    tags[i+1] = 'conflict' # KR 다음부터 언어 컬럼이므로 i+1
                    
            self.conflict_tree.insert("", "end", values=display_values, tags=tags)


    # 기존 sync_glossary_from_gsheet 함수를 아래 코드로 교체하세요.

    def sync_glossary_from_gsheet(self):
        """'구글 시트와 동기화' 버튼의 동작. (시트 이름 glossary로 고정)"""
        if not messagebox.askyesno("동기화 확인", "마스터 구글 시트의 최신 내용으로 로컬 용어집 DB를 동기화합니다.\n\n로컬 DB의 내용은 구글 시트 기준으로 덮어쓰거나 삭제될 수 있습니다. 계속하시겠습니까?", parent=self.root):
            return

        # 구글 시트 URL이나 ID를 입력받습니다.
        sheet_url_or_id = simpledialog.askstring("구글 시트 정보 입력", "구글 시트의 전체 URL 또는 스프레드시트 ID를 입력하세요:", initialvalue="1Ff7jFAmpgMLDFQ2S0pnT2sNkKfE__xW2jRnA5AQOZqY", parent=self.root)
        if not sheet_url_or_id: return

        # 시트 이름을 "glossary"로 고정
        sheet_name = "glossary"

        # 정규식을 이용해 URL에서 스프레드시트 ID만 추출
        match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', sheet_url_or_id)
        spreadsheet_id = match.group(1) if match else sheet_url_or_id

        threading.Thread(target=self._sync_gsheet_thread, args=(spreadsheet_id, sheet_name), daemon=True).start()
    
    
    def _sync_gsheet_thread(self, spreadsheet_id, sheet_name):
        """(스레드) 구글 시트와 로컬 DB 동기화 (STRING_ID 제거)"""
        try:
            self.update_status("동기화 시작: 구글 시트 데이터 가져오는 중...")
            
            from google.oauth2 import service_account
            from googleapiclient.discovery import build

            creds = service_account.Credentials.from_service_account_file('credentials.json', scopes=['https://www.googleapis.com/auth/spreadsheets.readonly'])
            service = build('sheets', 'v4', credentials=creds)
            range_name = sheet_name
            
            result = service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=range_name).execute()
            master_values = result.get('values', [])
            
            if not master_values or len(master_values) < 2:
                self.update_status("오류: 마스터 구글 시트에서 데이터를 찾을 수 없습니다.")
                return

            master_header = [h.lower() for h in master_values[0]]
            
            # STRING_ID가 있으면 무시하고 kr을 기준으로 처리
            master_data = {}
            for row in master_values[1:]:
                row_dict = dict(zip(master_header, row))
                kr = row_dict.get('kr')
                if kr and kr.strip():  # kr을 키로 사용
                    master_data[kr.strip()] = row_dict

            # 로컬 DB 데이터 로드 (kr 기준)
            self.update_status("로컬 캐시 DB와 데이터 비교 중...")
            conn = sqlite3.connect(self.translation_db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM glossary")
            local_data = {row['kr']: dict(row) for row in cursor.fetchall() if row['kr']}

            # 변경점 계산
            master_keys = set(master_data.keys())
            local_keys = set(local_data.keys())
            
            new_keys = master_keys - local_keys
            deleted_keys = local_keys - master_keys
            common_keys = master_keys.intersection(local_keys)
            
            updated_keys = set()
            for key in common_keys:
                # 간단한 비교
                master_row_str = str(sorted(master_data[key].items()))
                local_row_str = str(sorted(local_data[key].items()))
                if master_row_str != local_row_str:
                    updated_keys.add(key)
            
            if not any([new_keys, deleted_keys, updated_keys]):
                self.update_status("동기화 완료: 변경된 내용이 없습니다.")
                self.root.after(0, lambda: messagebox.showinfo("완료", "용어집이 이미 최신 상태입니다.", parent=self.root))
                conn.close()
                return
                    
            # 로컬 DB에 변경사항 적용
            self.update_status("로컬 캐시 DB에 변경사항 적용 중...")
            
            # 삭제
            if deleted_keys:
                cursor.executemany("DELETE FROM glossary WHERE kr = ?", [(key,) for key in deleted_keys])
            
            # 신규 및 변경 (INSERT OR REPLACE로 처리)
            keys_to_upsert = new_keys.union(updated_keys)
            upsert_data = []
            # STRING_ID 제외한 컬럼들
            db_cols = ["kr", "en", "cn", "tw", "th", "pt", "es", "de", "fr", "jp", "engine", "contributor", "update_at", "verified", "description"]
            
            for key in keys_to_upsert:
                row_data = master_data[key]
                db_values = tuple(row_data.get(col, None) for col in db_cols)
                upsert_data.append(db_values)

            if upsert_data:
                cursor.executemany(f"INSERT OR REPLACE INTO glossary ({', '.join(db_cols)}) VALUES ({','.join(['?']*len(db_cols))})", upsert_data)

            conn.commit()
            conn.close()

            # 완료 및 UI 새로고침
            self.root.after(0, self.load_glossary)
            self.update_status("동기화 완료!")
            summary = f"동기화가 완료되었습니다.\n\n- 신규: {len(new_keys)}개\n- 변경: {len(updated_keys)}개\n- 삭제: {len(deleted_keys)}개"
            self.root.after(0, lambda: messagebox.showinfo("동기화 완료", summary, parent=self.root))
            
        except Exception as e:
            self.update_status(f"동기화 오류: {e}")
            self.root.after(0, lambda: messagebox.showerror("오류", f"동기화 중 오류가 발생했습니다:\n{e}", parent=self.root))
            import traceback
            traceback.print_exc()

    def load_reference_from_gsheet(self):
        """구글 시트에서 시나리오 레퍼런스 데이터 로드"""
        try:
            # 1. 구글 시트 URL/ID 입력받기
            sheet_url_or_id = simpledialog.askstring(
                "구글 시트 정보 입력", 
                "시나리오 레퍼런스 구글 시트의 전체 URL 또는 스프레드시트 ID를 입력하세요:",
                parent=self.root
            )
            if not sheet_url_or_id:
                return

            # URL에서 스프레드시트 ID 추출
            match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', sheet_url_or_id)
            spreadsheet_id = match.group(1) if match else sheet_url_or_id

            # 2. 시트(탭) 이름 입력받기
            sheet_name = simpledialog.askstring(
                "시트 이름 입력", 
                "레퍼런스 데이터가 있는 시트(탭)의 정확한 이름을 입력하세요 (예: Sheet1):",
                parent=self.root
            )
            if not sheet_name:
                return

            # 3. 분석할 언어 미리 선택받기
            available_langs = ['EN', 'CN', 'TW', 'TH', 'PT', 'ES', 'DE', 'FR', 'JP']
            dialog = LanguageSelectionDialog(self.root, available_langs, "분석할 언어 선택")
            self.root.wait_window(dialog.top)
            
            target_language = dialog.selected_lang if hasattr(dialog, 'selected_lang') else None
            if not target_language:
                messagebox.showwarning("선택 취소", "언어를 선택하지 않아 작업을 취소합니다.")
                return

            # 4. 확인 메시지
            if not messagebox.askyesno(
                "구글 시트 로드 확인", 
                f"다음 설정으로 구글 시트에서 레퍼런스 데이터를 가져와 분석합니다:\n\n"
                f"• 스프레드시트 ID: {spreadsheet_id}\n"
                f"• 시트 이름: {sheet_name}\n"
                f"• 분석 언어: {target_language}\n\n"
                f"계속하시겠습니까?"
            ):
                return

            # 5. 별도 스레드에서 구글 시트 로드 및 분석 실행
            threading.Thread(
                target=self._load_reference_gsheet_thread, 
                args=(spreadsheet_id, sheet_name, target_language), 
                daemon=True
            ).start()

        except Exception as e:
            messagebox.showerror("오류", f"구글 시트 로드 중 오류 발생: {e}")
            import traceback
            traceback.print_exc()

    def _load_reference_gsheet_thread(self, spreadsheet_id, sheet_name, target_language):
        """(스레드) 구글 시트에서 시나리오 레퍼런스 데이터 로드 및 분석"""
        try:
            self.update_status("구글 시트에서 레퍼런스 데이터 가져오는 중...")
            
            # 1. 구글 시트 API 인증 및 데이터 가져오기
            from google.oauth2 import service_account
            from googleapiclient.discovery import build

            creds = service_account.Credentials.from_service_account_file(
                'credentials.json', 
                scopes=['https://www.googleapis.com/auth/spreadsheets.readonly']
            )
            service = build('sheets', 'v4', credentials=creds)
            
            # 전체 시트 데이터 가져오기
            range_name = sheet_name
            result = service.spreadsheets().values().get(
                spreadsheetId=spreadsheet_id, 
                range=range_name
            ).execute()
            
            values = result.get('values', [])
            
            if not values or len(values) < 2:
                self.update_status("오류: 구글 시트에서 데이터를 찾을 수 없습니다.")
                self.root.after(0, lambda: messagebox.showerror(
                    "데이터 없음", 
                    "구글 시트에 데이터가 없거나 형식이 올바르지 않습니다."
                ))
                return

            self.update_status("구글 시트 데이터 구조 분석 중...")
            
            # 2. 헤더 위치 자동 감지 (3~6행 범위에서 찾기)
            best_skiprows = None
            best_score = 0
            best_header = None
            
            for skiprows in range(min(6, len(values))):
                if skiprows >= len(values):
                    continue
                    
                try:
                    header = [str(h).strip() for h in values[skiprows]]
                    
                    # 필수 컬럼 점수 계산
                    score = 0
                    required_columns = ['STRING_ID', '#화자', 'KR']
                    for req_col in required_columns:
                        if req_col in header:
                            score += 10
                    
                    # 언어 컬럼 확인
                    lang_columns = ['EN', 'CN', 'TW', 'TH', 'PT', 'ES', 'DE', 'FR', 'JP']
                    for lang_col in lang_columns:
                        if lang_col in header:
                            score += 2
                    
                    # 대상 언어가 있는지 특별히 확인
                    if target_language in header:
                        score += 5
                    
                    if score > best_score:
                        best_score = score
                        best_skiprows = skiprows
                        best_header = header
                        
                except Exception as e:
                    print(f"헤더 분석 오류 (행 {skiprows}): {e}")
                    continue
            
            if best_skiprows is None or best_score < 15:  # 최소 점수 조건
                self.update_status("오류: 적절한 헤더를 찾을 수 없습니다.")
                self.root.after(0, lambda: messagebox.showerror(
                    "헤더 감지 실패", 
                    f"구글 시트에서 적절한 헤더를 찾을 수 없습니다.\n\n"
                    f"필요한 컬럼: STRING_ID, #화자, KR, {target_language}\n"
                    f"감지된 최고 점수: {best_score}/25"
                ))
                return

            self.update_status(f"헤더 감지 완료 (행 {best_skiprows + 1}, 점수: {best_score})")
            
            # 3. 데이터를 pandas DataFrame으로 변환
            import pandas as pd
            
            if best_skiprows + 1 >= len(values):
                raise ValueError("헤더 다음에 데이터가 없습니다.")
            
            data_rows = values[best_skiprows + 1:]  # 헤더 다음 행부터
            
            # 빈 행 제거 및 데이터 정리
            clean_data = []
            for row in data_rows:
                # 행이 완전히 비어있지 않은 경우만 포함
                if any(cell.strip() if isinstance(cell, str) else str(cell).strip() for cell in row if cell):
                    # 헤더 길이에 맞춰 행 길이 조정
                    while len(row) < len(best_header):
                        row.append('')
                    clean_data.append(row[:len(best_header)])  # 헤더 길이로 자르기
            
            if not clean_data:
                raise ValueError("헤더 다음에 유효한 데이터가 없습니다.")
            
            # DataFrame 생성
            df = pd.DataFrame(clean_data, columns=best_header)
            
            self.update_status(f"데이터 변환 완료: {len(df)}행 × {len(df.columns)}열")
            
            # 4. 시나리오 매니저 초기화 및 분석 실행
            if not self.scenario_manager:
                from scenario_manager import ScenarioTranslationManager
                self.scenario_manager = ScenarioTranslationManager(self.translation_db_path)
            
            self.update_status(f"{target_language} 언어로 시나리오 분석 실행 중...")
            
            # analyze_dataframe 메서드 사용
            analysis_result = self.scenario_manager.analyze_dataframe(df, target_language)
            
            # 5. 결과 처리
            if analysis_result:
                def update_ui_success():
                    self.update_speaker_list()
                    source_info = {
                        'type': 'gsheet',
                        'path': f"{spreadsheet_id}/{sheet_name}"
                    }
                    self.auto_save_reference_analysis(analysis_result, source_info, target_language)
                    # 성공 메시지 생성
                    total_speakers = len(analysis_result)
                    total_sentences = sum(data['total_sentences'] for data in analysis_result.values())
                    
                    summary = f"✅ 구글 시트 분석 완료!\n\n"
                    summary += f"📊 분석 결과:\n"
                    summary += f"• 화자 수: {total_speakers}명\n"
                    summary += f"• 총 문장 수: {total_sentences}개\n"
                    summary += f"• 분석 언어: {target_language}\n\n"
                    summary += f"📋 화자별 상세:\n"
                    
                    for speaker, data in analysis_result.items():
                        lang_count = data['languages'].get(target_language, {}).get('count', 0)
                        summary += f"• {speaker}: {lang_count}개 {target_language} 문장\n"
                    
                    messagebox.showinfo("구글 시트 분석 완료", summary)
                    self.update_status("구글 시트 레퍼런스 분석 완료")
                    
                    # 레퍼런스 파일 경로도 업데이트 (옵션)
                    self.ref_file_var.set(f"구글시트: {sheet_name}")
                
                self.root.after(0, update_ui_success)
            else:
                def update_ui_failure():
                    messagebox.showerror("분석 실패", "구글 시트 데이터 분석에 실패했습니다.")
                    self.update_status("구글 시트 분석 실패")
                
                self.root.after(0, update_ui_failure)
                
        except Exception as e:
            error_msg = f"구글 시트 로드 중 오류: {e}"
            self.update_status(error_msg)
            
            def show_error():
                messagebox.showerror("구글 시트 오류", f"구글 시트에서 레퍼런스 데이터를 가져오는 중 오류가 발생했습니다:\n\n{e}")
            
            self.root.after(0, show_error)
            
            import traceback
            traceback.print_exc()

    def on_conflict_row_selected(self, event):
        """충돌 목록에서 항목 선택 시, 해결 패널에 후보들을 표시"""
        selection = self.conflict_tree.selection()
        if not selection: return
        
        kr_text = self.conflict_tree.item(selection[0], "values")[0]
        self.conflict_kr_var.set(kr_text)
        
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT conflict_info FROM translation_memory WHERE kr_text=?", (kr_text,))
        result = cursor.fetchone()
        conn.close()
        
        if not result or not result[0]: return

        conflict_info = json.loads(result[0])
        
        for lang, combo in self.conflict_combos.items():
            if lang in conflict_info:
                candidates = conflict_info[lang] # {'번역1': 3, '번역2': 1} 형태의 dict
                # 빈도순으로 정렬된 후보 목록 생성
                sorted_candidates = sorted(candidates.items(), key=lambda item: item[1], reverse=True)
                display_list = [f"{text} ({count}회)" for text, count in sorted_candidates]
                
                combo.config(values=display_list, state="readonly")
                combo.set(display_list[0]) # 가장 빈도가 높은 것을 기본값으로 설정
                combo.real_values = [text for text, count in sorted_candidates] # 실제 값 저장
            else:
                combo.config(values=[], state="disabled")
                combo.set("")

    def resolve_selected_conflict(self):
        """사용자가 선택한 값으로 충돌을 해결하고 DB에 반영"""
        kr_text = self.conflict_kr_var.get()
        if not kr_text:
            messagebox.showwarning("선택 오류", "해결할 항목을 목록에서 선택하세요.")
            return

        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT translations FROM translation_memory WHERE kr_text=?", (kr_text,))
        current_translations = json.loads(cursor.fetchone()[0])

        # 콤보박스에서 선택된 값으로 번역문 업데이트
        for lang, combo in self.conflict_combos.items():
            if combo.cget("state") != "disabled":
                selected_index = combo.current()
                if selected_index != -1:
                    current_translations[lang] = combo.real_values[selected_index]
        
        # DB 업데이트: status를 'consolidated'로 변경하고, conflict_info를 null로, 확정된 번역문을 저장
        cursor.execute("""
            UPDATE translation_memory 
            SET translations=?, status='consolidated', conflict_info=NULL 
            WHERE kr_text=?
        """, (json.dumps(current_translations), kr_text))
        
        conn.commit()
        conn.close()

        messagebox.showinfo("해결 완료", f"'{kr_text}' 항목의 충돌이 해결되었습니다.")
        self.load_conflicts_to_view() # 충돌 목록 새로고침
        self.load_tm_view() # 마스터 TM 뷰도 새로고침

    def auto_resolve_all_conflicts(self):
        """모든 충돌 항목을 가장 빈도가 높은 번역으로 자동 해결"""
        if not messagebox.askyesno("경고", "모든 충돌 항목을 자동으로 해결하시겠습니까? 이 작업은 되돌릴 수 없습니다."):
            return
            
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT kr_text, translations, conflict_info FROM translation_memory WHERE status = 'conflict'")
        conflicts = cursor.fetchall()
        
        resolved_count = 0
        for kr_text, trans_json, conflict_json in conflicts:
            current_translations = json.loads(trans_json)
            conflict_info = json.loads(conflict_json)
            
            for lang, candidates in conflict_info.items():
                most_common = max(candidates, key=candidates.get)
                current_translations[lang] = most_common
                
            cursor.execute("""
                UPDATE translation_memory 
                SET translations=?, status='consolidated', conflict_info=NULL 
                WHERE kr_text=?
            """, (json.dumps(current_translations), kr_text))
            resolved_count += 1
            
        conn.commit()
        conn.close()
        
        messagebox.showinfo("자동 해결 완료", f"{resolved_count}개의 충돌 항목이 자동으로 해결되었습니다.")
        self.load_conflicts_to_view()
        self.load_tm_view()


    def setup_exclusion_tab(self):
        """규칙 기반 제외 목록 관리 탭 UI 구성"""
        main_frame = ttk.Frame(self.exclusion_tab, padding="10")
        main_frame.pack(fill="both", expand=True)

        # 1. 새 규칙 추가 프레임
        add_frame = ttk.LabelFrame(main_frame, text="새 규칙 추가")
        add_frame.pack(fill="x", pady=5)
        
        ttk.Label(add_frame, text="규칙 유형:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.rule_type_var = tk.StringVar(value="startswith")
        rule_types = ["startswith", "endswith", "contains", "equals", "length", "regex"]
        ttk.Combobox(add_frame, textvariable=self.rule_type_var, values=rule_types, state="readonly").grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(add_frame, text="적용 필드:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.rule_field_var = tk.StringVar(value="KR")
        # 나중에 다른 언어 필드도 추가할 수 있도록 확장 가능하게 구성
        rule_fields = ["KR", "STRING_ID"] + self.VISIBLE_LANGS
        ttk.Combobox(add_frame, textvariable=self.rule_field_var, values=rule_fields, state="readonly").grid(row=0, column=3, padx=5, pady=5, sticky="ew")
        
        ttk.Label(add_frame, text="값:").grid(row=0, column=4, padx=5, pady=5, sticky="w")
        self.rule_value_var = tk.StringVar()
        ttk.Entry(add_frame, textvariable=self.rule_value_var).grid(row=0, column=5, padx=5, pady=5, sticky="ew")

        ttk.Label(add_frame, text="설명:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.rule_desc_var = tk.StringVar()
        ttk.Entry(add_frame, textvariable=self.rule_desc_var).grid(row=1, column=1, columnspan=5, padx=5, pady=5, sticky="ew")
        
        ttk.Button(add_frame, text="규칙 추가", command=self.add_exclusion_rule).grid(row=0, column=6, rowspan=2, padx=10, pady=5, ipady=10)
        
        add_frame.grid_columnconfigure(5, weight=1)

        # 2. 규칙 목록 표시 프레임
        list_frame = ttk.LabelFrame(main_frame, text="제외 규칙 목록")
        list_frame.pack(fill="both", expand=True, pady=10)

        columns = ("description", "rule_type", "field", "value", "enabled")
        self.exclusion_rule_tree = ttk.Treeview(list_frame, columns=columns, show="headings")
        self.exclusion_rule_tree.heading("description", text="설명")
        self.exclusion_rule_tree.heading("rule_type", text="규칙 유형")
        self.exclusion_rule_tree.heading("field", text="적용 필드")
        self.exclusion_rule_tree.heading("value", text="값")
        self.exclusion_rule_tree.heading("enabled", text="활성화")
        
        self.exclusion_rule_tree.column("description", width=250)
        self.exclusion_rule_tree.column("rule_type", width=100)
        self.exclusion_rule_tree.column("field", width=100)
        self.exclusion_rule_tree.column("value", width=150)
        self.exclusion_rule_tree.column("enabled", width=80, anchor="center")
        
        self.exclusion_rule_tree.pack(side="left", fill="both", expand=True)
        
        # 3. 관리 버튼 프레임
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=5)
        
        ttk.Button(button_frame, text="활성화/비활성화", command=self.toggle_exclusion_rule).pack(side="left")
        ttk.Button(button_frame, text="규칙 삭제", command=self.delete_exclusion_rule).pack(side="left", padx=10)
        ttk.Button(button_frame, text="기본값으로 초기화", command=self.reset_default_rules).pack(side="right")

    def load_exclusion_rules(self):
        """DB에서 제외 규칙을 불러와 테이블에 표시"""
        self.exclusion_rule_tree.delete(*self.exclusion_rule_tree.get_children())
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT id, description, rule_type, field, value, is_enabled FROM exclusion_rules ORDER BY id")
        
        for row in cursor.fetchall():
            # is_enabled 값이 1이면 '활성화', 0이면 '비활성화'로 표시
            enabled_text = "활성화" if row[5] == 1 else "비활성화"
            # treeview에는 id를 내부 id로, 값에는 표시될 내용만 전달
            self.exclusion_rule_tree.insert("", "end", iid=row[0], values=(row[1], row[2], row[3], row[4], enabled_text))
        
        conn.close()

    def add_exclusion_rule(self):
        """새로운 제외 규칙을 DB에 추가"""
        desc = self.rule_desc_var.get().strip()
        rule_type = self.rule_type_var.get()
        field = self.rule_field_var.get()
        value = self.rule_value_var.get().strip()
        
        if not all([desc, rule_type, field, value]):
            messagebox.showwarning("입력 오류", "모든 필드를 입력해야 합니다.")
            return

        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        cursor.execute("INSERT INTO exclusion_rules (description, rule_type, field, value) VALUES (?, ?, ?, ?)",
                    (desc, rule_type, field, value))
        conn.commit()
        conn.close()
        
        # 입력 필드 초기화
        self.rule_desc_var.set("")
        self.rule_value_var.set("")
        
        self.load_exclusion_rules()

    def delete_exclusion_rule(self):
        """선택된 제외 규칙을 DB에서 삭제"""
        selected_items = self.exclusion_rule_tree.selection()
        if not selected_items: return

        if messagebox.askyesno("삭제 확인", f"{len(selected_items)}개의 규칙을 삭제하시겠습니까?"):
            conn = sqlite3.connect(self.translation_db_path)
            cursor = conn.cursor()
            for item_id in selected_items:
                cursor.execute("DELETE FROM exclusion_rules WHERE id = ?", (item_id,))
            conn.commit()
            conn.close()
            self.load_exclusion_rules()

    def toggle_exclusion_rule(self):
        """선택된 규칙의 활성화 상태를 변경"""
        selected_items = self.exclusion_rule_tree.selection()
        if not selected_items: return

        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        for item_id in selected_items:
            # 현재 상태를 확인하기 위해 is_enabled 값을 DB에서 직접 읽어옴
            cursor.execute("SELECT is_enabled FROM exclusion_rules WHERE id = ?", (item_id,))
            current_status = cursor.fetchone()[0]
            # 상태 토글 (1 -> 0, 0 -> 1)
            new_status = 1 - current_status
            cursor.execute("UPDATE exclusion_rules SET is_enabled = ? WHERE id = ?", (new_status, item_id))
        conn.commit()
        conn.close()
        self.load_exclusion_rules()

    def reset_default_rules(self):
        """모든 규칙을 삭제하고 기본 규칙 세트로 초기화"""
        if not messagebox.askyesno("초기화 확인", "모든 규칙을 삭제하고 기본값으로 되돌리시겠습니까? 이 작업은 되돌릴 수 없습니다."):
            return
            
        default_rules = [
            ('#으로 시작하는 KR 제외', 'startswith', 'KR', '#', 1),
            ('cs_로 시작하는 STRING_ID 제외', 'startswith', 'STRING_ID', 'cs_', 1),
            ('KR의 길이가 20 이상인 항목 제외', 'length', 'KR', '20', 0), # 기본 비활성화 예시
            ('\\n\\n으로 시작하는 언어 제외', 'startswith', 'KR', '\\n\\n', 1),
            ('[@...] 형식 제외', 'regex', 'KR', r'^\[@.\]\w*$', 1)
        ]
        
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM exclusion_rules") # 모든 규칙 삭제
        cursor.executemany("INSERT INTO exclusion_rules (description, rule_type, field, value, is_enabled) VALUES (?, ?, ?, ?, ?)", default_rules)
        conn.commit()
        conn.close()
        self.load_exclusion_rules()


    def _is_entry_excluded(self, entry, rules):
        """하나의 데이터 행(entry)이 제외 규칙에 해당하는지 검사"""
        for rule in rules:
            field_to_check = rule.get('field')
            if field_to_check not in entry:
                continue
            
            field_value = str(entry[field_to_check])
            rule_type = rule.get('type')
            rule_value = rule.get('value')
            
            try:
                if rule_type == "startswith" and field_value.startswith(rule_value): return True
                elif rule_type == "endswith" and field_value.endswith(rule_value): return True
                elif rule_type == "contains" and rule_value in field_value: return True
                elif rule_type == "equals" and field_value == rule_value: return True
                elif rule_type == "length" and len(field_value) > int(rule_value): return True
                elif rule_type == "regex" and re.search(rule_value, field_value): return True
            except Exception as e:
                print(f"규칙 적용 오류 (규칙 ID: {rule.get('id')}): {e}")
                continue # 규칙에 오류가 있으면 건너뜀
                
        return False


    def setup_tm_management_tab(self):
        """번역 메모리(TM) 관리 탭 UI 구성 (하위 탭 구조 적용)"""
        # TM 관리 탭의 메인 프레임
        main_tm_frame = ttk.Frame(self.tm_management_tab, padding="5")
        main_tm_frame.pack(fill="both", expand=True)
        
        # TM 관리 탭 내에 노트북(하위 탭) 생성
        tm_notebook = ttk.Notebook(main_tm_frame)
        tm_notebook.pack(fill="both", expand=True)

        # 탭 1: TM 조회 및 직접 편집
        view_edit_tab = ttk.Frame(tm_notebook)
        tm_notebook.add(view_edit_tab, text="TM 조회/편집")
        self.setup_tm_view_edit_tab(view_edit_tab) # UI 구성 함수 호출

        # 탭 2: Excel로 가져오기/업데이트
        import_tab = ttk.Frame(tm_notebook)
        tm_notebook.add(import_tab, text="Excel로 가져오기/업데이트")
        self.setup_excel_import_tab(import_tab) # UI 구성 함수 호출

        quality_tab = ttk.Frame(tm_notebook)
        tm_notebook.add(quality_tab, text="📊 데이터 품질")
        self.setup_data_quality_tab(quality_tab)
 

    def setup_data_quality_tab(self, parent_tab):
        """데이터 품질 관리 탭 설정"""
        main_frame = ttk.Frame(parent_tab, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # 안내 텍스트
        info_frame = ttk.LabelFrame(main_frame, text="ℹ️ 데이터 품질 관리")
        info_frame.pack(fill="x", pady=5)
        
        info_text = """TM과 용어집의 빈 칸을 자동으로 분석하고 보완합니다.
        
    📊 품질 분석: 현재 데이터의 완성도와 일관성을 분석
    🔧 자동 보완: 빈 번역을 패턴 기반으로 자동 채우기
    📄 품질 리포트: 상세한 분석 결과를 파일로 저장"""
        
        ttk.Label(info_frame, text=info_text, font=("맑은 고딕", 9), 
                justify="left").pack(padx=10, pady=10)
        
        # 버튼 프레임
        button_frame = ttk.LabelFrame(main_frame, text="🔧 도구")
        button_frame.pack(fill="x", pady=10)
        
        buttons_inner = ttk.Frame(button_frame, padding="10")
        buttons_inner.pack(fill="x")
        
        # 주요 기능 버튼들
        ttk.Button(buttons_inner, text="📊 품질 분석", 
                command=self.run_quality_analysis, 
                width=15).grid(row=0, column=0, padx=5, pady=5)
        
        ttk.Button(buttons_inner, text="🔧 자동 보완", 
                command=self.run_auto_improvement,
                width=15).grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Button(buttons_inner, text="📄 품질 리포트", 
                command=self.generate_quality_report,
                width=15).grid(row=0, column=2, padx=5, pady=5)
        
        # 고급 기능 (2행)
        ttk.Button(buttons_inner, text="🔍 일관성 검사", 
                command=self.check_translation_consistency,
                width=15).grid(row=1, column=0, padx=5, pady=5)
        
        ttk.Button(buttons_inner, text="🧹 중복 정리", 
                command=self.clean_duplicate_translations,
                width=15).grid(row=1, column=1, padx=5, pady=5)
        
        ttk.Button(buttons_inner, text="📈 품질 대시보드", 
                command=self.show_quality_dashboard,
                width=15).grid(row=1, column=2, padx=5, pady=5)
        
        # 결과 표시 영역
        results_frame = ttk.LabelFrame(main_frame, text="📋 분석 결과")
        results_frame.pack(fill="both", expand=True, pady=10)
        
        # 스크롤 가능한 텍스트 영역
        text_frame = ttk.Frame(results_frame)
        text_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.quality_results_text = tk.Text(text_frame, wrap="word", 
                                        font=("맑은 고딕", 9), height=12)
        quality_scroll = ttk.Scrollbar(text_frame, orient="vertical", 
                                    command=self.quality_results_text.yview)
        self.quality_results_text.configure(yscrollcommand=quality_scroll.set)
        
        self.quality_results_text.pack(side="left", fill="both", expand=True)
        quality_scroll.pack(side="right", fill="y")
        
        # 초기 메시지
        self.quality_results_text.insert("1.0", 
            "💡 품질 분석을 시작하려면 '📊 품질 분석' 버튼을 클릭하세요.\n\n"
            "이 도구는 다음과 같은 작업을 수행합니다:\n"
            "• TM과 용어집의 완성도 분석\n"
            "• 번역 일관성 검사\n"
            "• 빈 번역 자동 보완 제안\n"
            "• 데이터 품질 개선 권장사항 제공\n"
        )
        self.quality_results_text.config(state="disabled")


    def run_quality_analysis(self):
        """품질 분석 실행"""
        try:
            self.update_status("데이터 품질 분석 중...")
            
            # 결과 영역 초기화
            self.quality_results_text.config(state="normal")
            self.quality_results_text.delete("1.0", "end")
            self.quality_results_text.insert("1.0", "🔍 데이터 품질 분석을 시작합니다...\n\n")
            self.quality_results_text.config(state="disabled")
            self.root.update_idletasks()
            
            # 분석 실행
            manager = DataQualityManager(self.translation_db_path)
            analysis = manager.analyze_data_quality()
            
            # 결과 표시
            self.display_quality_analysis(analysis)
            self.update_status("품질 분석 완료")
            
        except Exception as e:
            self.update_status(f"품질 분석 오류: {e}")
            messagebox.showerror("분석 오류", f"품질 분석 중 오류 발생: {e}")

    def display_quality_analysis(self, analysis):
        """분석 결과를 텍스트 영역에 표시"""
        self.quality_results_text.config(state="normal")
        self.quality_results_text.delete("1.0", "end")
        
        # 요약 정보 생성
        tm_analysis = analysis['tm_analysis']
        glossary_analysis = analysis['glossary_analysis']
        consistency_analysis = analysis['consistency_analysis']
        
        result_text = f"""📊 데이터 품질 분석 결과
    {'='*50}

    🗃️ TM (번역 메모리) 현황:
    • 총 항목: {tm_analysis['total_entries']:,}개
    • 완전한 항목: {tm_analysis['complete_entries']:,}개 ({tm_analysis['complete_rate']:.1%})
    • 가장 완성도 높은 언어: {tm_analysis['most_complete_lang']}
    • 가장 완성도 낮은 언어: {tm_analysis['least_complete_lang']}

    언어별 완성도:
    """
        
        for lang, stats in tm_analysis['language_stats'].items():
            bar = "█" * int(stats['completeness_rate'] * 20)  # 간단한 바 그래프
            result_text += f"  {lang:>3}: {stats['completeness_rate']:>6.1%} {bar:<20} ({stats['filled']:,}개)\n"
        
        result_text += f"""
    📚 용어집 현황:
    • 총 항목: {glossary_analysis['total_entries']:,}개

    언어별 완성도:
    """
        
        for lang, stats in glossary_analysis['language_stats'].items():
            bar = "█" * int(stats['completeness_rate'] * 20)
            result_text += f"  {lang:>3}: {stats['completeness_rate']:>6.1%} {bar:<20} ({stats['filled']:,}개)\n"
        
        result_text += f"""
    🔍 번역 일관성:
    • 일관성 문제: {consistency_analysis['total_inconsistent']}개
    • 문제 언어: {', '.join(consistency_analysis['languages_with_issues']) if consistency_analysis['languages_with_issues'] else '없음'}

    💡 개선 권장사항:
    """
        
        for i, rec in enumerate(analysis['recommendations'], 1):
            result_text += f"{i}. {rec}\n"
        
        self.quality_results_text.insert("1.0", result_text)
        self.quality_results_text.config(state="disabled")

    def run_auto_improvement(self):
        """자동 개선 실행"""
        try:
            if not messagebox.askyesno("자동 보완 확인", 
                                    "TM/용어집의 빈 칸을 자동으로 보완하시겠습니까?\n"
                                    "먼저 시뮬레이션을 실행하여 결과를 확인합니다."):
                return
            
            self.update_status("자동 보완 시뮬레이션 중...")
            
            manager = DataQualityManager(self.translation_db_path)
            
            # 시뮬레이션 먼저 실행
            sim_results = manager.auto_fill_missing_translations(dry_run=True)
            
            sim_msg = f"""🔍 시뮬레이션 결과:

    TM 보완 가능: {sim_results['tm_filled']}개
    용어집 보완 가능: {sim_results['glossary_filled']}개

    보완 패턴:
    """
            for pattern, count in sim_results['fill_patterns'].items():
                sim_msg += f"• {pattern}: {count}개\n"
            
            sim_msg += "\n실제로 적용하시겠습니까?"
            
            if messagebox.askyesno("시뮬레이션 결과", sim_msg):
                self.update_status("자동 보완 실행 중...")
                
                # 실제 실행
                real_results = manager.auto_fill_missing_translations(dry_run=False)
                
                messagebox.showinfo("보완 완료", 
                                f"✅ 데이터 보완이 완료되었습니다!\n\n"
                                f"TM: {real_results['tm_filled']}개\n"
                                f"용어집: {real_results['glossary_filled']}개")
                
                # TM 메모리 재로드
                self.load_translation_memory()
                self.load_tm_view()
                self.update_status("자동 보완 완료 - TM 재로드됨")
                
        except Exception as e:
            self.update_status(f"자동 보완 오류: {e}")
            messagebox.showerror("보완 오류", f"자동 보완 중 오류 발생: {e}")

    def generate_quality_report(self):
        """품질 리포트 생성"""
        try:
            self.update_status("품질 리포트 생성 중...")
            
            manager = DataQualityManager(self.translation_db_path)
            analysis = manager.analyze_data_quality()
            report_file = manager.export_quality_report(analysis)
            
            messagebox.showinfo("리포트 생성 완료", 
                            f"📄 품질 리포트가 생성되었습니다:\n{report_file}\n\n"
                            f"리포트를 열어보시겠습니까?")
            
            # 리포트 파일 열기 (선택사항)
            if messagebox.askyesno("파일 열기", "생성된 리포트 파일을 열어보시겠습니까?"):
                import os
                os.startfile(report_file)
                
            self.update_status("품질 리포트 생성 완료")
            
        except Exception as e:
            self.update_status(f"리포트 생성 오류: {e}")
            messagebox.showerror("리포트 오류", f"리포트 생성 중 오류 발생: {e}")

    def check_translation_consistency(self):
        """번역 일관성 상세 검사"""
        messagebox.showinfo("준비 중", "상세 일관성 검사 기능은 2단계에서 구현됩니다.")

    def clean_duplicate_translations(self):
        """중복 번역 정리"""
        messagebox.showinfo("준비 중", "중복 번역 정리 기능은 2단계에서 구현됩니다.")

    def show_quality_dashboard(self):
        """품질 대시보드 표시"""
        messagebox.showinfo("준비 중", "품질 대시보드는 2단계에서 구현됩니다.")
        

    def debug_tm_status(self):
        """TM 상태 디버깅"""
        status_info = f"""TM 상태 정보:
        
    메모리상 TM 항목 수: {len(self.translation_memory)}
    DB TM 항목 수: {self.get_db_tm_count()}

    최근 5개 TM 항목:
    {list(self.translation_memory.keys())[:5]}

    분석 대상 첫 5개 KR:
    {[t["KR"] for t in self.pending_translations[:5]]}
    """
        messagebox.showinfo("TM 상태", status_info)

    def get_db_tm_count(self):
        """DB의 TM 항목 수 확인"""
        try:
            conn = sqlite3.connect(self.translation_db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM translation_memory")
            count = cursor.fetchone()[0]
            conn.close()
            return count
        except:
            return 0


    def setup_tm_view_edit_tab(self, parent_tab):
        """'TM 조회/편집' 하위 탭의 UI를 구성합니다. (컴팩트 버전)"""
        main_frame = ttk.Frame(parent_tab, padding="8")
        main_frame.pack(fill="both", expand=True)
        
        # 컨트롤 프레임을 좌우로 분할
        control_container = ttk.Frame(main_frame)
        control_container.pack(fill="x", pady=5)
        
        # 좌측: TM 관리 도구들
        left_tools = ttk.LabelFrame(control_container, text="TM 관리 도구")
        left_tools.pack(side="left", fill="both", expand=True, padx=(0, 5))
        
        # 기본 도구들 (2x3 그리드)
        tools_grid = ttk.Frame(left_tools, padding="5")
        tools_grid.pack(fill="x")
        
        # 첫 번째 행
        ttk.Button(tools_grid, text="🔍 TM 상태", command=self.debug_tm_status, width=12).grid(row=0, column=0, padx=2, pady=2, sticky="ew")
        ttk.Button(tools_grid, text="🔄 DB 구축", command=self.start_db_build, width=12).grid(row=0, column=1, padx=2, pady=2, sticky="ew")
        ttk.Button(tools_grid, text="🗑️ TM 정리", command=self.cleanup_tm_with_rules, width=12).grid(row=0, column=2, padx=2, pady=2, sticky="ew")
        
        # 두 번째 행: 용어집 관련 도구들
        ttk.Button(tools_grid, text="🔍 충돌검사", command=self.check_tm_glossary_conflicts, width=12).grid(row=1, column=0, padx=2, pady=2, sticky="ew")
        ttk.Button(tools_grid, text="📊 상세리포트", command=self.show_detailed_conflict_report, width=12).grid(row=1, column=1, padx=2, pady=2, sticky="ew")
        ttk.Button(tools_grid, text="🔧 용어집수정", command=self.update_tm_with_glossary, width=12).grid(row=1, column=2, padx=2, pady=2, sticky="ew")
        
        # 컬럼 너비 균등 분배
        for i in range(3):
            tools_grid.grid_columnconfigure(i, weight=1)
        
        # 빌드 모드 선택 (작게)
        mode_frame = ttk.Frame(left_tools, padding="2")
        mode_frame.pack(fill="x")
        
        ttk.Label(mode_frame, text="빌드모드:", font=("맑은 고딕", 8)).pack(side="left")
        ttk.Radiobutton(mode_frame, text="충돌해결", variable=self.db_build_mode_var, value="conflict").pack(side="left", padx=5)
        ttk.Radiobutton(mode_frame, text="빈칸채우기", variable=self.db_build_mode_var, value="fill_blanks").pack(side="left", padx=5)
        
        # 우측: 검색 도구
        search_tools = ttk.LabelFrame(control_container, text="TM 검색")
        search_tools.pack(side="right", fill="y", padx=(5, 0))
        
        search_inner = ttk.Frame(search_tools, padding="5")
        search_inner.pack(fill="both", expand=True)
        
        ttk.Label(search_inner, text="검색어:", font=("맑은 고딕", 9)).pack(anchor="w")
        self.tm_view_search_var = tk.StringVar()
        search_entry = ttk.Entry(search_inner, textvariable=self.tm_view_search_var, width=25)
        search_entry.pack(fill="x", pady=2)
        
        # 검색 옵션
        option_frame = ttk.Frame(search_inner)
        option_frame.pack(fill="x", pady=2)
        
        self.tm_search_exact_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(option_frame, text="완전일치", variable=self.tm_search_exact_var).pack(side="left")
        
        # 검색 버튼
        button_frame = ttk.Frame(search_inner)
        button_frame.pack(fill="x", pady=2)
        
        ttk.Button(button_frame, text="🔍", command=self.filter_tm_view, width=4).pack(side="left", padx=1)
        ttk.Button(button_frame, text="🔄", command=self.clear_tm_filter, width=4).pack(side="left", padx=1)
        
        # TM 목록 테이블
        list_frame = ttk.LabelFrame(main_frame, text="번역 메모리 내용")
        list_frame.pack(fill="both", expand=True, pady=5)

        base_columns = ["KR"] + self.VISIBLE_LANGS
        self.tm_view_tree = ttk.Treeview(list_frame, columns=base_columns, show="headings")
        
        # 컬럼 너비 최적화
        self.tm_view_tree.column("KR", width=200)  # 250 → 200
        for lang in self.VISIBLE_LANGS:
            self.tm_view_tree.column(lang, width=120)  # 150 → 120
            self.tm_view_tree.heading(lang, text=lang)
        
        self.tm_view_tree.heading("KR", text="한국어")

        v_scroll = ttk.Scrollbar(list_frame, orient="vertical", command=self.tm_view_tree.yview)
        h_scroll = ttk.Scrollbar(list_frame, orient="horizontal", command=self.tm_view_tree.xview)
        self.tm_view_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        
        self.tm_view_tree.pack(side="left", fill="both", expand=True)
        v_scroll.pack(side="right", fill="y")
        h_scroll.pack(side="bottom", fill="x")
        

    def check_tm_glossary_conflicts(self):
        """TM과 용어집 간 충돌 검사 (빠른 검사)"""
        if not self.glossary_matcher:
            messagebox.showwarning("용어집 없음", "용어집 매칭 시스템이 활성화되지 않았습니다.")
            return
        
        self.update_status("TM-용어집 충돌 검사 중...")
        
        try:
            conflicts = self._analyze_tm_glossary_conflicts()
            
            if not conflicts:
                messagebox.showinfo("검사 완료", "TM과 용어집 간 충돌이 발견되지 않았습니다.")
                self.update_status("충돌 검사 완료: 문제 없음")
                return
            
            # 간단한 요약 표시
            conflict_summary = f"""🔍 TM-용어집 충돌 검사 결과

    📊 전체 충돌 수: {len(conflicts)}개

    주요 충돌 유형:
    - 완전 불일치: {sum(1 for c in conflicts if c['conflict_type'] == 'complete_mismatch')}개
    - 부분 불일치: {sum(1 for c in conflicts if c['conflict_type'] == 'partial_mismatch')}개  
    - 누락된 용어: {sum(1 for c in conflicts if c['conflict_type'] == 'missing_term')}개

    📋 상위 5개 충돌:"""
            
            for i, conflict in enumerate(conflicts[:5], 1):
                conflict_summary += f"\n{i}. {conflict['kr']} → TM: {conflict['tm_translation'][:30]}..."
            
            if len(conflicts) > 5:
                conflict_summary += f"\n... 외 {len(conflicts) - 5}개 더"
            
            conflict_summary += "\n\n'📊 충돌 상세 리포트' 버튼으로 전체 결과를 확인하세요."
            
            messagebox.showinfo("충돌 검사 완료", conflict_summary)
            self.update_status(f"충돌 검사 완료: {len(conflicts)}개 충돌 발견")
            
            # 충돌 데이터를 임시 저장 (상세 리포트용)
            self._last_conflict_analysis = conflicts
            
        except Exception as e:
            self.update_status(f"충돌 검사 오류: {e}")
            messagebox.showerror("검사 오류", f"충돌 검사 중 오류 발생: {e}")

    def _analyze_tm_glossary_conflicts(self):
        """TM과 용어집 간 충돌 분석 (내부 함수)"""
        conflicts = []
        
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT kr_text, translations FROM translation_memory LIMIT 5000")  # 성능을 위해 제한
        
        for kr_text, trans_json in cursor.fetchall():
            try:
                translations = json.loads(trans_json)
                en_translation = translations.get("EN", "")
                
                if not en_translation:
                    continue
                
                # 용어집에서 관련 용어 찾기
                relevant_terms = self.glossary_matcher.find_relevant_terms(kr_text)
                
                if not relevant_terms:
                    continue
                
                # 각 관련 용어에 대해 충돌 검사
                for term in relevant_terms:
                    korean, english = term.split('→')
                    
                    conflict_info = self._check_single_term_conflict(
                        kr_text, en_translation, korean, english
                    )
                    
                    if conflict_info:
                        conflicts.append({
                            'kr': kr_text,
                            'tm_translation': en_translation,
                            'glossary_term': term,
                            'glossary_korean': korean,
                            'glossary_english': english,
                            'conflict_type': conflict_info['type'],
                            'severity': conflict_info['severity'],
                            'suggestion': conflict_info['suggestion']
                        })
            
            except Exception as e:
                print(f"충돌 분석 오류 (항목: {kr_text}): {e}")
                continue
        
        conn.close()
        
        # 심각도순으로 정렬
        severity_order = {'high': 3, 'medium': 2, 'low': 1}
        conflicts.sort(key=lambda x: severity_order.get(x['severity'], 1), reverse=True)
        
        return conflicts

    def _check_single_term_conflict(self, kr_text, tm_translation, glossary_korean, glossary_english):
        """개별 용어의 충돌 검사"""
        tm_lower = tm_translation.lower()
        glossary_lower = glossary_english.lower()
        
        # 용어집 용어가 한국어 텍스트에 포함되어 있는지 확인
        if glossary_korean not in kr_text:
            return None
        
        # 1. 완전 불일치: 용어집 번역이 TM에 전혀 없음
        if glossary_lower not in tm_lower:
            # 유사한 단어가 있는지 확인 (오타 가능성)
            similar_words = self._find_similar_words(glossary_english, tm_translation)
            
            if similar_words:
                return {
                    'type': 'complete_mismatch',
                    'severity': 'high',
                    'suggestion': f"'{similar_words[0]}' → '{glossary_english}'로 수정 권장"
                }
            else:
                return {
                    'type': 'complete_mismatch', 
                    'severity': 'high',
                    'suggestion': f"'{glossary_english}' 용어 추가 필요"
                }
        
        # 2. 부분 불일치: 용어집 번역이 있지만 다른 형태로 사용됨
        # (예: Bolt vs Vault - 완전히 다른 단어)
        words_in_tm = tm_translation.split()
        glossary_exact_match = any(word.lower() == glossary_lower for word in words_in_tm)
        
        if not glossary_exact_match:
            return {
                'type': 'partial_mismatch',
                'severity': 'medium', 
                'suggestion': f"용어 일관성 검토 필요"
            }
        
        return None  # 충돌 없음

    def _find_similar_words(self, target_word, text):
        """텍스트에서 목표 단어와 유사한 단어 찾기"""
        from difflib import SequenceMatcher
        
        words = text.split()
        similar_words = []
        
        for word in words:
            # 구두점 제거
            clean_word = re.sub(r'[^\w]', '', word)
            
            if len(clean_word) > 2:  # 너무 짧은 단어 제외
                similarity = SequenceMatcher(None, target_word.lower(), clean_word.lower()).ratio()
                if 0.6 <= similarity < 0.9:  # 유사하지만 다른 단어
                    similar_words.append((clean_word, similarity))
        
        # 유사도순 정렬
        similar_words.sort(key=lambda x: x[1], reverse=True)
        return [word for word, sim in similar_words[:3]]

    def show_detailed_conflict_report(self):
        """충돌 상세 리포트 창 표시"""
        if not hasattr(self, '_last_conflict_analysis') or not self._last_conflict_analysis:
            messagebox.showinfo("리포트 없음", "먼저 '용어집 충돌 검사'를 실행하세요.")
            return
        
        conflicts = self._last_conflict_analysis
        
        # 새 창 생성
        report_window = tk.Toplevel(self.root)
        report_window.title("📊 TM-용어집 충돌 상세 리포트")
        report_window.geometry("1000x700")
        report_window.transient(self.root)
        
        main_frame = ttk.Frame(report_window, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # 상단 요약
        summary_frame = ttk.LabelFrame(main_frame, text="📊 충돌 요약")
        summary_frame.pack(fill="x", pady=5)
        
        summary_text = f"""총 충돌 수: {len(conflicts)}개
    - 심각한 충돌 (High): {sum(1 for c in conflicts if c['severity'] == 'high')}개
    - 보통 충돌 (Medium): {sum(1 for c in conflicts if c['severity'] == 'medium')}개  
    - 경미한 충돌 (Low): {sum(1 for c in conflicts if c['severity'] == 'low')}개"""
        
        ttk.Label(summary_frame, text=summary_text, font=("맑은 고딕", 10)).pack(pady=5)
        
        # 충돌 목록 테이블
        list_frame = ttk.LabelFrame(main_frame, text="📋 충돌 상세 목록")
        list_frame.pack(fill="both", expand=True, pady=5)
        
        columns = ("심각도", "한국어", "TM 번역", "용어집 용어", "충돌 유형", "권장사항")
        conflict_tree = ttk.Treeview(list_frame, columns=columns, show="headings")
        
        # 컬럼 설정
        widths = {"심각도": 80, "한국어": 150, "TM 번역": 200, "용어집 용어": 120, "충돌 유형": 100, "권장사항": 200}
        for col in columns:
            conflict_tree.heading(col, text=col)
            conflict_tree.column(col, width=widths.get(col, 100))
        
        # 심각도별 색상 태그
        conflict_tree.tag_configure('high', foreground='red')
        conflict_tree.tag_configure('medium', foreground='orange') 
        conflict_tree.tag_configure('low', foreground='blue')
        
        # 데이터 입력
        for conflict in conflicts:
            conflict_tree.insert("", "end", values=(
                conflict['severity'].upper(),
                conflict['kr'][:30] + "..." if len(conflict['kr']) > 30 else conflict['kr'],
                conflict['tm_translation'][:40] + "..." if len(conflict['tm_translation']) > 40 else conflict['tm_translation'],
                conflict['glossary_term'],
                conflict['conflict_type'],
                conflict['suggestion'][:50] + "..." if len(conflict['suggestion']) > 50 else conflict['suggestion']
            ), tags=(conflict['severity'],))
        
        # 스크롤바
        v_scroll = ttk.Scrollbar(list_frame, orient="vertical", command=conflict_tree.yview)
        h_scroll = ttk.Scrollbar(list_frame, orient="horizontal", command=conflict_tree.xview)
        conflict_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        
        conflict_tree.pack(side="left", fill="both", expand=True)
        v_scroll.pack(side="right", fill="y")
        h_scroll.pack(side="bottom", fill="x")
        
        # 하단 버튼
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=10)
        
        ttk.Button(button_frame, text="📄 CSV로 내보내기", 
                command=lambda: self.export_conflicts_to_csv(conflicts)).pack(side="left")
        ttk.Button(button_frame, text="🔧 선택된 항목 수정", 
                command=lambda: self.fix_selected_conflicts(conflict_tree)).pack(side="left", padx=10)
        ttk.Button(button_frame, text="닫기", command=report_window.destroy).pack(side="right")

    def export_conflicts_to_csv(self, conflicts):
        """충돌 리포트를 CSV로 내보내기"""
        import csv
        from datetime import datetime
        
        file_path = filedialog.asksaveasfilename(
            title="충돌 리포트 저장",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialvalue=f"TM_Glossary_Conflicts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                
                # 헤더
                writer.writerow(['심각도', '한국어', 'TM 번역', '용어집 용어', '충돌 유형', '권장사항'])
                
                # 데이터
                for conflict in conflicts:
                    writer.writerow([
                        conflict['severity'],
                        conflict['kr'],
                        conflict['tm_translation'],
                        conflict['glossary_term'],
                        conflict['conflict_type'],
                        conflict['suggestion']
                    ])
            
            messagebox.showinfo("내보내기 완료", f"충돌 리포트가 저장되었습니다:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("내보내기 실패", f"CSV 저장 중 오류: {e}")

    def update_tm_with_glossary(self):
        """TM을 용어집 기준으로 업데이트"""
        if not messagebox.askyesno("TM 업데이트", 
                                "기존 TM을 용어집 기준으로 업데이트하시겠습니까?\n"
                                "이 작업은 되돌릴 수 없습니다."):
            return
        
        messagebox.showinfo("기능 준비 중", "이 기능은 2단계에서 구현될 예정입니다.\n먼저 '용어집 충돌 검사'로 문제를 파악해보세요.")

    def show_tm_update_report(self, updated_count, conflicts):
        """TM 업데이트 결과 리포트 (2단계에서 구현 예정)"""
        messagebox.showinfo("업데이트 완료", f"{updated_count}개 항목이 업데이트되었습니다.")

    def load_tm_view(self, search_term=None, exact_match=False):
        """DB의 TM을 읽어 관리 탭의 테이블에 표시 (개선된 필터링)"""
        self.tm_view_tree.delete(*self.tm_view_tree.get_children())
        
        # 매개변수가 없으면 현재 UI 상태에서 가져오기
        if search_term is None:
            search_term = self.tm_view_search_var.get().strip() if hasattr(self, 'tm_view_search_var') and self.tm_view_search_var else ""
            exact_match = self.tm_search_exact_var.get() if hasattr(self, 'tm_search_exact_var') and self.tm_search_exact_var else False
        
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()

        if search_term:
            if exact_match:
                cursor.execute("SELECT kr_text, translations FROM translation_memory WHERE kr_text = ?", (search_term,))
            else:
                cursor.execute("SELECT kr_text, translations FROM translation_memory WHERE kr_text LIKE ?", (f"%{search_term}%",))
        else:
            cursor.execute("SELECT kr_text, translations FROM translation_memory LIMIT 1000")  # 성능을 위해 제한 추가
        
        results = cursor.fetchall()
        
        for kr, trans_json in results:
            translations = json.loads(trans_json)
            values = [kr] + [translations.get(lang, "") for lang in self.VISIBLE_LANGS]
            self.tm_view_tree.insert("", "end", values=values)

        conn.close()
        
        # 상태 표시
        if search_term:
            match_type = "완전일치" if exact_match else "부분일치"
            self.update_status(f"TM 필터 적용: '{search_term}' ({match_type}) - {len(results)}개 결과")
        else:
            self.update_status(f"TM 전체 로드: {len(results)}개")


    def filter_tm_view(self):
        """TM 뷰 필터링 (버튼 방식)"""
        search_term = self.tm_view_search_var.get().strip()
        exact_match = self.tm_search_exact_var.get()
        self.load_tm_view(search_term=search_term, exact_match=exact_match)

    def clear_tm_filter(self):
        """TM 필터 초기화"""
        self.tm_view_search_var.set("")
        self.tm_search_exact_var.set(False)
        self.load_tm_view()
        

    def setup_excel_import_tab(self, parent_tab):
        """'Excel로 가져오기/업데이트' 하위 탭의 UI를 구성합니다."""
        main_frame = ttk.Frame(parent_tab, padding="10")
        main_frame.pack(fill="both", expand=True)

        # 1. 폴더 선택 및 파일 검색 프레임
        folder_frame = ttk.LabelFrame(main_frame, text="1. 파일 검색")
        folder_frame.pack(fill="x", pady=5)
        ttk.Label(folder_frame, text="엑셀 폴더:").pack(side="left", padx=5)
        ttk.Entry(folder_frame, textvariable=self.excel_import_folder_var, width=80).pack(side="left", fill="x", expand=True, padx=5)
        ttk.Button(folder_frame, text="폴더 선택", command=lambda: self.excel_import_folder_var.set(filedialog.askdirectory() or "")).pack(side="left", padx=5)
        ttk.Button(folder_frame, text="파일 검색 실행", command=self.search_excel_for_import).pack(side="left", padx=5)

        # 2. 파일 및 언어 선택 프레임
        selection_frame = ttk.LabelFrame(main_frame, text="2. 대상 선택")
        selection_frame.pack(fill="both", expand=True, pady=5)
        
        # 2-1. 파일 선택 리스트
        file_list_frame = ttk.Frame(selection_frame)
        file_list_frame.pack(side="left", fill="both", expand=True, padx=5)
        ttk.Label(file_list_frame, text="처리할 파일 선택:").pack(anchor="w")
        self.excel_files_checklist = ScrollableCheckList(file_list_frame, height=15)
        self.excel_files_checklist.pack(fill="both", expand=True)

        # 2-2. 언어 선택 리스트
        lang_list_frame = ttk.Frame(selection_frame)
        lang_list_frame.pack(side="left", fill="y", padx=5)
        ttk.Label(lang_list_frame, text="업데이트할 언어 선택:").pack(anchor="w")
        for lang in self.VISIBLE_LANGS:
            var = tk.BooleanVar(value=True)
            self.excel_import_lang_vars[lang] = var
            ttk.Checkbutton(lang_list_frame, text=lang, variable=var).pack(anchor="w")

        # 3. 실행 버튼 프레임
        action_frame = ttk.LabelFrame(main_frame, text="3. 실행")
        action_frame.pack(fill="x", pady=5)
        ttk.Button(action_frame, text="선택한 파일로 TM 업데이트 시작", command=self.start_excel_import, style="Accent.TButton").pack(pady=10)


    def import_glossary_from_gsheet(self):
        """'구글 시트에서 가져오기' 버튼의 동작"""
        # 구글 시트 URL이나 ID를 입력받습니다.
        sheet_url_or_id = simpledialog.askstring("구글 시트 정보 입력", "구글 시트의 전체 URL 또는 스프레드시트 ID를 입력하세요:", parent=self.root)
        if not sheet_url_or_id: return

        # URL에서 스프레드시트 ID 추출 (정규식 사용)
        match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', sheet_url_or_id)
        if match:
            spreadsheet_id = match.group(1)
        else:
            spreadsheet_id = sheet_url_or_id # ID 자체를 입력했을 경우

        # 시트 이름을 "glossary"로 고정
        sheet_name = "glossary"

        if messagebox.askyesno("가져오기 확인", "기존 용어집에 KR이 동일한 용어가 있으면, 구글 시트의 내용으로 덮어씁니다. 계속하시겠습니까?", parent=self.root):
            threading.Thread(target=self._import_gsheet_thread, args=(spreadsheet_id, sheet_name), daemon=True).start()


    def _import_gsheet_thread(self, spreadsheet_id, sheet_name):
        """(스레드) 구글 시트 API를 통해 새 구조의 용어집을 가져와 업데이트합니다."""
        try:
            self.update_status("구글 시트에서 데이터 인증 및 다운로드 중...")
            
            from google.oauth2 import service_account
            from googleapiclient.discovery import build

            SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
            SERVICE_ACCOUNT_FILE = 'dulcet-antler-462703-n8-d2fbdb362407.json' 

            creds = service_account.Credentials.from_service_account_file(
                    SERVICE_ACCOUNT_FILE, scopes=SCOPES)
            
            service = build('sheets', 'v4', credentials=creds)
            sheet = service.spreadsheets()
            
            range_name = f"{sheet_name}!A:Z"
            result = sheet.values().get(spreadsheetId=spreadsheet_id, range=range_name).execute()
            values = result.get('values', [])

            if not values or len(values) < 2:
                self.update_status("오류: 구글 시트에서 데이터를 찾을 수 없습니다.")
                return
                
            self.update_status("가져온 데이터 처리 및 DB 업데이트 중...")
            
            header = [h.lower() for h in values[0]] # 헤더는 소문자로 통일
            data_rows = values[1:]
            
            master_conn = sqlite3.connect(self.translation_db_path)
            master_cursor = master_conn.cursor()

            imported_count = 0
            for row in data_rows:
                row_data = dict(zip(header, row))
                string_id = row_data.get('string_id')

                if string_id and string_id.strip():
                    # INSERT OR REPLACE 쿼리를 위한 값 리스트 생성
                    # DB 컬럼 순서: string_id, kr, en, cn, tw, th, pt, es, de, fr, jp, engine, contributor, update_at, verified, description
                    db_values = (
                        string_id,
                        row_data.get('kr'), row_data.get('en'), row_data.get('cn'),
                        row_data.get('tw'), row_data.get('th'), row_data.get('pt'),
                        row_data.get('es'), row_data.get('de'), row_data.get('fr'),
                        row_data.get('jp'), row_data.get('engine'), row_data.get('contributor'),
                        row_data.get('update_at'), int(row_data.get('verified', 0) or 0),
                        row_data.get('description', '')
                    )
                    
                    master_cursor.execute("""
                        INSERT OR REPLACE INTO glossary 
                        (string_id, kr, en, cn, tw, th, pt, es, de, fr, jp, engine, contributor, update_at, verified, description)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, db_values)
                    imported_count += 1

            master_conn.commit()
            master_conn.close()

            self.root.after(0, self.load_glossary)
            self.update_status(f"구글 시트 가져오기 완료! {imported_count}개 항목 처리됨.")
            self.root.after(0, lambda: messagebox.showinfo("완료", f"구글 시트에서 {imported_count}개의 용어를 성공적으로 가져왔습니다."))

        except Exception as e:
            self.update_status(f"구글 시트 가져오기 오류: {e}")
            self.root.after(0, lambda: messagebox.showerror("API 오류", f"구글 시트 처리 중 오류가 발생했습니다:\n{e}"))
            import traceback
            traceback.print_exc()          
            

    def import_tm_from_folder(self):
        """'폴더에서 TM 가져오기' 버튼 클릭 시 실행될 함수"""
        folder_path = filedialog.askdirectory(title="번역 파일들이 있는 폴더(StringDB)를 선택하세요")
        if not folder_path:
            return
        
        # UI가 멈추지 않도록 별도 스레드에서 임포트 작업 실행
        threading.Thread(target=self._import_tm_thread, args=(folder_path,), daemon=True).start()


    def show_context_menu(self, event):
        """우클릭 컨텍스트 메뉴 표시"""
        # 클릭한 항목 선택
        item = self.translation_tree.identify_row(event.y)
        if item:
            self.translation_tree.selection_set(item)
            try:
                self.context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.context_menu.grab_release()

    def force_retranslate_selected(self):
        """선택된 항목을 강제로 재번역 (TM 무시)"""
        selected_items = self.translation_tree.selection()
        if not selected_items:
            messagebox.showwarning("선택 오류", "재번역할 항목을 선택하세요.")
            return
        
        if not messagebox.askyesno("재번역 확인", 
                                f"선택된 {len(selected_items)}개 항목을 API로 강제 재번역하시겠습니까?\n\n" +
                                "기존 번역 내용은 덮어쓰여지고 TM도 업데이트됩니다."):
            return
        
        # 선택된 항목들의 STRING_ID 수집
        selected_string_ids = []
        for item in selected_items:
            values = self.translation_tree.item(item, "values")
            if len(values) > 1:
                selected_string_ids.append(values[1])  # STRING_ID
        
        # 별도 스레드에서 재번역 실행
        threading.Thread(target=self._force_retranslate_thread, 
                        args=(selected_string_ids,), daemon=True).start()

    def _force_retranslate_thread(self, string_ids):
        """강제 재번역 스레드"""
        try:
            self.update_status("강제 재번역 시작...")
            
            # 재번역할 항목들 찾기
            items_to_retranslate = [
                trans for trans in self.pending_translations 
                if trans["STRING_ID"] in string_ids
            ]
            
            if not items_to_retranslate:
                self.update_status("재번역할 항목을 찾을 수 없습니다.")
                return
            
            # API 초기화
            if not DEEPL_API_KEY or DEEPL_API_KEY.startswith('여기에_'):
                self.root.after(0, lambda: messagebox.showerror("API 오류", "DeepL API 키가 설정되지 않았습니다."))
                return
            
            translator = deepl.Translator(DEEPL_API_KEY)
            
            # 각 항목을 강제로 재번역
            success_count = 0
            for i, trans in enumerate(items_to_retranslate):
                kr_text = trans["KR"]
                self.update_status(f"재번역 중 ({i+1}/{len(items_to_retranslate)}): {kr_text[:20]}...")
                
                try:
                    # 기존 번역 내용 백업
                    original_translations = trans["translations"].copy()
                    
                    # EN 번역 (TM 무시하고 API 직접 호출)
                    if self.translate_en_var.get():
                        en_result = self.translate_with_protection(kr_text, "EN-US", translator)
                        if en_result:
                            trans["translations"]["EN"] = en_result
                            trans["method"] = "재번역(API)"
                            trans["status"] = "[재번역완료]"
                            
                            # 다국어 번역도 필요하면
                            if self.translate_multi_var.get():
                                for lang in self.MULTI_LANG_GROUP:
                                    try:
                                        multi_result = self.translate_with_protection(
                                            en_result, LANG_CODES[lang][1], translator, source_lang="EN"
                                        )
                                        if multi_result:
                                            trans["translations"][lang] = multi_result
                                    except Exception as e:
                                        print(f"{lang} 재번역 오류: {e}")
                            
                            success_count += 1
                    
                except Exception as e:
                    print(f"재번역 오류 ({kr_text}): {e}")
                    trans["method"] = "재번역실패"
                    trans["status"] = "[실패]"
            
            # TM 업데이트
            if success_count > 0:
                self.update_translation_memory()
            
            # UI 업데이트
            def update_ui():
                # 재번역완료 필터가 체크되어 있는지 확인하고, 없으면 자동으로 체크
                if hasattr(self, 'filter_vars') and '재번역완료' in self.filter_vars:
                    self.filter_vars['재번역완료'].set(True)
                
                self.update_translation_table()
                self.update_status(f"재번역 완료: {success_count}/{len(items_to_retranslate)}개 성공")
                
            self.root.after(0, update_ui)
            
        except Exception as e:
            self.update_status(f"재번역 오류: {e}")

    def edit_translation_inline(self):
        """선택된 항목을 직접 편집"""
        selected_items = self.translation_tree.selection()
        if not selected_items:
            messagebox.showwarning("선택 오류", "편집할 항목을 선택하세요.")
            return
        
        if len(selected_items) > 1:
            messagebox.showwarning("선택 오류", "한 번에 하나의 항목만 편집할 수 있습니다.")
            return
        
        item = selected_items[0]
        values = self.translation_tree.item(item, "values")
        string_id = values[1]
        
        # pending_translations에서 해당 항목 찾기
        trans_item = next((t for t in self.pending_translations if t["STRING_ID"] == string_id), None)
        if not trans_item:
            messagebox.showerror("오류", "편집할 데이터를 찾을 수 없습니다.")
            return
        
        # 편집 다이얼로그 표시
        InlineEditDialog(self.root, trans_item, self.VISIBLE_LANGS, self.on_inline_edit_complete)

    def on_inline_edit_complete(self, updated_trans):
        """인라인 편집 완료 후 처리"""
        # TM 업데이트
        self.translation_memory[updated_trans["KR"]] = updated_trans["translations"].copy()
        
        # DB 업데이트
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        cursor.execute("""
            INSERT OR REPLACE INTO translation_memory 
            (kr_text, translations, source, confidence)
            VALUES (?, ?, ?, ?)
        """, (
            updated_trans["KR"],
            json.dumps(updated_trans["translations"]),
            "직접편집",
            1.0
        ))
        conn.commit()
        conn.close()
        
        # UI 업데이트
        self.update_translation_table()
        self.update_status("편집 내용이 저장되었습니다.")

    def remove_from_tm(self):
        """선택된 항목을 TM에서 삭제"""
        selected_items = self.translation_tree.selection()
        if not selected_items:
            messagebox.showwarning("선택 오류", "삭제할 항목을 선택하세요.")
            return
        
        kr_texts = []
        for item in selected_items:
            values = self.translation_tree.item(item, "values")
            kr_text = values[2]  # KR 컬럼
            kr_texts.append(kr_text)
        
        if not messagebox.askyesno("TM 삭제 확인", 
                                f"선택된 {len(kr_texts)}개 항목을 TM에서 삭제하시겠습니까?\n\n" +
                                "삭제 후 다음 번역 시 API를 통해 새로 번역됩니다."):
            return
        
        # 메모리와 DB에서 삭제
        deleted_count = 0
        for kr_text in kr_texts:
            if kr_text in self.translation_memory:
                del self.translation_memory[kr_text]
                deleted_count += 1
        
        # DB에서 삭제
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        cursor.executemany("DELETE FROM translation_memory WHERE kr_text = ?", 
                        [(kr,) for kr in kr_texts])
        conn.commit()
        conn.close()
        
        messagebox.showinfo("완료", f"{deleted_count}개 항목이 TM에서 삭제되었습니다.")
        self.update_status(f"TM에서 {deleted_count}개 항목 삭제됨")

    def view_tm_entry(self):
        """선택된 항목의 TM 정보 보기"""
        selected_items = self.translation_tree.selection()
        if not selected_items:
            messagebox.showwarning("선택 오류", "확인할 항목을 선택하세요.")
            return
        
        item = selected_items[0]
        values = self.translation_tree.item(item, "values")
        kr_text = values[2]  # KR 컬럼
        
        if kr_text in self.translation_memory:
            tm_data = self.translation_memory[kr_text]
            info_text = f"KR: {kr_text}\n\nTM 저장 내용:\n"
            for lang, trans in tm_data.items():
                info_text += f"{lang}: {trans}\n"
        else:
            info_text = f"KR: {kr_text}\n\n❌ TM에 저장된 내용이 없습니다."
        
        messagebox.showinfo("TM 정보", info_text)

    # 기존 _import_tm_thread 함수는 삭제하고, 아래 새로운 함수 3개를 클래스 내부에 추가하세요.
    def start_db_build(self):
        """DB 구축 시작 버튼의 동작. 선택된 모드에 따라 적절한 스레드를 실행."""
        folder_path = filedialog.askdirectory(title="소스 DB들이 있는 폴더를 선택하세요")
        if not folder_path:
            return
            
        mode = self.db_build_mode_var.get()
        
        if mode == 'conflict':
            target_func = self._run_conflict_build_mode
        elif mode == 'fill_blanks':
            target_func = self._run_fill_blanks_mode
        else:
            messagebox.showerror("오류", "알 수 없는 빌드 모드입니다.")
            return

        threading.Thread(target=target_func, args=(folder_path,), daemon=True).start()



    def _run_conflict_build_mode(self, folder_path):
        """(스레드) '충돌 우선 해결 모드'를 실행합니다. (진행 상황 표시 개선)"""
        try:
            self.update_status("충돌 해결 모드 시작: 소스 DB 분석 중...")
            self.progress_bar['value'] = 0

            conn = sqlite3.connect(self.translation_db_path)
            cursor = conn.cursor()
            
            cursor.execute("SELECT id, rule_type, field, value FROM exclusion_rules WHERE is_enabled = 1")
            active_rules = [{'id': r[0], 'type': r[1], 'field': r[2], 'value': r[3]} for r in cursor.fetchall()]

            raw_data = defaultdict(lambda: {lang: Counter() for lang in self.VISIBLE_LANGS})
            
            source_db_files = [os.path.join(root, file)
                            for root, _, files in os.walk(folder_path)
                            for file in files
                            if file.lower().startswith('string') and file.lower().endswith('.db')]

            if not source_db_files:
                self.update_status("오류: 폴더에 'String*.db' 파일이 없습니다.")
                conn.close()
                return

            # 1단계: 소스 DB 분석 (진행률 0% -> 50%)
            total_files = len(source_db_files)
            for i, db_path in enumerate(source_db_files):
                self.update_status(f"DB 분석 중 ({i+1}/{total_files}): {os.path.basename(db_path)}")
                self.progress_bar['value'] = ((i + 1) / total_files) * 50
                # (이하 로직은 기존과 동일)
                try:
                    source_conn = sqlite3.connect(db_path)
                    source_cursor = source_conn.cursor()
                    source_cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'String%'")
                    tables = [r[0] for r in source_cursor.fetchall()]
                    for table in tables:
                        source_cursor.execute(f'SELECT * FROM "{table}"')
                        cols = [desc[0] for desc in source_cursor.description]
                        for row in source_cursor.fetchall():
                            entry = dict(zip(cols, row))
                            if self._is_entry_excluded(entry, active_rules):
                                continue
                            kr_text = str(entry.get('KR', '')).strip()
                            if not kr_text: continue
                            for lang in self.VISIBLE_LANGS:
                                if lang in entry and pd.notna(entry[lang]):
                                    raw_data[kr_text][lang][str(entry[lang])] += 1
                    source_conn.close()
                except Exception as e:
                    print(f"소스 DB 처리 오류 {db_path}: {e}")

            # 2단계: 마스터 TM에 충돌 정보 기록 (진행률 50% -> 100%)
            self.update_status("마스터 TM에 충돌 정보 기록 중...")
            processed_count = 0
            total_kr = len(raw_data)
            
            if total_kr > 0:
                for i, (kr_text, lang_map) in enumerate(raw_data.items()):
                    # <<< 시작: 이 부분이 추가되었습니다 >>>
                    self.progress_bar['value'] = 50 + (((i + 1) / total_kr) * 50)
                    # <<< 종료: 이 부분이 추가되었습니다 >>>
                
                    is_conflict = any(len(counter) > 1 for counter in lang_map.values())
                    
                    if is_conflict:
                        status = "conflict"
                        conflict_info = {lang: dict(counter) for lang, counter in lang_map.items() if len(counter) > 1}
                        translations = {lang: counter.most_common(1)[0][0] if counter else "" for lang, counter in lang_map.items()}
                    else:
                        status = "consolidated"
                        conflict_info = None
                        translations = {lang: list(counter.keys())[0] if counter else "" for lang, counter in lang_map.items()}

                    cursor.execute("""
                        INSERT OR REPLACE INTO translation_memory (kr_text, translations, source, status, conflict_info)
                        VALUES (?, ?, ?, ?, ?)
                    """, (kr_text, json.dumps(translations), "DB Build (Conflict Mode)", status, json.dumps(conflict_info) if conflict_info else None))
                    processed_count += 1
            
            conn.commit()
            conn.close()

            # 완료 후 UI 새로고침
            self.root.after(0, self.load_translation_memory)
            self.root.after(0, self.load_tm_view)
            self.root.after(0, self.load_conflicts_to_view)
            self.update_status(f"충돌 해결 모드 완료. {processed_count}개 항목 처리됨.")
            self.root.after(0, lambda: messagebox.showinfo("완료", "충돌 감지가 완료되었습니다.\n[⚠️ 충돌 해결] 탭에서 결과를 확인하고 해결해주세요."))

        except Exception as e:
            self.update_status(f"오류: {e}")
            import traceback
            traceback.print_exc()


    # _run_fill_blanks_mode 함수 코드
    def _run_fill_blanks_mode(self, folder_path):
        """(스레드) '빈칸 채우기 모드'를 실행합니다."""
        try:
            self.update_status("빈칸 채우기 모드 시작: 소스 DB 분석 중...")
            self.progress_bar['value'] = 0
            
            # 1. 소스 DB에서 모든 항목을 수집하여 KR별로 최적의 번역본 하나로 병합
            all_source_entries = self._collect_entries_from_dbs(folder_path)
            self.progress_bar['value'] = 30
            
            # 2. 현재 마스터 TM의 데이터를 메모리로 로드
            self.update_status("기존 마스터 TM 데이터 로딩 중...")
            existing_master_entries = self._load_master_tm_for_merge()
            self.progress_bar['value'] = 40
            
            # 3. 두 데이터를 '비파괴적 업데이트' 방식으로 병합하고, 신규 항목에 ID 부여
            self.update_status("데이터 병합 및 ID 부여 중...")
            merged_entries, new_count, updated_count = self._merge_tm_entries(all_source_entries, existing_master_entries)
            self.progress_bar['value'] = 70

            # 4. 최종 병합 결과를 마스터 DB에 저장
            self.update_status("최종 결과를 마스터 TM에 저장 중...")
            self._save_merged_to_master_tm(merged_entries)
            
            # 5. 작업 완료 및 UI 새로고침
            self.root.after(0, self.load_translation_memory)
            self.root.after(0, self.load_tm_view)
            self.progress_bar['value'] = 100
            self.update_status(f"빈칸 채우기 완료! 신규 {new_count}개, 업데이트 {updated_count}개.")
            self.root.after(0, lambda: messagebox.showinfo("완료", f"빈칸 채우기 작업이 완료되었습니다.\n신규: {new_count}개, 업데이트: {updated_count}개"))

        except Exception as e:
            self.update_status(f"오류: {e}")
            import traceback
            traceback.print_exc()


    def _collect_entries_from_dbs(self, folder_path):
        """(헬퍼) 여러 소스 DB에서 모든 번역 항목을 수집합니다. (동적 테이블 조회 수정)"""
        kr_entries_map = defaultdict(lambda: {lang: [] for lang in self.VISIBLE_LANGS})
        
        source_db_files = [os.path.join(root, file)
                        for root, _, files in os.walk(folder_path)
                        for file in files
                        if file.lower().startswith('string') and file.lower().endswith('.db')]
        
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT id, rule_type, field, value FROM exclusion_rules WHERE is_enabled = 1")
        active_rules = [{'id': r[0], 'type': r[1], 'field': r[2], 'value': r[3]} for r in cursor.fetchall()]
        conn.close()

        for db_path in source_db_files:
            try:
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                
                # <<< 시작: 동적 테이블 조회 로직으로 수정 >>>
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'String%'")
                tables = [r[0] for r in cursor.fetchall()]

                for table in tables:
                    cursor.execute(f'SELECT * FROM "{table}"')
                    columns = [desc[0] for desc in cursor.description]
                    for row in cursor.fetchall():
                        entry = dict(zip(columns, row))
                        if self._is_entry_excluded(entry, active_rules):
                            continue
                        
                        kr_text = str(entry.get('KR', '')).strip()
                        if not kr_text: continue
                        
                        for lang in self.VISIBLE_LANGS:
                            if lang in entry and pd.notna(entry[lang]):
                                kr_entries_map[kr_text][lang].append(str(entry[lang]))
                # <<< 종료: 동적 테이블 조회 로직으로 수정 >>>
                                
                conn.close()
            except Exception as e:
                print(f"소스 DB 처리 오류 {db_path}: {e}")

        # ... (이후 병합 로직은 기존과 동일) ...
        all_merged_entries = []
        for kr, lang_values in kr_entries_map.items():
            merged_entry = {"KR": kr}
            for lang, values in lang_values.items():
                merged_entry[lang] = next((v for v in values if v.strip()), "")
            all_merged_entries.append(merged_entry)
            
        return all_merged_entries
  


    def _load_master_tm_for_merge(self):
        """(헬퍼) 병합을 위해 기존 마스터 TM을 KR을 키로 하는 딕셔너리로 로드합니다."""
        existing_entries = {}
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT kr_text, translations FROM translation_memory")
        for kr, trans_json in cursor.fetchall():
            entry = json.loads(trans_json)
            entry['KR'] = kr # KR 값 추가
            existing_entries[kr] = entry
        conn.close()
        return existing_entries

    def _merge_tm_entries(self, source_entries, master_entries):
        """(헬퍼) 소스 항목과 마스터 항목을 '비파괴적 업데이트' 방식으로 병합하고 ID를 부여합니다."""
        merged = master_entries.copy()
        new_count = 0
        updated_count = 0
        
        # 다음 ID 계산
        existing_ids = [v.get("STRING_ID", "") for v in master_entries.values()]
        nums = [int(i.replace("utext_", "")) for i in existing_ids if i and i.startswith("utext_")]
        next_num = max(nums + [0]) + 1
        
        for entry in source_entries:
            kr = entry["KR"]
            if kr in merged: # 이미 마스터에 있는 경우
                is_updated = False
                for lang in self.VISIBLE_LANGS:
                    # 마스터의 해당 언어가 비어있고, 소스에 값이 있는 경우에만 채워넣기
                    if not merged[kr].get(lang) and entry.get(lang):
                        merged[kr][lang] = entry[lang]
                        is_updated = True
                if is_updated:
                    updated_count += 1
            else: # 마스터에 없는 신규 항목인 경우
                entry["STRING_ID"] = f"utext_{next_num:05}"
                merged[kr] = entry
                next_num += 1
                new_count += 1
                
        return merged, new_count, updated_count

    def _save_merged_to_master_tm(self, merged_entries):
        """(헬퍼) 병합된 최종 결과를 DB에 다시 씁니다."""
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        
        # INSERT OR REPLACE를 사용하여 효율적으로 저장
        for kr, entry in merged_entries.items():
            # translations 컬럼에 넣을 JSON 데이터 준비
            translations_dict = {lang: entry.get(lang, "") for lang in self.VISIBLE_LANGS}
            translations_dict['STRING_ID'] = entry.get('STRING_ID', '')
            
            cursor.execute("""
                INSERT OR REPLACE INTO translation_memory (kr_text, translations, source, status)
                VALUES (?, ?, ?, ?)
            """, (kr, json.dumps(translations_dict), 'DB Merge', 'consolidated'))
            
        conn.commit()
        conn.close()         


    def cleanup_tm_with_rules(self):
        """'TM 정리하기' 버튼의 동작. 확인 후 정리 스레드를 실행합니다."""
        if messagebox.askyesno("TM 정리 확인", "현재 '제외 규칙'을 기준으로 마스터 TM 전체를 검사하여, 규칙에 위배되는 모든 항목을 영구적으로 삭제합니다.\n\n이 작업은 되돌릴 수 없습니다. 계속하시겠습니까?"):
            threading.Thread(target=self._cleanup_tm_thread, daemon=True).start()

    def _cleanup_tm_thread(self):
        """(스레드) 제외 규칙을 사용하여 마스터 TM을 정리합니다."""
        try:
            self.update_status("TM 정리 시작: 제외 규칙 및 TM 데이터 로드 중...")
            
            # 1. 활성화된 제외 규칙 로드
            conn = sqlite3.connect(self.translation_db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT id, rule_type, field, value FROM exclusion_rules WHERE is_enabled = 1")
            active_rules = [{'id': r[0], 'type': r[1], 'field': r[2], 'value': r[3]} for r in cursor.fetchall()]

            # 2. 마스터 TM의 모든 데이터 로드
            cursor.execute("SELECT kr_text, translations FROM translation_memory")
            all_tm_entries = cursor.fetchall()
            
            # 3. 삭제할 항목 식별
            self.update_status(f"TM 항목 검사 중... (총 {len(all_tm_entries)}개)")
            kr_to_delete = []
            for kr_text, trans_json in all_tm_entries:
                translations = json.loads(trans_json)
                entry_dict = {"KR": kr_text, **translations} # 검사를 위해 KR도 포함
                
                if self._is_entry_excluded(entry_dict, active_rules):
                    kr_to_delete.append(kr_text)
            
            # 4. 식별된 항목 삭제
            if kr_to_delete:
                self.update_status(f"삭제 작업 진행 중... ({len(kr_to_delete)}개 항목)")
                # 한 번의 쿼리로 여러 항목을 삭제하기 위해 튜플 리스트 생성
                cursor.executemany("DELETE FROM translation_memory WHERE kr_text = ?", [(kr,) for kr in kr_to_delete])
                conn.commit()
            
            conn.close()
            
            # 5. 완료 후 UI 새로고침
            self.root.after(0, self.load_translation_memory)
            self.root.after(0, self.load_tm_view)
            self.update_status(f"TM 정리 완료! 총 {len(kr_to_delete)}개 항목이 삭제되었습니다.")
            self.root.after(0, lambda: messagebox.showinfo("완료", f"TM 정리가 완료되었습니다.\n총 {len(kr_to_delete)}개 항목이 제외 규칙에 따라 삭제되었습니다."))
            
        except Exception as e:
            self.update_status(f"TM 정리 중 오류 발생: {e}")


    def search_excel_for_import(self):
        """'파일 검색 실행' 버튼의 동작. 지정된 폴더에서 엑셀 파일을 찾아 체크리스트에 표시."""
        folder = self.excel_import_folder_var.get()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("경고", "유효한 폴더를 선택하세요.")
            return
        
        self.excel_files_checklist.clear()
        self.excel_import_files.clear()
        
        found_files = []
        for root, _, files in os.walk(folder):
            for file in files:
                file_name_no_ext, ext = os.path.splitext(file)
                if ext.lower() == ".xlsx" and file_name_no_ext.lower().startswith("string"):
                    file_path = os.path.join(root, file)
                    self.excel_import_files.append((file, file_path))
                    found_files.append(file)

        if not found_files:
            messagebox.showinfo("알림", "조건에 맞는 엑셀 파일을 찾지 못했습니다.")
        else:
            for file_name in sorted(found_files):
                self.excel_files_checklist.add_item(file_name, checked=True)
            messagebox.showinfo("알림", f"{len(found_files)}개의 엑셀 파일을 찾았습니다. 목록에서 처리할 파일을 선택하세요.")

    def start_excel_import(self):
        """'업데이트 시작' 버튼의 동작. 선택된 파일과 언어로 TM 업데이트 스레드 실행."""
        checked_files_names = self.excel_files_checklist.get_checked_items()
        if not checked_files_names:
            messagebox.showwarning("선택 오류", "하나 이상의 엑셀 파일을 선택하세요.")
            return
            
        selected_langs = [lang for lang, var in self.excel_import_lang_vars.items() if var.get()]
        if not selected_langs:
            messagebox.showwarning("선택 오류", "하나 이상의 언어를 선택하세요.")
            return
            
        # (파일명, 파일경로) 튜플에서 선택된 파일명의 경로만 추출
        files_to_process = [path for name, path in self.excel_import_files if name in checked_files_names]
        
        # 실제 DB 업데이트 작업을 별도 스레드에서 실행
        threading.Thread(target=self._excel_import_thread, args=(files_to_process, selected_langs), daemon=True).start()


    def _excel_import_thread(self, files_to_process, selected_langs):
        """(스레드) 선택된 엑셀 파일을 읽어 마스터 TM을 업데이트합니다."""
        try:
            self.update_status("TM 업데이트 시작...")
            self.progress_bar['value'] = 0

            # 1. 활성화된 제외 규칙 미리 로드
            conn = sqlite3.connect(self.translation_db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT id, rule_type, field, value FROM exclusion_rules WHERE is_enabled = 1")
            active_rules = [{'id': r[0], 'type': r[1], 'field': r[2], 'value': r[3]} for r in cursor.fetchall()]

            total_files = len(files_to_process)
            processed_rows = 0

            # 2. 선택된 각 엑셀 파일을 순회하며 처리
            for i, file_path in enumerate(files_to_process):
                self.update_status(f"파일 처리 중 ({i+1}/{total_files}): {os.path.basename(file_path)}")
                self.progress_bar['value'] = (i / total_files) * 100
                
                try:
                    xls = pd.ExcelFile(file_path)
                    for sheet_name in xls.sheet_names:
                        if not sheet_name.lower().startswith('string'):
                            continue # 'string'으로 시작하는 시트만 처리
                        
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        if df.empty: continue
                        
                        # 3. 각 행을 순회하며 마스터 TM에 업데이트
                        for _, row in df.iterrows():
                            # 제외 규칙 적용
                            entry_dict = row.to_dict()
                            if self._is_entry_excluded(entry_dict, active_rules):
                                continue
                            
                            kr_text = str(row.get('KR', '')).strip()
                            if not kr_text: continue

                            # 4. 'Select -> Modify -> Replace' 전략으로 안전하게 업데이트
                            cursor.execute("SELECT translations FROM translation_memory WHERE kr_text=?", (kr_text,))
                            master_result = cursor.fetchone()

                            if master_result:
                                # 기존 항목: 선택된 언어만 업데이트
                                current_translations = json.loads(master_result[0])
                                for lang in selected_langs:
                                    if lang in row:
                                        current_translations[lang] = str(row[lang]) if pd.notna(row[lang]) else ""
                            else:
                                # 신규 항목: 모든 언어 정보로 새로 생성
                                current_translations = {}
                                for lang in self.VISIBLE_LANGS: # 모든 지원 언어에 대해
                                    if lang in row:
                                        current_translations[lang] = str(row[lang]) if pd.notna(row[lang]) else ""
                            
                            # DB에 최종본을 덮어쓰기 (없으면 새로 추가됨)
                            cursor.execute("""
                                INSERT OR REPLACE INTO translation_memory (kr_text, translations, source)
                                VALUES (?, ?, ?)
                            """, (kr_text, json.dumps(current_translations), "Excel Update"))
                            processed_rows += 1

                except Exception as e:
                    print(f"엑셀 파일 처리 오류 {file_path}: {e}")
            
            conn.commit()
            conn.close()

            # 5. 작업 완료 후 UI 새로고침
            self.root.after(0, self.load_translation_memory)
            self.root.after(0, self.load_tm_view)
            self.progress_bar['value'] = 100
            self.update_status(f"TM 업데이트 완료! {processed_rows}개 행이 처리되었습니다.")
            self.root.after(0, lambda: messagebox.showinfo("완료", f"선택된 엑셀 파일의 내용으로 마스터 TM이 업데이트되었습니다.\n(총 {processed_rows}개 행 처리)"))

        except Exception as e:
            self.update_status(f"TM 업데이트 중 오류 발생: {e}")
            
            

    def update_tm_from_excel(self):
        """'최신 번역(Excel)으로 업데이트' 버튼의 시작점"""
        file_path = filedialog.askopenfilename(
            title="업데이트할 엑셀 파일을 선택하세요",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not file_path:
            return

        dialog = LanguageSelectionDialog(self.root, self.VISIBLE_LANGS)
        self.root.wait_window(dialog.top)
        
        selected_langs = dialog.selected_langs
        
        if selected_langs:
            # 실제 비교 및 업데이트 작업을 별도 스레드에서 실행
            threading.Thread(target=self._update_tm_thread, args=(file_path, selected_langs), daemon=True).start()
        else:
            self.update_status("언어가 선택되지 않아 업데이트를 취소했습니다.")
        

    def _update_tm_thread(self, excel_path, selected_langs):
        """(스레드) 엑셀->스테이징DB->마스터TM 비교 및 미리보기 표시"""
        try:
            self.update_status("엑셀 파일 읽는 중...")
            # 1. 엑셀을 읽어 스테이징 DB (메모리상) 생성
            df = pd.read_excel(excel_path)
            
            staging_conn = sqlite3.connect(':memory:') # 메모리에 임시 DB 생성
            df.to_sql('staging_tm', staging_conn, if_exists='replace', index=False)
            staging_cursor = staging_conn.cursor()

            self.update_status("마스터 TM과 비교 중...")
            # 2. 마스터 TM과 스테이징 DB 비교
            master_conn = sqlite3.connect(self.translation_db_path)
            master_cursor = master_conn.cursor()

            new_entries = []
            updated_entries = []

            staging_cursor.execute("SELECT * FROM staging_tm")
            for row in staging_cursor.fetchall():
                # staging DB의 컬럼명과 값을 딕셔너리로 매핑
                staging_entry = dict(zip([desc[0] for desc in staging_cursor.description], row))
                kr_text = staging_entry.get('KR')
                if not kr_text: continue

                master_cursor.execute("SELECT translations FROM translation_memory WHERE kr_text=?", (kr_text,))
                master_result = master_cursor.fetchone()

                if not master_result:
                    # 마스터 TM에 없는 경우 -> 신규 항목
                    new_entries.append(staging_entry)
                else:
                    # 마스터 TM에 있는 경우 -> 변경점 확인
                    master_translations = json.loads(master_result[0])
                    changes = {}
                    for lang in selected_langs:
                        master_val = master_translations.get(lang, "")
                        staging_val = str(staging_entry.get(lang, ""))
                        if master_val != staging_val:
                            changes[lang] = {'old': master_val, 'new': staging_val}
                    
                    if changes:
                        updated_entries.append({'kr': kr_text, 'changes': changes})
            
            staging_conn.close()
            master_conn.close()
            
            if not new_entries and not updated_entries:
                self.update_status("마스터 TM과 비교하여 변경된 내용이 없습니다.")
                self.root.after(0, lambda: messagebox.showinfo("업데이트 정보", "새롭거나 변경된 번역 내용이 없습니다."))
                return

            # 3. 비교 결과를 미리보기 창에 표시 (UI 작업은 메인 스레드에서)
            self.root.after(0, self.show_update_preview, new_entries, updated_entries, selected_langs)

        except Exception as e:
            self.update_status(f"업데이트 중 오류 발생: {e}")
            print(f"업데이트 스레드 오류: {e}")

    def show_update_preview(self, new_entries, updated_entries, selected_langs):
        """비교 결과를 담은 미리보기 다이얼로그를 표시"""
        dialog = UpdatePreviewDialog(self.root, new_entries, updated_entries, self.VISIBLE_LANGS)
        self.root.wait_window(dialog.top)

        if dialog.confirmed:
            # 사용자가 '적용'을 눌렀으면 실제 DB 업데이트 실행
            self.apply_tm_updates(dialog.new_to_apply, dialog.updates_to_apply, selected_langs)

    def apply_tm_updates(self, new_entries, updated_entries, selected_langs):
        """미리보기에서 확정된 내용을 실제 마스터 TM DB에 반영"""
        try:
            self.update_status("마스터 TM 업데이트 중...")
            conn = sqlite3.connect(self.translation_db_path)
            cursor = conn.cursor()

            # 신규 항목 추가
            for entry in new_entries:
                kr_text = entry.get('KR')
                translations = {lang: str(entry.get(lang, '')) for lang in self.VISIBLE_LANGS}
                cursor.execute("""
                    INSERT OR REPLACE INTO translation_memory (kr_text, translations, source) 
                    VALUES (?, ?, ?)
                """, (kr_text, json.dumps(translations), "Excel Update"))

            # 기존 항목 업데이트
            for entry in updated_entries:
                kr_text = entry['kr']
                # 먼저 현재 DB의 번역문을 불러온다
                cursor.execute("SELECT translations FROM translation_memory WHERE kr_text=?", (kr_text,))
                current_translations = json.loads(cursor.fetchone()[0])
                # 변경점만 업데이트한다
                for lang, change in entry['changes'].items():
                    current_translations[lang] = change['new']
                
                cursor.execute("UPDATE translation_memory SET translations=? WHERE kr_text=?", 
                            (json.dumps(current_translations), kr_text))

            conn.commit()
            conn.close()
            
            # 메모리와 UI 새로고침
            self.load_translation_memory()
            self.load_tm_view()
            self.update_status("TM 업데이트 완료!")
            messagebox.showinfo("완료", "최신 번역 내용이 마스터 TM에 성공적으로 반영되었습니다.")

        except Exception as e:
            self.update_status(f"TM 반영 중 오류: {e}")
            print(f"TM 반영 오류: {e}")
            
    def update_tm_with_glossary(self):
        """TM을 용어집 기준으로 업데이트"""
        if not messagebox.askyesno("TM 업데이트", 
                                "기존 TM을 용어집 기준으로 업데이트하시겠습니까?\n"
                                "이 작업은 되돌릴 수 없습니다."):
            return
        
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        
        updated_count = 0
        conflicts = []
        
        # 모든 TM 항목 검사
        cursor.execute("SELECT kr_text, translations FROM translation_memory")
        for kr_text, trans_json in cursor.fetchall():
            translations = json.loads(trans_json)
            original_en = translations.get("EN", "")
            
            if not original_en:
                continue
            
            # 용어집 기반 수정 번역 생성
            corrected_en = self.correct_translation_with_glossary(kr_text, original_en)
            
            if corrected_en != original_en:
                # 변경사항 발견
                conflicts.append({
                    'kr': kr_text,
                    'old': original_en,
                    'new': corrected_en
                })
                
                # DB 업데이트
                translations["EN"] = corrected_en
                cursor.execute(
                    "UPDATE translation_memory SET translations = ? WHERE kr_text = ?",
                    (json.dumps(translations), kr_text)
                )
                updated_count += 1
        
        conn.commit()
        conn.close()
        
        # 결과 리포트
        self.show_tm_update_report(updated_count, conflicts)

    def correct_translation_with_glossary(self, kr_text, original_translation):
        """용어집을 기준으로 번역 수정 (개선된 버전)"""
        if not self.glossary_matcher:
            return original_translation
        
        corrected_translation = original_translation
        # find_relevant_terms는 "KR→EN" 문자열 리스트를 반환하므로, 이를 파싱합니다.
        relevant_terms_str = self.glossary_matcher.find_relevant_terms(kr_text)
        
        for term_str in relevant_terms_str:
            if '→' not in term_str: continue
            korean_term, english_term = term_str.split('→', 1)
            
            # 한국어 용어가 원문에 포함되어 있는지 확인
            if korean_term in kr_text:
                # 잘못 번역된 패턴 찾기
                wrong_patterns = self.find_incorrect_translations(korean_term, corrected_translation)
                
                if wrong_patterns:
                    print(f"🔄 용어집 교정: '{korean_term}'에 대해 잘못된 패턴 {wrong_patterns}을(를) '{english_term}'(으)로 수정합니다.")
                    for wrong in wrong_patterns:
                        # 단어 경계를 고려하여 교체 (정규식 사용)
                        corrected_translation = re.sub(r'\b' + re.escape(wrong) + r'\b', english_term, corrected_translation)

        return corrected_translation

    def find_incorrect_translations(self, korean_term: str, original_translation: str) -> List[str]:
        """
        번역문에서 특정 한국어 용어에 대한 잘못된 번역으로 의심되는 부분을 찾습니다.
        예: '카프카의 숲'(Forest of Kafka)에 대해 'Woods of Kafka'를 찾아 반환합니다.
        """
        # 용어집에서 정확한 영어 번역 가져오기
        glossary_entry = self.exact_matches.get(korean_term) or self.glossary.get(korean_term)
        if not glossary_entry:
            return []
        
        correct_translation = glossary_entry.get('english') or glossary_entry.get('en')
        if not correct_translation:
            return []
            
        # 정확한 번역이 이미 문장에 있으면, 잘못된 패턴이 없다고 가정 (단어 경계까지 확인)
        if re.search(r'\b' + re.escape(correct_translation) + r'\b', original_translation, re.IGNORECASE):
            return []

        # 유사하지만 다른 단어/구를 찾습니다.
        words_in_tm = original_translation.split()
        similar_words = []

        # 유사도 비교
        for word in words_in_tm:
            clean_word = re.sub(r'[^\w]', '', word)  # 구두점 제거
            if len(clean_word) < 3: continue

            similarity = SequenceMatcher(None, correct_translation.lower(), clean_word.lower()).ratio()
            
            # 유사도가 0.7 이상이지만 1.0 미만인 경우 (오타 또는 다른 단어)
            if 0.7 <= similarity < 1.0:
                similar_words.append(word)

        return list(set(similar_words))  # 중복 제거 후 반환

    def init_database(self):
        """데이터베이스 초기화 (새 DB 파일 생성 대응)"""
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        
        # 번역 메모리 테이블 (기존과 동일)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS translation_memory (
                kr_text TEXT PRIMARY KEY,
                translations TEXT,
                usage_count INTEGER DEFAULT 1,
                last_used TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                source TEXT,
                confidence REAL DEFAULT 1.0,
                status TEXT DEFAULT 'consolidated',
                conflict_info TEXT
            )
        """)
        
        # 🆕 용어집 테이블 생성/재생성 (안전한 방식)
        try:
            # 1. 기존 glossary 테이블이 존재하는지 확인
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='glossary'")
            table_exists = cursor.fetchone() is not None
            
            if table_exists:
                print("✅ 기존 용어집 테이블 발견 - 백업 및 재구성 시작")
                
                # 기존 테이블 백업
                cursor.execute("CREATE TABLE IF NOT EXISTS glossary_backup AS SELECT * FROM glossary WHERE 1=0")
                
                # 기존 데이터가 있으면 백업
                try:
                    cursor.execute("INSERT INTO glossary_backup SELECT * FROM glossary")
                    cursor.execute("SELECT COUNT(*) FROM glossary_backup")
                    backup_count = cursor.fetchone()[0]
                    print(f"✅ 기존 용어집 데이터를 백업했습니다: {backup_count}개 항목")
                except Exception as e:
                    print(f"⚠️ 백업 중 오류 (계속 진행): {e}")
                
                # 기존 테이블 삭제 후 새로 생성
                cursor.execute("DROP TABLE IF EXISTS glossary")
            else:
                print("ℹ️ 새로운 용어집 테이블 생성")
            
            # 새로운 용어집 테이블 생성 (STRING_ID 제거)
            cursor.execute("""
                CREATE TABLE glossary (
                    kr TEXT PRIMARY KEY,
                    en TEXT,
                    cn TEXT,
                    tw TEXT,
                    th TEXT,
                    pt TEXT,
                    es TEXT,
                    de TEXT,
                    fr TEXT,
                    jp TEXT,
                    engine TEXT,
                    contributor TEXT,
                    update_at TEXT,
                    verified INTEGER DEFAULT 0,
                    description TEXT
                )
            """)
            
            # 백업에서 데이터 복원 (STRING_ID 제외)
            if table_exists:
                try:
                    cursor.execute("""
                        INSERT OR REPLACE INTO glossary (kr, en, cn, tw, th, pt, es, de, fr, jp, engine, contributor, update_at, verified, description)
                        SELECT kr, en, cn, tw, th, pt, es, de, fr, jp, engine, contributor, update_at, verified, description 
                        FROM glossary_backup 
                        WHERE kr IS NOT NULL AND kr != ''
                    """)
                    
                    # 복원된 데이터 수 확인
                    cursor.execute("SELECT COUNT(*) FROM glossary")
                    restored_count = cursor.fetchone()[0]
                    print(f"✅ 용어집 데이터 복원 완료: {restored_count}개 항목")
                    
                    # 백업 테이블 삭제
                    cursor.execute("DROP TABLE IF EXISTS glossary_backup")
                    
                except Exception as e:
                    print(f"⚠️ 용어집 데이터 복원 중 오류: {e}")
                    print("   새로운 빈 용어집 테이블로 시작합니다.")
            
            print("✅ 용어집 테이블 초기화 완료")
            
        except Exception as e:
            print(f"❌ 용어집 테이블 초기화 오류: {e}")
            print("   기본 빈 테이블로 생성합니다.")
            
            # 오류 발생 시 기본 테이블만 생성
            cursor.execute("DROP TABLE IF EXISTS glossary")
            cursor.execute("""
                CREATE TABLE glossary (
                    kr TEXT PRIMARY KEY,
                    en TEXT,
                    cn TEXT,
                    tw TEXT,
                    th TEXT,
                    pt TEXT,
                    es TEXT,
                    de TEXT,
                    fr TEXT,
                    jp TEXT,
                    engine TEXT,
                    contributor TEXT,
                    update_at TEXT,
                    verified INTEGER DEFAULT 0,
                    description TEXT
                )
            """)
        
        # 나머지 테이블들 (기존과 동일)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS translation_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                string_id TEXT,
                kr_text TEXT,
                translations TEXT,
                translation_method TEXT,
                status TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS exclusion_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                description TEXT,
                rule_type TEXT NOT NULL,
                field TEXT NOT NULL,
                value TEXT NOT NULL,
                is_enabled INTEGER NOT NULL DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # 🆕 시나리오 번역 관련 테이블들 생성
        try:
            # 화자 정보 테이블
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS speakers (
                    name TEXT PRIMARY KEY,
                    gender TEXT,
                    tone TEXT,
                    style TEXT,
                    reference_count INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # 레퍼런스 데이터셋 테이블
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS reference_datasets (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE,
                    description TEXT,
                    source_type TEXT,
                    source_path TEXT,
                    target_language TEXT,
                    total_speakers INTEGER,
                    total_sentences INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    last_used TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # 레퍼런스 번역 데이터 테이블
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS reference_translations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    dataset_id INTEGER,
                    speaker_name TEXT,
                    korean_text TEXT,
                    target_text TEXT,
                    target_language TEXT,
                    string_id TEXT,
                    FOREIGN KEY (dataset_id) REFERENCES reference_datasets (id)
                )
            """)
            
            print("✅ 시나리오 번역 테이블 초기화 완료")
            
        except Exception as e:
            print(f"⚠️ 시나리오 번역 테이블 초기화 오류: {e}")
        
        # 기본 제외 규칙 생성 (테이블이 비어있을 때만)
        cursor.execute("SELECT COUNT(*) FROM exclusion_rules")
        if cursor.fetchone()[0] == 0:
            default_rules = [
                ('#으로 시작하는 KR 제외', 'startswith', 'KR', '#', 1),
                ('cs_로 시작하는 STRING_ID 제외', 'startswith', 'STRING_ID', 'cs_', 1),
                ('\\n\\n으로 시작하는 언어 제외', 'startswith', 'KR', '\\n\\n', 1),
                ('[@...] 형식 제외', 'regex', 'KR', r'^\[@.*\]$', 1)
            ]
            
            cursor.executemany("""
                INSERT INTO exclusion_rules (description, rule_type, field, value, is_enabled) 
                VALUES (?, ?, ?, ?, ?)
            """, default_rules)
            
            print(f"✅ 기본 제외 규칙 {len(default_rules)}개 생성 완료")
        
        conn.commit()
        conn.close()
        
        print("🎉 데이터베이스 초기화 완료!")

    def load_translation_memory(self):
        """번역 메모리 로드"""
        # 기존 unique_texts.db에서 로드
        if os.path.exists(self.unique_texts_db_path):
            try:
                conn = sqlite3.connect(self.unique_texts_db_path)
                cursor = conn.cursor()
                
                cursor.execute("SELECT KR, EN, CN, TW, JP, DE, FR, TH, PT, ES FROM unique_texts")
                rows = cursor.fetchall()
                
                for row in rows:
                    kr = row[0]
                    translations = {
                        "EN": row[1] or "",
                        "CN": row[2] or "",
                        "TW": row[3] or "",
                        "JP": row[4] or "",
                        "DE": row[5] or "",
                        "FR": row[6] or "",
                        "TH": row[7] or "",
                        "PT": row[8] or "",
                        "ES": row[9] or ""
                    }
                    self.translation_memory[kr] = translations
                
                conn.close()
                self.update_status(f"번역 메모리 로드 완료: {len(self.translation_memory)}개")
            except Exception as e:
                self.update_status(f"번역 메모리 로드 오류: {str(e)}")
        
        # smart_translations.db에서도 로드
        try:
            conn = sqlite3.connect(self.translation_db_path)
            cursor = conn.cursor()
            
            cursor.execute("SELECT kr_text, translations FROM translation_memory")
            rows = cursor.fetchall()
            
            for kr, trans_json in rows:
                translations = json.loads(trans_json)
                if kr in self.translation_memory:
                    # 병합
                    for lang, text in translations.items():
                        if text and not self.translation_memory[kr].get(lang):
                            self.translation_memory[kr][lang] = text
                else:
                    self.translation_memory[kr] = translations
            
            conn.close()
        except Exception as e:
            print(f"스마트 번역 DB 로드 오류: {str(e)}")
            
    def select_file(self):
        """파일 선택"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            
            
            
    def determine_status(self, kr_text):
        """텍스트 상태 결정"""
        # 기존 번역 확인
        if kr_text in self.translation_memory:
            # 모든 언어에 번역이 있는지 확인
            trans = self.translation_memory[kr_text]
            if all(trans.get(lang) for lang in ["EN", "JP"]):  # 주요 언어만 체크
                return "[확정]"
            else:
                return "[확인필요]"
        else:
            return "[신규]"
            
 
    def filter_translations(self):
        """번역 필터링"""
        self.update_translation_table()

        
    def update_stats_label(self):
        """통계 레이블 업데이트"""
        total = len(self.pending_translations)
        신규 = sum(1 for t in self.pending_translations if "[신규]" in t["status"])
        변경 = sum(1 for t in self.pending_translations if "[변경]" in t["status"])
        확인필요 = sum(1 for t in self.pending_translations if "[확인필요]" in t["status"])
        
        stats_text = f"전체: {total} | 신규: {신규} | 변경: {변경} | 확인필요: {확인필요}"
        
        if hasattr(self, 'stats_label'):
            self.stats_label.config(text=stats_text)
        else:
            self.stats_label = ttk.Label(self.stats_frame, text=stats_text)
            self.stats_label.pack()
            
    def on_tree_click(self, event):
        """트리 클릭 이벤트 (키보드 조작과 일관성 유지)"""
        region = self.translation_tree.identify_region(event.x, event.y)
        if region == "cell":
            column = self.translation_tree.identify_column(event.x)
            item = self.translation_tree.identify_row(event.y)
            
            if column == "#1" and item:  # 선택 컬럼 클릭
                # 현재 선택된 항목들이 여러 개인 경우
                selected_items = self.translation_tree.selection()
                if len(selected_items) > 1 and item in selected_items:
                    # 여러 항목이 선택된 상태에서 체크박스 클릭하면 모든 선택된 항목 토글
                    self.toggle_multiple_items(selected_items)
                else:
                    # 단일 항목만 토글
                    self.toggle_single_item(item)
            else:
                # 체크박스가 아닌 다른 영역 클릭 시에는 해당 항목을 선택
                if item:
                    self.translation_tree.selection_set(item)
                    self.translation_tree.focus(item)

    def toggle_single_item(self, item):
        """단일 항목의 체크 상태 토글"""
        current_state = self.check_states.get(item, True)
        new_state = not current_state
        self.check_states[item] = new_state
        
        values = list(self.translation_tree.item(item, "values"))
        values[0] = "☑" if new_state else "☐"
        self.translation_tree.item(item, values=values)
        
        self.update_select_all_checkbox()

    def toggle_multiple_items(self, selected_items):
        """여러 선택된 항목들의 체크 상태를 일괄 토글"""
        # 선택된 항목들의 현재 상태 확인
        current_states = [self.check_states.get(item, True) for item in selected_items]
        
        # 하나라도 체크되어 있으면 모두 해제, 모두 해제되어 있으면 모두 체크
        if any(current_states):
            new_state = False
            symbol = "☐"
        else:
            new_state = True
            symbol = "☑"
        
        # 모든 선택된 항목 업데이트
        for item in selected_items:
            self.check_states[item] = new_state
            values = list(self.translation_tree.item(item, "values"))
            values[0] = symbol
            self.translation_tree.item(item, values=values)
        
        self.update_select_all_checkbox()

    def analyze_translations(self):
        """번역 분석 (용어집 우선 적용 강화)"""
        do_en_trans = self.translate_en_var.get()
        do_multi_trans = self.translate_multi_var.get()
        do_cn_tw_trans = self.translate_cn_tw_var.get()

        target_langs = set()
        if do_en_trans:
            target_langs.add("EN")
        if do_multi_trans:
            target_langs.update(self.MULTI_LANG_GROUP)
        if do_cn_tw_trans:
            target_langs.update(["CN", "TW"])

        if not target_langs:
            messagebox.showinfo("정보", "분석할 언어 옵션을 선택해주세요.")
            return

        self.update_status("번역 분석 중... (용어집 우선 적용)")
        
        print(f"현재 TM에 {len(self.translation_memory)}개 항목 로드됨")
        print(f"현재 용어집에 {len(getattr(self, 'exact_matches', {}))}개 항목 로드됨")

        tm_used_count = 0
        glossary_used_count = 0
        api_needed_count = 0

        for trans in self.pending_translations:
            kr_text = trans["KR"]
            methods = set()
            needs_api = False

            # 🆕 1단계: 용어집 우선 확인
            if "EN" in target_langs:
                glossary_result = self.apply_glossary_first(kr_text)
                if glossary_result and glossary_result.get("EN"):
                    trans["translations"]["EN"] = glossary_result["EN"]
                    methods.add("용어집")
                    glossary_used_count += 1
                    print(f"용어집 적용: {kr_text} -> {glossary_result['EN']}")

            # 2단계: TM에서 확인하여 빈칸 채우기 (용어집과 충돌 체크)
            if kr_text in self.translation_memory:
                tm_entry = self.translation_memory[kr_text]
                filled_any = False
                for lang, text in tm_entry.items():
                    if lang in target_langs and not trans["translations"].get(lang) and text:
                        # 용어집과 충돌 체크
                        if self.validate_tm_with_glossary(kr_text, text, lang):
                            trans["translations"][lang] = text
                            filled_any = True
                            print(f"TM 사용 (검증됨): {kr_text} -> {lang}: {text}")
                        else:
                            print(f"TM 사용 거부 (용어집 충돌): {kr_text} -> {lang}: {text}")
                            needs_api = True  # 용어집과 충돌하면 API 재번역 필요
                if filled_any:
                    methods.add("DB")
                    tm_used_count += 1

            # 3단계: 여전히 빈 언어가 있는지 확인
            for lang in target_langs:
                if not trans["translations"].get(lang):
                    if lang in ["CN", "TW"] and do_cn_tw_trans:
                        methods.add("DB필요")
                    else:
                        needs_api = True

            if needs_api:
                methods.add("API필요")
                api_needed_count += 1

            # 번역 방법 결정
            if not methods:
                trans["method"] = "완료"
            else:
                trans["method"] = " / ".join(sorted(list(methods)))

        self.update_translation_table()
        self.update_status(f"분석 완료. 용어집: {glossary_used_count}, TM: {tm_used_count}, API필요: {api_needed_count}")
        
        # 용어집 적용 통계 표시
        if glossary_used_count > 0:
            messagebox.showinfo("분석 완료", 
                            f"📊 번역 분석 결과\n\n"
                            f"✅ 용어집 적용: {glossary_used_count}개\n"
                            f"💾 TM 사용: {tm_used_count}개\n"
                            f"🔌 API 필요: {api_needed_count}개\n\n"
                            f"용어집이 우선 적용되어 일관성이 향상됩니다.")
        
        
    def find_similar_translation(self, kr_text, threshold=0.9):
        """유사 번역 찾기"""
        best_match = None
        best_similarity = 0
        
        for saved_kr, translations in self.translation_memory.items():
            similarity = SequenceMatcher(None, kr_text, saved_kr).ratio()
            
            if similarity >= threshold and similarity > best_similarity:
                best_similarity = similarity
                best_match = {
                    "kr": saved_kr,
                    "translations": translations,
                    "similarity": similarity
                }
                
        return best_match
        
    def apply_glossary(self, kr_text):
        """용어집 적용"""
        # 간단한 구현 - 추후 고도화 가능
        if not self.glossary:
            return None
            
        # 용어집에 있는 단어가 포함되어 있는지 확인
        for kr_term, translations in self.glossary.items():
            if kr_term in kr_text:
                # 부분 번역 제공
                return {"EN": f"[{translations.get('EN', '')}...]"}
                
        return None
        
    def execute_translation(self):
        """번역 실행 (사전 조건 체크 포함)"""
        # 설정값 미리 확인
        do_en_trans = self.translate_en_var.get()
        do_multi_trans = self.translate_multi_var.get()
        do_cn_tw_trans = self.translate_cn_tw_var.get()
        
        if not do_en_trans and not do_multi_trans and not do_cn_tw_trans:
            messagebox.showwarning("경고", "번역 옵션을 선택하세요.")
            return

        # 다국어 번역 선행 조건 체크 (메인 스레드에서)
        if (do_multi_trans or do_cn_tw_trans) and not do_en_trans:
            # 처리할 항목 중 EN이 없는 항목 확인
            items_to_check = []
            for trans in self.pending_translations:
                item_id = self.find_item_id_by_string_id(trans["STRING_ID"])
                if self.check_states.get(item_id, True):
                    items_to_check.append(trans)
            
            # EN 값이 실제로 비어있거나 공백인 경우만 체크
            items_missing_en = [trans for trans in items_to_check 
                            if not trans["translations"].get("EN") or 
                                not str(trans["translations"].get("EN")).strip()]
            
            # EN이 없는 항목이 있는 경우에만 팝업
            if len(items_missing_en) > 0:
                missing_count = len(items_missing_en)
                existing_count = len(items_to_check) - missing_count
                
                message = f"🌍 다국어 번역 안내\n\n"
                
                if existing_count > 0:
                    message += f"✅ EN 번역이 있는 항목: {existing_count}개\n"
                    message += f"❌ EN 번역이 없는 항목: {missing_count}개\n\n"
                    message += f"EN이 없는 {missing_count}개 항목에 대해 먼저 영어 번역을 진행하시겠습니까?\n\n"
                else:
                    message += f"다국어 번역을 위해서는 영어(EN) 번역이 먼저 필요합니다.\n\n"
                    message += f"총 {missing_count}개 항목에 대해 다음 순서로 번역을 진행하시겠습니까?\n\n"
                
                message += "1️⃣ 한국어 → 영어(EN) 번역\n"
                message += "2️⃣ 영어(EN) → 다국어 번역\n\n"
                message += "📝 참고: EN 번역을 기준으로 다국어 번역이 더 정확합니다."
                
                if not messagebox.askyesno("다국어 번역 안내", message, icon="question"):
                    self.update_status("사용자가 다국어 번역을 취소했습니다.")
                    return
                
                # 사용자가 동의했으면 EN 번역도 활성화
                self.translate_en_var.set(True)
                self.update_status(f"✅ EN 번역 추가 - {missing_count}개 항목의 다국어 번역을 위한 선행 작업")
        
        # 번역 스레드 시작
        threading.Thread(target=self._execute_translation_thread, daemon=True).start()


    # def _execute_translation_thread(self):
    #     """
    #     [완전판] 시나리오, 용어 재조립, 다국어 번역을 모두 지원하는 실행 스레드
    #     """
    #     try:
    #         self.update_status("번역 준비 중...")
    #         self.progress_bar['value'] = 0
    #         self.root.update_idletasks()

    #         # --- 1. 설정값 가져오기 ---
    #         selected_engine = self.api_engine_var.get()
    #         do_en_trans = self.translate_en_var.get()
    #         do_multi_trans = self.translate_multi_var.get()
    #         do_cn_tw_trans = self.translate_cn_tw_var.get()
    #         use_scenario = self.scenario_translation_var.get()
    #         use_reassembly_pipeline = self.use_glossary_reassembly_var.get() and self.reassembly_pipeline

    #         # --- 2. 번역 대상 수집 ---
    #         items_to_translate = [
    #             trans for trans in self.pending_translations 
    #             if self.check_states.get(self.find_item_id_by_string_id(trans["STRING_ID"]), True)
    #         ]
    #         if not items_to_translate:
    #             self.update_status("번역할 항목이 없습니다.")
    #             return
    #         total_items = len(items_to_translate)

    #         # --- 3. 엔진 및 프롬프트 초기화 ---
    #         translator = None
    #         if selected_engine == "deepl":
    #             if not DEEPL_API_KEY or DEEPL_API_KEY.startswith('여기에_'):
    #                 self.root.after(0, lambda: messagebox.showerror("API 오류", "DeepL API 키가 설정되지 않았습니다."))
    #                 return
    #             translator = deepl.Translator(DEEPL_API_KEY)
    #         elif selected_engine == "llm" and (not OPENAI_API_KEY or OPENAI_API_KEY.startswith('여기에_')):
    #             self.root.after(0, lambda: messagebox.showerror("API 오류", "OpenAI API 키가 설정되지 않았습니다."))
    #             return
            
    #         base_prompt = self.get_llm_prompt() if not use_scenario else ""
    #         speaker_mapping = self.prepare_scenario_translation() if use_scenario else None

    #         # === 4. 메인 번역 루프 ===
    #         for i, trans in enumerate(items_to_translate):
    #             kr_text = trans["KR"]
    #             self.update_status(f"번역 중 ({i+1}/{total_items}): {kr_text[:20]}...")
    #             self.progress_bar['value'] = (i / total_items) * 100
    #             self.root.update_idletasks()

    #             translation_methods = set(trans['method'].split(' / ') if trans['method'] and trans['method'] != '번역없음' else [])
                
    #             # --- 4-1. 영어(EN) 번역 ---
    #             if do_en_trans and not trans["translations"].get("EN"):
    #                 en_result = None
                    
    #                 # [A] 용어 재조립 파이프라인 우선 실행
    #                 if use_reassembly_pipeline and selected_engine == 'llm':
    #                     try:
    #                         print(f"\n🚀'{kr_text}'에 대한 용어 재조립 파이프라인 실행...")
    #                         en_result = self.reassembly_pipeline.translate_and_reassemble(
    #                             kr_text,
    #                             self.call_gpt4o_mini_api,
    #                             base_prompt
    #                         )
    #                         if en_result:
    #                             translation_methods.add("용어재조립")
    #                     except Exception as e:
    #                         print(f"❌ 용어 재조립 파이프라인 오류: {e}")
    #                         en_result = None # 실패 시 폴백
                    
    #                 # [B] 폴백: 파이프라인 미사용 또는 실패 시 기존 방식 실행
    #                 if not en_result:
    #                     if use_scenario and selected_engine == 'llm' and speaker_mapping:
    #                         speaker = self.get_speaker_for_item_enhanced(trans, speaker_mapping)
    #                         if speaker:
    #                             scenario_result = self.translate_with_enhanced_scenario(kr_text, speaker)
    #                             if scenario_result:
    #                                 en_result = scenario_result.get("translation")
    #                                 translation_methods.add(scenario_result.get("method", "시나리오"))
                        
    #                     # [C] 일반 번역 (시나리오가 아니거나 화자가 없는 경우)
    #                     if not en_result:
    #                         if selected_engine == 'llm':
    #                             en_result = self.translate_with_llm(kr_text, base_prompt)
    #                             if en_result: translation_methods.add("LLM")
    #                         elif selected_engine == 'deepl':
    #                             en_result = self.translate_with_protection(kr_text, "EN-US", translator)
    #                             if en_result: translation_methods.add("DeepL")
    #                         elif selected_engine == 'azure':
    #                             en_result = self.translate_with_azure(kr_text, "en")
    #                             if en_result: translation_methods.add("Azure")

    #                 if en_result:
    #                     trans["translations"]["EN"] = en_result

    #             # --- 4-2. 다국어 번역 (영어 번역 기반) ---
    #             en_text = trans["translations"].get("EN")
    #             if en_text and (do_multi_trans or do_cn_tw_trans):
    #                 multi_lang_targets = []
    #                 if do_multi_trans: multi_lang_targets.extend(self.MULTI_LANG_GROUP)
    #                 if do_cn_tw_trans: multi_lang_targets.extend(["CN", "TW"])
                    
    #                 for lang in multi_lang_targets:
    #                     if not trans["translations"].get(lang): # 해당 언어 번역이 비어있을 때만 실행
    #                         try:
    #                             multi_result = None
    #                             if selected_engine == 'deepl' and translator:
    #                                 multi_result = self.translate_with_protection(en_text, LANG_CODES[lang][1], translator, source_lang="EN")
    #                             elif selected_engine == 'azure':
    #                                 multi_result = self.translate_with_azure(en_text, LANG_CODES[lang][0], source_lang="en")
    #                             elif selected_engine == 'llm':
    #                                 multi_prompt = f"Translate the following English text to {LANG_CODES[lang][2]}:\n\n{en_text}"
    #                                 multi_result = self.translate_with_llm(en_text, multi_prompt)
                                
    #                             if multi_result:
    #                                 trans["translations"][lang] = multi_result
    #                                 if "다국어" not in translation_methods:
    #                                     translation_methods.add("다국어")
    #                         except Exception as e:
    #                             print(f"❌ {lang} 다국어 번역 오류: {e}")

    #             # --- 4-3. 결과 정리 ---
    #             trans["method"] = " / ".join(sorted(list(translation_methods))) if translation_methods else "변경없음"
    #             trans["status"] = "[완료]" if "변경없음" not in trans["method"] else trans["status"]

    #         # === 5. 최종 처리 ===
    #         self.progress_bar['value'] = 100
    #         self.update_translation_memory()
            
    #         # UI 업데이트는 메인 스레드에서 실행
    #         def final_update():
    #             self.update_translation_table()
    #             self.update_stats_label()
    #             self.update_status(f"🎉 번역 완료! {total_items}개 항목 처리됨.")
    #             messagebox.showinfo("번역 완료", f"총 {total_items}개 항목의 번역 작업이 완료되었습니다.")
            
    #         self.root.after(0, final_update)
            
    #     except Exception as e:
    #         self.update_status(f"번역 스레드 오류: {e}")
    #         import traceback
    #         traceback.print_exc()
    #         self.root.after(0, lambda: self.progress_bar.configure(value=0))
            

    def _execute_translation_thread(self):
        """[배치 처리 적용 최종판] 시나리오, 용어 재조립, 다국어 번역을 모두 지원하는 실행 스레드"""
        try:
            self.update_status("번역 준비 중...")
            self.progress_bar['value'] = 0
            
            # --- 1. 설정값 가져오기 ---
            selected_engine = self.api_engine_var.get()
            do_en_trans = self.translate_en_var.get()
            do_multi_trans = self.translate_multi_var.get()
            do_cn_tw_trans = self.translate_cn_tw_var.get()
            use_scenario = self.scenario_translation_var.get()
            use_reassembly_pipeline = self.use_glossary_reassembly_var.get() and self.reassembly_pipeline

            # --- 2. 번역 대상 수집 ---
            items_to_translate = [
                trans for trans in self.pending_translations 
                if self.check_states.get(self.find_item_id_by_string_id(trans["STRING_ID"]), True)
            ]
            if not items_to_translate:
                self.update_status("번역할 항목이 없습니다."); return
                
            total_items = len(items_to_translate)

            # --- 3. 엔진 및 프롬프트 초기화 ---
            translator = None
            if selected_engine == "deepl":
                if not DEEPL_API_KEY or DEEPL_API_KEY.startswith('여기에_'):
                    self.root.after(0, lambda: messagebox.showerror("API 오류", "DeepL API 키가 설정되지 않았습니다."))
                    return
                translator = deepl.Translator(DEEPL_API_KEY)
            elif selected_engine == "llm" and (not OPENAI_API_KEY or OPENAI_API_KEY.startswith('여기에_')):
                 self.root.after(0, lambda: messagebox.showerror("API 오류", "OpenAI API 키가 설정되지 않았습니다."))
                 return
            
            base_prompt = self.get_llm_prompt()
            speaker_mapping = self.prepare_scenario_translation() if use_scenario else None

            # --- 4. 배치 처리 실행 (파이프라인 사용 시) ---
            if use_reassembly_pipeline and selected_engine == 'llm':
                self.update_status("1/3: 텍스트 분리 및 분석 중...")
                unique_text_fragments = set()
                item_split_map = {}
                for i, trans in enumerate(items_to_translate):
                    self.progress_bar['value'] = (i / total_items) * 30
                    split_parts = self.reassembly_pipeline.split_text(trans["KR"])
                    item_split_map[trans["STRING_ID"]] = split_parts
                    for part in split_parts:
                        if part['type'] == 'text' and part['content'].strip():
                            unique_text_fragments.add(part['content'])
                
                self.update_status(f"2/3: {len(unique_text_fragments)}개 조각 배치 번역 중...")
                translated_fragments = {}
                fragments_list = list(unique_text_fragments)
                for i, fragment in enumerate(fragments_list):
                    self.progress_bar['value'] = 30 + (i / max(1, len(fragments_list))) * 40
                    translated_fragments[fragment] = self.call_gpt4o_mini_api(fragment, base_prompt)
                
                self.update_status("3/3: 번역된 조각 재조립 중...")
                for i, trans in enumerate(items_to_translate):
                    self.progress_bar['value'] = 70 + (i / total_items) * 30
                    split_parts = item_split_map.get(trans["STRING_ID"], [])
                    
                    if not split_parts or (len(split_parts) == 1 and split_parts[0]['type']=='text'):
                        trans['translations']['EN'] = self.call_gpt4o_mini_api(trans['KR'], base_prompt)
                        trans['method'] = 'LLM(일반)'
                    else:
                        reassembly_parts = []
                        for part in split_parts:
                            if part['type'] == 'text':
                                part['translation'] = translated_fragments.get(part['content'], part['content'])
                            reassembly_parts.append(part)
                        
                        en_result = self.reassembly_pipeline._reassemble_with_llm(trans['KR'], reassembly_parts, self.call_gpt4o_mini_api)
                        if en_result:
                            trans['translations']['EN'] = en_result
                            trans['method'] = '용어재조립(배치)'
                    trans['status'] = '[완료]'

            # --- 5. 폴백: 기존 방식 실행 (파이프라인 미사용 시) ---
            else:
                if use_reassembly_pipeline and selected_engine != 'llm':
                     self.root.after(0, lambda: messagebox.showinfo("알림", "배치 처리는 '용어 재조립' 옵션과 'LLM' 엔진 선택 시에만 동작합니다.\n기존 방식으로 번역을 진행합니다."))

                for i, trans in enumerate(items_to_translate):
                    kr_text = trans["KR"]
                    self.update_status(f"번역 중 ({i+1}/{total_items}): {kr_text[:20]}...")
                    self.progress_bar['value'] = (i / total_items) * 100
                    
                    translation_methods = set(trans['method'].split(' / ') if trans['method'] and trans['method'] != '번역없음' else [])
                    
                    if do_en_trans and not trans["translations"].get("EN"):
                        en_result = None
                        if use_scenario and selected_engine == 'llm' and speaker_mapping:
                            speaker = self.get_speaker_for_item_enhanced(trans, speaker_mapping)
                            if speaker:
                                scenario_result = self.translate_with_enhanced_scenario(kr_text, speaker)
                                if scenario_result:
                                    en_result = scenario_result.get("translation")
                                    translation_methods.add(scenario_result.get("method", "시나리오"))
                        
                        if not en_result:
                            if selected_engine == 'llm':
                                en_result = self.translate_with_llm(kr_text, base_prompt)
                                if en_result: translation_methods.add("LLM")
                            elif selected_engine == 'deepl':
                                en_result = self.translate_with_protection(kr_text, "EN-US", translator)
                                if en_result: translation_methods.add("DeepL")
                            elif selected_engine == 'azure':
                                en_result = self.translate_with_azure(kr_text, "en")
                                if en_result: translation_methods.add("Azure")

                        if en_result:
                            trans["translations"]["EN"] = en_result
                    
                    en_text = trans["translations"].get("EN")
                    if en_text and (do_multi_trans or do_cn_tw_trans):
                        multi_lang_targets = []
                        if do_multi_trans: multi_lang_targets.extend(self.MULTI_LANG_GROUP)
                        if do_cn_tw_trans: multi_lang_targets.extend(["CN", "TW"])
                        
                        for lang in multi_lang_targets:
                            if not trans["translations"].get(lang):
                                try:
                                    multi_result = None
                                    if selected_engine == 'deepl' and translator:
                                        multi_result = self.translate_with_protection(en_text, LANG_CODES[lang][1], translator, source_lang="EN")
                                    elif selected_engine == 'azure':
                                        multi_result = self.translate_with_azure(en_text, LANG_CODES[lang][0], source_lang="en")
                                    elif selected_engine == 'llm':
                                        multi_prompt = f"Translate the following English text to {LANG_CODES[lang][2]}:\n\n{en_text}"
                                        multi_result = self.translate_with_llm(en_text, multi_prompt)
                                    
                                    if multi_result:
                                        trans["translations"][lang] = multi_result
                                        if "다국어" not in translation_methods:
                                            translation_methods.add("다국어")
                                except Exception as e:
                                    print(f"❌ {lang} 다국어 번역 오류: {e}")
                    
                    trans["method"] = " / ".join(sorted(list(translation_methods))) if translation_methods else "변경없음"
                    trans["status"] = "[완료]" if "변경없음" not in trans["method"] else trans["status"]

            # --- 최종 처리 ---
            self.progress_bar['value'] = 100
            self.update_translation_memory()
            
            def final_update():
                self.update_translation_table()
                self.update_stats_label()
                self.update_status(f"🎉 번역 완료! {total_items}개 항목 처리됨.")
                messagebox.showinfo("번역 완료", f"총 {total_items}개 항목의 번역 작업이 완료되었습니다.")
            
            self.root.after(0, final_update)
            
        except Exception as e:
            self.update_status(f"번역 스레드 오류: {e}")
            import traceback
            traceback.print_exc()
            self.root.after(0, lambda: self.progress_bar.configure(value=0))

    def apply_glossary_first(self, kr_text):
        """용어집 기반 직접 번역 시도 (개선된 버전)"""
        if not kr_text or not kr_text.strip():
            return None
        
        try:
            # 1. 완전 일치 확인
            if hasattr(self, 'exact_matches') and kr_text in self.exact_matches:
                glossary_item = self.exact_matches[kr_text]
                english = glossary_item.get('english', '')
                if english and english.strip():
                    print(f"✅ 용어집 완전 일치: {kr_text} → {english}")
                    return {"EN": english}
            
            # 2. self.glossary에서도 확인 (백업)
            if hasattr(self, 'glossary') and kr_text in self.glossary:
                glossary_item = self.glossary[kr_text]
                english = glossary_item.get('en', '')
                if english and english.strip():
                    print(f"✅ 용어집 백업 매칭: {kr_text} → {english}")
                    return {"EN": english}
            
            # 3. GlossaryMatcher 사용 (부분 매칭)
            if self.glossary_matcher:
                relevant_terms = self.glossary_matcher.find_relevant_terms(kr_text, max_terms=1)
                if relevant_terms:
                    # "한국어→영어" 형태를 파싱
                    for term in relevant_terms:
                        if '→' in term:
                            korean, english = term.split('→', 1)
                            if korean.strip() == kr_text:  # 정확히 일치하는 경우만
                                print(f"✅ 용어집 매처 일치: {kr_text} → {english}")
                                return {"EN": english}
            
            print(f"ℹ️ 용어집에서 찾을 수 없음: {kr_text}")
            return None
            
        except Exception as e:
            print(f"❌ 용어집 적용 오류: {e}")
            return None

    def validate_tm_with_glossary(self, kr_text, tm_translation, target_lang):
        """TM 번역이 용어집과 일치하는지 검증 (개선된 버전)"""
        if not self.glossary_matcher or target_lang != "EN":
            return True
        
        try:
            # 1. 완전 일치 확인
            if hasattr(self, 'exact_matches') and kr_text in self.exact_matches:
                glossary_english = self.exact_matches[kr_text].get('english', '').strip()
                if glossary_english and glossary_english.lower() != tm_translation.lower():
                    print(f"⚠️ TM-용어집 불일치 감지: {kr_text}")
                    print(f"   TM: {tm_translation}")
                    print(f"   용어집: {glossary_english}")
                    return False
            
            # 2. 부분 매칭으로 관련 용어 확인
            relevant_terms = self.glossary_matcher.find_relevant_terms(kr_text, max_terms=5)
            for term in relevant_terms:
                if '→' in term:
                    korean, english = term.split('→', 1)
                    korean, english = korean.strip(), english.strip()
                    
                    # 한국어가 포함되어 있으면서 영어 번역이 다른 경우
                    if korean in kr_text and english.lower() not in tm_translation.lower():
                        print(f"⚠️ TM-용어집 부분 불일치: {korean} → TM에 {english} 없음")
                        return False
            
            return True
            
        except Exception as e:
            print(f"❌ TM-용어집 검증 오류: {e}")
            return True  # 오류 시 기본적으로 허용
        
    def translate_with_llm_protected(self, text, prompt, target_lang=None):
        """
        LLM 번역에 특수 태그 보호 기능 추가
        
        Args:
            text: 번역할 텍스트
            prompt: LLM 프롬프트
            target_lang: 대상 언어 (선택사항, 로깅용)
        
        Returns:
            번역된 텍스트 또는 None (실패 시)
        """
        if not text or not text.strip():
            return text
        
        try:
            # 1. 텍스트 보호 (특수 태그 처리)
            protected_text = text
            protection_map = {}
            
            if self.text_protector:
                try:
                    self.text_protector.reset()
                    protected_text, protection_map = self.text_protector.protect_text(text)
                    
                    # 보호된 텍스트가 번역할 내용이 있는지 확인
                    text_without_placeholders = re.sub(r'<PROTECTED_\d+>', '', protected_text).strip()
                    if not text_without_placeholders:
                        print(f"LLM 보호 번역: 번역할 내용 없음, 원본 반환")
                        return text
                    
                    # 프롬프트에서 원본 텍스트를 보호된 텍스트로 교체
                    if text != protected_text:
                        protected_prompt = prompt.replace(text, protected_text)
                        print(f"🛡️ LLM 특수 태그 보호 적용: {len(protection_map)}개 태그")
                    else:
                        protected_prompt = prompt
                        
                except Exception as e:
                    print(f"LLM 보호 설정 오류, 기본 방식 사용: {e}")
                    protected_text = text
                    protected_prompt = prompt
                    protection_map = {}
            else:
                # text_protector가 없으면 기본 방식
                protected_prompt = prompt
            
            # 2. LLM 번역 실행 (기존 translate_with_llm 로직 사용)
            try:
                import openai
                
                client = openai.OpenAI(api_key=OPENAI_API_KEY)
                
                response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {
                            "role": "system", 
                            "content": "You are a professional translator. Return only the translation, no explanations."
                        },
                        {
                            "role": "user", 
                            "content": protected_prompt
                        }
                    ],
                    max_tokens=min(800, len(protected_text) * 3),
                    temperature=0.2,
                    top_p=0.9,
                    frequency_penalty=0.1,
                    presence_penalty=0.1,
                    timeout=25
                )
                
                if response.choices and len(response.choices) > 0:
                    translated_text = response.choices[0].message.content.strip()
                    
                    # 3. 기본 후처리
                    translated_text = self.post_process_llm_result(translated_text, protected_text)
                    
                    # 4. 보호된 요소 복원
                    if protection_map and self.text_protector:
                        try:
                            restored_text = self.text_protector.restore_text(translated_text, protection_map)
                            print(f"✅ LLM 보호 번역 성공: {text[:30]}... → {target_lang or 'Unknown'}")
                            return restored_text
                        except Exception as e:
                            print(f"LLM 보호 복원 오류: {e}")
                            return translated_text
                    else:
                        print(f"✅ LLM 일반 번역 성공: {text[:30]}... → {target_lang or 'Unknown'}")
                        return translated_text
                else:
                    print(f"LLM API 응답 없음")
                    return None
                    
            except Exception as e:
                print(f"LLM 보호 번역 API 오류: {e}")
                return None
        
        except Exception as e:
            print(f"LLM 보호 번역 전체 오류: {e}")
            # 폴백: 기본 LLM 번역 시도
            try:
                return self.translate_with_llm(text, prompt)
            except:
                return None

    def translate_with_llm_batch_protected(self, texts, base_prompt, target_lang=None):
        """
        여러 텍스트의 배치 LLM 번역 (보호 기능 포함) - 향후 확장용
        
        Args:
            texts: 번역할 텍스트 리스트
            base_prompt: 기본 프롬프트
            target_lang: 대상 언어
        
        Returns:
            번역 결과 리스트
        """
        if not texts:
            return []
        
        # 단순 구현: 개별 번역
        results = []
        for text in texts:
            full_prompt = f"{base_prompt}\n\nText to translate: {text}\nTranslation:"
            result = self.translate_with_llm_protected(text, full_prompt, target_lang)
            results.append(result)
        
        return results

    def translate_with_azure(self, text, target_lang, source_lang="auto"):
        """
        Azure Translator API를 사용한 번역 (특수 태그 보호 기능 포함)
        
        Args:
            text: 번역할 텍스트
            target_lang: 대상 언어 코드 (예: "en", "ko", "th")
            source_lang: 소스 언어 코드 (기본: "auto")
        
        Returns:
            번역된 텍스트 또는 None (실패 시)
        """
        if not text or not text.strip():
            return text
        
        try:
            # Azure API 키 확인
            if not AZURE_API_KEY or AZURE_API_KEY.startswith('여기에_'):
                print("Azure API 키가 설정되지 않았습니다.")
                return None
            
            # 특수 태그 보호 (사용자 설정에 따라)
            protected_text = text
            protection_map = {}
            
            if self.protect_tags_var.get() and self.text_protector:
                try:
                    self.text_protector.reset()
                    protected_text, protection_map = self.text_protector.protect_text(text)
                    
                    # 보호된 텍스트가 번역할 내용이 있는지 확인
                    text_without_placeholders = re.sub(r'<PROTECTED_\d+>', '', protected_text).strip()
                    if not text_without_placeholders:
                        return text  # 번역할 내용이 없으면 원본 반환
                        
                except Exception as e:
                    print(f"Azure 번역 보호 설정 오류: {e}")
                    protected_text = text
                    protection_map = {}
            
            # Azure Translator API 호출
            endpoint = f"https://api.cognitive.microsofttranslator.com/translate"
            headers = {
                'Ocp-Apim-Subscription-Key': AZURE_API_KEY,
                'Ocp-Apim-Subscription-Region': AZURE_REGION,
                'Content-Type': 'application/json',
                'X-ClientTraceId': str(uuid.uuid4())
            }
            
            # API 요청 파라미터
            params = {
                'api-version': '3.0',
                'to': target_lang
            }
            
            # 소스 언어 지정 (auto가 아닌 경우)
            if source_lang and source_lang.lower() != "auto":
                params['from'] = source_lang
            
            # 요청 바디
            body = [{
                'text': protected_text
            }]
            
            # API 호출
            response = requests.post(endpoint, params=params, headers=headers, json=body, timeout=30)
            
            if response.status_code == 200:
                result_json = response.json()
                
                if result_json and len(result_json) > 0:
                    translated_text = result_json[0]['translations'][0]['text']
                    
                    # 보호된 요소 복원
                    if protection_map and self.text_protector:
                        try:
                            restored_text = self.text_protector.restore_text(translated_text, protection_map)
                            return restored_text
                        except Exception as e:
                            print(f"Azure 번역 복원 오류: {e}")
                            return translated_text
                    
                    return translated_text
                else:
                    print(f"Azure API 응답 형식 오류: {result_json}")
                    return None
            else:
                error_info = "알 수 없는 오류"
                try:
                    error_json = response.json()
                    if 'error' in error_json:
                        error_info = error_json['error'].get('message', str(error_json['error']))
                except:
                    error_info = f"HTTP {response.status_code}: {response.text[:100]}"
                
                print(f"Azure 번역 API 오류: {error_info}")
                return None
                
        except requests.exceptions.Timeout:
            print("Azure 번역 API 타임아웃")
            return None
        except requests.exceptions.RequestException as e:
            print(f"Azure 번역 네트워크 오류: {e}")
            return None
        except Exception as e:
            print(f"Azure 번역 예외 오류: {e}")
            return None

    def translate_with_azure_protected(self, text, target_lang, source_lang="auto"):
        """
        Azure 번역에 특수 태그 보호 기능 추가 (별도 함수)
        
        Args:
            text: 번역할 텍스트
            target_lang: 대상 언어 코드
            source_lang: 소스 언어 코드
        
        Returns:
            번역된 텍스트 또는 None
        """
        # 강제로 보호 기능 활성화
        original_protect_setting = self.protect_tags_var.get()
        self.protect_tags_var.set(True)
        
        try:
            result = self.translate_with_azure(text, target_lang, source_lang)
            return result
        finally:
            # 원래 설정 복원
            self.protect_tags_var.set(original_protect_setting)

   
    def check_multilang_prerequisites(self):
        """다국어 번역 선행 조건 체크 (개별 함수로 분리)"""
        do_en_trans = self.translate_en_var.get()
        do_multi_trans = self.translate_multi_var.get()
        do_cn_tw_trans = self.translate_cn_tw_var.get()
        
        # EN이 체크되지 않았지만 다국어가 체크된 경우
        if (do_multi_trans or do_cn_tw_trans) and not do_en_trans:
            return False, "EN 번역이 필요합니다"
        
        # 선택된 항목 중 EN이 비어있는 항목 체크
        empty_en_count = 0
        total_selected = 0
        
        for trans in self.pending_translations:
            item_id = self.find_item_id_by_string_id(trans["STRING_ID"])
            if self.check_states.get(item_id, True):
                total_selected += 1
                if not trans["translations"].get("EN"):
                    empty_en_count += 1
        
        if empty_en_count > 0 and (do_multi_trans or do_cn_tw_trans):
            return False, f"{empty_en_count}개 항목에 EN 번역이 없습니다"
        
        return True, "조건 충족"


    def get_speaker_for_item(self, trans_item, speaker_mapping):
        """번역 항목의 화자 정보 가져오기 (최적화된 버전)"""
        if not speaker_mapping:
            return None
            
        string_id = trans_item["STRING_ID"]
        
        # speaker_mapping에서 해당 STRING_ID가 포함된 화자 찾기
        for speaker, string_ids in speaker_mapping.items():
            if string_id in string_ids:
                return speaker
        
        return None


    def get_speaker_for_translation(self, trans_item):
        """번역 대상 항목의 화자 정보 가져오기"""
        try:
            # 현재 번역 파일에서 해당 STRING_ID의 화자 정보 찾기
            file_path = self.file_path_var.get()
            if not file_path:
                return None
                
            # 캐시된 데이터에서 찾기 (성능 개선)
            if not hasattr(self, '_speaker_cache'):
                self._speaker_cache = {}
                df = pd.read_excel(file_path, skiprows=3)
                if '#화자' in df.columns and 'STRING_ID' in df.columns:
                    for _, row in df.iterrows():
                        if not pd.isna(row['STRING_ID']) and not pd.isna(row['#화자']):
                            self._speaker_cache[str(row['STRING_ID'])] = str(row['#화자']).strip()
            
            return self._speaker_cache.get(trans_item['STRING_ID'])
            
        except Exception as e:
            print(f"화자 정보 조회 오류: {e}")
            return None


    def debug_file_structure(self):
        """레퍼런스 파일 구조 디버깅"""
        ref_file = self.ref_file_var.get()
        if not ref_file:
            messagebox.showwarning("파일 없음", "레퍼런스 파일을 먼저 선택하세요.")
            return
        
        try:
            # 여러 skiprows 옵션으로 시도
            for skip_rows in [0, 1, 2, 3, 4]:
                try:
                    df = pd.read_excel(ref_file, skiprows=skip_rows, nrows=5)  # 상위 5행만 읽기
                    
                    debug_info = f"=== skiprows={skip_rows} ===\n"
                    debug_info += f"컬럼명: {list(df.columns)}\n"
                    debug_info += f"첫 번째 행: {df.iloc[0].to_dict() if len(df) > 0 else 'Empty'}\n"
                    debug_info += f"데이터 타입: {df.dtypes.to_dict()}\n\n"
                    
                    print(debug_info)
                    
                    # 언어 컬럼 찾기
                    possible_lang_cols = []
                    for col in df.columns:
                        col_str = str(col).strip().upper()
                        if col_str in ['EN', 'CN', 'TW', 'TH', 'PT', 'ES', 'DE', 'FR', 'JP']:
                            possible_lang_cols.append(col)
                    
                    if possible_lang_cols:
                        debug_info += f"발견된 언어 컬럼: {possible_lang_cols}\n"
                        break
                        
                except Exception as e:
                    print(f"skiprows={skip_rows} 실패: {e}")
                    continue
            
            # 결과를 사용자에게 표시
            result_dialog = tk.Toplevel(self.root)
            result_dialog.title("파일 구조 분석 결과")
            result_dialog.geometry("800x600")
            
            text_widget = tk.Text(result_dialog, wrap="word", font=("Consolas", 10))
            text_widget.pack(fill="both", expand=True, padx=10, pady=10)
            text_widget.insert("1.0", debug_info)
            
            ttk.Button(result_dialog, text="닫기", command=result_dialog.destroy).pack(pady=5)
            
        except Exception as e:
            messagebox.showerror("디버깅 오류", f"파일 분석 중 오류: {e}")
            

    def translate_with_protection(self, text, target_lang, translator, source_lang="KO"):
        """특수 태그를 보호하면서 번역하는 함수"""
            # 사용자가 특수 태그 보호를 비활성화했으면 일반 번역
        if not self.protect_tags_var.get():
            return translator.translate_text(text, target_lang=target_lang).text
        
        # 복잡한 마크업 옵션이 활성화되어 있고 색상 태그가 있으면
        if self.complex_markup_var.get() and '[#' in text:
            return self.translate_complex_markup(text, target_lang, translator)
        
        try:
            # 1단계: 텍스트 보호
            self.text_protector.reset()
            protected_text, protection_map = self.text_protector.protect_text(text)
            
            print(f"원본: {text}")
            print(f"보호된 텍스트: {protected_text}")
            print(f"보호 맵: {protection_map}")
            
            # 2단계: 보호된 텍스트가 번역할 내용이 있는지 확인
            # 플레이스홀더만 있고 실제 텍스트가 없으면 번역하지 않음
            text_without_placeholders = re.sub(r'<PROTECTED_\d+>', '', protected_text).strip()
            if not text_without_placeholders:
                # 번역할 텍스트가 없으면 원본 반환
                return text
            
            # 3단계: 번역 실행
            translated = translator.translate_text(
                protected_text, 
                target_lang=target_lang,
                source_lang=source_lang if source_lang != "KO" else None
            ).text
            
            print(f"번역된 텍스트: {translated}")
            
            # 4단계: 보호된 요소 복원
            restored_text = self.text_protector.restore_text(translated, protection_map)
            
            print(f"복원된 텍스트: {restored_text}")
            
            return restored_text
            
        except Exception as e:
            print(f"보호된 번역 오류: {e}")
            # 오류 시 기존 방식으로 폴백
            try:
                return translator.translate_text(text, target_lang=target_lang).text
            except:
                return text

    def translate_complex_markup(self, text, target_lang, translator):
        """복잡한 마크업이 있는 텍스트의 고급 번역"""
        try:
            # 색상 태그가 있는 복잡한 텍스트 처리
            translatable_parts = self.text_protector.extract_translatable_parts(text)
            
            if not translatable_parts:
                # 번역 가능한 부분이 없으면 보호된 번역 시도
                return self.translate_with_protection(text, target_lang, translator)
            
            # 각 번역 가능한 부분을 개별 번역
            result_text = text
            for part in reversed(translatable_parts):  # 뒤에서부터 처리
                try:
                    translated_part = translator.translate_text(
                        part['text'], target_lang=target_lang
                    ).text
                    
                    # 원본 텍스트에서 해당 부분을 번역된 텍스트로 교체
                    result_text = (result_text[:part['start']] + 
                                translated_part + 
                                result_text[part['end']:])
                except Exception as e:
                    print(f"부분 번역 오류 ({part['text']}): {e}")
                    
            return result_text
            
        except Exception as e:
            print(f"복잡한 마크업 번역 오류: {e}")
            return self.translate_with_protection(text, target_lang, translator)

    def show_efficiency_report(self):
        """번역 효율성 종합 리포트 표시"""
        summary = self.metrics.get_session_summary()
        
        # 새 창 생성
        report_window = tk.Toplevel(self.root)
        report_window.title("🔍 번역 효율성 분석 리포트")
        report_window.geometry("800x700")
        report_window.transient(self.root)
        
        # 메인 프레임
        main_frame = ttk.Frame(report_window, padding="15")
        main_frame.pack(fill="both", expand=True)
        
        # 탭 생성
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill="both", expand=True)
        
        # 탭 1: 요약 정보
        self.create_summary_tab(notebook, summary)
        
        # 탭 2: API 사용량 분석
        self.create_api_usage_tab(notebook, summary)
        
        # 탭 3: TM 효율성
        self.create_tm_efficiency_tab(notebook, summary)
        
        # 탭 4: 비용 분석
        self.create_cost_analysis_tab(notebook, summary)

    def create_summary_tab(self, notebook, summary):
        """요약 정보 탭"""
        tab = ttk.Frame(notebook)
        notebook.add(tab, text="📊 요약")
        
        # 스크롤 가능한 텍스트 위젯
        text_frame = ttk.Frame(tab)
        text_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        text_widget = tk.Text(text_frame, wrap="word", font=("Consolas", 10))
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        # 리포트 내용 생성
        report_content = f"""
    🎯 번역 세션 효율성 리포트
    {'='*50}

    ⏱️ 세션 정보:
    • 총 소요 시간: {summary['session_duration']:.1f}초
    • 총 API 호출: {summary['total_api_calls']}회
    • 예상 비용: ${summary['total_cost']:.4f}

    📈 TM 효율성:
    • TM 히트율: {summary['tm_hit_rate']*100:.1f}%
    • TM 완전 히트: {summary['detailed_metrics']['tm_usage']['hits']}개
    • TM 부분 히트: {summary['detailed_metrics']['tm_usage']['partial_hits']}개
    • TM 미스: {summary['detailed_metrics']['tm_usage']['misses']}개

    🔄 중복 방지 효율성:
    • 중복 방지된 호출: {summary['api_efficiency']}회
    • 고유 텍스트 수: {len(summary['detailed_metrics']['duplicate_prevention']['unique_texts'])}개

    📋 번역 결과:
    • 성공: {summary['detailed_metrics']['translation_results']['success']}개
    • 실패: {summary['detailed_metrics']['translation_results']['failures']}개
    • 건너뜀: {summary['detailed_metrics']['translation_results']['skipped']}개

    💡 효율성 분석:
    """
        
        # 효율성 등급 계산
        if summary['tm_hit_rate'] > 0.8:
            report_content += "   ✅ TM 활용도: 우수 (80% 이상)\n"
        elif summary['tm_hit_rate'] > 0.5:
            report_content += "   ⚠️ TM 활용도: 보통 (50-80%)\n"
        else:
            report_content += "   ❌ TM 활용도: 낮음 (50% 미만)\n"
            
        if summary['api_efficiency'] > 10:
            report_content += "   ✅ 중복 방지: 우수 (10회 이상 절약)\n"
        else:
            report_content += "   ℹ️ 중복 방지: 일반\n"
            
        avg_time_per_item = summary['session_duration'] / max(1, summary['detailed_metrics']['translation_results']['success'])
        if avg_time_per_item < 2:
            report_content += "   ✅ 처리 속도: 우수 (2초/항목 미만)\n"
        else:
            report_content += "   ⚠️ 처리 속도: 개선 필요\n"

        text_widget.insert("1.0", report_content)
        text_widget.config(state="disabled")
        
        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def create_api_usage_tab(self, notebook, summary):
        """API 사용량 분석 탭"""
        tab = ttk.Frame(notebook)
        notebook.add(tab, text="🔌 API 사용량")
        
        # API별 사용량 테이블
        columns = ("API", "호출횟수", "총 문자수", "평균 응답시간", "예상 비용")
        tree = ttk.Treeview(tab, columns=columns, show="headings")
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120, anchor="center")
        
        # 데이터 입력
        for api_name, data in summary['detailed_metrics']['api_calls'].items():
            if data['count'] > 0:
                avg_response = data['time'] / data['count'] if data['count'] > 0 else 0
                tree.insert("", "end", values=(
                    api_name,
                    data['count'],
                    f"{data['chars']:,}",
                    f"{avg_response:.2f}초",
                    f"${data['cost']:.4f}"
                ))
        
        tree.pack(fill="both", expand=True, padx=10, pady=10)

    def create_tm_efficiency_tab(self, notebook, summary):
        """TM 효율성 탭"""
        tab = ttk.Frame(notebook)
        notebook.add(tab, text="💾 TM 효율성")
        
        # TM 통계 시각화 (간단한 바 차트)
        canvas = tk.Canvas(tab, height=200, bg="white")
        canvas.pack(fill="x", padx=20, pady=20)
        
        tm_data = summary['detailed_metrics']['tm_usage']
        total = sum(tm_data.values())
        
        if total > 0:
            # 간단한 바 차트 그리기
            canvas_width = 600
            bar_height = 40
            
            hits_width = (tm_data['hits'] / total) * canvas_width
            partial_width = (tm_data['partial_hits'] / total) * canvas_width
            miss_width = (tm_data['misses'] / total) * canvas_width
            
            # 히트 바
            canvas.create_rectangle(50, 50, 50 + hits_width, 50 + bar_height, fill="green", tags="hits")
            canvas.create_text(50 + hits_width/2, 70, text=f"완전 히트: {tm_data['hits']}")
            
            # 부분 히트 바
            canvas.create_rectangle(50 + hits_width, 50, 50 + hits_width + partial_width, 50 + bar_height, fill="orange", tags="partial")
            canvas.create_text(50 + hits_width + partial_width/2, 70, text=f"부분 히트: {tm_data['partial_hits']}")
            
            # 미스 바
            canvas.create_rectangle(50 + hits_width + partial_width, 50, 50 + hits_width + partial_width + miss_width, 50 + bar_height, fill="red", tags="miss")
            canvas.create_text(50 + hits_width + partial_width + miss_width/2, 70, text=f"미스: {tm_data['misses']}")

    def create_cost_analysis_tab(self, notebook, summary):
        """비용 분석 탭"""
        tab = ttk.Frame(notebook)
        notebook.add(tab, text="💰 비용 분석")
        
        cost_frame = ttk.LabelFrame(tab, text="예상 비용 분석")
        cost_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        total_cost = summary['total_cost']
        
        # 월간/연간 예상 비용 (현재 세션 기준 추정)
        monthly_estimate = total_cost * 30  # 일일 사용량 기준 월간 추정
        yearly_estimate = monthly_estimate * 12
        
        cost_info = f"""
    💸 이번 세션 비용: ${total_cost:.4f}

    📅 예상 비용 (현재 사용량 기준):
    • 월간 예상: ${monthly_estimate:.2f}
    • 연간 예상: ${yearly_estimate:.2f}

    💡 비용 절약 팁:
    • TM 히트율을 높이면 API 비용을 {(1-summary['tm_hit_rate'])*100:.0f}% 절약 가능
    • 중복 텍스트 사전 제거로 추가 절약 가능
    • 배치 처리로 API 호출 최적화 고려
    """
        
        cost_label = ttk.Label(cost_frame, text=cost_info, justify="left", font=("맑은 고딕", 10))
        cost_label.pack(pady=20)
        


    def show_translation_report(self, report_data):
        """번역 작업이 끝난 후 결과 요약 창을 표시합니다."""
        TranslationReportDialog(self.root, "번역 결과 요약", report_data)

            
    # --- find_item_id_by_string_id 헬퍼 함수 추가 ---
    def find_item_id_by_string_id(self, string_id):
        """STRING_ID로 Treeview의 item id를 찾습니다. (오류 처리 강화)"""
        try:
            for item_id in self.translation_tree.get_children():
                values = self.translation_tree.item(item_id, "values")
                if len(values) > 1 and values[1] == string_id:
                    return item_id
        except Exception as e:
            print(f"find_item_id_by_string_id 오류: {e}")
        return None

    #--- 시나리오 번역 용 
    def analyze_reference_data_smart(self):
        """스마트 레퍼런스 데이터 분석 (헤더 자동 감지)"""
        ref_file = self.ref_file_var.get()
        if not ref_file:
            messagebox.showwarning("파일 없음", "레퍼런스 파일을 먼저 선택하세요.")
            return
        
        try:
            self.update_status("헤더 위치 자동 감지 중...")
            
            # 3~6행 범위에서 헤더 찾기
            best_skiprows = None
            best_score = 0
            best_columns = None
            
            for skiprows in range(3, 7):  # 3~6행 시도 (0부터 시작하므로 4~7행째)
                try:
                    df = pd.read_excel(ref_file, skiprows=skiprows, nrows=5)
                    
                    # 필수 컬럼 점수 계산
                    score = 0
                    columns = list(df.columns)
                    
                    # 필수 컬럼 확인
                    required_columns = ['STRING_ID', '#화자', 'KR']
                    for req_col in required_columns:
                        if req_col in columns:
                            score += 10
                    
                    # 언어 컬럼 확인
                    lang_columns = ['EN', 'CN', 'TW', 'TH', 'PT', 'ES', 'DE', 'FR', 'JP']
                    for lang_col in lang_columns:
                        if lang_col in columns:
                            score += 2
                    
                    # 데이터 유효성 확인 (빈 값이 너무 많으면 감점)
                    if len(df) > 0:
                        empty_ratio = df.isnull().sum().sum() / (len(df) * len(df.columns))
                        if empty_ratio < 0.5:  # 빈 값이 50% 미만이면 가점
                            score += 5
                    
                    print(f"skiprows={skiprows}, score={score}, columns={columns}")
                    
                    if score > best_score:
                        best_score = score
                        best_skiprows = skiprows
                        best_columns = columns
                        
                except Exception as e:
                    print(f"skiprows={skiprows} 실패: {e}")
                    continue
            
            if best_skiprows is None or best_score < 10:
                messagebox.showerror("헤더 감지 실패", 
                                "적절한 헤더 행을 찾을 수 없습니다.\n\n" +
                                "'수동 매핑' 버튼을 사용해주세요.")
                return
            
            self.update_status(f"헤더 감지 완료: {best_skiprows+1}행 (점수: {best_score})")
            
            # 사용할 언어 선택
            available_langs = [col for col in best_columns if col in ['EN', 'CN', 'TW', 'TH', 'PT', 'ES', 'DE', 'FR', 'JP']]
            
            if not available_langs:
                messagebox.showerror("언어 컬럼 없음", "번역 가능한 언어 컬럼을 찾을 수 없습니다.")
                return
            
            # 언어 선택 다이얼로그
            if len(available_langs) == 1:
                target_lang = available_langs[0]
            else:
                dialog = LanguageSelectionDialog(self.root, available_langs, "분석할 언어 선택")
                self.root.wait_window(dialog.top)
                target_lang = dialog.selected_lang if hasattr(dialog, 'selected_lang') else None
                
                if not target_lang:
                    return
            
            # 실제 분석 실행
            if not self.scenario_manager:
                from scenario_manager import ScenarioTranslationManager
                self.scenario_manager = ScenarioTranslationManager(self.translation_db_path)
            
            self.update_status(f"{target_lang} 언어로 분석 중...")
            analysis_result = self.scenario_manager.analyze_reference_data(ref_file, target_lang, best_skiprows)
            
            if analysis_result:
                self.update_speaker_list()
                summary = f"✅ 자동 분석 완료! (헤더: {best_skiprows+1}행)\n\n"
                total_speakers = len(analysis_result)
                total_sentences = sum(data['total_sentences'] for data in analysis_result.values())
                
                summary += f"화자 수: {total_speakers}명\n"
                summary += f"총 문장 수: {total_sentences}개\n\n"
                summary += "화자별 상세:\n"
                
                for speaker, data in analysis_result.items():
                    lang_count = data['languages'].get(target_lang, {}).get('count', 0)
                    summary += f"• {speaker}: {lang_count}개 {target_lang} 문장\n"
                    
                messagebox.showinfo("분석 완료", summary)
                source_info = {
                    'type': 'file',
                    'path': ref_file
                }
                self.auto_save_reference_analysis(analysis_result, source_info, target_lang)
                self.update_status("자동 분석 완료")
            else:
                messagebox.showerror("분석 실패", "데이터 분석에 실패했습니다.")
                
        except Exception as e:
            messagebox.showerror("오류", f"자동 분석 중 오류: {e}")
            import traceback
            traceback.print_exc()


    def select_analysis_language_flexible(self, file_path):
        """유연한 언어 선택 (여러 skiprows 시도)"""
        available_langs = []
        working_skiprows = 0
        
        # 여러 skiprows 값으로 시도하여 언어 컬럼 찾기
        for skiprows in [3, 4, 5]:
            try:
                df = pd.read_excel(file_path, skiprows=skiprows, nrows=1)
                
                # 언어 컬럼 찾기
                langs = []
                for col in df.columns:
                    col_str = str(col).strip().upper()
                    if col_str in ['EN', 'CN', 'TW', 'TH', 'PT', 'ES', 'DE', 'FR', 'JP']:
                        langs.append(col_str)
                
                if langs:
                    available_langs = langs
                    working_skiprows = skiprows
                    break
                    
            except Exception as e:
                print(f"skiprows={skiprows} 언어 감지 실패: {e}")
                continue
        
        if not available_langs:
            messagebox.showerror("언어 없음", "번역 가능한 언어 컬럼을 찾을 수 없습니다.")
            return None
        
        # skiprows 저장
        self.last_skiprows = working_skiprows
        
        print(f"언어 감지 성공: skiprows={working_skiprows}, 언어={available_langs}")
        
        # 언어 선택 다이얼로그
        dialog = LanguageSelectionDialog(self.root, available_langs, "분석할 언어 선택")
        self.root.wait_window(dialog.top)
        
        return dialog.selected_lang if hasattr(dialog, 'selected_lang') else None


    def select_analysis_language(self, file_path):
        """분석할 언어 선택 다이얼로그 (간소화된 버전)"""
        try:
            # skiprows=3으로 고정 (파일 구조가 확인됨)
            df = pd.read_excel(file_path, skiprows=3, nrows=1)
            
            # 언어 컬럼 찾기
            available_langs = []
            for col in df.columns:
                col_str = str(col).strip().upper()
                if col_str in ['EN', 'CN', 'TW', 'TH', 'PT', 'ES', 'DE', 'FR', 'JP']:
                    available_langs.append(col_str)
            
            if not available_langs:
                messagebox.showerror("언어 없음", f"번역 가능한 언어 컬럼이 없습니다.\n발견된 컬럼: {list(df.columns)}")
                return None
            
            # skiprows 저장
            self.last_skiprows = 3
            
            # 언어 선택 다이얼로그
            dialog = LanguageSelectionDialog(self.root, available_langs, "분석할 언어 선택")
            self.root.wait_window(dialog.top)
            
            return dialog.selected_lang if hasattr(dialog, 'selected_lang') else None
            
        except Exception as e:
            messagebox.showerror("오류", f"파일 분석 중 오류: {e}")
            return None
        

    def debug_file_structure_detailed(self):
        """레퍼런스 파일 구조 상세 디버깅"""
        ref_file = self.ref_file_var.get()
        if not ref_file:
            messagebox.showwarning("파일 없음", "레퍼런스 파일을 먼저 선택하세요.")
            return
        
        try:
            debug_info = "=== 파일 구조 상세 분석 ===\n\n"
            
            # 여러 skiprows 옵션으로 시도
            for skip_rows in range(6):  # 0~5까지 시도
                try:
                    df = pd.read_excel(ref_file, skiprows=skip_rows, nrows=10)
                    
                    debug_info += f"📊 skiprows={skip_rows} 결과:\n"
                    debug_info += f"   데이터프레임 크기: {df.shape}\n"
                    debug_info += f"   컬럼 개수: {len(df.columns)}\n\n"
                    
                    # 컬럼명 상세 분석
                    debug_info += f"   📋 컬럼명 목록:\n"
                    for i, col in enumerate(df.columns):
                        debug_info += f"      [{i}] '{col}' (타입: {type(col).__name__})\n"
                    
                    # 필수 컬럼 검사
                    required_columns = ['KR', '#화자']
                    found_required = []
                    for req_col in required_columns:
                        matches = [col for col in df.columns if str(col).strip() == req_col]
                        if matches:
                            found_required.append(req_col)
                    
                    debug_info += f"\n   ✅ 발견된 필수 컬럼: {found_required}\n"
                    
                    # 언어 컬럼 검사
                    lang_columns = []
                    for col in df.columns:
                        col_str = str(col).strip().upper()
                        if col_str in ['EN', 'CN', 'TW', 'TH', 'PT', 'ES', 'DE', 'FR', 'JP']:
                            lang_columns.append(col_str)
                    
                    debug_info += f"   🌍 발견된 언어 컬럼: {lang_columns}\n"
                    
                    # 샘플 데이터
                    if len(df) > 0:
                        debug_info += f"\n   📝 첫 번째 행 데이터:\n"
                        for col in df.columns:
                            value = df[col].iloc[0] if len(df) > 0 else "N/A"
                            debug_info += f"      {col}: '{value}'\n"
                    
                    debug_info += "\n" + "="*50 + "\n\n"
                    
                except Exception as e:
                    debug_info += f"❌ skiprows={skip_rows} 실패: {e}\n\n"
                    continue
            
            # 결과를 새 창에 표시
            self.show_debug_result(debug_info)
            
        except Exception as e:
            messagebox.showerror("디버깅 오류", f"파일 분석 중 오류: {e}")

    def show_debug_result(self, debug_info):
        """디버깅 결과를 새 창에 표시"""
        result_dialog = tk.Toplevel(self.root)
        result_dialog.title("📊 파일 구조 상세 분석 결과")
        result_dialog.geometry("900x700")
        result_dialog.transient(self.root)
        
        main_frame = ttk.Frame(result_dialog, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # 텍스트 위젯 (스크롤 가능)
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill="both", expand=True)
        
        text_widget = tk.Text(text_frame, wrap="word", font=("Consolas", 9))
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        text_widget.insert("1.0", debug_info)
        text_widget.config(state="disabled")
        
        # 버튼 프레임
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=10)
        
        ttk.Button(button_frame, text="🔧 수동 컬럼 매핑", command=lambda: self.manual_column_mapping_dialog()).pack(side="left")
        ttk.Button(button_frame, text="📋 클립보드 복사", command=lambda: self.copy_to_clipboard(debug_info)).pack(side="left", padx=5)
        ttk.Button(button_frame, text="닫기", command=result_dialog.destroy).pack(side="right")

    def copy_to_clipboard(self, text):
        """텍스트를 클립보드에 복사"""
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            messagebox.showinfo("복사 완료", "분석 결과가 클립보드에 복사되었습니다.")
        except Exception as e:
            messagebox.showerror("복사 실패", f"클립보드 복사 중 오류: {e}")
            

    
    def manual_column_mapping_dialog(self):
        """간단한 수동 컬럼 매핑 다이얼로그"""
        ref_file = self.ref_file_var.get()
        if not ref_file:
            messagebox.showwarning("파일 없음", "레퍼런스 파일이 선택되지 않았습니다.")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("🔧 수동 헤더 설정")
        dialog.geometry("500x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        main_frame = ttk.Frame(dialog, padding="15")
        main_frame.pack(fill="both", expand=True)
        
        ttk.Label(main_frame, text="헤더가 있는 행 번호를 지정하세요.", font=("맑은 고딕", 10, "bold")).pack(pady=10)
        
        # skiprows 설정
        skiprows_frame = ttk.LabelFrame(main_frame, text="헤더 행 설정")
        skiprows_frame.pack(fill="x", pady=10)
        
        info_frame = ttk.Frame(skiprows_frame)
        info_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(info_frame, text="헤더 행 번호:").grid(row=0, column=0, sticky="w")
        skiprows_var = tk.IntVar(value=4)  # 기본값 4행
        skiprows_spin = ttk.Spinbox(info_frame, from_=1, to=10, textvariable=skiprows_var, width=10)
        skiprows_spin.grid(row=0, column=1, padx=5)
        ttk.Label(info_frame, text="행").grid(row=0, column=2, sticky="w")
        
        ttk.Label(info_frame, text="(예: 4행에 헤더가 있으면 4 입력)").grid(row=1, column=0, columnspan=3, sticky="w", pady=2)
        
        # 미리보기
        preview_frame = ttk.LabelFrame(main_frame, text="미리보기")
        preview_frame.pack(fill="both", expand=True, pady=5)
        
        preview_text = tk.Text(preview_frame, height=8, wrap="word", font=("Consolas", 9))
        preview_scroll = ttk.Scrollbar(preview_frame, orient="vertical", command=preview_text.yview)
        preview_text.configure(yscrollcommand=preview_scroll.set)
        
        preview_text.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        preview_scroll.pack(side="right", fill="y", pady=5)
        
        def update_preview():
            try:
                skiprows = skiprows_var.get() - 1  # 사용자는 1부터, pandas는 0부터
                df = pd.read_excel(ref_file, skiprows=skiprows, nrows=3)
                
                preview_info = f"헤더 행: {skiprows_var.get()}\n"
                preview_info += f"컬럼 개수: {len(df.columns)}\n\n"
                preview_info += "발견된 컬럼:\n"
                
                for i, col in enumerate(df.columns):
                    preview_info += f"{i+1:2d}. {col}\n"
                
                # 필수 컬럼 체크
                required = ['STRING_ID', '#화자', 'KR']
                found_required = [col for col in required if col in df.columns]
                missing_required = [col for col in required if col not in df.columns]
                
                preview_info += f"\n✅ 발견된 필수 컬럼: {found_required}\n"
                if missing_required:
                    preview_info += f"❌ 누락된 필수 컬럼: {missing_required}\n"
                
                # 언어 컬럼 체크
                lang_cols = [col for col in df.columns if col in ['EN', 'CN', 'TW', 'TH', 'PT', 'ES', 'DE', 'FR', 'JP']]
                preview_info += f"🌍 언어 컬럼: {lang_cols}\n"
                
                preview_text.delete("1.0", "end")
                preview_text.insert("1.0", preview_info)
                
            except Exception as e:
                preview_text.delete("1.0", "end")
                preview_text.insert("1.0", f"미리보기 오류: {e}")
        
        # 초기 미리보기
        update_preview()
        
        # skiprows 변경 시 미리보기 업데이트
        skiprows_spin.configure(command=update_preview)
        
        # 버튼
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=10)
        
        ttk.Button(button_frame, text="🔄 미리보기 새로고침", command=update_preview).pack(side="left")
        
        def apply_and_analyze():
            try:
                skiprows = skiprows_var.get() - 1
                df = pd.read_excel(ref_file, skiprows=skiprows, nrows=1)
                
                # 필수 컬럼 확인
                required = ['STRING_ID', '#화자', 'KR']
                missing = [col for col in required if col not in df.columns]
                
                if missing:
                    messagebox.showerror("필수 컬럼 없음", f"다음 필수 컬럼이 없습니다: {missing}")
                    return
                
                dialog.destroy()
                
                # 분석 실행
                self.analyze_with_skiprows(ref_file, skiprows)
                
            except Exception as e:
                messagebox.showerror("오류", f"분석 실행 오류: {e}")
        
        ttk.Button(button_frame, text="취소", command=dialog.destroy).pack(side="right", padx=5)
        ttk.Button(button_frame, text="✅ 분석 실행", command=apply_and_analyze).pack(side="right")

    def analyze_with_skiprows(self, file_path, skiprows):
        """지정된 skiprows로 분석 실행"""
        try:
            # 언어 선택
            df = pd.read_excel(file_path, skiprows=skiprows, nrows=1)
            available_langs = [col for col in df.columns if col in ['EN', 'CN', 'TW', 'TH', 'PT', 'ES', 'DE', 'FR', 'JP']]
            
            if not available_langs:
                messagebox.showerror("언어 없음", "번역 가능한 언어 컬럼이 없습니다.")
                return
            
            if len(available_langs) == 1:
                target_lang = available_langs[0]
            else:
                dialog = LanguageSelectionDialog(self.root, available_langs, "분석할 언어 선택")
                self.root.wait_window(dialog.top)
                target_lang = dialog.selected_lang if hasattr(dialog, 'selected_lang') else None
                
                if not target_lang:
                    return
            
            # 분석 실행
            if not self.scenario_manager:
                from scenario_manager import ScenarioTranslationManager
                self.scenario_manager = ScenarioTranslationManager(self.translation_db_path)
            
            self.update_status(f"{target_lang} 언어로 분석 중...")
            analysis_result = self.scenario_manager.analyze_reference_data(file_path, target_lang, skiprows)
            
            if analysis_result:
                self.update_speaker_list()
                summary = f"✅ 수동 설정 분석 완료!\n\n"
                summary += f"화자 수: {len(analysis_result)}명\n"
                summary += f"총 문장 수: {sum(data['total_sentences'] for data in analysis_result.values())}개\n\n"
                
                for speaker, data in analysis_result.items():
                    lang_count = data['languages'].get(target_lang, {}).get('count', 0)
                    summary += f"• {speaker}: {lang_count}개 {target_lang} 문장\n"
                    
                messagebox.showinfo("분석 완료", summary)
                
                source_info = {
                    'type': 'file', 
                    'path': file_path
                }
                self.auto_save_reference_analysis(analysis_result, source_info, target_lang)
                
                self.update_status("수동 설정 분석 완료")
            else:
                messagebox.showerror("분석 실패", "데이터 분석에 실패했습니다.")
                
        except Exception as e:
            messagebox.showerror("오류", f"분석 중 오류: {e}")
            import traceback
            traceback.print_exc()


    def find_language_columns(self, columns):
        """컬럼명에서 언어 코드 찾기 (유연한 매칭)"""
        available_langs = []
        
        for col in columns:
            col_str = str(col).strip().upper()
            
            # 정확한 매칭
            if col_str in ['EN', 'CN', 'TW', 'TH', 'PT', 'ES', 'DE', 'FR', 'JP']:
                available_langs.append(col_str)
            # 부분 매칭 (예: "English", "EN_US" 등)
            elif any(lang in col_str for lang in ['EN', 'ENGLISH']):
                available_langs.append('EN')
            elif any(lang in col_str for lang in ['CN', 'CHINESE', 'ZH_CN']):
                available_langs.append('CN')
            elif any(lang in col_str for lang in ['TW', 'ZH_TW', 'TRADITIONAL']):
                available_langs.append('TW')
            # 필요에 따라 다른 언어도 추가
        
        return list(set(available_langs))  # 중복 제거

    def show_manual_column_mapping(self, file_path):
        """수동 컬럼 매핑 다이얼로그"""
        try:
            # 첫 5행 정도 읽어서 사용자에게 보여주기
            df = pd.read_excel(file_path, nrows=5)
            
            dialog = tk.Toplevel(self.root)
            dialog.title("수동 컬럼 매핑")
            dialog.geometry("600x400")
            dialog.transient(self.root)
            dialog.grab_set()
            
            ttk.Label(dialog, text="파일의 컬럼 구조를 확인하고 언어 컬럼을 수동으로 지정하세요.").pack(pady=10)
            
            # 컬럼 정보 표시
            info_frame = ttk.LabelFrame(dialog, text="발견된 컬럼")
            info_frame.pack(fill="both", expand=True, padx=10, pady=5)
            
            text_widget = tk.Text(info_frame, height=10, wrap="word")
            text_widget.pack(fill="both", expand=True, padx=5, pady=5)
            
            column_info = f"컬럼 목록: {list(df.columns)}\n\n"
            column_info += "첫 번째 행 데이터:\n"
            for col in df.columns:
                column_info += f"{col}: {df[col].iloc[0] if len(df) > 0 else 'N/A'}\n"
            
            text_widget.insert("1.0", column_info)
            text_widget.config(state="disabled")
            
            # 수동 매핑 프레임
            mapping_frame = ttk.LabelFrame(dialog, text="언어 컬럼 매핑")
            mapping_frame.pack(fill="x", padx=10, pady=5)
            
            ttk.Label(mapping_frame, text="영어 컬럼:").grid(row=0, column=0, sticky="w", padx=5)
            en_combo = ttk.Combobox(mapping_frame, values=list(df.columns), state="readonly")
            en_combo.grid(row=0, column=1, sticky="ew", padx=5)
            
            mapping_frame.grid_columnconfigure(1, weight=1)
            
            # 버튼
            button_frame = ttk.Frame(dialog)
            button_frame.pack(fill="x", padx=10, pady=5)
            
            def apply_mapping():
                en_col = en_combo.get()
                if en_col:
                    # 임시로 EN 컬럼 매핑 저장
                    self.manual_column_mapping = {'EN': en_col}
                    dialog.destroy()
                else:
                    messagebox.showwarning("선택 오류", "최소한 영어 컬럼은 선택해야 합니다.")
            
            ttk.Button(button_frame, text="적용", command=apply_mapping).pack(side="right", padx=5)
            ttk.Button(button_frame, text="취소", command=dialog.destroy).pack(side="right")
            
        except Exception as e:
            messagebox.showerror("오류", f"수동 매핑 다이얼로그 오류: {e}")


    def add_speaker(self):
        """새 화자 추가 (안전한 초기화)"""
        try:
            self.ensure_scenario_manager()
            if not self.scenario_manager:
                messagebox.showerror("오류", "시나리오 매니저가 초기화되지 않았습니다.")
                return
                
            from dialogs.speaker_dialog import SpeakerEditDialog
            SpeakerEditDialog(self.root, None, self.on_speaker_saved)
        except Exception as e:
            messagebox.showerror("오류", f"화자 추가 중 오류: {e}")

    def edit_speaker(self):
        """선택된 화자 편집 (안전한 초기화)"""
        try:
            if not hasattr(self, 'speaker_tree'):
                return
                
            selected = self.speaker_tree.selection()
            if not selected:
                messagebox.showwarning("선택 없음", "편집할 화자를 선택하세요.")
                return
            
            self.ensure_scenario_manager()
            if not self.scenario_manager:
                messagebox.showerror("오류", "시나리오 매니저가 초기화되지 않았습니다.")
                return
            
            speaker_name = self.speaker_tree.item(selected[0])['values'][0]
            if speaker_name in self.scenario_manager.speakers:
                speaker = self.scenario_manager.speakers[speaker_name]
                from dialogs.speaker_dialog import SpeakerEditDialog
                SpeakerEditDialog(self.root, speaker, self.on_speaker_saved)
        except Exception as e:
            messagebox.showerror("오류", f"화자 편집 중 오류: {e}")

    def delete_speaker(self):
        """선택된 화자 삭제 (안전한 초기화)"""
        try:
            if not hasattr(self, 'speaker_tree'):
                return
                
            selected = self.speaker_tree.selection()
            if not selected:
                messagebox.showwarning("선택 없음", "삭제할 화자를 선택하세요.")
                return
            
            self.ensure_scenario_manager()
            if not self.scenario_manager:
                messagebox.showerror("오류", "시나리오 매니저가 초기화되지 않았습니다.")
                return
            
            speaker_name = self.speaker_tree.item(selected[0])['values'][0]
            if messagebox.askyesno("삭제 확인", f"'{speaker_name}' 화자를 삭제하시겠습니까?"):
                # DB에서 삭제
                import sqlite3
                conn = sqlite3.connect(self.translation_db_path)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM speakers WHERE name = ?", (speaker_name,))
                conn.commit()
                conn.close()
                
                # 메모리에서 삭제
                if speaker_name in self.scenario_manager.speakers:
                    del self.scenario_manager.speakers[speaker_name]
                
                self.refresh_speaker_list()
                self.update_status(f"화자 '{speaker_name}' 삭제됨")
        except Exception as e:
            messagebox.showerror("오류", f"화자 삭제 중 오류: {e}")

    def on_speaker_saved(self, speaker):
        """화자 저장 완료 콜백 (안전한 초기화)"""
        try:
            self.ensure_scenario_manager()
            if self.scenario_manager:
                self.scenario_manager.save_speaker(speaker)
                self.refresh_speaker_list()
                self.update_status(f"화자 '{speaker.name}' 저장됨")
        except Exception as e:
            print(f"화자 저장 콜백 오류: {e}")
            messagebox.showerror("오류", f"화자 저장 중 오류: {e}")
            
                    
    def batch_translate(self, api_needed):
        """배치 API 번역"""
        # DeepL API 사용 (기존 코드 참조)
        try:
            translator = deepl.Translator(DEEPL_API_KEY)
            
            for item in api_needed:
                trans = item["trans"]
                kr_text = trans["KR"]
                
                for lang in item["langs"]:
                    if lang in LANG_CODES:
                        target_lang = LANG_CODES[lang][1]
                        
                        try:
                            result = translator.translate_text(kr_text, target_lang=target_lang)
                            trans["translations"][lang] = result.text
                            trans["method"] = "API"
                        except Exception as e:
                            print(f"번역 오류 ({lang}): {str(e)}")
                            
        except Exception as e:
            print(f"DeepL API 오류: {str(e)}")
            
            
    def update_translation_memory(self):
        """번역 메모리 업데이트 (메모리 동기화 포함)"""
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        
        updated_items = []
        
        for trans in self.pending_translations:
            if trans["translations"]:
                # DB에 저장
                cursor.execute("""
                    INSERT OR REPLACE INTO translation_memory 
                    (kr_text, translations, source, confidence)
                    VALUES (?, ?, ?, ?)
                """, (
                    trans["KR"],
                    json.dumps(trans["translations"]),
                    trans["method"],
                    1.0 if trans["method"] == "완전일치" else 0.8
                ))
                
                # 메모리에도 즉시 반영 (핵심 수정)
                self.translation_memory[trans["KR"]] = trans["translations"].copy()
                updated_items.append(trans["KR"])
                
                # 이력 저장
                cursor.execute("""
                    INSERT INTO translation_history
                    (string_id, kr_text, translations, translation_method, status)
                    VALUES (?, ?, ?, ?, ?)
                """, (
                    trans["STRING_ID"],
                    trans["KR"],
                    json.dumps(trans["translations"]),
                    trans["method"],
                    trans["status"]
                ))
                
        conn.commit()
        conn.close()
        
        if updated_items:
            self.update_status(f"TM 업데이트 완료: {len(updated_items)}개 항목")
            print(f"TM에 추가된 항목들: {updated_items[:5]}...")  # 디버깅용
        

    def save_results(self):
        """(새로운 로직) 번역된 내용을 원본 엑셀 파일에 직접 업데이트합니다."""
        # 1. 대상 파일 경로 가져오기
        file_path = self.file_path_var.get()
        if not file_path:
            messagebox.showwarning("경고", "원본 파일을 먼저 로드하세요.")
            return

        if not messagebox.askyesno("저장 확인", f"번역 결과를 원본 파일에 직접 덮어씁니다.\n파일: {os.path.basename(file_path)}\n\n계속하시겠습니까?"):
            return

        try:
            self.update_status("엑셀 파일 여는 중...")
            # 2. openpyxl로 엑셀 워크북 열기
            workbook = openpyxl.load_workbook(file_path)
            # 우선 첫 번째 시트를 대상으로 작업 (필요시 시트 선택 로직 추가 가능)
            worksheet = workbook.active
            
            # 3. 헤더 위치 및 컬럼 인덱스 동적으로 찾기
            self.update_status("엑셀 구조 분석 중...")
            header_row_index = -1
            string_id_col_index = -1
            lang_col_indices = {}

            # 일반적으로 헤더는 상단 5행 안에 위치
            for r_idx in range(1, 6):
                for c_idx in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=r_idx, column=c_idx).value
                    if isinstance(cell_value, str) and cell_value.strip().upper() == "STRING_ID":
                        header_row_index = r_idx
                        string_id_col_index = c_idx
                        break
                if header_row_index != -1:
                    break
            
            if header_row_index == -1:
                messagebox.showerror("오류", "엑셀 시트에서 'STRING_ID' 헤더를 찾을 수 없습니다.")
                workbook.close()
                return
                
            # 찾은 헤더 행에서 각 언어 컬럼 위치 찾기
            for c_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row_index, column=c_idx).value
                if isinstance(cell_value, str):
                    lang_code = cell_value.strip().upper()
                    if lang_code in self.VISIBLE_LANGS:
                        lang_col_indices[lang_code] = c_idx

            # 4. 번역된 데이터를 빠른 조회를 위해 딕셔너리로 변환
            translations_map = {item['STRING_ID']: item['translations'] for item in self.pending_translations if item.get("translations")}

            # 5. 엑셀 시트를 순회하며 셀 업데이트
            self.update_status("번역 내용 업데이트 중...")
            updated_rows = 0
            # 헤더 다음 행부터 시작
            for r_idx in range(header_row_index + 1, worksheet.max_row + 1):
                string_id = worksheet.cell(row=r_idx, column=string_id_col_index).value
                if not string_id: continue
                
                string_id = str(string_id).strip()
                
                # 번역된 데이터가 있는지 확인
                if string_id in translations_map:
                    row_updated = False
                    translations = translations_map[string_id]
                    
                    # 각 언어에 대해 셀 업데이트
                    for lang, col_idx in lang_col_indices.items():
                        if lang in translations:
                            new_value = translations[lang]
                            worksheet.cell(row=r_idx, column=col_idx).value = new_value
                            row_updated = True
                    
                    if row_updated:
                        updated_rows += 1
            
            # 6. 변경된 내용을 원본 파일에 덮어쓰기
            self.update_status("파일 저장 중...")
            try:
                workbook.save(file_path)
                workbook.close()
                self.update_status(f"결과 저장 완료! 총 {updated_rows}개 행의 번역이 적용되었습니다.")
                messagebox.showinfo("저장 완료", f"'{os.path.basename(file_path)}' 파일에 번역 내용이 성공적으로 적용되었습니다.\n(총 {updated_rows}개 행 업데이트)")
            except PermissionError:
                workbook.close()
                self.update_status("오류: 파일을 다른 프로그램에서 사용 중입니다.")
                messagebox.showerror("저장 실패", "파일을 저장할 수 없습니다. 엑셀에서 파일이 열려있는지 확인하고, 닫은 후 다시 시도해주세요.")
                
        except Exception as e:
            self.update_status(f"결과 저장 중 오류 발생: {e}")
            messagebox.showerror("저장 오류", f"결과를 저장하는 동안 예기치 않은 오류가 발생했습니다:\n{e}")
            import traceback
            traceback.print_exc()
   
    def search_history(self):
        """번역 이력 검색"""
        search_text = self.history_search_var.get().strip()
        
        # 트리 초기화
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
            
        conn = sqlite3.connect(self.translation_db_path)
        cursor = conn.cursor()
        
        if search_text:
            cursor.execute("""
                SELECT created_at, string_id, kr_text, translation_method, status
                FROM translation_history
                WHERE kr_text LIKE ? OR string_id LIKE ?
                ORDER BY created_at DESC
                LIMIT 100
            """, (f"%{search_text}%", f"%{search_text}%"))
        else:
            cursor.execute("""
                SELECT created_at, string_id, kr_text, translation_method, status
                FROM translation_history
                ORDER BY created_at DESC
                LIMIT 100
            """)
            
        for row in cursor.fetchall():
            self.history_tree.insert("", "end", values=row)
            
        conn.close()
        
    def update_status(self, message):
        """상태 메시지 업데이트 (이모지 포함)"""
        self.status_label.config(text=message)
        self.root.update_idletasks()

    #시나리오 번역 관련
    def prepare_scenario_translation(self):
        """시나리오 번역 준비 - 화자 정보 매핑 및 검증"""
        try:
            if self.scenario_manager:
                available_datasets = self.scenario_manager.get_available_datasets()
                if available_datasets:
                    # 가장 최근 사용된 데이터셋 자동 제안
                    latest_dataset = available_datasets[0]  # last_used DESC로 정렬됨
                    
                    if messagebox.askyesno("레퍼런스 데이터셋 발견", 
                                        f"저장된 레퍼런스 데이터셋을 발견했습니다:\n\n"
                                        f"• 이름: {latest_dataset['name']}\n"
                                        f"• 화자: {latest_dataset['total_speakers']}명\n"
                                        f"• 문장: {latest_dataset['total_sentences']}개\n"
                                        f"• 언어: {latest_dataset['target_language']}\n\n"
                                        f"이 데이터셋을 사용하시겠습니까?"):
                        
                        # 데이터셋 로드
                        loaded_data = self.scenario_manager.load_reference_dataset(latest_dataset['name'])
                        if loaded_data:
                            self.update_status(f"레퍼런스 데이터셋 '{latest_dataset['name']}' 로드됨")
                            messagebox.showinfo("데이터셋 로드 완료", 
                                            f"저장된 레퍼런스 데이터가 활성화되었습니다.\n"
                                            f"기존 화자 정보와 번역 패턴을 사용합니다.")
            
            file_path = self.file_path_var.get()
            
            if not file_path:
                raise ValueError("번역 대상 파일이 선택되지 않았습니다.")
            
            # 현재 파일에서 화자 정보 추출
            df = pd.read_excel(file_path, skiprows=3)
            
            if '#화자' not in df.columns:
                raise ValueError("번역 파일에 '#화자' 컬럼이 없습니다.")
            
            # 화자별 문장 수 집계
            speaker_stats = {}
            unknown_speakers = set()
            
            for _, row in df.iterrows():
                if pd.isna(row.get('STRING_ID')) or pd.isna(row.get('#화자')):
                    continue
                    
                speaker = str(row['#화자']).strip()
                string_id = str(row['STRING_ID'])
                
                if speaker not in speaker_stats:
                    speaker_stats[speaker] = []
                speaker_stats[speaker].append(string_id)
                
                # 등록되지 않은 화자 체크
                if speaker not in self.scenario_manager.speakers:
                    unknown_speakers.add(speaker)
            
            # 알 수 없는 화자가 있으면 사용자에게 알림
            if unknown_speakers:
                self.handle_unknown_speakers(unknown_speakers, speaker_stats)
            
            return speaker_stats
            
        except Exception as e:
            messagebox.showerror("준비 오류", f"시나리오 번역 준비 중 오류 발생: {e}")
            return None

    def handle_unknown_speakers(self, unknown_speakers, speaker_stats):
        """알 수 없는 화자 처리"""
        unknown_list = list(unknown_speakers)
        
        message = f"등록되지 않은 화자가 {len(unknown_list)}명 발견되었습니다:\n\n"
        for speaker in unknown_list[:5]:  # 최대 5명만 표시
            count = len(speaker_stats.get(speaker, []))
            message += f"• {speaker}: {count}개 문장\n"
        
        if len(unknown_list) > 5:
            message += f"... 외 {len(unknown_list) - 5}명\n"
        
        message += "\n어떻게 처리하시겠습니까?"
        
        # 선택 다이얼로그
        choice = messagebox.askyesnocancel(
            "알 수 없는 화자 발견", 
            message + "\n\n예: 자동으로 기본 프로필 생성\n아니오: 수동으로 화자 추가\n취소: 번역 중단"
        )
        
        if choice is True:  # 자동 생성
            self.auto_create_speaker_profiles(unknown_speakers)
        elif choice is False:  # 수동 추가
            self.manual_add_speakers(unknown_speakers)
        else:  # 취소
            raise ValueError("사용자가 번역을 취소했습니다.")

    def auto_create_speaker_profiles(self, unknown_speakers):
        """알 수 없는 화자들의 기본 프로필 자동 생성"""
        from scenario_manager import SpeakerProfile
        
        for speaker_name in unknown_speakers:
            # 기본 프로필 생성 (이름 기반 간단 추론)
            default_profile = SpeakerProfile(
                name=speaker_name,
                gender=self.infer_gender_from_name(speaker_name),
                tone="보통",
                style="일반적인 말투, 자동 생성된 프로필",
                examples=[]
            )
            
            self.scenario_manager.save_speaker(default_profile)
            
        self.update_speaker_list()
        self.update_status(f"{len(unknown_speakers)}개 화자 프로필이 자동 생성되었습니다.")

    def infer_gender_from_name(self, name):
        """이름으로부터 성별 간단 추론"""
        name_lower = name.lower()
        
        # 간단한 키워드 기반 추론
        male_keywords = ['king', 'prince', 'duke', 'sir', 'mr', 'lord', 'knight', '왕', '왕자', '공작']
        female_keywords = ['queen', 'princess', 'duchess', 'lady', 'ms', 'mrs', 'miss', '여왕', '공주', '공작부인']
        
        for keyword in male_keywords:
            if keyword in name_lower:
                return "남성"
        
        for keyword in female_keywords:
            if keyword in name_lower:
                return "여성"
        
        return "중성"

    def manual_add_speakers(self, unknown_speakers):
        """수동으로 화자 추가 안내"""
        messagebox.showinfo(
            "수동 추가 안내", 
            f"화자 관리 탭에서 다음 화자들을 추가해주세요:\n\n" + 
            "\n".join(f"• {speaker}" for speaker in list(unknown_speakers)[:10]) +
            "\n\n추가 완료 후 번역을 다시 실행하세요."
        )
        raise ValueError("화자를 수동으로 추가한 후 다시 시도하세요.")


# analysis_btn_frame에 버튼 추가


    def show_latest_debug_log(self):
        """최신 디버그 로그 표시"""
        try:
            import os
            import glob
            
            log_dir = "debug_logs"
            if not os.path.exists(log_dir):
                messagebox.showinfo("로그 없음", "디버그 로그 폴더가 없습니다.")
                return
            
            # 가장 최신 로그 파일 찾기
            log_files = glob.glob(os.path.join(log_dir, "analysis_debug_*.txt"))
            if not log_files:
                messagebox.showinfo("로그 없음", "디버그 로그 파일이 없습니다.")
                return
            
            latest_log = max(log_files, key=os.path.getctime)
            
            # 로그 내용 읽기
            with open(latest_log, 'r', encoding='utf-8') as f:
                log_content = f.read()
            
            # 로그 뷰어 창 생성
            log_window = tk.Toplevel(self.root)
            log_window.title(f"📜 디버그 로그 - {os.path.basename(latest_log)}")
            log_window.geometry("1000x700")
            log_window.transient(self.root)
            
            main_frame = ttk.Frame(log_window, padding="10")
            main_frame.pack(fill="both", expand=True)
            
            # 텍스트 위젯
            text_frame = ttk.Frame(main_frame)
            text_frame.pack(fill="both", expand=True)
            
            text_widget = tk.Text(text_frame, wrap="word", font=("Consolas", 9))
            scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=text_widget.yview)
            text_widget.configure(yscrollcommand=scrollbar.set)
            
            text_widget.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            text_widget.insert("1.0", log_content)
            text_widget.config(state="disabled")
            
            # 버튼
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(fill="x", pady=10)
            
            ttk.Button(button_frame, text="📂 로그 폴더 열기", 
                    command=lambda: os.startfile(log_dir)).pack(side="left")
            ttk.Button(button_frame, text="📋 클립보드 복사", 
                    command=lambda: self.copy_to_clipboard(log_content)).pack(side="left", padx=5)
            ttk.Button(button_frame, text="닫기", command=log_window.destroy).pack(side="right")
            
        except Exception as e:
            messagebox.showerror("오류", f"디버그 로그 표시 중 오류: {e}")

    def setup_reference_persistence(self):
        """레퍼런스 데이터 영구 저장 시스템 설정 (안전한 초기화)"""
        try:
            self.ensure_scenario_manager()
            if self.scenario_manager and hasattr(self.scenario_manager, 'init_reference_tables'):
                self.scenario_manager.init_reference_tables()
            else:
                print("경고: 시나리오 매니저가 초기화되지 않아 레퍼런스 테이블 초기화를 건너뜁니다.")
        except Exception as e:
            print(f"레퍼런스 영구 저장 시스템 설정 오류: {e}")

    def auto_save_reference_analysis(self, analysis_result, source_info, target_language):
        """분석 완료 후 자동 저장 제안"""
        if not analysis_result:
            return
        
        # 자동 이름 생성
        from datetime import datetime
        auto_name = f"REF_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        # 사용자에게 저장 여부 확인
        save_dialog = tk.Toplevel(self.root)
        save_dialog.title("레퍼런스 데이터 저장")
        save_dialog.geometry("500x300")
        save_dialog.transient(self.root)
        save_dialog.grab_set()
        
        main_frame = ttk.Frame(save_dialog, padding="20")
        main_frame.pack(fill="both", expand=True)
        
        ttk.Label(main_frame, text="분석 완료된 레퍼런스 데이터를 저장하시겠습니까?", 
                font=("맑은 고딕", 10, "bold")).pack(pady=10)
        
        info_text = f"""• 화자 수: {len(analysis_result)}명
    • 총 문장 수: {sum(data['total_sentences'] for data in analysis_result.values())}개
    • 분석 언어: {target_language}
    • 소스: {source_info['path']}

    저장하면 다음 번역 시 레퍼런스를 다시 선택하지 않아도 됩니다."""
        
        ttk.Label(main_frame, text=info_text, justify="left").pack(pady=10)
        
        # 데이터셋 이름 입력
        name_frame = ttk.Frame(main_frame)
        name_frame.pack(fill="x", pady=10)
        
        ttk.Label(name_frame, text="데이터셋 이름:").pack(side="left")
        name_var = tk.StringVar(value=auto_name)
        ttk.Entry(name_frame, textvariable=name_var, width=30).pack(side="left", padx=5, fill="x", expand=True)
        
        result = {'save': False, 'name': ''}
        
        def save_and_close():
            result['save'] = True
            result['name'] = name_var.get().strip()
            save_dialog.destroy()
        
        def cancel_and_close():
            save_dialog.destroy()
        
        # 버튼
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=10)
        
        ttk.Button(button_frame, text="저장", command=save_and_close).pack(side="right", padx=5)
        ttk.Button(button_frame, text="저장하지 않음", command=cancel_and_close).pack(side="right")
        
        # 다이얼로그 대기
        self.root.wait_window(save_dialog)
        
        # 저장 실행
        if result['save'] and result['name']:
            # 덮어쓰기 확인 콜백 함수 정의
            def confirm_overwrite(message):
                return messagebox.askyesno("기존 데이터셋 발견", message)
            
            success = self.scenario_manager.save_reference_dataset(
                result['name'], source_info, analysis_result, target_language, confirm_overwrite
            )
            if success:
                self.update_status(f"레퍼런스 데이터셋 '{result['name']}' 저장 완료")
                messagebox.showinfo("저장 완료", f"레퍼런스 데이터셋이 성공적으로 저장되었습니다.\n이름: {result['name']}")
            else:
                messagebox.showerror("저장 실패", "레퍼런스 데이터셋 저장에 실패했습니다.")

    def show_reference_dataset_manager(self):
        """레퍼런스 데이터셋 관리 창"""
        if not self.scenario_manager:
            messagebox.showwarning("시나리오 매니저 없음", "시나리오 번역 기능이 초기화되지 않았습니다.")
            return
        
        manager_window = tk.Toplevel(self.root)
        manager_window.title("레퍼런스 데이터셋 관리")
        manager_window.geometry("800x600")
        manager_window.transient(self.root)
        
        main_frame = ttk.Frame(manager_window, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # 상단 버튼
        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill="x", pady=5)
        
        ttk.Button(top_frame, text="🔄 새로고침", command=lambda: refresh_list()).pack(side="left")
        ttk.Button(top_frame, text="🗑️ 삭제", command=lambda: delete_selected()).pack(side="left", padx=5)
        ttk.Button(top_frame, text="📋 불러오기", command=lambda: load_selected()).pack(side="left", padx=5)
        
        # 데이터셋 목록
        columns = ("이름", "설명", "소스", "언어", "화자수", "문장수", "생성일", "최근사용")
        tree = ttk.Treeview(main_frame, columns=columns, show="headings")
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100)
        
        tree.pack(fill="both", expand=True, pady=10)
        
        def refresh_list():
            # 기존 항목 삭제
            for item in tree.get_children():
                tree.delete(item)
            
            # 데이터셋 목록 로드
            datasets = self.scenario_manager.get_available_datasets()
            for dataset in datasets:
                tree.insert("", "end", values=(
                    dataset['name'],
                    dataset['description'][:30] + "..." if len(dataset['description']) > 30 else dataset['description'],
                    dataset['source_type'],
                    dataset['target_language'],
                    dataset['total_speakers'],
                    dataset['total_sentences'],
                    dataset['created_at'][:10],  # 날짜만
                    dataset['last_used'][:10]    # 날짜만
                ))
        
        def delete_selected():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("선택 없음", "삭제할 데이터셋을 선택하세요.")
                return
            
            dataset_name = tree.item(selected[0])['values'][0]
            if messagebox.askyesno("삭제 확인", f"'{dataset_name}' 데이터셋을 삭제하시겠습니까?"):
                # 삭제 실행
                try:
                    conn = sqlite3.connect(self.scenario_manager.db_path)
                    cursor = conn.cursor()
                    cursor.execute("SELECT id FROM reference_datasets WHERE name = ?", (dataset_name,))
                    dataset_id = cursor.fetchone()[0]
                    cursor.execute("DELETE FROM reference_translations WHERE dataset_id = ?", (dataset_id,))
                    cursor.execute("DELETE FROM reference_datasets WHERE id = ?", (dataset_id,))
                    conn.commit()
                    conn.close()
                    refresh_list()
                    messagebox.showinfo("삭제 완료", f"'{dataset_name}' 데이터셋이 삭제되었습니다.")
                except Exception as e:
                    messagebox.showerror("삭제 실패", f"데이터셋 삭제 중 오류: {e}")
        
        def load_selected():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("선택 없음", "불러올 데이터셋을 선택하세요.")
                return
            
            dataset_name = tree.item(selected[0])['values'][0]
            loaded_data = self.scenario_manager.load_reference_dataset(dataset_name)
            
            if loaded_data:
                # 시나리오 번역 활성화 및 설정
                self.scenario_translation_var.set(True)
                self.on_scenario_option_changed()
                
                messagebox.showinfo("불러오기 완료", 
                                f"'{dataset_name}' 데이터셋이 활성화되었습니다.\n"
                                f"시나리오 번역에서 사용할 수 있습니다.")
                manager_window.destroy()
            else:
                messagebox.showerror("불러오기 실패", "데이터셋을 불러오는데 실패했습니다.")
        # 초기 목록 로드
        refresh_list()
        
        ttk.Button(main_frame, text="닫기", command=manager_window.destroy).pack(pady=10)


    def translate_with_enhanced_scenario(self, kr_text, speaker_name):
        """일관성 개선된 시나리오 번역"""
        try:
            if not self.scenario_manager:
                return None
            
            # 1. 향상된 프롬프트 생성 (유사 레퍼런스 기반)
            enhanced_prompt = self.scenario_manager.generate_enhanced_speaker_prompt(
                speaker_name, kr_text, "EN"
            )
            
            # 2. LLM 번역 실행
            translation_result = self.translate_with_llm(kr_text, enhanced_prompt)
            
            if not translation_result:
                return None
            
            # 3. 일관성 검증
            consistency_check = self.scenario_manager.validate_translation_consistency(
                speaker_name, kr_text, translation_result, "EN"
            )
            
            # 4. 일관성이 낮으면 재번역 시도
            if not consistency_check["is_consistent"] and consistency_check["confidence"] < 0.5:
                print(f"일관성 부족으로 재번역 시도: {kr_text[:30]}...")
                
                # 더 강화된 프롬프트로 재시도
                refined_prompt = enhanced_prompt + f"""

    이전 번역이 '{speaker_name}' 캐릭터의 기존 패턴과 일치하지 않습니다.
    다음 점을 개선하여 다시 번역하세요:
    {chr(10).join(['- ' + suggestion for suggestion in consistency_check["suggestions"]])}

    더 일관된 번역:"""
                
                refined_result = self.translate_with_llm(kr_text, refined_prompt)
                if refined_result:
                    translation_result = refined_result
            
            return {
                "translation": translation_result,
                "consistency": consistency_check,
                "method": f"시나리오({speaker_name})" + ("_검증" if not consistency_check["is_consistent"] else "")
            }
            
        except Exception as e:
            print(f"향상된 시나리오 번역 오류: {e}")
            return None

    def get_speaker_for_item_enhanced(self, trans_item, speaker_mapping=None):
        """화자 정보 가져오기 (캐싱 및 오류 처리 강화)"""
        try:
            string_id = trans_item["STRING_ID"]
            
            # 1. 매핑에서 찾기
            if speaker_mapping:
                for speaker, string_ids in speaker_mapping.items():
                    if string_id in string_ids:
                        return speaker
            
            # 2. 원본 파일에서 찾기 (캐싱)
            if not hasattr(self, '_speaker_cache'):
                self._speaker_cache = {}
                try:
                    file_path = self.file_path_var.get()
                    if file_path:
                        df = pd.read_excel(file_path, skiprows=3)
                        if '#화자' in df.columns and 'STRING_ID' in df.columns:
                            for _, row in df.iterrows():
                                if not pd.isna(row['STRING_ID']) and not pd.isna(row['#화자']):
                                    self._speaker_cache[str(row['STRING_ID'])] = str(row['#화자']).strip()
                except Exception as e:
                    print(f"화자 캐시 생성 오류: {e}")
            
            return self._speaker_cache.get(string_id)
            
        except Exception as e:
            print(f"화자 정보 조회 오류: {e}")
            return None

    def fix_selected_conflicts(self, conflict_tree):
        """충돌 리포트 창에서 선택된 항목을 가장 빈도가 높은 번역으로 자동 해결합니다."""
        selected_items = conflict_tree.selection()
        if not selected_items:
            messagebox.showwarning("선택 오류", "해결할 충돌 항목을 리포트 목록에서 선택하세요.", parent=conflict_tree.winfo_toplevel())
            return

        # 여러 개가 선택되어도 첫 번째 항목만 처리 (또는 반복문으로 여러 개 처리 가능)
        selected_item_id = selected_items[0]
        values = conflict_tree.item(selected_item_id, "values")
        # 컬럼 순서: "심각도", "한국어", "TM 번역", ...
        kr_text = values[1] 

        if not messagebox.askyesno("자동 해결 확인", 
                                f"선택된 '{kr_text[:25]}...' 항목을\n가장 빈도가 높은 번역으로 자동 해결하시겠습니까?\n\n이 작업은 되돌릴 수 없습니다.",
                                parent=conflict_tree.winfo_toplevel()):
            return
        
        try:
            self.update_status(f"'{kr_text[:20]}...' 자동 해결 중...")
            conn = sqlite3.connect(self.translation_db_path)
            cursor = conn.cursor()
            
            # DB에서 현재 충돌 정보 가져오기
            cursor.execute("SELECT translations, conflict_info FROM translation_memory WHERE kr_text = ?", (kr_text,))
            result = cursor.fetchone()
            if not result or not result[1]:
                messagebox.showerror("오류", "DB에서 해당 항목의 충돌 정보를 찾을 수 없습니다.", parent=conflict_tree.winfo_toplevel())
                conn.close()
                return
                
            current_translations = json.loads(result[0])
            conflict_info = json.loads(result[1])

            # 가장 빈도가 높은 번역으로 충돌 해결
            for lang, candidates in conflict_info.items():
                if isinstance(candidates, dict) and candidates:
                    most_common_translation = max(candidates, key=candidates.get)
                    current_translations[lang] = most_common_translation
            
            # DB 업데이트
            cursor.execute("""
                UPDATE translation_memory 
                SET translations=?, status='consolidated', conflict_info=NULL 
                WHERE kr_text=?
            """, (json.dumps(current_translations), kr_text))
            
            conn.commit()
            conn.close()

            # UI 업데이트
            conflict_tree.delete(selected_item_id) # 리포트 창에서 해결된 항목 제거
            self.load_conflicts_to_view() # 메인 탭의 충돌 목록 새로고침
            self.load_tm_view() # TM 조회 탭도 새로고침
            
            messagebox.showinfo("해결 완료", f"'{kr_text[:25]}...' 항목이 성공적으로 해결되었습니다.", parent=conflict_tree.winfo_toplevel())
            self.update_status("선택된 충돌 항목 해결 완료.")

        except Exception as e:
            messagebox.showerror("자동 해결 오류", f"오류 발생: {e}", parent=conflict_tree.winfo_toplevel())
            self.update_status(f"자동 해결 오류: {e}")


    def translate_texts_batch(texts_to_translate, progress_callback=None):
        """
        기존 번역 함수를 배치 처리로 개선
        main.py에서 기존 번역 루프를 이것으로 교체
        """
        if not texts_to_translate:
            return {}
        
        print(f"🚀 개선된 배치 번역 시작: {len(texts_to_translate)}개 항목")
        
        # 1. 용어집 일관성 수정 (처음 한 번만)
        if hasattr(glossary_matcher, 'fix_inconsistent_translations'):
            glossary_matcher.fix_inconsistent_translations()
        
        # 2. 텍스트 목록 준비
        texts_list = list(texts_to_translate.keys())
        
        # 3. 배치 번역 실행
        if hasattr(pipeline, 'batch_translate_texts'):
            translated_list = pipeline.batch_translate_texts(
                texts_list, 
                translate_with_llm,  # 기존 함수 사용
                "Translate Korean game content to natural English:"
            )
        else:
            # 기존 방식으로 폴백
            translated_list = []
            for i, text in enumerate(texts_list):
                result = pipeline.translate_and_reassemble(text, translate_with_llm, "Translate:")
                translated_list.append(result)
                
                if progress_callback:
                    progress_callback(i + 1, len(texts_list))
        
        # 4. 결과 매핑
        results = {}
        for i, original_text in enumerate(texts_list):
            if i < len(translated_list):
                results[original_text] = translated_list[i]
            else:
                results[original_text] = original_text
        
        print(f"✅ 배치 번역 완료")
        return results


def run_smart_translation_manager(parent=None):
    """스마트 번역 관리자 실행"""
    if parent:
        # 자식 창으로는 드래그앤드랍 기능을 적용하기 복잡하므로 일단 제외합니다.
        root = tk.Toplevel(parent)
    else:
        root = TkinterDnD.Tk()
        
    app = SmartTranslationManager(root)
    
    if not parent:
        root.mainloop()
    
    return app


if __name__ == "__main__":
    run_smart_translation_manager()