# 파일명: file_handler.py

import os
import json
import pandas as pd
import openpyxl
from tkinter import messagebox # UI 피드백은 여전히 메인에서 처리해야 하지만, 일부 의존성은 남을 수 있음

def check_config_files():
    """
    실행에 필요한 설정 파일(.env, credentials.json)이 있는지 확인하고
    없는 파일의 목록을 반환합니다.
    """
    missing_files = []
    if not os.path.exists('.env'):
        missing_files.append('.env (API 키 설정 파일)')
    if not os.path.exists('credentials.json'):
        missing_files.append('credentials.json (구글 시트 인증 파일)')
    return missing_files

def create_config_templates():
    """
    .env와 credentials.json의 템플릿 파일을 생성합니다.
    """
    # .env 템플릿
    if not os.path.exists('.env'):
        env_template = """DEEPL_API_KEY=여기에_DeepL_API_키를_입력하세요
AZURE_API_KEY=여기에_Azure_API_키를_입력하세요
AZURE_REGION=koreacentral
OPENAI_API_KEY=여기에_OpenAI_API_키를_입력하세요
"""
        with open('.env', 'w', encoding='utf-8') as f:
            f.write(env_template)

    # credentials.json 템플릿
    if not os.path.exists('credentials.json'):
        cred_template = {
            "type": "service_account",
            "project_id": "여기에_프로젝트_ID_입력",
            "private_key_id": "여기에_private_key_id_입력",
            "private_key": "-----BEGIN PRIVATE KEY-----\n여기에_private_key_입력\n-----END PRIVATE KEY-----\n",
            "client_email": "여기에_client_email_입력",
            "client_id": "여기에_client_id_입력",
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
            "client_x509_cert_url": "여기에_client_x509_cert_url_입력"
        }
        with open('credentials.json', 'w', encoding='utf-8') as f:
            json.dump(cred_template, f, indent=2, ensure_ascii=False)

def load_data_from_excel(file_path):
    """
    지정된 엑셀 파일 경로에서 데이터를 읽어 DataFrame으로 반환합니다.
    헤더는 4번째 줄(skiprows=3)에 있다고 가정합니다.
    """
    if not file_path or not os.path.exists(file_path):
        raise FileNotFoundError("엑셀 파일을 찾을 수 없습니다.")

    df = pd.read_excel(file_path, skiprows=3, engine='openpyxl')
    if "STRING_ID" not in df.columns or "KR" not in df.columns:
        raise ValueError("엑셀 파일에 'STRING_ID'와 'KR' 컬럼이 반드시 필요합니다.")
    
    return df

def save_results_to_excel(file_path, pending_translations, visible_langs):
    """
    번역된 내용을 원본 엑셀 파일에 직접 업데이트합니다.
    성공적으로 업데이트된 행의 수를 반환합니다.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active

        # 헤더 위치 및 컬럼 인덱스 동적 찾기
        header_row_index = -1
        string_id_col_index = -1
        lang_col_indices = {}

        for r_idx in range(1, 6):
            for c_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=r_idx, column=c_idx).value
                if isinstance(cell_value, str) and cell_value.strip().upper() == "STRING_ID":
                    header_row_index = r_idx
                    string_id_col_index = c_idx
                    break
            if header_row_index != -1:
                break
        
        if header_row_index == -1:
            raise ValueError("엑셀 시트에서 'STRING_ID' 헤더를 찾을 수 없습니다.")
        
        for c_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=header_row_index, column=c_idx).value
            if isinstance(cell_value, str):
                lang_code = cell_value.strip().upper()
                if lang_code in visible_langs:
                    lang_col_indices[lang_code] = c_idx

        # 번역 데이터를 {STRING_ID: {lang: text}} 맵으로 변환
        translations_map = {item['STRING_ID']: item['translations'] for item in pending_translations if item.get("translations")}
        
        updated_rows = 0
        for r_idx in range(header_row_index + 1, worksheet.max_row + 1):
            string_id_val = worksheet.cell(row=r_idx, column=string_id_col_index).value
            if not string_id_val: continue
            
            string_id = str(string_id_val).strip()
            if string_id in translations_map:
                row_updated = False
                translations = translations_map[string_id]
                for lang, col_idx in lang_col_indices.items():
                    if lang in translations:
                        worksheet.cell(row=r_idx, column=col_idx).value = translations[lang]
                        row_updated = True
                if row_updated:
                    updated_rows += 1
        
        workbook.save(file_path)
        workbook.close()
        return updated_rows

    except PermissionError:
        workbook.close()
        # 오류 메시지는 호출한 쪽에서 처리하도록 예외를 다시 발생시킴
        raise PermissionError("파일 저장 실패: 다른 프로그램에서 사용 중일 수 있습니다.")
    except Exception as e:
        # 다른 모든 예외도 다시 발생시켜 호출한 쪽에서 처리
        raise e