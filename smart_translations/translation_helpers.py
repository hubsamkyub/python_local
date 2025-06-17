import re
import json
from difflib import SequenceMatcher
import requests
import uuid
import deepl
import openai
from collections import Counter
from typing import Dict, List, Tuple, Optional


class TranslationHelper:
    """번역 관련 헬퍼 함수 모음"""
    
    @staticmethod
    def azure_translate(text: str, from_lang: str, to_langs: List[str], 
                       api_key: str, region: str = "koreacentral") -> Dict[str, str]:
        """Azure 번역 API 호출"""
        endpoint = "https://api.cognitive.microsofttranslator.com"
        path = '/translate'
        constructed_url = endpoint + path
        
        params = {
            'api-version': '3.0',
            'from': from_lang,
            'to': to_langs
        }
        
        headers = {
            'Ocp-Apim-Subscription-Key': api_key,
            'Ocp-Apim-Subscription-Region': region,
            'Content-type': 'application/json',
            'X-ClientTraceId': str(uuid.uuid4())
        }
        
        body = [{'text': text}]
        
        try:
            response = requests.post(constructed_url, params=params, 
                                   headers=headers, json=body, timeout=10)
            response.raise_for_status()
            
            result = response.json()
            translations = {}
            
            for translation in result[0]['translations']:
                lang_code = translation['to']
                translations[lang_code] = translation['text']
                
            return translations
            
        except Exception as e:
            print(f"Azure 번역 오류: {str(e)}")
            return {}
    
    @staticmethod
    def deepl_batch_translate(texts: List[str], target_langs: List[str], 
                             api_key: str) -> Dict[str, Dict[str, str]]:
        """DeepL 배치 번역"""
        try:
            translator = deepl.Translator(api_key)
            results = {}
            
            for text in texts:
                results[text] = {}
                
                for lang in target_langs:
                    try:
                        trans_result = translator.translate_text(text, target_lang=lang)
                        results[text][lang] = trans_result.text
                    except Exception as e:
                        print(f"DeepL 번역 오류 ({text[:20]}... -> {lang}): {str(e)}")
                        results[text][lang] = ""
                        
            return results
            
        except Exception as e:
            print(f"DeepL 초기화 오류: {str(e)}")
            return {}


    # translation_helpers.py 파일의 refine_with_llm 함수를 아래 코드로 교체하세요.
    @staticmethod
    def refine_with_llm(text_to_refine: str, target_lang_code: str, custom_prompt: str,
                        api_key: str) -> Optional[str]:
        """LLM(GPT)을 사용하여 번역문 후편집 및 교정 (텍스트 길이에 따라 모델 분기)"""
        if not api_key or "여기에" in api_key:
            print("OpenAI API 키가 설정되지 않았습니다.")
            return None

        openai.api_key = api_key

        # <<< 시작: 텍스트 길이에 따른 모델 선택 로직 >>>
        # 기준이 되는 글자 수를 설정합니다. 이 값은 얼마든지 조정 가능합니다.
        # 한국어는 보통 한 단어가 2~5자, 문장은 10자 이상인 경우가 많습니다.
        CHARACTER_THRESHOLD = 15

        if len(text_to_refine) <= CHARACTER_THRESHOLD:
            model_to_use = "gpt-4o-mini"  # 짧은 텍스트(단어 등)용 모델
        else:
            model_to_use = "gpt-4o"      # 긴 텍스트(문장 등)용 모델

        # 터미널에 어떤 모델이 사용되었는지 출력하여 확인을 돕습니다.
        print(f"길이: {len(text_to_refine)}, 모델: {model_to_use} -> '{text_to_refine[:30]}...' 후편집 중")
        # <<< 종료: 텍스트 길이에 따른 모델 선택 로직 >>>

        # LLM에 전달할 시스템 메시지 (역할 부여)
        system_message = f"""
        You are a professional localizer and post-editor for game text.
        Your task is to refine a machine-translated text into natural, fluent {target_lang_code}.
        Follow the user's custom instructions precisely.
        """

        # 사용자 요청 메시지
        user_message = f"""
        [Custom Instructions]:
        {custom_prompt}

        [Machine-Translated Text to Refine]:
        "{text_to_refine}"

        Please provide only the refined text, without any additional explanations or introductions.
        """

        try:
            response = openai.chat.completions.create(
                model=model_to_use,  # <<< 수정: 선택된 모델 변수 사용 >>>
                messages=[
                    {"role": "system", "content": system_message},
                    {"role": "user", "content": user_message}
                ],
                temperature=0.5,
                max_tokens=2048
            )
            refined_text = response.choices[0].message.content.strip()

            if refined_text.startswith('"') and refined_text.endswith('"'):
                refined_text = refined_text[1:-1]
                
            return refined_text
        except Exception as e:
            print(f"OpenAI API 호출 오류 (모델: {model_to_use}): {str(e)}")
            return None

    
    @staticmethod
    def calculate_similarity(text1: str, text2: str) -> float:
        """두 텍스트의 유사도 계산 (0~1)"""
        # 기본 유사도
        basic_similarity = SequenceMatcher(None, text1, text2).ratio()
        
        # 단어 기반 유사도
        words1 = set(text1.split())
        words2 = set(text2.split())
        
        if not words1 or not words2:
            return basic_similarity
            
        word_similarity = len(words1 & words2) / len(words1 | words2)
        
        # 문자 n-gram 유사도
        ngram_similarity = TranslationHelper._ngram_similarity(text1, text2, n=3)
        
        # 가중 평균
        return (basic_similarity * 0.4 + word_similarity * 0.3 + ngram_similarity * 0.3)
    
    @staticmethod
    def _ngram_similarity(text1: str, text2: str, n: int = 3) -> float:
        """N-gram 기반 유사도"""
        def get_ngrams(text, n):
            return set(text[i:i+n] for i in range(len(text) - n + 1))
            
        if len(text1) < n or len(text2) < n:
            return 0.0
            
        ngrams1 = get_ngrams(text1, n)
        ngrams2 = get_ngrams(text2, n)
        
        if not ngrams1 or not ngrams2:
            return 0.0
            
        return len(ngrams1 & ngrams2) / len(ngrams1 | ngrams2)
    
    @staticmethod
    def find_best_matches(target_text: str, candidates: List[Tuple[str, Dict]], 
                         threshold: float = 0.7, top_k: int = 5) -> List[Dict]:
        """가장 유사한 번역 찾기"""
        matches = []
        
        for candidate_text, translations in candidates:
            similarity = TranslationHelper.calculate_similarity(target_text, candidate_text)
            
            if similarity >= threshold:
                matches.append({
                    'text': candidate_text,
                    'translations': translations,
                    'similarity': similarity
                })
                
        # 유사도 순으로 정렬
        matches.sort(key=lambda x: x['similarity'], reverse=True)
        
        return matches[:top_k]
    
    @staticmethod
    def extract_patterns(texts: List[str]) -> Dict[str, List[str]]:
        """텍스트 패턴 추출 (버튼, 메시지, 다이얼로그 등)"""
        patterns = {
            'button': [],
            'dialog': [],
            'menu': [],
            'message': [],
            'error': [],
            'confirm': [],
            'tooltip': [],
            'placeholder': []
        }
        
        # 패턴 매칭 규칙
        pattern_rules = {
            'button': [r'버튼', r'클릭', r'누르', r'선택'],
            'dialog': [r'대화', r'창', r'팝업'],
            'menu': [r'메뉴', r'옵션', r'설정'],
            'message': [r'메시지', r'알림', r'안내'],
            'error': [r'오류', r'에러', r'실패', r'문제'],
            'confirm': [r'확인', r'예', r'아니오', r'취소'],
            'tooltip': [r'도움말', r'설명', r'툴팁'],
            'placeholder': [r'입력', r'검색', r'여기에']
        }
        
        for text in texts:
            for pattern_type, keywords in pattern_rules.items():
                for keyword in keywords:
                    if keyword in text:
                        patterns[pattern_type].append(text)
                        break
                        
        return patterns
    
    @staticmethod
    def apply_glossary_rules(text: str, glossary: Dict[str, Dict[str, str]], 
                            target_lang: str) -> Optional[str]:
        """용어집 기반 번역 적용"""
        translated = text
        applied_terms = []
        
        # 긴 용어부터 적용 (짧은 용어가 긴 용어의 일부인 경우 방지)
        sorted_terms = sorted(glossary.keys(), key=len, reverse=True)
        
        for kr_term in sorted_terms:
            if kr_term in translated and target_lang in glossary[kr_term]:
                target_term = glossary[kr_term][target_lang]
                translated = translated.replace(kr_term, target_term)
                applied_terms.append(kr_term)
                
        # 변경사항이 있으면 반환
        if applied_terms:
            return translated
        return None
    
    @staticmethod
    def validate_translation(original: str, translated: str, lang: str) -> Dict[str, bool]:
        """번역 유효성 검증"""
        validations = {
            'length_check': True,
            'placeholder_check': True,
            'number_check': True,
            'punctuation_check': True,
            'tag_check': True
        }
        
        # 1. 길이 체크 (너무 짧거나 긴 번역)
        if lang in ['CN', 'TW', 'JP']:  # 아시아 언어는 일반적으로 짧음
            if len(translated) > len(original) * 1.5:
                validations['length_check'] = False
        else:  # 유럽 언어는 일반적으로 김
            if len(translated) > len(original) * 2.5 or len(translated) < len(original) * 0.5:
                validations['length_check'] = False
                
        # 2. 플레이스홀더 체크 ({0}, {1}, %s 등)
        original_placeholders = re.findall(r'\{[0-9]+\}|%[sd]', original)
        translated_placeholders = re.findall(r'\{[0-9]+\}|%[sd]', translated)
        if sorted(original_placeholders) != sorted(translated_placeholders):
            validations['placeholder_check'] = False
            
        # 3. 숫자 일치 체크
        original_numbers = re.findall(r'\d+', original)
        translated_numbers = re.findall(r'\d+', translated)
        if original_numbers != translated_numbers:
            validations['number_check'] = False
            
        # 4. 특수 구두점 체크 (물음표, 느낌표 등)
        if original.count('?') != translated.count('?'):
            validations['punctuation_check'] = False
        if original.count('!') != translated.count('!'):
            validations['punctuation_check'] = False
            
        # 5. HTML/XML 태그 체크
        original_tags = re.findall(r'<[^>]+>', original)
        translated_tags = re.findall(r'<[^>]+>', translated)
        if original_tags != translated_tags:
            validations['tag_check'] = False
            
        return validations
    
    @staticmethod
    def generate_translation_report(translations: List[Dict]) -> Dict:
        """번역 리포트 생성"""
        report = {
            'total': len(translations),
            'by_status': Counter(),
            'by_method': Counter(),
            'by_language': Counter(),
            'issues': [],
            'cn_tw_requests': []
        }
        
        for trans in translations:
            # 상태별 집계
            status = trans.get('status', '').strip('[]')
            report['by_status'][status] += 1
            
            # 번역 방법별 집계
            method = trans.get('method', 'unknown')
            report['by_method'][method] += 1
            
            # 언어별 집계
            for lang in trans.get('translations', {}):
                if trans['translations'][lang]:
                    report['by_language'][lang] += 1
                    
            # CN/TW 번역 요청 수집
            if status in ['신규', '변경'] and ('CN' not in trans['translations'] or 'TW' not in trans['translations']):
                report['cn_tw_requests'].append({
                    'string_id': trans['STRING_ID'],
                    'kr_text': trans['KR'],
                    'status': status
                })
                
            # 검증 이슈 확인
            for lang, text in trans.get('translations', {}).items():
                if text:
                    validations = TranslationHelper.validate_translation(trans['KR'], text, lang)
                    if not all(validations.values()):
                        report['issues'].append({
                            'string_id': trans['STRING_ID'],
                            'kr_text': trans['KR'],
                            'language': lang,
                            'translation': text,
                            'failed_checks': [k for k, v in validations.items() if not v]
                        })
                        
        return report
    
    @staticmethod
    def export_for_publisher(translations: List[Dict], output_path: str):
        """퍼블리셔용 번역 요청서 생성"""
        import pandas as pd
        
        # CN/TW 번역이 필요한 항목만 필터링
        publisher_data = []
        
        for trans in translations:
            if trans['status'] in ['[신규]', '[변경]']:
                publisher_data.append({
                    'STRING_ID': trans['STRING_ID'],
                    'KR': trans['KR'],
                    '상태': trans['status'],
                    'CN': '',
                    'TW': '',
                    '참고_EN': trans['translations'].get('EN', ''),
                    '참고_JP': trans['translations'].get('JP', ''),
                    '비고': ''
                })
                
        if publisher_data:
            df = pd.DataFrame(publisher_data)
            
            # Excel 파일로 저장
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='번역요청', index=False)
                
                # 워크시트 가져오기
                worksheet = writer.sheets['번역요청']
                
                # 스타일 적용
                from openpyxl.styles import PatternFill, Font, Alignment
                
                # 헤더 스타일
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                    
                # 상태별 색상
                for row in range(2, len(publisher_data) + 2):
                    status = worksheet.cell(row=row, column=3).value
                    
                    if "[신규]" in status:
                        worksheet.cell(row=row, column=3).font = Font(color="FF0000", bold=True)
                    elif "[변경]" in status:
                        worksheet.cell(row=row, column=3).font = Font(color="FFA500", bold=True)
                        
                # 컬럼 너비 자동 조정
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                            
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                    
            return len(publisher_data)
            
        return 0