# tools/db_compare_manager.py

import os
import openpyxl
from PyQt5.QtWidgets import QMessageBox
from ..tools.translate.translation_db_manager import TranslationDBManager

class DBCompareManager:
    # __init__에서 config를 인자로 받도록 수정합니다.
    def __init__(self, parent=None, config=None):
        self.parent = parent
        self.db_manager = TranslationDBManager(parent)
        self.config = config

    def compare_databases_and_export(self, db1_name, db2_name):
        conn1 = self.db_manager.get_connection(db1_name)
        conn2 = self.db_manager.get_connection(db2_name)

        if not conn1 or not conn2:
            return

        try:
            c1 = conn1.cursor()
            c2 = conn2.cursor()

            c1.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'")
            tables1 = {row[0] for row in c1.fetchall()}
            c2.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'")
            tables2 = {row[0] for row in c2.fetchall()}
            common_tables = sorted(list(tables1.intersection(tables2)))

            # =====================================================================================
            # 하드코딩된 경로와 파일 이름 대신, self.config에서 값을 읽어옵니다.
            # =====================================================================================
            output_folder = self.config.get('Paths', 'output_folder', fallback='output')
            if not os.path.exists(output_folder):
                os.makedirs(output_folder)
            
            filename = self.config.get('Filenames', 'comparison_output_filename', fallback='db_comparison.xlsx')
            output_path = os.path.join(output_folder, filename)

            wb = openpyxl.Workbook()
            # 기본 시트는 삭제
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            for table in common_tables:
                # 테이블 이름에 부적합한 문자가 있으면 제거/교체
                safe_sheet_name = "".join(c if c.isalnum() else '_' for c in table)[:31]
                ws = wb.create_sheet(title=safe_sheet_name)
                
                c1.execute(f"PRAGMA table_info({table})")
                headers = [info[1] for info in c1.fetchall()]
                ws.append(headers + ["Status"] + headers)

                c1.execute(f"SELECT * FROM {table}")
                data1 = {row[0]: row for row in c1.fetchall()}
                c2.execute(f"SELECT * FROM {table}")
                data2 = {row[0]: row for row in c2.fetchall()}

                all_ids = sorted(list(set(data1.keys()) | set(data2.keys())))

                for id_val in all_ids:
                    row1 = data1.get(id_val, [''] * len(headers))
                    row2 = data2.get(id_val, [''] * len(headers))
                    
                    status = ""
                    if id_val in data1 and id_val in data2:
                        status = "Same" if row1 == row2 else "Different"
                    elif id_val in data1:
                        status = f"Only in {db1_name}"
                    elif id_val in data2:
                        status = f"Only in {db2_name}"
                    
                    ws.append(list(row1) + [status] + list(row2))

            wb.save(output_path)
            QMessageBox.information(self.parent, "완료", f"DB 비교 완료.\n결과가 '{output_path}'에 저장되었습니다.")

        except openpyxl.utils.exceptions.IllegalCharacterError:
            QMessageBox.critical(self.parent, "오류", "시트 이름에 사용할 수 없는 문자가 테이블 이름에 포함되어 있습니다.")
        except Exception as e:
            QMessageBox.critical(self.parent, "오류", f"DB 비교 중 오류 발생: {e}")
        finally:
            if conn1:
                conn1.close()
            if conn2:
                conn2.close()