# 파일명: file_handler.py

import os
import json
import pandas as pd
import openpyxl
from tkinter import messagebox # UI 피드백은 여전히 메인에서 처리해야 하지만, 일부 의존성은 남을 수 있음

# 애플리케이션에 필요한 환경 변수 목록
REQUIRED_ENV_VARS = {
    "OPENAI_API_KEY": "sk-...",
    "AZURE_TRANSLATOR_API_KEY": "YOUR_AZURE_KEY",
    "AZURE_TRANSLATOR_REGION": "YOUR_AZURE_REGION",
    "GEMINI_API_KEY": "AIzaSy...",
}

# .env 템플릿 파일의 내용
ENV_TEMPLATE = """# 스마트 번역 매니저 API 키 설정 파일
# 아래 각 항목에 자신의 API 키를 입력하고 저장하세요.
# 키를 입력한 후에는 프로그램을 다시 시작해야 합니다.
# '#'으로 시작하는 줄은 주석으로, 프로그램에 영향을 주지 않습니다.

"""
def check_and_create_env_file():
    """
    .env 파일의 존재를 확인하고, 없으면 템플릿 파일을 생성합니다.
    
    Returns:
        bool: 파일이 이미 존재하거나 처리가 필요 없으면 True,
              새 파일이 생성되어 사용자가 수정해야 하면 False를 반환합니다.
    """
    if os.path.exists(".env"):
        # 파일이 이미 존재하면 아무것도 하지 않고 성공(True)을 반환
        return True

    print("INFO: .env 파일이 존재하지 않아 새로 생성합니다.")
    
    # 템플릿 내용 생성
    template_content = ENV_TEMPLATE
    for key, placeholder in REQUIRED_ENV_VARS.items():
        template_content += f"{key}={placeholder}\n"

    try:
        # .env 파일 쓰기
        with open(".env", "w", encoding="utf-8") as f:
            f.write(template_content)
        # 파일을 새로 생성했으므로, 사용자에게 알리고 프로그램을 종료하도록 유도 (False 반환)
        return False
    except IOError as e:
        # 파일 생성 중 오류 발생 시
        messagebox.showerror("파일 생성 오류", f".env 파일을 생성하는 데 실패했습니다:\n{e}")
        return False

# === [추가된 부분 끝] ===

def check_config_files():
    """
    실행에 필요한 설정 파일(.env, credentials.json)이 있는지 확인하고
    없는 파일의 목록을 반환합니다.
    """
    missing_files = []
    if not os.path.exists('.env'):
        missing_files.append('.env (API 키 설정 파일)')
    if not os.path.exists('credentials.json'):
        missing_files.append('credentials.json (구글 시트 인증 파일)')
    return missing_files

def create_config_templates():
    """설정 파일 템플릿을 생성합니다."""
    # .env 파일 생성은 check_and_create_env_file에서 처리하므로 여기서는 호출만 하거나 비워둘 수 있습니다.
    check_and_create_env_file()
    
    # credentials.json은 사용자가 직접 다운로드해야 하므로 템플릿을 만들지 않습니다.
    # 대신, 사용법을 안내하는 텍스트 파일을 만들 수 있습니다.
    creds_readme = """
    이 파일은 구글 시트 연동에 필요한 'credentials.json' 파일에 대한 안내입니다.

    1. 구글 클라우드 플랫폼(GCP)에 접속하여 프로젝트를 생성합니다.
    2. 'API 및 서비스' -> '사용자 인증 정보' 메뉴로 이동합니다.
    3. '+ 사용자 인증 정보 만들기' -> '서비스 계정'을 선택하여 새 서비스 계정을 만듭니다.
    4. 생성된 서비스 계정 이메일을 복사한 뒤, 연동할 구글 시트의 '공유' 설정에 들어가
       이 이메일을 추가하고 '편집자' 또는 '뷰어' 권한을 부여합니다.
    5. 다시 구글 클라우드 콘솔로 돌아와, 생성된 서비스 계정을 클릭하고 '키' 탭으로 이동합니다.
    6. '키 추가' -> '새 키 만들기'를 선택하고, 키 유형은 'JSON'으로 하여 다운로드합니다.
    7. 다운로드한 JSON 파일의 이름을 'credentials.json'으로 변경하고,
       이 프로그램의 실행 파일(.exe)이 있는 폴더에 넣어주세요.
    """
    try:
        with open("credentials.json_안내.txt", "w", encoding="utf-8") as f:
            f.write(creds_readme)
    except IOError:
        pass

def load_data_from_excel(file_path):
    """
    엑셀 파일에서 데이터를 로드하고 유효성을 검사합니다.
    헤더는 4번째 줄에 있다고 가정합니다.
    """
    if not file_path or not os.path.exists(file_path):
        raise FileNotFoundError("엑셀 파일이 선택되지 않았거나 존재하지 않습니다.")

    try:
        df = pd.read_excel(file_path, skiprows=3)
        if "STRING_ID" not in df.columns or "KR" not in df.columns:
            raise ValueError("엑셀 파일에 'STRING_ID'와 'KR' 컬럼이 반드시 필요합니다.")
        return df
    except Exception as e:
        raise ValueError(f"엑셀 파일을 읽는 중 오류가 발생했습니다: {e}")

def save_results_to_excel(file_path, pending_translations, visible_langs):
    """
    번역된 내용을 원본 엑셀 파일에 직접 업데이트합니다.
    성공적으로 업데이트된 행의 수를 반환합니다.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active

        # 헤더 위치 및 컬럼 인덱스 동적 찾기
        header_row_index = -1
        string_id_col_index = -1
        lang_col_indices = {}

        for r_idx in range(1, 6):
            for c_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=r_idx, column=c_idx).value
                if isinstance(cell_value, str) and cell_value.strip().upper() == "STRING_ID":
                    header_row_index = r_idx
                    string_id_col_index = c_idx
                    break
            if header_row_index != -1:
                break
        
        if header_row_index == -1:
            raise ValueError("엑셀 시트에서 'STRING_ID' 헤더를 찾을 수 없습니다.")
        
        for c_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=header_row_index, column=c_idx).value
            if isinstance(cell_value, str):
                lang_code = cell_value.strip().upper()
                if lang_code in visible_langs:
                    lang_col_indices[lang_code] = c_idx

        # 번역 데이터를 {STRING_ID: {lang: text}} 맵으로 변환
        translations_map = {item['STRING_ID']: item['translations'] for item in pending_translations if item.get("translations")}
        
        updated_rows = 0
        for r_idx in range(header_row_index + 1, worksheet.max_row + 1):
            string_id_val = worksheet.cell(row=r_idx, column=string_id_col_index).value
            if not string_id_val: continue
            
            string_id = str(string_id_val).strip()
            if string_id in translations_map:
                row_updated = False
                translations = translations_map[string_id]
                for lang, col_idx in lang_col_indices.items():
                    if lang in translations:
                        worksheet.cell(row=r_idx, column=col_idx).value = translations[lang]
                        row_updated = True
                if row_updated:
                    updated_rows += 1
        
        workbook.save(file_path)
        workbook.close()
        return updated_rows

    except PermissionError:
        workbook.close()
        # 오류 메시지는 호출한 쪽에서 처리하도록 예외를 다시 발생시킴
        raise PermissionError("파일 저장 실패: 다른 프로그램에서 사용 중일 수 있습니다.")
    except Exception as e:
        # 다른 모든 예외도 다시 발생시켜 호출한 쪽에서 처리
        raise e