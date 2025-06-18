import os
import logging
import importlib
import sys
import json
import pandas as pd
import tkinter as tk
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
from tkinter import ttk, messagebox, filedialog, messagebox, Toplevel, Label, simpledialog
import subprocess
import sqlite3
import tempfile
import msvcrt
import multiprocessing
import time
import threading
import hashlib

# 현재 스크립트 파일이 있는 디렉토리 경로 (Excel_Data_Search_46_clean 폴더)
# PyInstaller 실행 파일과 일반 Python 스크립트 모두 지원
if getattr(sys, 'frozen', False):
    # PyInstaller로 빌드된 실행 파일인 경우
    current_dir = os.path.dirname(sys.executable)
else:
    # 일반 Python 스크립트인 경우
    current_dir = os.path.dirname(os.path.abspath(__file__))

# 현재 디렉토리를 모듈 검색 경로에 추가
if current_dir not in sys.path:
    sys.path.append(current_dir)

def get_application_dir():
    """애플리케이션 실행 디렉토리를 반환 (PyInstaller 지원)"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))


# 내부 참조
from utils.string_db_utils import load_or_build_string_db, should_rebuild_string_db
from tools.reward_search_popup import RewardSearchPopup
from tools.string_search_popup import StringSearchPopup
from tools.db_viewer import open_db_viewer_with_excel_match
from tools.column_search_popup import ColumnSearchPopup
from utils.cache_utils import (load_cached_data, update_db_cache, hash_paths, build_table_sheet_index,
                         get_file_mtime, save_cache)
from utils.config_utils import load_config, save_config, CONFIG_FILE, get_preset_paths, load_search_history, save_search_history
from utils.excel_utils import ExcelFileManager
from tools.quest_search_popup import QuestSearchPopup
from tools.npc_creator import NPCCreatorPopup
from tools.translate_tool_main import TranslationAutomationTool
from tools.excel_diff_tool import ExcelDiffTool
from tools.box_search_popup import BoxListPopup
from tools.npc_search_popup import NPCListPopup

# 로깅 설정
logging.basicConfig(
    filename='excel_search_debug.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filemode='w'
)

class LoadingPopup:
    """비동기 안전하게 상태를 업데이트하는 로딩 팝업 클래스"""
    def __init__(self, parent, title="로딩 중...", message="⏳ 준비 중입니다..."):
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.geometry("400x150")
        self.top.grab_set()
        self.label = tk.Label(self.top, text=message, font=("Arial", 12))
        self.label.pack(expand=True, pady=20)
        
        # 진행 상황 표시용
        self.progress_label = tk.Label(self.top, text="0%", font=("Arial", 10))
        self.progress_label.pack()

    def update_message(self, text):
        self.label.config(text=text)
        self.top.update_idletasks()

    def update_progress(self, current, total):
        percent = int((current / total) * 100) if total else 0
        self.progress_label.config(text=f"{percent}% ({current}/{total})")
        self.top.update_idletasks()

    def close(self):
        self.top.destroy()
        

class RibbonMenu(tk.Frame):
    """리본 메뉴 클래스"""
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.config(bg="#f0f0f0", relief="raised", bd=1)
        
        self.tabs = {}
        self.buttons = {}
        self.current_tab = None
        self.tab_frame = None
        self.content_frame = None
        
        self._init_ui()
    
    def _init_ui(self):
        """UI 초기화"""
        # 탭 버튼 영역
        self.tab_frame = tk.Frame(self, bg="#f0f0f0")
        self.tab_frame.pack(side="top", fill="x")
        
        # 컨텐츠 영역
        self.content_frame = tk.Frame(self, bg="#f5f5f5", height=80)
        self.content_frame.pack(side="top", fill="x", pady=(0, 2))
    
    def add_tab(self, name, text):
        """리본 메뉴 탭 추가"""
        # 탭 버튼 생성
        if name in self.tabs:
            return False
        
        tab_btn = tk.Button(
            self.tab_frame, 
            text=text, 
            relief="flat", 
            bg="#f0f0f0", 
            padx=10,
            command=lambda: self.select_tab(name)
        )
        tab_btn.pack(side="left", padx=2)
        
        # 컨텐츠 영역 생성
        content = tk.Frame(self.content_frame, bg="#f5f5f5")
        content.pack_forget()  # 처음에는 숨김
        
        # 저장
        self.tabs[name] = {
            "button": tab_btn,
            "content": content,
            "text": text
        }
        
        # 처음 탭이면 바로 선택
        if len(self.tabs) == 1:
            self.select_tab(name)
        
        return True
    
    def select_tab(self, name):
        """탭 선택"""
        if name not in self.tabs:
            return False
        
        # 이전 탭 비활성화
        if self.current_tab and self.current_tab in self.tabs:
            self.tabs[self.current_tab]["button"].config(
                relief="flat", 
                bg="#f0f0f0"
            )
            self.tabs[self.current_tab]["content"].pack_forget()
        
        # 새 탭 활성화
        self.tabs[name]["button"].config(
            relief="sunken",
            bg="#e0e0e0"
        )
        self.tabs[name]["content"].pack(fill="both", expand=True)
        
        self.current_tab = name
        return True
    
    def add_button(self, tab_name, name, text, icon=None, command=None, row=0, col=0, tooltip=None):
        """리본 메뉴 버튼 추가"""
        if tab_name not in self.tabs:
            return False
        
        content = self.tabs[tab_name]["content"]
        
        # 버튼 위젯 생성
        button_frame = tk.Frame(content, bg="#f5f5f5")
        button_frame.grid(row=row, column=col, padx=5, pady=3)
        
        button = tk.Button(
            button_frame,
            text=text,
            compound="top",  # 아이콘 위, 텍스트 아래
            command=command,
            relief="flat",
            bg="#f5f5f5",
            width=12,
            height=2
        )
        
        # 아이콘 설정 (텍스트로 대체)
        if icon:
            button.config(text=f"{icon}\n{text}")
        
        button.pack(pady=2)
        
        # 툴팁 설정 (개선된 버전)
        if tooltip:
            # 툴팁 객체를 저장할 변수
            tip = None
            
            def show_tooltip(event):
                nonlocal tip
                # 버튼 위치 계산
                x = button.winfo_rootx() + button.winfo_width() + 5
                y = button.winfo_rooty()  # 버튼 옆에 툴팁 표시
                
                # 팁 윈도우 생성
                tip = tk.Toplevel(button)
                tip.wm_overrideredirect(True)
                tip.wm_geometry(f"+{x}+{y}")
                
                label = tk.Label(
                    tip, 
                    text=tooltip, 
                    bg="#ffffcc", 
                    relief="solid", 
                    borderwidth=1,
                    padx=3,
                    pady=2,
                    font=("맑은 고딕", 9)
                )
                label.pack()
                
            def hide_tooltip(event):
                nonlocal tip
                if tip:
                    tip.destroy()
                    tip = None
            
            button.bind("<Enter>", show_tooltip)
            button.bind("<Leave>", hide_tooltip)
        
        # 저장
        if name not in self.buttons:
            self.buttons[name] = {}
        
        self.buttons[name][tab_name] = button
        return True

class ExcelSearchApp:
    def __init__(self, root, data_path=""):
        logging.debug("===== 애플리케이션 초기화 시작 =====")
        # 로깅 설정을 수정하여 더 자세한 로그 기록 (기존 로그 설정 대체)
        logging.basicConfig(
            filename='excel_search_debug.log',
            level=logging.DEBUG,  # 로그 레벨을 DEBUG로 설정
            format='%(asctime)s - %(levelname)s - %(message)s',
            filemode='w'  # 파일을 덮어쓰는 모드
        )
        
        self.root = root
        self.root.title("📊 엑셀 조건 검색기 (통합 UI)")
        self.root.geometry("1600x800")

        # 설정 로드
        self.config = load_config(CONFIG_FILE)
        
        # 경로 정규화 및 처리
        if not data_path:
            data_path = self.config.get("data_path", "")
        
        # 경로 정규화 - 슬래시를 OS에 맞게 변환
        data_path = os.path.normpath(data_path) if data_path else ""
        self.data_path = tk.StringVar(value=data_path)
        
        cache_id = hashlib.md5(data_path.encode()).hexdigest()[:8]
        cache_dir = os.path.join(get_application_dir(), ".cache", cache_id)
        self.db_path = tk.StringVar(value=cache_dir)
        
        # 프리셋 경로 처리
        self.preset_paths = get_preset_paths(self.config)
        self._table_filter_after_id = None

        self.selected_table = tk.StringVar()
        self.status_var = tk.StringVar()
        self.table_list = []
        self.global_search_history = load_search_history()
        self.reward_search_history = load_search_history("reward")

        self.existing_db_files = set()  # DB 캐시용 집합
        self.excel_cache = {}  # 엑셀 캐시 초기화
        self.db_cache = {}  # DB 캐시 초기화
        self.db_schemas = {}  # DB 스키마 캐시
        self.filtered_results = []  # 검색 결과 캐시

        self.build_ui()

        if self.data_path.get():
            self.load_initial_cache()
            
    def build_ui(self):
        # 상단 경로 설정 프레임
        self.build_path_ui()

        # 중간 좌측: 테이블 필터 및 목록
        self.build_table_list_ui()

        # 중간 우측: 컬럼 조건 및 검색 결과
        self.build_search_ui()

        # 하단 상태 표시
        tk.Label(self.root, textvariable=self.status_var).pack(fill="x")

    def build_path_ui(self):
        path_frame = tk.Frame(self.root)
        path_frame.pack(fill="x", padx=10, pady=5)

        # 좌측: 통합된 경로 입력 UI
        left_frame = tk.Frame(path_frame)
        left_frame.pack(side="left", fill="x", expand=True)
        
        tk.Label(left_frame, text="📁 데이터 경로").grid(row=0, column=0)
        tk.Entry(left_frame, textvariable=self.data_path, width=70).grid(row=0, column=1)
        tk.Button(left_frame, text="선택", command=self.select_data_folder).grid(row=0, column=2)
        
        # 우측: 프리셋 관련 버튼들
        right_frame = tk.Frame(path_frame)
        right_frame.pack(side="right", fill="x", padx=(20, 0))
        
        # 프리셋 버튼 프레임
        self.preset_button_frame = tk.Frame(right_frame)
        self.preset_button_frame.pack(side="left", padx=(0, 10))
        
        # 프리셋 저장 버튼
        tk.Button(right_frame, text="📌 프리셋 저장", command=self.save_preset).pack(side="left")
        
        # 프리셋 버튼 생성
        self.build_preset_buttons()

    def build_table_list_ui(self):
        # 상단 프레임
        frame = tk.Frame(self.root)
        frame.pack(fill="x", padx=10, pady=(5, 0))

        # 리본 메뉴 추가
        self.ribbon = RibbonMenu(self.root)
        self.ribbon.pack(fill="x", padx=10, pady=5)
        
        # 기본 탭 추가
        self.ribbon.add_tab("main", "메인")
        self.ribbon.add_tab("search", "검색 도구")
        self.ribbon.add_tab("translation", "번역 도구")
        # self.ribbon.add_tab("advanced", "고급 기능")
        
        # 메인 탭 버튼
        self.ribbon.add_button("main", "refresh", "최신화", icon="🔃", 
                            command=self.refresh_all_cache, row=0, col=0, 
                            tooltip="캐시를 최신 상태로 업데이트합니다")
        # 메인 탭에 전체 재구축 버튼 추가
        self.ribbon.add_button("main", "db_view", "DB 보기", icon="🔎", 
                            command=self.open_db_viewer, row=0, col=1,
                            tooltip="선택한 테이블의 DB를 조회합니다")
        
        self.ribbon.add_button("main", "excel_diff", "엑셀 비교", icon="📊", 
                            command=self.open_excel_diff_tool, row=0, col=2,
                            tooltip="엑셀 파일 데이터 비교 도구")
        
        self.ribbon.add_button("main", "rebuild_all", "전체 재구축", icon="🔄", 
                            command=self.rebuild_all_cache_and_index, row=0, col=3, 
                            tooltip="모든 캐시와 인덱스를 완전히 재구축합니다")


        # 검색 도구 탭 버튼
        self.ribbon.add_button("search", "string_search", "String 검색기", icon="🧾", 
                            command=self.open_string_searcher, row=0, col=0,
                            tooltip="문자열 데이터 검색 도구")
        self.ribbon.add_button("search", "reward_search", "Reward 검색기", icon="🎁", 
                            command=self.open_reward_searcher, row=0, col=1,
                            tooltip="보상 데이터 검색 도구")
        self.ribbon.add_button("search", "box_seracher", "BOX 검색", icon="📝", 
                    command=self.open_box_list, row=0, col=2,
                    tooltip="Box 검색 도구 (베타)")
        # 리본 메뉴에 NPC 검색 버튼 추가 (기존 코드에서 추가)
        self.ribbon.add_button("search", "npc_searcher", "NPC 검색", icon="👤", 
                            command=self.open_npc_search_popup, row=0, col=3,
                            tooltip="NPC 검색 도구 (베타)")
        # self.ribbon.add_button("search", "column_search", "컬럼 검색기", icon="🧩", 
        #                     command=self.open_column_searcher, row=0, col=2,
        #                     tooltip="특정 컬럼 값으로 검색")
        # self.ribbon.add_button("search", "quest_search", "퀘스트 검색", icon="📝", 
        #                     command=self.open_quest_searcher, row=0, col=3,
        #                     tooltip="퀘스트 데이터 검색 도구 (베타)")

        # 번역 도구 탭 버튼
        self.ribbon.add_button("translation", "trans_sync", "CN/TW 동기화", icon="🔁", 
                            command=self.open_translation_sync_tool, row=0, col=0,
                            tooltip="CN/TW 번역 동기화 데이터베이스")
        self.ribbon.add_button("translation", "trans_extract", "번역 파일 추출", icon="🛠️", 
                            command=self.open_translation_automation_tool, row=0, col=1,
                            tooltip="DB 기반 번역 파일 추출 도구")
        self.ribbon.add_button("translation", "unique_text", "고유 텍스트 관리", icon="🛠️", 
                            command=self.open_unique_text_manager, row=0, col=2,
                            tooltip="고유 텍스트 관리 도구")
        
        # 고급 기능 탭 버튼
        # self.ribbon.add_button("advanced", "npc_creator", "NPC 생성기", icon="👤", 
        #                     command=self.open_npc_creator, row=0, col=0,
        #                     tooltip="NPC 생성 도구 (베타)")
        # self.ribbon.add_button("advanced", "relation_analyzer", "관계 분석기", icon="🔗", 
        #                     command=self.open_relationship_analyzer, row=0, col=1,
        #                     tooltip="데이터 관계 분석 도구 (베타)")
        
        # 테이블 필터 UI 부분
        filter_frame = tk.Frame(self.root)
        filter_frame.pack(fill="x", padx=10, pady=(0, 5))
        
        tk.Label(filter_frame, text="테이블 필터:").pack(side="left")
        self.table_filter_var = tk.StringVar()
        self.table_filter_var.trace_add("write", self._on_table_filter_changed)
        tk.Entry(filter_frame, textvariable=self.table_filter_var, width=30).pack(side="left", padx=(0, 10))
        tk.Button(filter_frame, text="초기화", command=self.reset_table_filter).pack(side="left")

        # 테이블 목록 UI
        list_frame = tk.Frame(self.root)
        list_frame.pack(fill="y", side="left", padx=10, pady=5)
        
        tk.Label(list_frame, text="테이블 목록").pack(anchor="w")
        self.table_listbox = tk.Listbox(list_frame, height=20, width=30)
        self.table_listbox.pack(fill="y", expand=True)
        self.table_listbox.bind("<<ListboxSelect>>", self.on_table_selected)
        
        
    def build_search_ui(self):

        search_frame = tk.Frame(self.root)
        search_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # 우측: 컬럼 입력, 검색 버튼, 결과
        right_frame = tk.Frame(search_frame)
        right_frame.pack(side="left", fill="both", expand=True)

        # 컬럼 입력 영역 (스크롤 포함)
        column_container = tk.Frame(right_frame)
        column_container.pack(fill="x", padx=5, pady=(0, 2))  # 위 여백 줄임

        self.column_canvas = tk.Canvas(column_container)
        self.column_canvas.pack(side="left", fill="both", expand=True)

        scrollbar = tk.Scrollbar(column_container, orient="vertical", command=self.column_canvas.yview)
        scrollbar.pack(side="right", fill="y")

        self.column_canvas.configure(yscrollcommand=scrollbar.set)

        self.column_frame = tk.Frame(self.column_canvas)
        self.column_canvas.create_window((0, 0), window=self.column_frame, anchor="nw")

        self.column_frame.bind(
            "<Configure>",
            lambda e: self.column_canvas.configure(scrollregion=self.column_canvas.bbox("all"))
        )

        # 마우스 휠 스크롤 바인딩
        self.column_canvas.bind("<Enter>", lambda e: self._bind_mousewheel(self.column_canvas))
        self.column_canvas.bind("<Leave>", lambda e: self._unbind_mousewheel(self.column_canvas))

        self.column_entries = {}

        # 검색 버튼
        self.search_button = tk.Button(right_frame, text="검색", command=self.run_search)
        self.search_button.pack(pady=(3, 5))

        # 검색 결과 필터 입력창과 버튼들을 TreeView 위에 정렬
        filter_frame = tk.Frame(right_frame)
        filter_frame.pack(anchor="ne", padx=5, pady=(0, 2))

        self.result_filter_var = tk.StringVar()
        filter_entry = tk.Entry(filter_frame, textvariable=self.result_filter_var, width=25)
        filter_entry.pack(side=tk.LEFT)

        tk.Button(filter_frame, text="결과 필터", command=self.filter_results).pack(side=tk.LEFT, padx=(5, 0))
        tk.Button(filter_frame, text="필터 초기화", command=self.restore_all_results).pack(side=tk.LEFT, padx=(5, 0))

        # 검색 결과 TreeView와 스크롤바를 담을 프레임
        tree_frame = tk.Frame(right_frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # 수직 스크롤바 생성
        tree_scrollbar = tk.Scrollbar(tree_frame, orient="vertical")
        tree_scrollbar.pack(fill="y", side="right")

        # TreeView 생성 및 스크롤바 연결
        self.tree = ttk.Treeview(tree_frame, columns=("파일", "시트", "결과"), show="headings", yscrollcommand=tree_scrollbar.set)
        tree_scrollbar.config(command=self.tree.yview)
        
        self.tree.heading("파일", text="파일명")
        self.tree.heading("시트", text="시트명")
        self.tree.heading("결과", text="데이터")
        
        # 여기에 컬럼 너비 설정 추가
        self.tree.column("파일", width=150, stretch=False)  # 파일명 컬럼 너비 축소
        self.tree.column("시트", width=150, stretch=False)   # 시트명 컬럼 너비 축소
        self.tree.column("결과", width=350)  # 데이터 컬럼 너비 확장
        
        self.tree.pack(fill="both", expand=True, padx=5, pady=5)
        self.tree.bind("<Double-1>", self.open_excel_from_result)
        self.tree.bind("<<TreeviewSelect>>", self.on_result_select)
        
    def on_result_select(self, event):
        """검색 결과 항목 선택 시 해당 값을 각 컬럼 라벨에 표시"""
        item = self.tree.focus()
        if not item:
            return
        
        values = self.tree.item(item, "values")
        if len(values) < 3:
            return
        
        # 결과 데이터에서 컬럼:값 쌍 추출
        preview = values[2]
        pairs = preview.split(" | ")
        
        # 현재 선택된 테이블 확인
        table_name = self.selected_table.get()
        if not table_name:
            return
        
        # column_value_labels 존재 확인
        if not hasattr(self, 'column_value_labels'):
            self.column_value_labels = {}  # 없으면 초기화
            return
        
        # 모든 라벨 텍스트 초기화
        for label in self.column_value_labels.values():
            label.config(text="")
        
        # 결과에서 컬럼:값 쌍을 파싱하여 해당 라벨에 값 표시
        for pair in pairs:
            if ":" not in pair:
                continue
                
            col_name, value = pair.split(":", 1)
            col_name = col_name.strip()
            value = value.strip()
            
            # 해당 컬럼에 대한 라벨이 있으면 값 설정 (대소문자 구분 없이)
            for label_col, label in self.column_value_labels.items():
                if label_col.upper() == col_name.upper():
                    label.config(text=value)
                    break

    def _on_table_filter_changed(self, *args):
        """테이블 필터 변경 시 지연 처리"""
        if self._table_filter_after_id:
            self.root.after_cancel(self._table_filter_after_id)
        self._table_filter_after_id = self.root.after(100, self.update_table_list)

    def _on_mousewheel(self, event):
        """마우스 휠 이벤트 처리"""
        self.column_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _bind_mousewheel(self, widget):
        """마우스 휠 바인딩"""
        widget.bind_all("<MouseWheel>", self._on_mousewheel)

    def _unbind_mousewheel(self, widget):
        """마우스 휠 바인딩 해제"""
        widget.unbind_all("<MouseWheel>")

    def filter_results(self):
        """검색 결과 필터링 함수"""
        keyword = self.result_filter_var.get().strip().lower()
        self.tree.delete(*self.tree.get_children())
        for file, sheet, data in self.filtered_results:
            if keyword in data.lower():
                self.tree.insert("", "end", values=(file, sheet, data))

    def restore_all_results(self):
        """검색 결과 필터링 초기화"""
        self.tree.delete(*self.tree.get_children())
        for file, sheet, data in self.filtered_results:
            self.tree.insert("", "end", values=(file, sheet, data))


    def save_preset(self):
        """프리셋 저장 기능"""
        name = simpledialog.askstring("프리셋 저장", "프리셋 이름을 입력하세요:")
        if not name:
            return

        # 항상 최신 설정 파일을 로드하여 사용
        current_config = load_config(CONFIG_FILE)
        presets = current_config.get("presets", {})

        # 같은 이름의 프리셋이 있는지 확인
        existing_preset_id = None
        for pid, preset in presets.items():
            if preset.get("name") == name:
                existing_preset_id = pid
                break
        
        # 같은 이름의 프리셋이 있으면 덮어쓰기
        if existing_preset_id is not None:
            confirm = messagebox.askyesno(
                "프리셋 덮어쓰기", 
                f"같은 이름의 프리셋 '{name}'이 이미 존재합니다. 덮어쓰시겠습니까?"
            )
            if confirm:
                # 기존 프리셋 덮어쓰기
                presets[existing_preset_id]["data"] = self.data_path.get()
                    
                # 전체 설정 업데이트
                current_config["presets"] = presets
                save_config(CONFIG_FILE, current_config)
                
                # 현재 객체의 설정도 업데이트
                self.config = current_config
                
                self.show_info("저장 완료", f"프리셋 '{name}'이 업데이트되었습니다.")
                self.build_preset_buttons()
                return
            else:
                # 취소하면 함수 종료
                return
        
        # 새 프리셋 저장 (기존 프리셋이 없을 경우)
        if len(presets) >= 3:
            self.show_warning("제한 초과", "프리셋은 최대 3개까지 저장 가능합니다.")
            return

        # 새 ID 생성
        if not presets:
            preset_id = "1"
        else:
            used_ids = [int(k) for k in presets.keys()]
            preset_id = str(self._find_next_available_id(used_ids))
                    
        # 새 프리셋 저장
        presets[preset_id] = {
            "name": name,
            "data": self.data_path.get()
        }

        # 설정 전체 업데이트 (중요!)
        current_config["presets"] = presets
        save_config(CONFIG_FILE, current_config)
        
        # 현재 객체의 설정도 업데이트
        self.config = current_config
        
        self.show_info("저장 완료", f"프리셋 '{name}'이 저장되었습니다.")
        self.build_preset_buttons()


    def _find_next_available_id(self, used_ids):
        """사용되지 않은 가장 작은 ID 값을 찾습니다"""
        used_ids.sort()
        next_id = 1
        
        for id_num in used_ids:
            if id_num == next_id:
                next_id += 1
            elif id_num > next_id:
                break
        
        return next_id

    def build_preset_buttons(self):
        """프리셋 버튼 생성"""
        config = load_config(CONFIG_FILE)
        presets = config.get("presets", {})

        for widget in self.preset_button_frame.winfo_children():
            widget.destroy()

        for preset_id, preset in presets.items():
            tk.Button(
                self.preset_button_frame,
                text=preset["name"],
                command=lambda p=preset: self.load_preset(p)
            ).pack(side="left", padx=3)

    def load_preset(self, preset):
        """프리셋 로드"""
        data_path = preset.get("data", "")
        if not os.path.exists(data_path):
            self.show_error("경로 오류", f"프리셋 경로를 찾을 수 없습니다.\n\n경로: {data_path}")
            return
            
        old_path = self.data_path.get()
        self.data_path.set(data_path)
        
        # DB 재빌드 필요성 체크 후 필요한 경우에만 로딩 시작
        if self.check_if_rebuild_needed(old_path, data_path):
            self.refresh_all_cache()
        else:
            # DB 재빌드 없이 UI 갱신만 진행
            self.update_ui_after_loading()

    def reset_table_filter(self):
        """테이블 필터 초기화"""
        self.table_filter_var.set("")

    def select_data_folder(self):
        """데이터 폴더 선택"""
        folder = filedialog.askdirectory()
        if folder:
            self.data_path.set(folder)
            self.save_paths()

    def save_paths(self):
        """경로 설정 저장 - 프리셋 정보 유지"""
        # 최신 설정 로드
        current_config = load_config(CONFIG_FILE)
        
        # 데이터 경로만 업데이트하고 나머지 설정은 유지
        current_config["data_path"] = self.data_path.get()
        
        # 설정 저장
        save_config(CONFIG_FILE, current_config)
        
        # 로컬 설정 객체 업데이트
        self.config = current_config

    def check_if_rebuild_needed(self, old_path, new_path):
        """이전 경로와 새 경로를 비교하여 DB 재빌드가 필요한지 확인"""
        # 1. 경로가 같으면 재빌드 필요 없음
        if old_path == new_path:
            return False
            
        try:
            # 캐시 파일 경로 계산
            cache_dir = self.get_cache_dir(new_path)
            excel_cache_path = os.path.join(cache_dir, "excel_cache.json")
            db_path = os.path.join(cache_dir, "string_data.db")
            mtime_cache_path = db_path + ".mtime.json"
            
            # 기존 string_db_utils의 함수 활용
            return should_rebuild_string_db(excel_cache_path, db_path, mtime_cache_path)
        except Exception as e:
            # 오류 발생 시 안전하게 재빌드 진행
            return True
        
    def update_ui_after_loading(self):
        """캐시 로딩 후 UI 갱신 작업을 수행"""
        self.status_var.set("준비 완료")
        
        if hasattr(self, 'search_button'):
            self.search_button.configure(state="normal")
        
        # 테이블 목록 업데이트
        self.update_table_list()

    def get_cache_dir(self, data_path):
        """데이터 경로에 대한 캐시 디렉토리 계산"""
        path_hash = hashlib.md5(data_path.encode()).hexdigest()[:8]
        return os.path.join(get_application_dir(), ".cache", path_hash)

    def load_initial_cache(self):
        data_folder = self.data_path.get()
        data_folder = os.path.normpath(data_folder) if data_folder else ""

        if not data_folder or not os.path.exists(data_folder):
            self.show_warning("경로 오류", f"경로가 올바르지 않습니다.\n데이터: {data_folder}")
            return

        # 캐시 파일 경로 (엑셀과 DB 폴더 동일)
        cache_id = hash_paths(data_folder, data_folder)
        cache_dir = os.path.join(get_application_dir(), ".cache", cache_id)
        excel_cache_path = os.path.join(cache_dir, "excel_cache.json")
        db_cache_path = os.path.join(cache_dir, "db_cache.json")
        index_path = os.path.join(cache_dir, "table_sheet_index.json")
        string_dbs_mtime_path = os.path.join(cache_dir, "string_dbs.mtime.json")

        # 캐시 파일 존재 여부 확인
        excel_cache_exists = os.path.exists(excel_cache_path)
        db_cache_exists = os.path.exists(db_cache_path)
        index_exists = os.path.exists(index_path)
        string_dbs_mtime_exists = os.path.exists(string_dbs_mtime_path)

        # 모두 존재하면 바로 캐시 로드
        if excel_cache_exists and db_cache_exists and index_exists and string_dbs_mtime_exists:
            self.excel_cache = load_cached_data(excel_cache_path)
            self.db_cache = load_cached_data(db_cache_path)
            
            # 테이블 목록 추출
            db_tables = []
            for db_file, db_info in self.db_cache.items():
                if 'tables' in db_info:
                    for table in db_info['tables']:
                        db_tables.append(table)
            self.existing_db_files = set(db_tables)  
            
            self.update_table_list()
            self.status_var.set("✅ 캐시 로드 완료 (최신)")
            self.init_string_db()
        else:
            # 없으면 정상 로딩 프로세스 진행
            self.refresh_all_cache()


    def clean_cache(self):
        """더 이상 존재하지 않는 파일들을 캐시에서 제거"""
        data_folder = self.data_path.get()
        if not data_folder or not os.path.exists(data_folder):
            self.show_warning("경로 오류", "유효한 데이터 경로가 지정되지 않았습니다.")
            return
            
        # 로딩 팝업 생성
        loading_popup = LoadingPopup(self.root, title="캐시 정리 중", message="⏳ 캐시 검사 중...")
        
        def update_status(message):
            """백그라운드 스레드에서 UI 업데이트를 위한 함수"""
            self.root.after(0, lambda msg=message: loading_popup.update_message(msg))
            
        def clean_task():
            try:
                # 1. 캐시 파일 경로 계산
                cache_id = hash_paths(data_folder, data_folder)
                cache_dir = os.path.join(get_application_dir(), ".cache", cache_id)
                excel_cache_path = os.path.join(cache_dir, "excel_cache.json")
                db_cache_path = os.path.join(cache_dir, "db_cache.json")
                index_path = os.path.join(cache_dir, "table_sheet_index.json")
                
                # 2. 엑셀 캐시 정리
                update_status("엑셀 캐시 정리 중...")
                cleaned = False
                
                if os.path.exists(excel_cache_path):
                    excel_cache = load_cached_data(excel_cache_path)
                    original_count = len(excel_cache)
                    new_cache = {}
                    
                    # 존재하는 파일만 유지
                    for rel_path, info in excel_cache.items():
                        full_path = os.path.join(data_folder, rel_path)
                        if os.path.exists(full_path):
                            new_cache[rel_path] = info
                        else:
                            logging.info(f"캐시에서 제거: {rel_path} (파일 없음)")
                            cleaned = True
                    
                    # 변경된 경우에만 저장
                    if cleaned:
                        save_cache(excel_cache_path, new_cache)
                        logging.info(f"엑셀 캐시 정리 완료: {original_count} -> {len(new_cache)} 항목")
                        self.excel_cache = new_cache
                
                # 3. 인덱스 파일 재생성
                if cleaned and os.path.exists(index_path):
                    update_status("테이블-시트 인덱스 재생성 중...")
                    try:
                        os.remove(index_path)
                        logging.info("기존 인덱스 파일 삭제")
                    except Exception as e:
                        logging.error(f"인덱스 파일 삭제 실패: {e}")
                    
                    # 인덱스 재구축
                    build_table_sheet_index(data_folder, data_folder, status_callback=update_status)
                
                # 4. 완료 처리
                def on_complete():
                    loading_popup.close()
                    if cleaned:
                        self.show_info("캐시 정리 완료", "존재하지 않는 파일들이 캐시에서 제거되었습니다.")
                        # UI 갱신
                        self.update_table_list()
                    else:
                        self.show_info("캐시 검사 완료", "모든 캐시 항목이 유효합니다.")
                    
                self.root.after(0, on_complete)
                
            except Exception as e:
                # 오류 발생 시 처리
                error_msg = str(e)
                logging.error(f"캐시 정리 오류: {error_msg}")
                import traceback
                logging.error(traceback.format_exc())
                
                def on_error():
                    loading_popup.close()
                    self.show_error("캐시 정리 오류", f"처리 중 오류가 발생했습니다: {error_msg}")
                    
                self.root.after(0, on_error)
        
        # 백그라운드 스레드에서 실행
        thread = threading.Thread(target=clean_task)
        thread.daemon = True
        thread.start()


    def refresh_all_cache(self):
        """전체 캐시 새로고침: 엑셀 캐시, DB 캐시, 문자열 DB를 모두 업데이트"""
        data_folder = self.data_path.get()
        
        if not os.path.exists(data_folder):
            messagebox.showwarning("경로 오류", f"데이터 경로가 올바르지 않습니다.\n경로: {data_folder}")
            return
        
        # 로딩 팝업 생성
        loading_popup = LoadingPopup(self.root, title="캐시 업데이트", message="⏳ 캐시 업데이트 준비 중...")
        
        def update_status(message):
            """백그라운드 스레드에서 UI 업데이트를 위한 함수"""
            self.root.after(0, lambda msg=message: loading_popup.update_message(msg))

        def update_progress(current, total):
            """진행률 업데이트를 위한 함수"""
            self.root.after(0, lambda curr=current, tot=total: loading_popup.update_progress(curr, tot))
                
        # 커스텀 콜백 함수 정의
        def excel_progress_callback(idx, total, filename):
            """엑셀 분석 진행 상황을 팝업 UI에 표시"""
            status_msg = f"엑셀 분석 중 [{idx}/{total}]: {filename}"
            update_status(status_msg)
            update_progress(idx, total)
        
        def update_task():
            try:
                # 1. 엑셀 캐시 업데이트
                pk_override_dict = ExcelFileManager.load_pk_overrides()
                update_status("1/3: 엑셀 파일 분석 준비 중...")                
                cache_id = hash_paths(data_folder, data_folder)
                cache_dir = os.path.join(get_application_dir(), ".cache", cache_id)
                os.makedirs(cache_dir, exist_ok=True)

                # update_excel_cache 함수 오버라이드 또는 래핑
                cache_path = os.path.join(cache_dir, "excel_cache.json")
                old_cache = load_cached_data(cache_path)
                new_cache = {}
                files_to_scan = []

                for root, _, files in os.walk(data_folder):
                
                    for file in files:
                        # 엑셀 파일만 처리
                        if file.endswith(".xlsx") and not file.startswith("~$"):
                            # 누락된 변수 정의 부분 추가
                            path = os.path.join(root, file)
                            
                            # 파일 존재 확인 추가 - 중요!
                            if not os.path.exists(path):
                                logging.warning(f"파일이 존재하지 않음: {path}")
                                continue
                                
                            rel_path = os.path.relpath(path, data_folder).replace("\\", "/")
                            mtime = get_file_mtime(path)
                                            
                            # 캐시 상태 확인
                            if rel_path not in old_cache or old_cache[rel_path]["mtime"] != mtime:
                                files_to_scan.append((rel_path, path, mtime, data_folder))
                            else:
                                new_cache[rel_path] = old_cache[rel_path]
                
                scan_count = len(files_to_scan)
                total_count = len(new_cache) + scan_count
                update_status(f"스캔 완료: {scan_count}/{total_count} 파일 변경됨")
                update_progress(0, scan_count)
                
                # 변경된 파일만 분석
                for idx, (rel_path, path, mtime, db_folder) in enumerate(files_to_scan):
                    file_name = os.path.basename(path)
                    update_status(f"[{idx+1}/{scan_count}] 분석 중: {file_name}")
                    update_progress(idx+1, scan_count)
                    
                    
                    sheets_result = ExcelFileManager.analyze_excel_file(path, db_folder, pk_override_dict=pk_override_dict)
                    if sheets_result:
                        new_cache[rel_path] = {
                            "path": path,
                            "mtime": mtime,
                            "sheets": sheets_result
                        }
                
                # 캐시 저장
                update_status("엑셀 캐시 저장 중...")
                save_cache(cache_path, new_cache)
                self.excel_cache = new_cache
                
                # 2. DB 캐시 업데이트
                update_status("2/3: DB 파일 분석 중...")
                cache_base_dir = os.path.join(get_application_dir(), ".cache")
                db_folder_path, folder_path, db_cache = update_db_cache(data_folder, data_folder, cache_base_dir)
                self.db_cache = db_cache
                
                # 테이블 목록 추출
                update_status("테이블 목록 추출 중...")
                self.existing_db_files = set()
                for db_file, db_info in self.db_cache.items():
                    if 'tables' in db_info:
                        for table in db_info['tables']:
                            self.existing_db_files.add(table)
            
                
                # 3. 문자열 DB 업데이트 
                update_status("3/3: 문자열 DB 업데이트 중...")
                
                excel_cache_path = os.path.join(cache_dir, "excel_cache.json")
                db_path = os.path.join(cache_dir, "string_dbs")
                
                # 여기서 문자열 관련 파일만 처리
                load_or_build_string_db(excel_cache_path, db_path, data_folder)
                
                # 완료 후 UI 업데이트 함수
                def on_complete():
                    loading_popup.close()
                    self.update_table_list()
                    self.status_var.set("✅ 모든 캐시 업데이트 완료")
                
                # UI 업데이트는 메인 스레드에서 수행
                self.root.after(0, on_complete)
                
            except Exception as e:
                # 오류 메시지 저장
                error_msg = str(e)
                
                # 오류 발생 시 처리 함수
                def on_error():
                    loading_popup.close()
                    messagebox.showerror("캐시 업데이트 오류", error_msg)
                    self.status_var.set(f"❌ 캐시 업데이트 실패: {error_msg}")
                
                # 메인 스레드에서 오류 처리
                self.root.after(0, on_error)
        
        # 백그라운드 스레드에서 처리
        thread = threading.Thread(target=update_task)
        thread.daemon = True
        thread.start()


    def rebuild_table_index(self):
        """테이블-시트 인덱스를 강제로 재구축합니다"""
        data_folder = self.data_path.get()
        if not data_folder or not os.path.exists(data_folder):
            self.show_warning("경로 오류", "유효한 데이터 경로가 설정되지 않았습니다.")
            return
            
        # 로딩 팝업 생성
        loading_popup = LoadingPopup(self.root, title="인덱스 재구축", message="⏳ 테이블-시트 인덱스 재구축 중...")

        def update_status(message):
            """백그라운드 스레드에서 UI 업데이트를 위한 함수"""
            self.root.after(0, lambda msg=message: loading_popup.update_message(msg))
            
        def rebuild_task():
            try:
                # 1. 현재 캐시 디렉토리와 인덱스 파일 경로 계산
                cache_id = hash_paths(data_folder, data_folder)
                cache_dir = os.path.join(get_application_dir(), ".cache", cache_id)
                excel_cache_path = os.path.join(cache_dir, "excel_cache.json")
                index_path = os.path.join(cache_dir, "table_sheet_index.json")
                
                # 2. 현재 엑셀 캐시 로드
                excel_cache = load_cached_data(excel_cache_path)
                if not excel_cache:
                    raise Exception("엑셀 캐시를 찾을 수 없습니다. 먼저 캐시를 새로고침하세요.")
                
                # 3. 인덱스 파일 삭제
                if os.path.exists(index_path):
                    try:
                        os.remove(index_path)
                        logging.info(f"기존 인덱스 파일 삭제: {index_path}")
                    except Exception as e:
                        logging.error(f"인덱스 파일 삭제 실패: {e}")
                
                # 4. 엑셀 캐시에서 존재하지 않는 파일 정리
                new_cache = {}
                removed_count = 0
                
                update_status("캐시 확인 중...")
                for rel_path, info in excel_cache.items():
                    full_path = os.path.join(data_folder, rel_path)
                    if os.path.exists(full_path):
                        new_cache[rel_path] = info
                    else:
                        logging.info(f"캐시에서 제거: {rel_path} (파일 없음)")
                        removed_count += 1
                
                # 5. 정리된 캐시 저장
                if removed_count > 0:
                    update_status(f"{removed_count}개 파일 제거됨. 캐시 저장 중...")
                    save_cache(excel_cache_path, new_cache)
                    self.excel_cache = new_cache
                
                # 6. 테이블-시트 인덱스 재구축
                update_status("테이블-시트 인덱스 재구축 중...")
                
                # 진행 상태 업데이트 콜백 함수
                def status_callback(message):
                    update_status(message)
                
                # 인덱스 재구축 함수 호출
                build_table_sheet_index(data_folder, data_folder, status_callback=update_status)
                
                # 7. 완료 처리
                def on_complete():
                    loading_popup.close()
                    self.show_info("인덱스 재구축 완료", "테이블-시트 인덱스가 성공적으로 재구축되었습니다.")
                    # UI 갱신
                    self.update_table_list()
                
                self.root.after(0, on_complete)
                
            except Exception as e:
                # 오류 발생 시 처리
                error_msg = str(e)
                logging.error(f"인덱스 재구축 오류: {error_msg}")
                import traceback
                logging.error(traceback.format_exc())
                
                def on_error():
                    loading_popup.close()
                    self.show_error("인덱스 재구축 오류", f"처리 중 오류가 발생했습니다: {error_msg}")
                    
                self.root.after(0, on_error)
        
        # 백그라운드 스레드에서 실행
        thread = threading.Thread(target=rebuild_task)
        thread.daemon = True
        thread.start()


    def rebuild_all_cache_and_index(self):
        """캐시와 인덱스를 모두 강제로 재구축"""
        result = messagebox.askyesno(
            "전체 캐시 재구축", 
            "모든 캐시와 인덱스를 완전히 재구축하시겠습니까?\n"
            "이 과정은 시간이 다소 소요될 수 있습니다."
        )
        
        if not result:
            return
            
        data_folder = self.data_path.get()
        if not data_folder or not os.path.exists(data_folder):
            self.show_warning("경로 오류", "유효한 데이터 경로가 설정되지 않았습니다.")
            return
        
        # 로딩 팝업 생성
        loading_popup = LoadingPopup(self.root, title="전체 캐시 재구축", message="⏳ 캐시 디렉토리 초기화 중...")
        
        def update_status(message):
            """백그라운드 스레드에서 UI 업데이트를 위한 함수"""
            self.root.after(0, lambda msg=message: loading_popup.update_message(msg))
            
        def rebuild_task():
            try:
                # 1. 캐시 디렉토리 경로 계산
                cache_id = hash_paths(data_folder, data_folder)
                cache_dir = os.path.join(get_application_dir(), ".cache", cache_id)
                
                # 2. 캐시 디렉토리 삭제 후 재생성
                update_status("기존 캐시 디렉토리 삭제 중...")
                if os.path.exists(cache_dir):
                    import shutil
                    try:
                        shutil.rmtree(cache_dir)
                        logging.info(f"캐시 디렉토리 삭제 완료: {cache_dir}")
                    except Exception as e:
                        logging.error(f"캐시 디렉토리 삭제 실패: {e}")
                
                os.makedirs(cache_dir, exist_ok=True)
                
                # 3. 엑셀 캐시 재구축
                update_status("엑셀 캐시 재구축 중...")
                
                # 4. 완료 후 UI 업데이트 함수
                def on_complete():
                    loading_popup.close()
                    
                    # refresh_all_cache 함수 호출 (새 스레드에서)
                    self.refresh_all_cache()
                    
                # UI 업데이트는 메인 스레드에서 수행
                self.root.after(0, on_complete)
                
            except Exception as e:
                # 오류 발생 시 처리
                error_msg = str(e)
                logging.error(f"전체 캐시 재구축 오류: {error_msg}")
                import traceback
                logging.error(traceback.format_exc())
                
                def on_error():
                    loading_popup.close()
                    self.show_error("전체 캐시 재구축 오류", f"처리 중 오류가 발생했습니다: {error_msg}")
                    
                self.root.after(0, on_error)
        
        # 백그라운드 스레드에서 실행
        thread = threading.Thread(target=rebuild_task)
        thread.daemon = True
        thread.start()


    def update_table_list(self):
        """테이블 목록 업데이트: DB 파일 기반 테이블 이름 필터링"""
        filter_text = self.table_filter_var.get().lower()
        table_names = []

        # DB 파일 기반으로 테이블 목록 필터링
        for table in self.existing_db_files:
            if not filter_text or filter_text in table.lower():
                table_names.append(table)
        
        # 테이블 목록 업데이트
        self.table_listbox.delete(0, tk.END)
        for name in sorted(table_names):
            self.table_listbox.insert(tk.END, name)

    def init_string_db(self):
        """문자열 DB 초기화"""
        data_folder = self.data_path.get()
        # 경로 유효성 검사 추가
        if not data_folder or not os.path.exists(data_folder):
            messagebox.showerror("경로 오류", "유효한 데이터 경로가 설정되지 않았습니다.")
            return
        
        # 경로 정규화
        data_folder = os.path.normpath(data_folder)
        cache_id = hash_paths(data_folder, data_folder)

        excel_cache_path = os.path.join(get_application_dir(), ".cache", cache_id, "excel_cache.json")
        db_path = os.path.join(get_application_dir(), ".cache", cache_id, "string_dbs")

        popup = LoadingPopup(self.root, title="DB 로딩 중", message="🔄 문자열 검색용 DB 초기화 중...")

        # 진행 상황 업데이트 함수
        def update_progress(message):
            # UI 스레드에서 실행하도록 after 메소드 사용
            self.root.after(0, lambda: popup.update_message(message))

        def task():
            try:
                # 콜백 함수 전달
                load_or_build_string_db(excel_cache_path, db_path, data_folder, progress_callback=update_progress)
                
                # UI 스레드에서 상태 업데이트
                self.root.after(0, lambda: self.status_var.set("✅ DB 로딩 완료"))
            except Exception as e:
                # UI 스레드에서 오류 메시지 표시
                self.root.after(0, lambda error=e: messagebox.showerror("DB 로딩 실패", str(error)))
            finally:
                self.root.after(0, popup.close)

        # 백그라운드 스레드에서 실행
        t = threading.Thread(target=task)
        t.daemon = True
        t.start()

    def run_search(self):
        """검색 실행: table_sheet_index와 조건 기반으로 데이터 검색"""
        table = self.selected_table.get()
        if not table:
            messagebox.showwarning("테이블 선택", "먼저 테이블을 선택해주세요.")
            return
        
        # 조건 수집
        conditions = self.get_search_conditions()
        if not conditions:
            messagebox.showwarning("검색 조건", "최소 하나 이상의 검색 조건을 입력하세요.")
            return
        
        # 진행 상태 표시
        progress_popup = self.create_search_progress_popup()
        
        # 검색 실행 (스레드로)
        thread = threading.Thread(target=lambda: self.perform_search(table, conditions, progress_popup))
        thread.daemon = True
        thread.start()
        
    def get_search_conditions(self):
        """검색 조건 수집"""
        conditions = {
            col: entry.get().strip()
            for col, entry in self.column_entries.items()
            if entry.get().strip()
        }
        return conditions
        
    def create_search_progress_popup(self):
        """검색 진행 상태 팝업 생성"""
        progress_popup = tk.Toplevel(self.root)
        progress_popup.title("검색 중")
        progress_popup.geometry("350x100")
        progress_label = tk.Label(progress_popup, text="데이터 검색 준비 중...")
        progress_label.pack(pady=20)
        progress_popup.grab_set()
        self.root.update_idletasks()
        return progress_popup
    
    def perform_search(self, table, conditions, progress_popup):
        """검색 작업 수행"""
        data_folder = self.data_path.get()
        cache_id = hash_paths(data_folder, data_folder)
        index_path = os.path.join(get_application_dir(), ".cache", cache_id, "table_sheet_index.json")
    
        # 로그 추가 - 검색 시작
        logging.info(f"===== 테이블 '{table}' 검색 시작 =====")
        logging.info(f"검색 조건: {conditions}")
        
        # 인덱스 확인 및 생성
        if not self.ensure_table_index(index_path, data_folder, progress_popup):
            return
        
        # 인덱스 로드
        table_index = self.load_table_index(index_path, progress_popup)
        if not table_index:
            return
        
        # 검색 수행
        search_results = self.search_in_sheets(table, table_index, conditions, data_folder, progress_popup)
        
        # 결과 업데이트 (UI 스레드에서)
        self.root.after(0, lambda: self.update_search_results(search_results, progress_popup))
    
    
    def ensure_table_index(self, index_path, data_folder, progress_popup):
        """테이블 인덱스가 존재하는지 확인하고 없으면 생성"""
        if not os.path.exists(index_path):
            def update_status(msg):
                self.root.after(0, lambda: self.update_progress_label(progress_popup, msg))
            
            try:
                # 이 부분을 수정 - status_callback 매개변수 추가
                build_table_sheet_index(data_folder, data_folder, status_callback=update_status)
                return True
            except Exception as e:
                error_msg = str(e)
                self.root.after(0, lambda err=error_msg, popup=progress_popup: [
                    popup.destroy(),
                    messagebox.showerror("인덱스 오류", f"테이블-시트 인덱스 생성 실패: {err}")
                ])
                return False
        return True
    

    def update_progress_label(self, popup, text):
        """진행 상태 라벨 업데이트"""
        for widget in popup.winfo_children():
            if isinstance(widget, tk.Label):
                widget.config(text=text)
                break
                

    def load_table_index(self, index_path, progress_popup):
        """테이블 인덱스 로드 및 유효성 검증"""
        try:
            if not os.path.exists(index_path):
                logging.warning(f"인덱스 파일이 존재하지 않음: {index_path}")
                return None
                    
            with open(index_path, "r", encoding="utf-8") as f:
                index_data = json.load(f)
                        
            # 새 형식과 이전 형식 모두 지원
            if isinstance(index_data, dict) and "index" in index_data:
                table_index = index_data["index"]
            else:
                # 이전 형식 (직접 테이블 인덱스만 저장)
                table_index = index_data
            
            # 인덱스 유효성 검증 부분 제거 또는 다음과 같이 수정
            # 방법 1: 유효성 검증 부분 완전히 제거
            
            # 방법 2: 다른 방식으로 변수 확인 (권장)
            validate_index = False  # 기본값으로 검증 비활성화
            try:
                # 변수가 있으면 확인, 없으면 기본값 사용
                validate_index = getattr(self, 'validate_index', False)
            except:
                pass
                
            if validate_index:
                # 이후 코드는 그대로 유지
                self.root.after(0, lambda popup=progress_popup: self.update_progress_label(
                    popup, "인덱스 유효성 검증 중..."))
                
                data_folder = self.data_path.get()
                for table, sheets in table_index.items():
                    valid_sheets = []
                    for rel_path, sheet in sheets:
                        full_path = os.path.join(data_folder, rel_path)
                        if os.path.exists(full_path):
                            valid_sheets.append((rel_path, sheet))
                    table_index[table] = valid_sheets
            
            # 성공적으로 로드한 경우 테이블 인덱스 반환
            return table_index
            
        except Exception as e:
            error_msg = str(e)
            self.root.after(0, lambda err=error_msg, popup=progress_popup: [
                popup.destroy(),
                messagebox.showerror("인덱스 오류", f"테이블-시트 인덱스 로드 실패: {err}")
            ])
            return None  # 예외 발생 시 None 반환


    def search_in_sheets(self, table, table_index, conditions, data_folder, progress_popup):
        """모든 관련 시트에서 검색 수행"""
        sheet_list = table_index.get(table, [])
        search_results = []
        hit_count = 0
        missing_files = 0  # 존재하지 않는 파일 카운터
        missing_paths = set()  # 존재하지 않는 파일 경로 집합
        
        # 로그 추가
        logging.info(f"테이블 '{table}' 검색 시작: {len(sheet_list)} 시트 발견")
        
        # 테이블 관련 시트에서 검색
        for idx, (rel_path, sheet) in enumerate(sheet_list):
            # 진행 상태 업데이트
            self.root.after(0, lambda i=idx, t=len(sheet_list): 
                self.update_progress_label(progress_popup, f"검색 중... ({i+1}/{t})"))
            
            try:
                full_path = os.path.join(data_folder, rel_path)
                if not os.path.exists(full_path):
                    logging.warning(f"파일이 존재하지 않습니다: {full_path}")
                    missing_files += 1
                    missing_paths.add(rel_path)
                    continue
                
                # 헤더 행 정보 확인
                header_row = self.get_sheet_header_row(rel_path, sheet, data_folder)
                
                # 데이터 로드 및 필터링
                with pd.ExcelFile(full_path) as xls:
                    df = pd.read_excel(xls, sheet_name=sheet, header=header_row)
                    if df.empty:
                        continue
                        
                    # 여기가 중요! - filtered_rows 변수 정의 부분이 누락됨
                    filtered_rows = self.filter_dataframe(df, conditions)
                    
                    # 빈 결과면 다음 시트로 넘어감
                    if filtered_rows.empty:
                        continue
                    
                    # 결과 저장
                    for _, row in filtered_rows.iterrows():
                        preview = self.format_row_preview(row, df.columns)
                        file_name = os.path.basename(full_path)
                        
                        search_results.append((file_name, sheet, preview, full_path))
                        hit_count += 1
                    
            except Exception as e:
                logging.error(f"검색 오류: {rel_path} / {sheet}: {e}")
        
        # 찾을 수 없는 파일이 많으면 인덱스 재구축 제안
        if missing_files > 0:
            missing_paths_str = "\n".join(list(missing_paths)[:5])
            if len(missing_paths) > 5:
                missing_paths_str += f"\n... 외 {len(missing_paths) - 5}개 파일"
                
            logging.warning(f"검색 중 {missing_files}개의 파일을 찾을 수 없습니다. 인덱스 재구축이 필요합니다.")
            
            # 메인 스레드에서 메시지 표시
            def show_rebuild_message():
                result = messagebox.askyesno(
                    "인덱스 재구축 필요", 
                    f"검색 중 {missing_files}개의 파일을 찾을 수 없습니다.\n"
                    f"이는 테이블-시트 인덱스가 최신 상태가 아님을 의미합니다.\n\n"
                    f"예시 경로:\n{missing_paths_str}\n\n"
                    f"테이블-시트 인덱스를 재구축하시겠습니까?"
                )
                if result:
                    self.rebuild_table_index()
            
            if missing_files >= 3:  # 3개 이상의 파일이 없는 경우에만 메시지 표시
                self.root.after(0, show_rebuild_message)
        
        return {"results": search_results, "hit_count": hit_count}

    def format_row_preview_with_db_names(self, row, excel_columns, db_columns_map):
        """행 데이터를 DB 컬럼명으로 변환하여 미리보기 포맷팅"""
        formatted_parts = []
        
        # Excel 컬럼 위치에 따라 DB 컬럼 매핑 시도
        for i, col in enumerate(excel_columns):
            if i >= len(self.db_columns):
                break  # DB 컬럼 수보다 많으면 중단
                
            # 실제 DB 컬럼명 사용
            db_col = self.db_columns[i]
            value = row[col]
            
            # 숫자 포맷팅
            if isinstance(value, float) and value.is_integer():
                value = int(value)
            
            # 결과에 추가
            formatted_parts.append(f"{db_col}: {value}")
        
        return " | ".join(formatted_parts)

    def get_sheet_header_row(self, rel_path, sheet, data_folder):
        """시트의 헤더 행 가져오기 (개선된 버전)"""
        try:
            cache_id = hash_paths(data_folder, data_folder)
            cache_path = os.path.join(".cache", cache_id, "excel_cache.json")
            logging.info(f"헤더 행 조회: {rel_path} / {sheet}")
            
            cache = load_cached_data(cache_path)
            if not cache:
                logging.warning("엑셀 캐시가 비어있습니다. 기본 헤더 행(1) 사용")
                return 1  # 기본값 수정: 첫 번째 행을 헤더로 시도 (0 기반 인덱스)
            
            # 경로 정규화 - 여러 형태로 시도
            paths_to_try = [
                rel_path,
                os.path.normpath(rel_path),
                rel_path.replace('\\', '/'),
                rel_path.replace('/', '\\'),
                os.path.basename(rel_path)  # 파일명만으로도 시도
            ]
            
            for path in paths_to_try:
                if path in cache and "sheets" in cache[path]:
                    sheet_meta = cache[path]["sheets"].get(sheet, {})
                    if "header_row" in sheet_meta:
                        header_row = sheet_meta["header_row"]
                        return header_row
            
            # 여기에 도달한다면 캐시에서 찾지 못한 것
            logging.warning(f"캐시에서 헤더 행을 찾지 못함: {rel_path} / {sheet}, 기본값(1) 사용")
            # 기본값 변경: 이전의 2(3번째 행) 대신 1(2번째 행) 사용
            return 1  
        except Exception as e:
            logging.error(f"헤더 행 조회 실패: {e}")
            import traceback
            logging.error(traceback.format_exc())
            return 1  # 오류 시 안전한 기본값

    def filter_dataframe(self, df, conditions):
        """조건에 따라 데이터프레임 필터링: 부분 일치 검색 지원"""
        mask = pd.Series([True] * len(df))
        # 로그 추가 - 필터링 시작
        logging.info(f"데이터프레임 필터링 시작 (행 수: {len(df)})")
        
        for col, val in conditions.items():
            if col in df.columns:
                try:
                    col_series = df[col].copy()
                    
                    # 데이터 샘플 확인 로그
                    sample_values = col_series.head(3).tolist()
                    logging.info(f"컬럼 '{col}' 필터링: 값='{val}', 샘플 데이터={sample_values}")
                    
                    # 숫자 컬럼이고 숫자 입력인 경우 정확히 일치하는 조건 적용
                    if val.isdigit() and pd.api.types.is_numeric_dtype(col_series):
                        mask &= col_series.fillna(0).astype(int) == int(val)
                        logging.info(f"숫자 필터링 적용: '{col}' == {val}")
                    else:
                        # 문자열 변환 시 NaN 값을 빈 문자열로 대체
                        str_series = col_series.fillna('').astype(str)
                        val_lower = val.lower()  # 검색어를 소문자로 변환
                        mask &= str_series.str.lower().str.contains(val_lower, regex=False, na=False)
                        logging.info(f"문자열 필터링 적용: '{col}' contains '{val_lower}'")
                
                    # 적용 후 일치하는 행 수 확인
                    matching_rows = sum(mask)
                    logging.info(f"필터 '{col}:{val}' 적용 후 일치하는 행 수: {matching_rows}")
                
                except Exception as e:
                    logging.error(f"컬럼 '{col}' 필터링 오류: {e}")
                    import traceback
                    logging.error(traceback.format_exc())
        
            else:
                logging.warning(f"컬럼 '{col}'이 데이터프레임에 없습니다! 사용 가능한 컬럼: {list(df.columns)}")    
    
        filtered_df = df[mask]
        logging.info(f"최종 필터링 결과: {len(filtered_df)}행")
        return df[mask]
    
    #행 데이터 미리보기 형식화
    def format_row_preview(self, row, columns):
        return " | ".join(
            f"{c}: {int(row[c]) if isinstance(row[c], float) and row[c].is_integer() else row[c]}"
            for c in columns if c in row
        )
        
    def update_search_results(self, search_data, progress_popup):
        """검색 결과 UI 업데이트"""
        # 결과 출력
        self.tree.delete(*self.tree.get_children())
        self.filtered_results = []
        
        for file_name, sheet, preview, full_path in search_data["results"]:
            self.tree.insert("", "end", values=(file_name, sheet, preview), tags=(full_path,))
            self.filtered_results.append((file_name, sheet, preview))
        
        # 결과 상태 업데이트
        self.status_var.set(f"✅ 검색 완료: {search_data['hit_count']}건 발견")
        progress_popup.destroy()
        
    def open_db_viewer(self):
        table = self.selected_table.get()
        if not table:
            messagebox.showwarning("선택 필요", "DB 보기 전에 테이블을 선택해주세요.")
            return
        
        # 만약 excel_cache가 비어있다면 로드
        if not self.excel_cache:
            data_folder = self.data_path.get()
            cache_id = hash_paths(data_folder, data_folder)
            cache_path = os.path.join(".cache", cache_id, "excel_cache.json")
            self.excel_cache = load_cached_data(cache_path)
        
        open_db_viewer_with_excel_match(self.root, self.data_path.get(), table, self.excel_cache, self.data_path.get())

    def open_string_searcher(self):
        try:
            # UI 모듈을 찾을 수 있도록 현재 디렉토리를 경로에 추가
            import sys
            if current_dir not in sys.path:
                sys.path.append(current_dir)
                
            # 만약 excel_cache가 비어있다면 로드
            data_folder = self.data_path.get()
            if not data_folder:  # 비어 있으면 기본값 설정
                data_folder = "."  # 현재 디렉토리를 기본값으로 사용
                self.show_warning("경로 오류", "데이터 경로가 설정되지 않았습니다. 현재 폴더를 사용합니다.")
            
            # 디버깅 로그 추가
            logging.debug(f"String 검색기 초기화 시작: 경로={data_folder}")
            
            StringSearchPopup(
                self.root,
                data_folder,
                data_folder,
                self.excel_cache
            )
        except Exception as e:
            # 상세한 오류 메시지를 로그와 메시지 박스에 표시
            import traceback
            tb = traceback.format_exc()
            self.show_error("String 검색기 오류", f"String 검색기를 열 수 없습니다: {str(e)}")
            logging.error(f"String 검색기 오류: {str(e)}\n{tb}")

    def open_reward_searcher(self):
        # 만약 excel_cache가 비어있다면 로드
        if not self.excel_cache:
            data_folder = self.data_path.get()
            cache_id = hash_paths(data_folder, data_folder)
            cache_path = os.path.join(".cache", cache_id, "excel_cache.json")
            self.excel_cache = load_cached_data(cache_path)
            
        RewardSearchPopup(
            self.root,
            self.data_path.get(),
            self.data_path.get(),  # DB 경로와 엑셀 경로 동일
            self.excel_cache
        )

    def open_column_searcher(self):
        # 만약 excel_cache가 비어있다면 로드
        if not self.excel_cache:
            data_folder = self.data_path.get()
            cache_id = hash_paths(data_folder, data_folder)
            cache_path = os.path.join(".cache", cache_id, "excel_cache.json")
            self.excel_cache = load_cached_data(cache_path)
            
        ColumnSearchPopup(self.root, self.excel_cache)

    def open_quest_searcher(self):
        # 만약 excel_cache가 비어있다면 로드
        if not self.excel_cache:
            data_folder = self.data_path.get()
            cache_id = hash_paths(data_folder, data_folder)
            cache_path = os.path.join(".cache", cache_id, "excel_cache.json")
            self.excel_cache = load_cached_data(cache_path)
            
        QuestSearchPopup(
            self.root,
            self.data_path.get(),
            self.data_path.get(),  # DB 경로와 엑셀 경로 동일
            self.excel_cache
        )
        
    def open_npc_creator(self):
        """NPC 생성기 팝업 열기"""
        NPCCreatorPopup(
            self.root,
            self.data_path.get(),
            self.data_path.get(),  # DB 경로와 엑셀 경로 동일
            self.excel_cache
        )

    def open_relationship_analyzer(self):
        from tools.table_relationship_analyzer import TableRelationshipAnalyzerApp

        window = tk.Toplevel(self.root)
        TableRelationshipAnalyzerApp(window)


    def open_excel_diff_tool(self):
        """엑셀 데이터 비교 도구 열기"""
        try:
            from tools.excel_diff_tool import open_excel_diff_tool
            open_excel_diff_tool(self.root, self.data_path.get())
        except Exception as e:
            self.show_error("Excel Diff 도구 오류", f"엑셀 비교 도구를 열 수 없습니다: {str(e)}")
            logging.error(f"Excel Diff 도구 오류: {str(e)}")


    def open_box_list(self):
        """전체 Box 목록을 직접 엽니다."""
        # 만약 excel_cache가 비어있다면 로드
        if not self.excel_cache:
            data_folder = self.data_path.get()
            cache_id = hash_paths(data_folder, data_folder)
            cache_path = os.path.join(".cache", cache_id, "excel_cache.json")
            self.excel_cache = load_cached_data(cache_path)
        
        # typecode_mapping 로드
        typecode_mapping = []
        try:
            # 매핑 파일 경로 찾기
            mapping_path = os.path.join(self.data_path.get(), "typecode_mapping.json")
            
            # 다른 위치도 확인
            if not os.path.exists(mapping_path):
                mapping_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "typecode_mapping.json")
            
            if not os.path.exists(mapping_path):
                mapping_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "typecode_mapping.json")
            
            if not os.path.exists(mapping_path):
                documents_path = os.path.join(os.path.expanduser("~"), "Documents")
                mapping_path = os.path.join(documents_path, "typecode_mapping.json")
            
            # 파일이 존재하면 로드
            if os.path.exists(mapping_path):
                with open(mapping_path, 'r', encoding='utf-8') as f:
                    typecode_mapping = json.load(f)
        except Exception as e:
            print(f"타입코드 매핑 로드 오류: {e}")
            typecode_mapping = []
        
        # BoxListPopup 직접 생성
        BoxListPopup(
            self.root,
            self.data_path.get(),
            self.data_path.get(),  # DB 경로와 엑셀 경로 동일
            typecode_mapping,
            self.excel_cache
        )


    # NPC 검색 기능 열기 메서드 추가
    def open_npc_search_popup(self):
        """전체 NPC 목록을 직접 엽니다."""
        # 만약 excel_cache가 비어있다면 로드
        if not self.excel_cache:
            data_folder = self.data_path.get()
            cache_id = hash_paths(data_folder, data_folder)
            cache_path = os.path.join(".cache", cache_id, "excel_cache.json")
            self.excel_cache = load_cached_data(cache_path)
        
                # typecode_mapping 로드
        typecode_mapping = []
        try:
            # 매핑 파일 경로 찾기
            mapping_path = os.path.join(self.data_path.get(), "typecode_mapping.json")
            
            # 다른 위치도 확인
            if not os.path.exists(mapping_path):
                mapping_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "typecode_mapping.json")
            
            if not os.path.exists(mapping_path):
                mapping_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "typecode_mapping.json")
            
            if not os.path.exists(mapping_path):
                documents_path = os.path.join(os.path.expanduser("~"), "Documents")
                mapping_path = os.path.join(documents_path, "typecode_mapping.json")
            
            # 파일이 존재하면 로드
            if os.path.exists(mapping_path):
                with open(mapping_path, 'r', encoding='utf-8') as f:
                    typecode_mapping = json.load(f)
        except Exception as e:
            print(f"타입코드 매핑 로드 오류: {e}")
            typecode_mapping = []

        NPCListPopup(
            self.root,
            self.data_path.get(),
            self.data_path.get(),
            typecode_mapping,  # Box 검색과 동일하게 타입코드 맵핑 전달
            self.excel_cache
        )


    def on_table_selected(self, event):
        selection = self.table_listbox.curselection()
        if not selection:
            return
        table_name = self.table_listbox.get(selection[0])
        self.selected_table.set(table_name)

        db_columns = self.get_db_columns(table_name)
        for widget in self.column_frame.winfo_children():
            widget.destroy()
        self.column_entries.clear()
        
        # 결과 표시용 라벨 딕셔너리 추가
        self.column_value_labels = {}

        for idx, col in enumerate(db_columns):
            # 행 프레임 생성 (각 컬럼별 UI 요소를 담는 컨테이너)
            row_frame = tk.Frame(self.column_frame)
            row_frame.grid(row=idx, column=0, sticky="w", pady=2)
            
            # 컬럼 라벨
            tk.Label(row_frame, text=f"{col}:", width=15, anchor="e").pack(side="left")
            
            # 입력 필드
            entry = tk.Entry(row_frame, width=20)
            entry.pack(side="left", padx=5)
            self.column_entries[col] = entry
            
            # 파이프 구분자
            tk.Label(row_frame, text="|").pack(side="left", padx=5)
            
            # 결과 표시 라벨 (선택된 값 표시)
            value_label = tk.Label(row_frame, text="", width=30, anchor="w", fg="blue")
            value_label.pack(side="left", padx=5)
            self.column_value_labels[col] = value_label


    def get_db_columns(self, table_name):
        # DB 스키마 캐싱 적용
        if table_name in self.db_schemas:
            return self.db_schemas[table_name]
        
        # 통합된 경로 사용
        db_file = os.path.join(self.data_path.get(), f"{table_name}.db")
        
        # 파일 존재 여부 및 크기 확인
        if not os.path.exists(db_file) or os.path.getsize(db_file) == 0:
            return []
        
        try:
            conn = sqlite3.connect(db_file)
            cur = conn.cursor()
            cur.execute(f"PRAGMA table_info({table_name})")
            columns = [row[1] for row in cur.fetchall()]
            conn.close()
            self.db_schemas[table_name] = columns
            return columns
        except Exception as e:
            logging.error(f"DB 컬럼 조회 오류: {db_file} - {e}")
            
            # 오류 발생 후 0KB 파일이 생성되었다면 삭제
            if os.path.exists(db_file) and os.path.getsize(db_file) == 0:
                try:
                    os.remove(db_file)
                    logging.info(f"빈 DB 파일 삭제: {db_file}")
                except Exception as del_e:
                    logging.error(f"빈 DB 파일 삭제 실패: {db_file} - {del_e}")
            
            return []
        
    def open_excel_from_result(self, event):
        """검색 결과 더블 클릭 시 엑셀 파일 열기"""
        item = self.tree.focus()
        if not item:
            return
        
        tags = self.tree.item(item, "tags")
        if not tags:
            return

        path = tags[0]
        if not os.path.exists(path):
            return
            
        values = self.tree.item(item, "values")
        if len(values) < 3:
            return

        sheet, preview = values[1], values[2]
        try:
            # preview에서 첫 번째 key:value 쌍 추출
            if " | " in preview:
                first_pair = preview.split(" | ")[0]
                if ":" in first_pair:
                    key, val = first_pair.split(":", 1)
                    key, val = key.strip(), val.strip()
                    
                    # 엑셀 캐시 전달하여 명확한 경로로 실행
                    try:                        
                        ExcelFileManager.highlight_excel_by_value(path, sheet, key, val, self.excel_cache)
                    except Exception as highlight_err:
                        os.startfile(path)
                else:
                    os.startfile(path)
            else:
                os.startfile(path)
        except Exception as e:
            logging.error(f"엑셀 열기 오류: {path} / {sheet}: {e}")
            try:
                # 기본 방식으로 파일 열기 시도
                os.startfile(path)
            except Exception as e2:
                logging.error(f"기본 열기 오류: {e2}")

    # 빠른 시트 컬럼 확인 함수
    def sheet_has_column(self, path, sheet, column):
        from openpyxl import load_workbook        
        try:
            wb = load_workbook(path, read_only=True)
            if sheet not in wb.sheetnames:
                return False
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                if column in row:
                    return True
        except Exception as e:
            logging.error(f"시트 컬럼 확인 오류: {path} / {sheet}: {e}")
        return False

    def on_closing(self):
        save_search_history(self.global_search_history)
        self.root.destroy()


    def open_translation_sync_tool(self):
        try:
            # tools 폴더 경로를 시스템 경로에 추가
            tools_path = os.path.join(get_application_dir(), "tools", "translate")
            if tools_path not in sys.path:
                sys.path.insert(0, tools_path)  # 맨 앞에 추가하여 우선순위 높임
            
            module = importlib.import_module("translation_sync_extension")
            importlib.reload(module)
            module.TranslationSyncExtension(tk.Toplevel(self.root))
        except Exception as e:
            messagebox.showerror("에러", f"동기화 도구 실행 실패: {e}")


    def open_translation_automation_tool(self):
        try:
            # 이미 상단에서 임포트된 클래스 사용
            TranslationAutomationTool(tk.Toplevel(self.root))
        except Exception as e:
            messagebox.showerror("에러", f"번역 자동화 툴 실행 실패: {e}")


    def open_unique_text_manager(self):
        try:
            translate_path = os.path.join(get_application_dir(), "tools", "translate")
            if translate_path not in sys.path:
                sys.path.insert(0, translate_path)  # 맨 앞에 추가하여 우선순위 높임
            
            # 모듈 임포트 시도
            module = importlib.import_module("translate_setup_add_unique_text_tab")
            importlib.reload(module)
            module.run_unique_text_manager(self.root)
        except Exception as e:
            import traceback
            traceback_str = traceback.format_exc()
            messagebox.showerror("에러", f"고유 텍스트 관리자 실행 실패:\n{e}\n\n{traceback_str}")


    def _show_message(self, message_type, title, message, parent=None, log_level=None):
        """
        통합 메시지 박스 표시 함수
        
        Args:
            message_type: 메시지 박스 타입 ('info', 'warning', 'error', 'yesno')
            title: 메시지 박스 제목
            message: 표시할 메시지
            parent: 부모 윈도우 (기본값: None - 루트 윈도우 사용)
            log_level: 로깅 레벨 (None이면 로깅하지 않음)
        
        Returns:
            yesno 타입인 경우 사용자 선택 결과(True/False), 아니면 None
        """
        if parent is None:
            # 진행 창이 있으면 진행 창을, 아니면 루트 윈도우를 부모로 설정
            parent = getattr(self, 'progress_window', None) or self.root
        
        # 로깅 처리
        if log_level:
            log_func = getattr(logging, log_level.lower(), logging.info)
            log_func(f"메시지 박스 ({message_type}): {title} - {message}")
        
        # 진행 창이 있으면 상태 업데이트
        if hasattr(self, '_update_progress') and getattr(self, 'progress_window', None):
            success = message_type != 'error'
            self._update_progress(f"{title}: {message}", success=success)
        
        # 메시지 박스 타입에 따라 다른 함수 호출
        if message_type == 'info':
            messagebox.showinfo(title, message, parent=parent)
            return None
        elif message_type == 'warning':
            messagebox.showwarning(title, message, parent=parent)
            return None
        elif message_type == 'error':
            messagebox.showerror(title, message, parent=parent)
            return None
        elif message_type == 'yesno':
            return messagebox.askyesno(title, message, parent=parent)
        else:
            # 지원하지 않는 타입은 기본적으로 info로 처리
            logging.warning(f"지원하지 않는 메시지 타입: {message_type}, info로 대체합니다.")
            messagebox.showinfo(title, message, parent=parent)
            return None

    def show_info(self, title, message, parent=None):
        """정보 메시지 표시"""
        return self._show_message('info', title, message, parent, 'info')

    def show_warning(self, title, message, parent=None):
        """경고 메시지 표시"""
        return self._show_message('warning', title, message, parent, 'warning')

    def show_error(self, title, message, parent=None):
        """오류 메시지 표시"""
        return self._show_message('error', title, message, parent, 'error')

    def show_confirm(self, title, message, parent=None):
        """확인/취소 메시지 표시, 사용자 선택 결과 반환"""
        return self._show_message('yesno', title, message, parent, 'info')

if __name__ == "__main__":
    multiprocessing.freeze_support()
    lockfile = os.path.join(tempfile.gettempdir(), "ExcelSearchTool.lock")
    try:
        fp = open(lockfile, "w")
        msvcrt.locking(fp.fileno(), msvcrt.LK_NBLCK, 1)
    except OSError:
        print("⚠ 이미 실행 중입니다.")
        sys.exit(0)

    # 설정 로드
    config_path = CONFIG_FILE
    config = load_config(config_path)
    data_path = config.get("data_path", "")

    root = tk.Tk()
    app = ExcelSearchApp(root, data_path)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()