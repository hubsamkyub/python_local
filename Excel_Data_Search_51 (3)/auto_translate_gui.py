import traceback
import shutil
import tempfile
import pythoncom
import win32com.client
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import openpyxl
import os
import requests
import uuid
import deepl
import sqlite3
import json
import time
from datetime import datetime
import hashlib
import gspread
from google.oauth2.service_account import Credentials


# 번역 API 키
GOOGLE_SHEET_ID = "19v86VBcbzEzzF4I6g8RgCthostf3YFTAQaFFDjQuYlk"
SERVICE_ACCOUNT_FILE = "dulcet-antler-462703-n8-d2fbdb362407.json"
DEEPL_API_KEY = "e5cd9a97-3319-41f7-aeb3-3db8491a8b1a:fx"
AZURE_API_KEY = "cc919cc9-c499-4c2c-bb13-b661ce5cec53"

LANG_CODES = {
    "EN": ("en", "EN"),      # (Azure, DeepL)
    "CN": ("zh-Hans", "ZH"), # CN: 간체
    "TW": ("zh-Hant", "ZH"), # TW: 번체
    "TH": ("th", "TH"),
    "PT": ("pt", "PT"),
    "ES": ("es", "ES"),
    "DE": ("de", "DE"),
    "FR": ("fr", "FR"),
}

TRANSLATE_ENGINES = ["DeepL", "Azure"]

def azure_translate(text, from_lang, to_langs, key, location="koreacentral"):
    endpoint = "https://api.cognitive.microsofttranslator.com"
    path = '/translate'
    constructed_url = endpoint + path

    params = {
        'api-version': '3.0',
        'from': from_lang,
        'to': to_langs
    }
    headers = {
        'Ocp-Apim-Subscription-Key': key,
        'Ocp-Apim-Subscription-Region': location,
        'Content-type': 'application/json',
        'X-ClientTraceId': str(uuid.uuid4())
    }
    body = [{'text': text}]
    response = requests.post(constructed_url, params=params, headers=headers, json=body, timeout=10)
    response.raise_for_status()
    return response.json()

def deepl_multi_translate(text, target_langs, auth_key):
    client = deepl.Translator(auth_key)
    result = {}
    for lang in target_langs:
        try:
            trans = client.translate_text(text, target_lang=lang)
            result[lang] = trans.text
        except Exception as e:
            result[lang] = f"[번역실패:{e}]"
    return result

def pass_translate_rule(kr):
    if not kr or str(kr).strip() == "":
        return True
    return False

import gspread
from google.oauth2.service_account import Credentials

class GoogleSheetsSyncGspread:
    """gspread 기반 Google Sheets 동기화 클래스"""
    def __init__(self, sheet_id, json_path, worksheet_index=0):
        self.sheet_id = sheet_id
        self.json_path = json_path
        self.worksheet_index = worksheet_index
        self.gc = None
        self.sheet = None
        self.ws = None

    def connect(self):
        scopes = ['https://www.googleapis.com/auth/spreadsheets']
        credentials = Credentials.from_service_account_file(self.json_path, scopes=scopes)
        self.gc = gspread.authorize(credentials)
        self.sheet = self.gc.open_by_key(self.sheet_id)
        self.ws = self.sheet.get_worksheet(self.worksheet_index)

    def download_sheet_data(self):
        """Google Sheets에서 데이터 다운로드 (헤더 기반 dict 반환)"""
        if self.ws is None:
            self.connect()
        records = self.ws.get_all_records()  # 리스트[dict] (헤더 기준)
        return records

    def upload_translations(self, translations):
        """번역 데이터를 Google Sheets에 업로드 (append)"""
        if self.ws is None:
            self.connect()

        # 업로드할 컬럼 순서에 맞게 데이터 추출
        # [KR_TEXT, EN, CN, ... , engine, contributor, updated_at, verified]
        for trans in translations:
            row = [
                trans.get('kr_text', ''),
                trans.get('EN', ''),
                trans.get('CN', ''),
                trans.get('TW', ''),
                trans.get('TH', ''),
                trans.get('PT', ''),
                trans.get('ES', ''),
                trans.get('DE', ''),
                trans.get('FR', ''),
                trans.get('engine', ''),
                trans.get('contributor', 'anonymous'),
                trans.get('updated_at', ''),
                trans.get('verified', 'FALSE'),
            ]
            self.ws.append_row(row, value_input_option="RAW")
        return True


class ConflictResolutionDialog:
    """번역 충돌 해결 다이얼로그"""
    def __init__(self, parent, conflicts):
        self.conflicts = conflicts
        self.resolved = []
        self.current_index = 0
        self.use_all_local = False
        self.use_all_server = False
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("번역 충돌 해결")
        self.dialog.geometry("600x500")
        
        self.create_widgets()
        self.show_conflict()
        
    def create_widgets(self):
        # 상단 정보
        self.info_label = ttk.Label(self.dialog, text="", font=('', 11, 'bold'))
        self.info_label.pack(pady=10)
        
        # 원본 텍스트
        kr_frame = ttk.LabelFrame(self.dialog, text="원본 텍스트")
        kr_frame.pack(fill='x', padx=20, pady=5)
        
        self.kr_text_label = ttk.Label(kr_frame, text="", wraplength=550)
        self.kr_text_label.pack(padx=10, pady=10)
        
        # 선택 옵션
        self.choice_var = tk.StringVar(value="local")
        
        # 내 번역
        local_frame = ttk.LabelFrame(self.dialog, text="내 번역")
        local_frame.pack(fill='x', padx=20, pady=5)
        
        ttk.Radiobutton(local_frame, text="이 번역 사용", 
                       variable=self.choice_var, value="local").pack(anchor='w', padx=10, pady=5)
        
        self.local_trans_label = ttk.Label(local_frame, text="")
        self.local_trans_label.pack(padx=30, pady=5, anchor='w')
        
        # 서버 번역
        server_frame = ttk.LabelFrame(self.dialog, text="서버 번역")
        server_frame.pack(fill='x', padx=20, pady=5)
        
        ttk.Radiobutton(server_frame, text="이 번역 사용", 
                       variable=self.choice_var, value="server").pack(anchor='w', padx=10, pady=5)
        
        self.server_trans_label = ttk.Label(server_frame, text="")
        self.server_trans_label.pack(padx=30, pady=5, anchor='w')
        
        self.server_info_label = ttk.Label(server_frame, text="", font=('', 9))
        self.server_info_label.pack(padx=30, pady=2, anchor='w')
        
        # 일괄 적용 옵션
        batch_frame = ttk.Frame(self.dialog)
        batch_frame.pack(fill='x', padx=20, pady=10)
        
        ttk.Button(batch_frame, text="남은 항목 모두 내 번역 사용", 
                  command=self.use_all_local_trans).pack(side='left', padx=5)
        ttk.Button(batch_frame, text="남은 항목 모두 서버 번역 사용", 
                  command=self.use_all_server_trans).pack(side='left', padx=5)
        
        # 하단 버튼
        button_frame = ttk.Frame(self.dialog)
        button_frame.pack(side='bottom', pady=20)
        
        ttk.Button(button_frame, text="이전", command=self.prev_conflict).pack(side='left', padx=5)
        ttk.Button(button_frame, text="다음", command=self.next_conflict).pack(side='left', padx=5)
        ttk.Button(button_frame, text="완료", command=self.finish).pack(side='left', padx=20)
        
    def show_conflict(self):
        if self.current_index >= len(self.conflicts):
            return
            
        conflict = self.conflicts[self.current_index]
        
        # 정보 업데이트
        self.info_label.config(text=f"충돌 {self.current_index + 1} / {len(self.conflicts)}")
        self.kr_text_label.config(text=conflict['kr'])
        
        # 번역 정보 표시
        local_trans = json.loads(conflict['local'])
        server_trans = json.loads(conflict['server'])
        
        local_text = "\n".join([f"{lang}: {trans}" for lang, trans in local_trans.items()])
        server_text = "\n".join([f"{lang}: {trans}" for lang, trans in server_trans.items()])
        
        self.local_trans_label.config(text=local_text)
        self.server_trans_label.config(text=server_text)
        self.server_info_label.config(text=f"작성자: {conflict.get('contributor', 'unknown')} | {conflict.get('updated_at', '')}")
        
        # 일괄 적용 중이면 자동 선택
        if self.use_all_local:
            self.choice_var.set("local")
        elif self.use_all_server:
            self.choice_var.set("server")
            
    def next_conflict(self):
        self.save_current_choice()
        if self.current_index < len(self.conflicts) - 1:
            self.current_index += 1
            self.show_conflict()
            
    def prev_conflict(self):
        self.save_current_choice()
        if self.current_index > 0:
            self.current_index -= 1
            self.show_conflict()
            
    def save_current_choice(self):
        if self.current_index < len(self.conflicts):
            conflict = self.conflicts[self.current_index]
            choice = self.choice_var.get()
            
            resolved = {
                'kr_text': conflict['kr'],
                'translations': conflict['local'] if choice == 'local' else conflict['server'],
                'choice': choice
            }
            
            # 기존 선택 덮어쓰기
            if self.current_index < len(self.resolved):
                self.resolved[self.current_index] = resolved
            else:
                self.resolved.append(resolved)
                
    def use_all_local_trans(self):
        self.use_all_local = True
        self.use_all_server = False
        messagebox.showinfo("알림", "남은 항목에 모두 내 번역을 사용합니다.")
        
    def use_all_server_trans(self):
        self.use_all_local = False
        self.use_all_server = True
        messagebox.showinfo("알림", "남은 항목에 모두 서버 번역을 사용합니다.")
        
    def finish(self):
        self.save_current_choice()
        
        # 남은 충돌 자동 해결
        while self.current_index < len(self.conflicts) - 1:
            self.current_index += 1
            if self.use_all_local:
                self.choice_var.set("local")
            elif self.use_all_server:
                self.choice_var.set("server")
            self.save_current_choice()
            
        self.dialog.destroy()
        
    def get_resolved_translations(self):
        self.dialog.wait_window()
        return self.resolved


def ensure_translation_table_columns(conn):
    """translations 테이블에 누락 컬럼이 있으면 자동으로 추가"""
    required_columns = {
        "contributor": "TEXT DEFAULT 'anonymous'",
        "engine": "TEXT DEFAULT ''",
        "updated_at": "TEXT DEFAULT ''",
        "verified": "BOOLEAN DEFAULT 0",
        "synced_at": "TIMESTAMP",
        "created_at": "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
    }
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(translations)")
    existing_columns = [row[1] for row in cursor.fetchall()]

    for col, col_type in required_columns.items():
        if col not in existing_columns:
            try:
                cursor.execute(f"ALTER TABLE translations ADD COLUMN {col} {col_type}")
            except Exception as e:
                print(f"[DB] 컬럼 추가 오류: {col}: {e}")
    conn.commit()


class AutoTranslateGUI:
    def __init__(self, master):
        self.master = master
        master.title("자동 번역 도구 (DeepL/Azure) + Google Sheets 동기화")
        master.geometry("900x750")

        # Google Sheets 동기화 객체
        self.sheets_sync = GoogleSheetsSyncGspread(GOOGLE_SHEET_ID, SERVICE_ACCOUNT_FILE)
        
        # 파일 선택
        self.file_path = tk.StringVar()
        ttk.Label(master, text="엑셀 파일 선택:").pack(pady=(10,0), anchor='w', padx=10)
        frame1 = ttk.Frame(master)
        frame1.pack(fill='x', padx=10)
        ttk.Entry(frame1, textvariable=self.file_path, width=75).pack(side='left', padx=5)
        ttk.Button(frame1, text="찾아보기", command=self.browse_file).pack(side='left', padx=5)

        # DB 상태 표시
        db_status_frame = ttk.Frame(master)
        db_status_frame.pack(fill='x', padx=10, pady=5)
        
        self.db_status_label = ttk.Label(db_status_frame, text="DB 상태: ● 로컬 DB")
        self.db_status_label.pack(side='left', padx=5)
        
        self.sync_status_label = ttk.Label(db_status_frame, text="(동기화 필요)")
        self.sync_status_label.pack(side='left', padx=5)
        
        ttk.Button(db_status_frame, text="DB 동기화", command=self.sync_database).pack(side='right', padx=5)
        ttk.Button(db_status_frame, text="대기 중인 번역 업로드", command=self.upload_pending).pack(side='right', padx=5)

        # 번역 언어 선택
        self.langs = list(LANG_CODES.keys())
        self.selected_langs = {lang: tk.BooleanVar(value=(lang=="EN")) for lang in self.langs}
        ttk.Label(master, text="번역 언어:").pack(anchor='w', padx=10, pady=(10,0))
        frame2 = ttk.Frame(master)
        frame2.pack(fill='x', padx=10)
        for lang in self.langs:
            ttk.Checkbutton(frame2, text=lang, variable=self.selected_langs[lang]).pack(side='left', padx=5)

        # 번역 엔진 선택
        ttk.Label(master, text="번역 엔진:").pack(anchor='w', padx=10, pady=(10,0))
        self.engine_var = tk.StringVar(value=TRANSLATE_ENGINES[0])
        frame3 = ttk.Frame(master)
        frame3.pack(fill='x', padx=10)
        for eng in TRANSLATE_ENGINES:
            ttk.Radiobutton(frame3, text=eng, value=eng, variable=self.engine_var).pack(side='left', padx=5)

        # API 키 입력
        ttk.Label(master, text="DeepL API Key / Azure API Key:").pack(anchor='w', padx=10, pady=(10,0))
        frame4 = ttk.Frame(master)
        frame4.pack(fill='x', padx=10)
        self.deepl_key = tk.StringVar()
        self.azure_key = tk.StringVar()
        self.azure_region = tk.StringVar(value="koreacentral")
        ttk.Label(frame4, text="DeepL:").pack(side='left', padx=3)
        ttk.Entry(frame4, textvariable=self.deepl_key, width=30, show="*").pack(side='left', padx=3)
        ttk.Label(frame4, text="Azure:").pack(side='left', padx=3)
        ttk.Entry(frame4, textvariable=self.azure_key, width=30, show="*").pack(side='left', padx=3)
        ttk.Label(frame4, text="Region:").pack(side='left', padx=3)
        ttk.Entry(frame4, textvariable=self.azure_region, width=14).pack(side='left', padx=3)

        # 미리보기/실행 버튼
        btn_frame = ttk.Frame(master)
        btn_frame.pack(fill='x', pady=10, padx=10)
        ttk.Button(btn_frame, text="번역 대상 미리보기", command=self.preview).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="자동 번역 실행", command=self.start_translate).pack(side='left', padx=5)
        
        # 전체 선택/해제 체크박스와 통계 정보
        preview_control_frame = ttk.Frame(master)
        preview_control_frame.pack(fill='x', padx=10)
        
        self.select_all_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(preview_control_frame, text="전체 선택", 
                        variable=self.select_all_var, 
                        command=self.toggle_all_checks).pack(side='left', padx=5)
        
        self.stats_label = ttk.Label(preview_control_frame, text="")
        self.stats_label.pack(side='left', padx=20)
        
        # 미리보기 테이블 프레임 (스크롤바 포함)
        tree_frame = ttk.Frame(master)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # 미리보기 테이블
        columns = ('check', 'STRING_ID', 'KR') + tuple(self.langs)
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='tree headings', height=15)
        
        # 컬럼 설정
        self.tree.column('#0', width=0, stretch=False)
        self.tree.column('check', width=50, stretch=False)
        self.tree.column('STRING_ID', width=150, stretch=False)
        self.tree.column('KR', width=200)
        
        self.tree.heading('check', text='선택')
        self.tree.heading('STRING_ID', text='STRING_ID')
        self.tree.heading('KR', text='KR')
        
        for lang in self.langs:
            self.tree.column(lang, width=90)
            self.tree.heading(lang, text=lang)
        
        # 스크롤바 추가
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # 배치
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # 체크박스 상태 저장용 딕셔너리
        self.check_states = {}
        
        # 트리뷰 클릭 이벤트 바인딩
        self.tree.bind('<Button-1>', self.on_tree_click)

        # 로그창
        self.log_text = tk.Text(master, height=8)
        self.log_text.pack(fill='both', padx=10, pady=5, expand=True)
        
        # DB 초기화
        self.init_cache_db()
        self.update_db_status()

    def browse_file(self):
        fpath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if fpath:
            self.file_path.set(fpath)

    def preview(self):
        self.tree.delete(*self.tree.get_children())
        self.check_states.clear()
        path = self.file_path.get()
        if not path or not os.path.exists(path):
            messagebox.showerror("오류", "엑셀 파일을 선택하세요.")
            return
        
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            messagebox.showerror("오류", f"파일을 열 수 없습니다: {e}")
            return
            
        ws = wb.active
        header_row = 4
        header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]
        cols = {cell.value: idx for idx, cell in enumerate(header_cells, 1)}
        
        kr_col = cols.get("KR")
        if not kr_col:
            messagebox.showerror("오류", f"{header_row}행에 'KR' 컬럼명이 없습니다.")
            wb.close()
            return
        
        string_id_col = cols.get("STRING_ID", 1)
        sel_lang_cols = {lang: cols.get(lang) for lang in self.langs if self.selected_langs[lang].get()}
        
        total_chars = 0
        translation_needed = 0
        shown = 0
        
        for row in ws.iter_rows(min_row=header_row+1, values_only=True):
            if not row or kr_col-1 >= len(row) or not row[kr_col-1]:
                continue
            
            kr_text = str(row[kr_col-1])
            string_id = str(row[string_id_col-1]) if string_id_col-1 < len(row) else f"Row_{shown+1}"
            
            # 번역이 필요한지 확인
            needs_translation = False
            if not pass_translate_rule(kr_text):
                for lang, idx in sel_lang_cols.items():
                    if idx and (idx-1 >= len(row) or not row[idx-1]):
                        needs_translation = True
                        break
            
            # 트리에 아이템 추가 (번역 필요 여부와 관계없이)
            row_data = ['☑' if needs_translation else '☐', string_id, kr_text]
            for lang in self.langs:
                if lang in sel_lang_cols:
                    idx = sel_lang_cols[lang]
                    row_data.append(row[idx-1] if idx and idx-1 < len(row) else "")
                else:
                    row_data.append("")
            
            # 아이템 추가 및 스타일 설정
            item = self.tree.insert('', 'end', values=row_data)
            self.check_states[item] = needs_translation
            
            # 번역이 필요한 항목은 굵게 표시
            if needs_translation:
                self.tree.item(item, tags=('needs_translation',))
                total_chars += len(kr_text)
                translation_needed += 1
            else:
                self.tree.item(item, tags=('already_translated',))
            
            shown += 1
            if shown >= 500:
                self.tree.insert('', 'end', values=['', '...', f'(추가 항목 있음)', '', '', '', '', '', '', ''])
                break
        
        # 태그 스타일 설정
        self.tree.tag_configure('needs_translation', foreground='black')
        self.tree.tag_configure('already_translated', foreground='gray')
        
        wb.close()
        
        # 통계 정보 업데이트
        self.stats_label.config(text=f"전체: {shown}개 | 번역 필요: {translation_needed}개 | 총 {total_chars:,}자")

    def on_tree_click(self, event):
        """트리뷰 클릭 이벤트 처리"""
        region = self.tree.identify_region(event.x, event.y)
        if region == "cell":
            column = self.tree.identify_column(event.x)
            item = self.tree.identify_row(event.y)
            
            # 체크박스 컬럼 클릭 시
            if column == '#1' and item:
                current_state = self.check_states.get(item, True)
                new_state = not current_state
                self.check_states[item] = new_state
                
                # 체크박스 표시 업데이트
                values = list(self.tree.item(item, 'values'))
                values[0] = '☑' if new_state else '☐'
                self.tree.item(item, values=values)
                
                # 전체 선택 체크박스 상태 업데이트
                self.update_select_all_state()

    def toggle_all_checks(self):
        """전체 선택/해제"""
        check_all = self.select_all_var.get()
        
        for item in self.tree.get_children():
            values = list(self.tree.item(item, 'values'))
            if values[1] != '...':
                self.check_states[item] = check_all
                values[0] = '☑' if check_all else '☐'
                self.tree.item(item, values=values)

    def update_select_all_state(self):
        """전체 선택 체크박스 상태 업데이트"""
        all_checked = all(self.check_states.get(item, True) for item in self.tree.get_children() 
                         if self.tree.item(item, 'values')[1] != '...')
        self.select_all_var.set(all_checked)

    def get_excluded_string_ids(self):
        """체크 해제된 STRING_ID 목록 반환"""
        excluded = []
        for item in self.tree.get_children():
            if not self.check_states.get(item, True):
                values = self.tree.item(item, 'values')
                if values[1] != '...':
                    excluded.append(values[1])
        return excluded

    def log(self, msg):
        self.log_text.insert('end', msg+'\n')
        self.log_text.see('end')
        self.master.update_idletasks()

    def start_translate(self):
        # 번역 시작 전 동기화 확인
        if messagebox.askyesno("동기화", "최신 번역 DB를 다운로드하시겠습니까?\n(권장: 예)"):
            self.sync_database()
        
        threading.Thread(target=self._translate_proc, daemon=True).start()

    def _translate_proc(self):
        path = self.file_path.get()
        if not path or not os.path.exists(path):
            self.log("엑셀 파일을 선택하세요.")
            return
        
        # 파일이 이미 열려있는지 확인
        try:
            with open(path, 'r+b') as f:
                pass
        except IOError:
            self.log("파일이 다른 프로그램에서 사용 중입니다. Excel을 닫고 다시 시도하세요.")
            messagebox.showerror("오류", "파일이 다른 프로그램에서 사용 중입니다.\nExcel을 닫고 다시 시도하세요.")
            return
        
        # 제외할 STRING_ID 목록 가져오기
        excluded_ids = self.get_excluded_string_ids() if hasattr(self, 'check_states') else []
        
        excel = None
        wb = None
        
        try:
            # COM 초기화
            pythoncom.CoInitialize()
            
            # 임시 파일로 복사
            temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(temp_fd)
            shutil.copy2(path, temp_path)
            
            # Excel 애플리케이션 시작
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # 임시 파일로 작업
            wb = excel.Workbooks.Open(temp_path)
            ws = wb.ActiveSheet
            
            header_row = 4
            # 헤더 읽기
            cols = {}
            for col in range(1, min(ws.UsedRange.Columns.Count + 1, 100)):
                val = ws.Cells(header_row, col).Value
                if val:
                    cols[val] = col
            
            kr_col = cols.get("KR")
            if not kr_col:
                self.log("'KR' 컬럼을 찾을 수 없습니다.")
                return
            
            sel_langs = [lang for lang in self.langs if self.selected_langs[lang].get()]
            sel_lang_cols = {lang: cols.get(lang) for lang in sel_langs}
            
            engine = self.engine_var.get()
            deepl_key = self.deepl_key.get().strip() or DEEPL_API_KEY
            azure_key = self.azure_key.get().strip() or AZURE_API_KEY
            azure_region = self.azure_region.get().strip()
            
            # 실제 데이터가 있는 마지막 행 찾기
            last_row = ws.Cells(ws.Rows.Count, kr_col).End(-4162).Row
            total = last_row - header_row
            done = 0
            
            self.log(f"번역 시작: 총 {total}개 행")
            
            for row_num in range(header_row + 1, last_row + 1):
                try:
                    kr = ws.Cells(row_num, kr_col).Value
                    if pass_translate_rule(kr):
                        continue
                    
                    # STRING_ID 확인 및 제외 처리
                    string_id_col_idx = cols.get("STRING_ID", 1)
                    string_id = ws.Cells(row_num, string_id_col_idx).Value
                    if str(string_id) in excluded_ids:
                        self.log(f"[제외] STRING_ID: {string_id}")
                        done += 1
                        continue
                        
                    targets = []
                    col_idx_map = {}
                    for lang in sel_langs:
                        idx = sel_lang_cols[lang]
                        if not idx or ws.Cells(row_num, idx).Value:
                            continue
                        targets.append(lang)
                        col_idx_map[lang] = idx
                        
                    if not targets:
                        done += 1
                        continue
                        
                    # 캐시 확인
                    cached = self.get_cached_translation(kr, engine)
                    
                    if cached:
                        for lang in targets:
                            if lang in cached:
                                idx = col_idx_map[lang]
                                ws.Cells(row_num, idx).Value = cached[lang]
                                self.log(f"[{done+1}/{total}] {lang} 번역(캐시): {kr[:16]} → {cached[lang][:16]}")
                    else:
                        translation_results = {}
                        
                        if engine == "DeepL":
                            if not deepl_key:
                                self.log("DeepL API Key 입력 필요")
                                return
                            trans_results = deepl_multi_translate(kr, [LANG_CODES[l][1] for l in targets], deepl_key)
                            for lang in targets:
                                idx = col_idx_map[lang]
                                deepl_code = LANG_CODES[lang][1]
                                translated_text = trans_results.get(deepl_code, "")
                                ws.Cells(row_num, idx).Value = translated_text
                                translation_results[lang] = translated_text
                                self.log(f"[{done+1}/{total}] {lang} 번역(API): {kr[:16]} → {translated_text[:16]}")
                        else:  # Azure
                            if not azure_key:
                                self.log("Azure API Key 입력 필요")
                                return
                            codes = [LANG_CODES[l][0] for l in targets]
                            result = azure_translate(kr, "ko", codes, azure_key, azure_region)
                            for t in result[0]['translations']:
                                code = t['to']
                                lang = None
                                for k, v in LANG_CODES.items():
                                    if v[0] == code and k in targets:
                                        lang = k
                                        break
                                if lang:
                                    idx = col_idx_map[lang]
                                    ws.Cells(row_num, idx).Value = t['text']
                                    translation_results[lang] = t['text']
                                    self.log(f"[{done+1}/{total}] {lang} 번역(API): {kr[:16]} → {t['text'][:16]}")
                        
                        if translation_results:
                            self.save_translation(kr, translation_results, engine)
                            
                except Exception as e:
                    self.log(f"행 {row_num} 처리 오류: {e}")
                    continue
                
                done += 1
                
                # 100행마다 저장 (메모리 및 안정성)
                if done % 100 == 0:
                    wb.Save()
                    self.log(f"중간 저장 완료: {done}행 처리됨")
            
            # 최종 저장
            wb.Save()
            wb.Close()
            excel.Quit()
            
            # 원본 파일로 복사
            try:
                shutil.move(temp_path, path)
                self.log(f"완료! 원본 파일에 저장됨: {path}")
            except Exception as e:
                self.log(f"원본 파일 덮어쓰기 실패: {e}")
                self.log(f"임시 파일 위치: {temp_path}")
                messagebox.showwarning("경고", f"원본 파일을 덮어쓸 수 없습니다.\n임시 파일: {temp_path}")
            
            # 대기 중인 번역 개수 업데이트
            self.update_db_status()
            
        except Exception as e:
            self.log(f"전체 오류: {e}")
        finally:
            # 정리 작업
            try:
                if wb:
                    wb.Close(SaveChanges=False)
            except:
                pass
            
            try:
                if excel:
                    excel.Quit()
            except:
                pass
            
            # COM 해제
            try:
                pythoncom.CoUninitialize()
            except:
                pass
            
            # 임시 파일 정리
            try:
                if 'temp_path' in locals() and os.path.exists(temp_path):
                    os.remove(temp_path)
            except:
                pass
            
            # 가비지 컬렉션
            import gc
            gc.collect()

    def init_cache_db(self):
        """번역 캐시 DB 초기화"""
        self.db_path = "auto_translate.db"
        conn = sqlite3.connect(self.db_path)
        ensure_translation_table_columns(conn)
        cursor = conn.cursor()
        
        # 메인 번역 테이블
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS translations (
                kr_text TEXT PRIMARY KEY,
                translations TEXT,
                engine TEXT,
                contributor TEXT DEFAULT 'anonymous',
                verified BOOLEAN DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                synced_at TIMESTAMP
            )
        ''')
        
        # 업로드 대기 테이블
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS pending_uploads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kr_text TEXT,
                translations TEXT,
                engine TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 동기화 메타데이터
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS sync_metadata (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        ''')
        
        conn.commit()
        conn.close()


    def get_cached_translation(self, kr_text, engine):
        """캐시된 번역 결과 조회"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT translations FROM translations WHERE kr_text = ? AND engine = ?",
            (kr_text, engine)
        )
        result = cursor.fetchone()
        conn.close()
        if result:
            return json.loads(result[0])
        return None

    def save_translation(self, kr_text, translations, engine):
        """번역 결과를 DB에 저장"""
        conn = sqlite3.connect(self.db_path)
        ensure_translation_table_columns(conn)
        cursor = conn.cursor()
        
        # 메인 테이블에 저장
        cursor.execute(
            "INSERT OR REPLACE INTO translations (kr_text, translations, engine) VALUES (?, ?, ?)",
            (kr_text, json.dumps(translations), engine)
        )
        
        # 대기 테이블에도 추가
        cursor.execute(
            "INSERT INTO pending_uploads (kr_text, translations, engine) VALUES (?, ?, ?)",
            (kr_text, json.dumps(translations), engine)
        )
        
        conn.commit()
        conn.close()

    def sync_database(self):
        """Google Sheets와 로컬 DB 동기화"""
        self.log("Google Sheets 동기화 시작...")
        
        try:
            self.log("구글 시트 객체 생성 시도...")
            sheet_data = self.sheets_sync.download_sheet_data()
            self.log(f"시트 데이터 개수: {len(sheet_data)}")
            
            if not sheet_data:
                self.log("Google Sheets에서 데이터를 가져올 수 없습니다.")
                return
            
            self.log(f"Google Sheets에서 {len(sheet_data)}개 항목 다운로드 완료")
            
            # 로컬 DB와 병합
            conn = sqlite3.connect(self.db_path)
            ensure_translation_table_columns(conn)
            cursor = conn.cursor()
            
            conflicts = []
            updated = 0
            new_items = 0
            
            for row in sheet_data:
                kr_text = row.get('KR', '')
                if not kr_text:
                    continue
                
                # 번역 데이터 구성
                translations = {}
                for lang in self.langs:
                    if row.get(lang):
                        translations[lang] = row.get(lang)
                
                if not translations:
                    continue
                
                engine = row.get('ENGINE', 'unknown')
                contributor = row.get('CONTRIBUTOR', 'unknown')
                updated_at = row.get('UPDATED_AT', '')
                verified = row.get('VERIFIED', 'FALSE') == 'TRUE'
                
                # 기존 번역 확인
                cursor.execute("SELECT translations FROM translations WHERE kr_text = ?", (kr_text,))
                existing = cursor.fetchone()
                
                if existing:
                    existing_trans = json.loads(existing[0])
                    if existing_trans != translations:
                        # 충돌 발생
                        conflicts.append({
                            'kr': kr_text,
                            'local': json.dumps(existing_trans),
                            'server': json.dumps(translations),
                            'contributor': contributor,
                            'updated_at': updated_at
                        })
                    else:
                        updated += 1
                else:
                    # 새 항목 추가
                    cursor.execute('''
                        INSERT INTO translations 
                        (kr_text, translations, engine, contributor, verified, synced_at)
                        VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    ''', (kr_text, json.dumps(translations), engine, contributor, verified))
                    new_items += 1
            
            # 충돌 해결
            if conflicts:
                self.log(f"{len(conflicts)}개 충돌 발견. 해결 중...")
                dialog = ConflictResolutionDialog(self.master, conflicts)
                resolved = dialog.get_resolved_translations()
                
                for item in resolved:
                    if item['choice'] == 'server':
                        cursor.execute(
                            "INSERT OR REPLACE INTO translations (kr_text, translations, synced_at) VALUES (?, ?, CURRENT_TIMESTAMP)",
                            (item['kr_text'], item['translations'])
                        )
                        updated += 1
            
            # 마지막 동기화 시간 저장
            cursor.execute(
                "INSERT OR REPLACE INTO sync_metadata (key, value) VALUES ('last_sync', ?)",
                (datetime.now().strftime('%Y-%m-%d %H:%M:%S'),)
            )
            
            conn.commit()
            conn.close()
            
            self.log(f"동기화 완료: 신규 {new_items}개, 업데이트 {updated}개")
            self.update_db_status()
            
        except Exception as e:
            tb = traceback.format_exc()
            self.log(f"동기화 오류: {e}\n{tb}")
            messagebox.showerror("오류", f"동기화 중 오류가 발생했습니다: {e}\n{tb}")

    def upload_pending(self):
        """대기 중인 번역을 Google Sheets에 업로드"""
        conn = sqlite3.connect(self.db_path)
        ensure_translation_table_columns(conn)
        cursor = conn.cursor()
        
        # 대기 중인 번역 조회
        cursor.execute("SELECT DISTINCT kr_text, translations, engine FROM pending_uploads")
        pending = cursor.fetchall()
        
        if not pending:
            messagebox.showinfo("알림", "업로드할 번역이 없습니다.")
            return
        
        self.log(f"{len(pending)}개 번역 업로드 시작...")
        
        # Google Sheets 형식으로 변환
        upload_data = []
        for kr_text, translations_json, engine in pending:
            translations = json.loads(translations_json)
            item = {
                'kr_text': kr_text,
                'engine': engine,
                'contributor': 'anonymous'  # 추후 사용자 설정으로 변경 가능
            }
            item.update(translations)
            upload_data.append(item)
        
        # 업로드
        success = self.sheets_sync.upload_translations(upload_data)
        
        if success:
            # 업로드 성공 시 대기 테이블 비우기
            cursor.execute("DELETE FROM pending_uploads")
            conn.commit()
            self.log("업로드 완료!")
            messagebox.showinfo("완료", f"{len(pending)}개 번역이 업로드되었습니다.")
        else:
            self.log("업로드 실패!")
            messagebox.showerror("오류", "업로드 중 오류가 발생했습니다.")
        
        conn.close()
        self.update_db_status()

    def update_db_status(self):
        """DB 상태 업데이트"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 전체 번역 수
        cursor.execute("SELECT COUNT(*) FROM translations")
        total = cursor.fetchone()[0]
        
        # 대기 중인 번역 수
        cursor.execute("SELECT COUNT(DISTINCT kr_text) FROM pending_uploads")
        pending = cursor.fetchone()[0]
        
        # 마지막 동기화 시간
        cursor.execute("SELECT value FROM sync_metadata WHERE key = 'last_sync'")
        last_sync = cursor.fetchone()
        
        conn.close()
        
        # UI 업데이트
        self.db_status_label.config(text=f"DB 상태: ● 로컬 캐시 {total:,}개")
        
        if pending > 0:
            self.sync_status_label.config(text=f"(대기 중: {pending}개)")
        elif last_sync:
            sync_time = datetime.strptime(last_sync[0], '%Y-%m-%d %H:%M:%S')
            time_diff = datetime.now() - sync_time
            
            if time_diff.days > 0:
                time_str = f"{time_diff.days}일 전"
            elif time_diff.seconds > 3600:
                time_str = f"{time_diff.seconds // 3600}시간 전"
            else:
                time_str = f"{time_diff.seconds // 60}분 전"
            
            self.sync_status_label.config(text=f"(최종 동기화: {time_str})")
        else:
            self.sync_status_label.config(text="(동기화 필요)")

if __name__ == "__main__":
    root = tk.Tk()
    app = AutoTranslateGUI(root)
    root.mainloop()