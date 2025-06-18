import os
import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import pandas as pd
import json
from datetime import datetime

class NPCCreatorPopup:
    def __init__(self, parent, excel_path, db_path, excel_cache=None):
        self.parent = parent
        self.excel_path = excel_path
        self.db_path = db_path
        self.excel_cache = excel_cache or {}
        
        # 팝업 창 생성
        self.popup = tk.Toplevel(parent)
        self.popup.title("NPC 생성기")
        self.popup.geometry("1400x800")
        self.popup.grab_set()
        
        # 히스토리 및 프리셋 데이터 로드
        self.history_file = os.path.join(".cache", "npc_creator_history.json")
        self.presets_file = os.path.join(".cache", "npc_model_presets.json")
        self.npc_history = self.load_history()
        self.model_presets = self.load_presets()
        
        # 데이터 필드 초기화
        self.unique_id_var = tk.StringVar()
        self.base_group_id_var = tk.StringVar()
        self.name_var = tk.StringVar()
        self.nickname_var = tk.StringVar()
        self.event_group_id_var = tk.StringVar()
        self.model_hero_id_var = tk.StringVar()
        self.size_x_var = tk.StringVar(value="3")
        self.size_y_var = tk.StringVar(value="3")
        
        # 경로 관련 변수
        # 엑셀 파일 변수
        self.excel_file_var = tk.StringVar(value=os.path.join(excel_path, "HeroTemplate@_AutoGen.xlsx"))
        import warnings
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        
        # 카테고리 타입 체크박스 변수
        # CategoryType 변수 정의 변경
        self.category_types = [
            ("CommonHero", 10),
            ("LimitBreakHero", 11),
            ("Monster", 20),
            ("BossMonster", 21),
            ("NPC", 30),
            ("NoShadowNPC", 31),
            ("NoBlockNPC", 32),
            ("MarkNPC", 51),
            ("ETC", 99)
        ]
        self.category_var = tk.IntVar(value=30)  # 기본값 NPC(30)
        
        # 배치 관련 변수
        self.map_id_var = tk.StringVar()
        self.map_spawn_group_id_var = tk.StringVar(value="0")
        self.helper_name_var = tk.StringVar()
        self.hero_id_var = tk.StringVar()
        self.direction_id_var = tk.IntVar(value=0)
        self.show_condition_tid_var = tk.StringVar(value="0")
        self.hide_condition_tid_var = tk.StringVar(value="0")
        
        # 현재 선택된 NPC 정보
        self.selected_npc = None
        
        # UI 구성
        self.build_ui()
        
        # 초기 데이터 로드
        self.load_npc_list()
    
    def load_history(self):
        """NPC 생성 히스토리 로드"""
        if os.path.exists(self.history_file):
            try:
                with open(self.history_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return []
        return []
    
    def save_history(self, npc_data):
        """NPC 생성 히스토리 저장"""
        # 디렉토리 확인 및 생성
        os.makedirs(os.path.dirname(self.history_file), exist_ok=True)
        
        # 히스토리에 저장할 데이터 준비
        category_name = next((name for name, value in self.category_types if value == npc_data["CategoryType"]), "Unknown")
    
        history_entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "UniqueID": npc_data["UniqueID"],
            "Name": npc_data["Name"],
            "NickName": npc_data["NickName"],
            "CategoryType": npc_data["CategoryType"],
            "CategoryName": category_name,  # 표시용 이름 추가
            "BaseHeroID": npc_data["BaseHeroID"],
            "EventGroupID": npc_data["EventGroupID"],
            "ModelID": npc_data["ModelID"],
            "SizeX": npc_data["SizeX"],
            "SizeY": npc_data["SizeY"]
        }
        
        # 히스토리에 추가
        self.npc_history.insert(0, history_entry)
        
        # 히스토리 크기 제한 (최대 100개)
        if len(self.npc_history) > 100:
            self.npc_history = self.npc_history[:100]
        
        # 파일에 저장
        with open(self.history_file, 'w', encoding='utf-8') as f:
            json.dump(self.npc_history, f, ensure_ascii=False, indent=2)
        
        # 히스토리 목록 업데이트
        self.update_history_list()
    
    def load_presets(self):
        """모델 프리셋 로드"""
        if os.path.exists(self.presets_file):
            try:
                with open(self.presets_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return []
        return []
    
    def save_presets(self):
        """모델 프리셋 저장"""
        # 디렉토리 확인 및 생성
        os.makedirs(os.path.dirname(self.presets_file), exist_ok=True)
        
        # 파일에 저장
        with open(self.presets_file, 'w', encoding='utf-8') as f:
            json.dump(self.model_presets, f, ensure_ascii=False, indent=2)
    
    def build_ui(self):
        """UI 구성"""
        # 노트북 (탭) 생성
        self.notebook = ttk.Notebook(self.popup)        
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 생성 탭
        self.create_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.create_tab, text="NPC 생성")
        
        # 배치 탭
        self.placement_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.placement_tab, text="NPC 배치")
        
        # 생성 탭 구성
        self.build_create_tab()
        
        # 배치 탭 구성
        self.build_placement_tab()
                
    
    def build_create_tab(self):
        """NPC 생성 탭 구성"""
        create_frame = ttk.Frame(self.create_tab)
        create_frame.pack(fill=tk.BOTH, expand=True)
        
        # 좌측 프레임 (입력 필드들)
        left_frame = ttk.Frame(create_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        # 상단: 엑셀 파일 경로 설정
        path_frame = ttk.LabelFrame(left_frame, text="엑셀 파일 경로")
        path_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Entry(path_frame, textvariable=self.excel_file_var, width=50).pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)
        ttk.Button(path_frame, text="찾기", command=self.select_excel_file).pack(side=tk.LEFT, padx=5, pady=5)
        
        # 중간: 입력 필드들
        input_frame = ttk.LabelFrame(left_frame, text="NPC 기본 정보")
        input_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 제목
        title_label = ttk.Label(input_frame, text="NPC 기본 정보 입력", font=("Arial", 12, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=10, sticky="w")
        
        # 입력 필드들
        current_row = 1
        
        # UniqueID
        ttk.Label(input_frame, text="UniqueID:").grid(row=current_row, column=0, sticky="w", pady=5)
        id_frame = ttk.Frame(input_frame)
        id_frame.grid(row=current_row, column=1, sticky="w")
        ttk.Entry(id_frame, textvariable=self.unique_id_var, width=15).pack(side=tk.LEFT, padx=(0,5))
        ttk.Button(id_frame, text="유효성 검사", command=self.validate_unique_id).pack(side=tk.LEFT, padx=(0,5))
        ttk.Button(id_frame, text="NPC 검색", command=self.show_hero_search_popup).pack(side=tk.LEFT)
        current_row += 1
        
        # BaseGroupID
        ttk.Label(input_frame, text="BaseHeroID:").grid(row=current_row, column=0, sticky="w", pady=5)
        ttk.Entry(input_frame, textvariable=self.base_group_id_var, width=15).grid(row=current_row, column=1, sticky="w")
        current_row += 1
        
        # Name
        ttk.Label(input_frame, text="Name:").grid(row=current_row, column=0, sticky="w", pady=5)
        ttk.Entry(input_frame, textvariable=self.name_var, width=30).grid(row=current_row, column=1, sticky="w")
        current_row += 1
        
        # NickName
        ttk.Label(input_frame, text="NickName:").grid(row=current_row, column=0, sticky="w", pady=5)
        ttk.Entry(input_frame, textvariable=self.nickname_var, width=30).grid(row=current_row, column=1, sticky="w")
        current_row += 1
        
        # CategoryType UI 변경 (라디오 버튼)
        ttk.Label(input_frame, text="CategoryType:").grid(row=current_row, column=0, sticky="w", pady=5)
        cat_frame = ttk.Frame(input_frame)
        cat_frame.grid(row=current_row, column=1, sticky="w")
        
        # 라디오 버튼 생성 (2열로 배치)
        for i, (cat_text, cat_value) in enumerate(self.category_types):
            row = i // 5  # 한 줄에 5개씩
            col = i % 5
            ttk.Radiobutton(
                cat_frame, 
                text=cat_text, 
                variable=self.category_var, 
                value=cat_value
            ).grid(row=row, column=col, padx=5, sticky="w")
        current_row += 1
            
        # EventGroupID
        ttk.Label(input_frame, text="EventGroupID:").grid(row=current_row, column=0, sticky="w", pady=5)
        ttk.Entry(input_frame, textvariable=self.event_group_id_var, width=15).grid(row=current_row, column=1, sticky="w")
        current_row += 1
        
        # ModelHero ID
        ttk.Label(input_frame, text="외형 선택(ModelHero ID):").grid(row=current_row, column=0, sticky="w", pady=5)
        model_frame = ttk.Frame(input_frame)
        model_frame.grid(row=current_row, column=1, sticky="w")
        ttk.Entry(model_frame, textvariable=self.model_hero_id_var, width=15).pack(side=tk.LEFT, padx=(0,5))
        ttk.Button(model_frame, text="모델 검증", command=self.validate_model_id).pack(side=tk.LEFT, padx=(0,5))
        ttk.Button(model_frame, text="프리셋", command=self.show_model_presets).pack(side=tk.LEFT)
        current_row += 1
        
        # Size X, Y
        size_frame = ttk.Frame(input_frame)
        size_frame.grid(row=current_row, column=0, columnspan=2, sticky="w", pady=5)
        ttk.Label(size_frame, text="SizeX:").pack(side=tk.LEFT, padx=(0,5))
        ttk.Entry(size_frame, textvariable=self.size_x_var, width=5).pack(side=tk.LEFT, padx=(0,15))
        ttk.Label(size_frame, text="SizeY:").pack(side=tk.LEFT, padx=(0,5))
        ttk.Entry(size_frame, textvariable=self.size_y_var, width=5).pack(side=tk.LEFT)
        current_row += 1
        
        # 버튼들
        button_frame = ttk.Frame(input_frame)
        button_frame.grid(row=current_row, column=0, columnspan=2, pady=20)
        ttk.Button(button_frame, text="NPC 생성", command=self.create_npc, width=15).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="입력값 초기화", command=self.clear_input_fields, width=15).pack(side=tk.LEFT, padx=10)
        
        # 우측 프레임 (히스토리 및 NPC 목록)
        right_frame = ttk.Frame(create_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        # 히스토리
        history_frame = ttk.LabelFrame(right_frame, text="NPC 생성 히스토리")
        history_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 히스토리 목록
        columns = ("UniqueID", "Name", "CategoryType", "EventGroupID")
        self.history_tree = ttk.Treeview(history_frame, columns=columns, show="headings", height=10)
        
        for col in columns:
            self.history_tree.heading(col, text=col)
            self.history_tree.column(col, width=80)
        
        self.history_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 히스토리 버튼
        history_btn_frame = ttk.Frame(history_frame)
        history_btn_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(history_btn_frame, text="히스토리 불러오기", command=self.load_from_history).pack(side=tk.LEFT, padx=5)
        
        # NPC 목록
        npc_list_frame = ttk.LabelFrame(right_frame, text="NPC 목록 정보")
        npc_list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # NPC 목록 트리뷰
        npc_columns = ("상태", "UniqueID", "Name", "CategoryType", "EventGroupID", "ModelID")
        self.npc_tree = ttk.Treeview(npc_list_frame, columns=npc_columns, show="headings", height=15)
        
        for col in npc_columns:
            self.npc_tree.heading(col, text=col)
            self.npc_tree.column(col, width=80)
        
        # 스크롤바 추가
        npc_scroll = ttk.Scrollbar(npc_list_frame, orient="vertical", command=self.npc_tree.yview)
        self.npc_tree.configure(yscrollcommand=npc_scroll.set)
        
        self.npc_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        npc_scroll.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 5), pady=5)

        # NPC 목록 버튼
        npc_btn_frame = ttk.Frame(npc_list_frame)
        npc_btn_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Button(npc_btn_frame, text="새로고침", command=self.load_npc_list).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(npc_btn_frame, text="선택적용", command=self.load_from_npc_list).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(npc_btn_frame, text="NPC제거", command=self.remove_npc).grid(row=1, column=0, padx=5, pady=5)
        ttk.Button(npc_btn_frame, text="NPC활성", command=self.activate_npc).grid(row=1, column=1, padx=5, pady=5)

                
        # 이벤트 바인딩
        self.unique_id_var.trace_add("write", self.on_unique_id_change)
        self.npc_tree.bind("<Double-1>", lambda e: self.load_from_npc_list())
        self.history_tree.bind("<Double-1>", lambda e: self.load_from_history())
        
        # 히스토리 목록 업데이트
        self.update_history_list()
    
    def build_placement_tab(self):
        """NPC 배치 탭 구성"""
        placement_frame = ttk.Frame(self.placement_tab)
        placement_frame.pack(fill=tk.BOTH, expand=True)
        
        # 좌측 프레임 (NPC 목록)
        left_frame = ttk.Frame(placement_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # NPC 목록
        npc_list_frame = ttk.LabelFrame(left_frame, text="NPC 목록")
        npc_list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # NPC 목록 트리뷰 (배치 탭용)
        npc_columns = ("UniqueID", "Name", "CategoryType")
        self.placement_npc_tree = ttk.Treeview(npc_list_frame, columns=npc_columns, show="headings", height=20)
        
        for col in npc_columns:
            self.placement_npc_tree.heading(col, text=col)
            self.placement_npc_tree.column(col, width=80)
        
        # 스크롤바 추가
        npc_scroll = ttk.Scrollbar(npc_list_frame, orient="vertical", command=self.placement_npc_tree.yview)
        self.placement_npc_tree.configure(yscrollcommand=npc_scroll.set)
        
        self.placement_npc_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        npc_scroll.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 5), pady=5)
        
        # 새로고침 버튼
        ttk.Button(npc_list_frame, text="새로고침", command=self.load_placement_npc_list).pack(padx=5, pady=5)
        
        # 우측 프레임 (배치 설정)
        right_frame = ttk.Frame(placement_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        # 위치 설정 프레임
        location_frame = ttk.LabelFrame(right_frame, text="위치 설정 (MapSpawn@_AutoGen)")
        location_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # MapID
        current_row = 0
        ttk.Label(location_frame, text="MapID:").grid(row=current_row, column=0, sticky="w", pady=5)
        map_id_frame = ttk.Frame(location_frame)
        map_id_frame.grid(row=current_row, column=1, sticky="w")
        ttk.Entry(map_id_frame, textvariable=self.map_id_var, width=15).pack(side=tk.LEFT, padx=(0,5))
        ttk.Button(map_id_frame, text="템플릿 목록", command=self.show_map_templates).pack(side=tk.LEFT)
        current_row += 1
        
        # MapSpawnGroupID
        ttk.Label(location_frame, text="MapSpawnGroupID:").grid(row=current_row, column=0, sticky="w", pady=5)
        ttk.Entry(location_frame, textvariable=self.map_spawn_group_id_var, width=15).grid(row=current_row, column=1, sticky="w")
        current_row += 1
        
        # HelperName
        ttk.Label(location_frame, text="HelperName:").grid(row=current_row, column=0, sticky="w", pady=5)
        ttk.Entry(location_frame, textvariable=self.helper_name_var, width=30).grid(row=current_row, column=1, sticky="w")
        current_row += 1
        
        # HeroID
        ttk.Label(location_frame, text="HeroID:").grid(row=current_row, column=0, sticky="w", pady=5)
        ttk.Entry(location_frame, textvariable=self.hero_id_var, width=15).grid(row=current_row, column=1, sticky="w")
        current_row += 1
        
        # DirectionID
        ttk.Label(location_frame, text="DirectionID:").grid(row=current_row, column=0, sticky="w", pady=5)
        dir_frame = ttk.Frame(location_frame)
        dir_frame.grid(row=current_row, column=1, sticky="w")
        directions = [("서쪽", 0), ("북쪽", 1), ("동쪽", 2), ("남쪽", 3)]
        for i, (text, value) in enumerate(directions):
            ttk.Radiobutton(dir_frame, text=text, variable=self.direction_id_var, value=value).grid(row=0, column=i, padx=5)
        current_row += 1
        
        # ShowConditionTID
        ttk.Label(location_frame, text="ShowConditionTID:").grid(row=current_row, column=0, sticky="w", pady=5)
        show_frame = ttk.Frame(location_frame)
        show_frame.grid(row=current_row, column=1, sticky="w")
        ttk.Entry(show_frame, textvariable=self.show_condition_tid_var, width=15).pack(side=tk.LEFT, padx=(0,5))
        ttk.Button(show_frame, text="유효성 검사", command=lambda: self.validate_condition_tid("show")).pack(side=tk.LEFT)
        ttk.Button(show_frame, text="Condition보기", command=lambda: self.show_condition_search_popup("show")).pack(side=tk.LEFT, padx=(5,0))
        current_row += 1
        
        # HideConditionTID
        ttk.Label(location_frame, text="HideConditionTID:").grid(row=current_row, column=0, sticky="w", pady=5)
        hide_frame = ttk.Frame(location_frame)
        hide_frame.grid(row=current_row, column=1, sticky="w")
        ttk.Entry(hide_frame, textvariable=self.hide_condition_tid_var, width=15).pack(side=tk.LEFT, padx=(0,5))
        ttk.Button(hide_frame, text="유효성 검사", command=lambda: self.validate_condition_tid("hide")).pack(side=tk.LEFT)
        ttk.Button(hide_frame, text="Condition보기", command=lambda: self.show_condition_search_popup("hide")).pack(side=tk.LEFT, padx=(5,0))
        current_row += 1
        
        # 배치 버튼
        ttk.Button(location_frame, text="맵에 배치하기", command=self.place_npc, width=20).grid(row=current_row, column=0, columnspan=2, pady=20)
        
        # 이벤트 정보 프레임 (UI만 구성)
        event_frame = ttk.LabelFrame(right_frame, text="이벤트 정보")
        event_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 이벤트 정보 버튼들
        ttk.Button(event_frame, text="이벤트 연결하기", state=tk.DISABLED).pack(pady=5, padx=10, anchor="w")
        ttk.Button(event_frame, text="대화 연결하기", state=tk.DISABLED).pack(pady=5, padx=10, anchor="w")
        ttk.Button(event_frame, text="이벤트 조건 설정하기", state=tk.DISABLED).pack(pady=5, padx=10, anchor="w")
        
        # NPC 배치 정보 목록 프레임 (하단)
        self.build_placement_info_tab()
        
        # 이벤트 바인딩
        self.placement_npc_tree.bind("<<TreeviewSelect>>", self.on_placement_npc_selected)
        
        # 페이지 로드 시 데이터 로드
        self.load_placement_npc_list()
        self.load_placement_info_list()
    
    #NPC 배치 하단 정보
    def build_placement_info_tab(self):
        """NPC 배치 정보 목록 구성"""
        info_frame = ttk.LabelFrame(self.placement_tab, text="NPC 배치 정보")
        info_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # NPC 배치 정보 트리뷰 - 컬럼 추가
        columns = ("상태", "MapID", "HelperName", "HeroID", "ShowConditionTID", "HideConditionTID")
        self.placement_info_tree = ttk.Treeview(info_frame, columns=columns, show="headings", height=15)

        for col, width in zip(columns, [60, 80, 180, 80, 120, 120]):
            self.placement_info_tree.heading(col, text=col)
            self.placement_info_tree.column(col, width=width)

        # 스크롤바
        scroll = ttk.Scrollbar(info_frame, orient="vertical", command=self.placement_info_tree.yview)
        self.placement_info_tree.configure(yscrollcommand=scroll.set)

        self.placement_info_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 버튼 프레임
        btn_frame = ttk.Frame(info_frame)
        btn_frame.pack(fill=tk.X, pady=5)

        ttk.Button(btn_frame, text="새로고침", command=self.load_placement_info_list).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(btn_frame, text="선택적용", command=self.load_from_placement_info_list).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(btn_frame, text="NPC제거", command=self.remove_placement_info).grid(row=1, column=0, padx=5, pady=5)
        ttk.Button(btn_frame, text="NPC활성", command=self.activate_placement_info).grid(row=1, column=1, padx=5, pady=5)

        # 이벤트 바인딩
        self.placement_info_tree.bind("<Double-1>", lambda e: self.load_from_placement_info_list())
        
    #MapSpawn 엑셀 파일의 구조 정보 가져오기
    def get_excel_structure_for_mapspawn(self, excel_file):        
        try:
            # excel_cache에서 정보 확인
            if self.excel_cache and "MapSpawn" in self.excel_cache:
                header_row = self.excel_cache["MapSpawn"].get("header_row")
                if header_row is not None:
                    return {
                        "header_row": header_row,
                        "sample_row": header_row + 1  # 샘플 데이터는 헤더 바로 다음 행
                    }
            
            # 캐시에 정보가 없으면 엑셀 파일 분석
            df = pd.read_excel(excel_file, sheet_name="MapSpawn", header=None)
            
            # 헤더 행 찾기 (MapID, HeroID 등 주요 컬럼이 있는 행)
            header_row = None
            for idx, row in df.iterrows():
                # 주요 컬럼명 포함된 행 찾기
                cols = [str(val).strip() for val in row.values if pd.notna(val)]
                if "MapID" in cols and "HeroID" in cols:
                    header_row = idx
                    break
            
            if header_row is None:
                # 헤더를 찾지 못하면 기본값 사용
                header_row = 0  # 1번째 행 (0-based)
                
            # 찾은 정보 캐시에 저장
            if not self.excel_cache:
                self.excel_cache = {}
            if "MapSpawn" not in self.excel_cache:
                self.excel_cache["MapSpawn"] = {}
            
            self.excel_cache["MapSpawn"]["header_row"] = header_row
            
            return {
                "header_row": header_row,
                "sample_row": header_row + 1  # 샘플 데이터는 헤더 바로 다음 행
            }
        except Exception as e:
            print(f"MapSpawn 엑셀 구조 분석 오류: {str(e)}")
            # 오류 발생 시 기본값 반환
            return {
                "header_row": 0,  # 기본값: 1번째 행 (0-based)
                "sample_row": 1   # 기본값: 2번째 행 (0-based)
            }


    #NPC 배치 정보 목록 로드
    def load_placement_info_list(self):
        """NPC 배치 정보 목록 로드"""
        # 트리뷰 초기화
        self.placement_info_tree.delete(*self.placement_info_tree.get_children())
        
        # 파일 경로
        map_spawn_file = os.path.join(self.excel_path, "MapSpawn@_AutoGen.xlsx")
        if not os.path.exists(map_spawn_file):
            return
        
        try:
            # 경고 무시
            import warnings
            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
            
            # 시트 이름 설정
            sheet_name = self.get_sheet_name_from_filename(map_spawn_file)
            
            # 엑셀 파일 로드
            workbook = openpyxl.load_workbook(map_spawn_file, read_only=True)
            
            # 시트 선택
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                # 시트가 없으면 첫 번째 시트 사용
                if len(workbook.sheetnames) > 0:
                    sheet = workbook[workbook.sheetnames[0]]
                else:
                    return
            
            # 헤더 찾기 (실제 컬럼 명이 있는 행)
            header_row = None
            header_values = []
            header_idx = 0
            
            for idx, row in enumerate(sheet.rows):
                row_values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
                # MapID와 HeroID가 모두 있는 행이 헤더 행
                if "MapID" in row_values and "HeroID" in row_values:
                    header_row = row
                    header_values = row_values
                    header_idx = idx
                    break
            
            if header_row is None:
                messagebox.showwarning("경고", "MapSpawn 파일에서 헤더를 찾을 수 없습니다.")
                return
            
            # 필요한 컬럼 인덱스 찾기
            col_indices = {
                "상태": 0,  # 첫 번째 열은 항상 상태로 간주
                "MapID": -1,
                "HelperName": -1,
                "HeroID": -1,
                "ShowConditionTID": -1,
                "HideConditionTID": -1
            }
            
            for idx, value in enumerate(header_values):
                if value == "MapID":
                    col_indices["MapID"] = idx
                elif value == "HelperName":
                    col_indices["HelperName"] = idx
                elif value == "HeroID":
                    col_indices["HeroID"] = idx
                elif value == "ShowConditionTID":
                    col_indices["ShowConditionTID"] = idx
                elif value == "HideConditionTID":
                    col_indices["HideConditionTID"] = idx
            
            # 필수 컬럼이 없으면 경고
            if col_indices["MapID"] == -1 or col_indices["HeroID"] == -1:
                messagebox.showwarning("경고", "MapSpawn 파일에서 필수 컬럼(MapID, HeroID)을 찾을 수 없습니다.")
                return
            
            # 데이터 행 처리 (헤더 다음 행부터)
            rows = list(sheet.rows)
            for row_idx in range(header_idx + 1, len(rows)):
                row = rows[row_idx]
                
                # 행의 모든 셀이 비어있는지 확인
                if all(cell.value is None for cell in row):
                    continue
                
                try:
                    # 상태 열 확인 (첫 번째 열)
                    status_val = row[0].value
                    status = "비활성" if status_val and str(status_val).startswith("#") else "활성"
                    
                    # 필수 데이터 확인
                    if col_indices["MapID"] >= len(row) or col_indices["HeroID"] >= len(row):
                        continue
                    
                    map_id = row[col_indices["MapID"]].value
                    hero_id = row[col_indices["HeroID"]].value
                    
                    # 필수 값 유효성 검사
                    if map_id is None or hero_id is None:
                        continue
                    
                    # HelperName은 선택적
                    helper_name = ""
                    if col_indices["HelperName"] != -1 and col_indices["HelperName"] < len(row):
                        helper_name = row[col_indices["HelperName"]].value or ""
                    
                    # ShowConditionTID와 HideConditionTID
                    show_condition = ""
                    if col_indices["ShowConditionTID"] != -1 and col_indices["ShowConditionTID"] < len(row):
                        show_condition = row[col_indices["ShowConditionTID"]].value or "0"
                    
                    hide_condition = ""
                    if col_indices["HideConditionTID"] != -1 and col_indices["HideConditionTID"] < len(row):
                        hide_condition = row[col_indices["HideConditionTID"]].value or "0"
                    
                    # 트리뷰에 추가 (컬럼 순서: 상태, MapID, HelperName, HeroID, ShowConditionTID, HideConditionTID)
                    self.placement_info_tree.insert("", "end", values=(
                        status,
                        map_id,
                        helper_name,
                        hero_id,
                        show_condition,
                        hide_condition
                    ))
                except Exception as e:
                    # 개별 행 처리 오류는 건너뛰기
                    print(f"행 {row_idx} 처리 오류: {str(e)}")
                    continue
            
            # 워크북 닫기
            workbook.close()
            
        except Exception as e:
            messagebox.showerror("오류", f"NPC 배치 정보 로드 중 오류 발생: {str(e)}")
            import traceback
            traceback.print_exc()  # 디버깅용 상세 오류 출력
    

    #NPC 배치 정보 목록에서 데이터 로드
    def load_from_placement_info_list(self):        
        selected = self.placement_info_tree.selection()
        if not selected:
            messagebox.showwarning("선택 필요", "로드할 NPC 배치 정보를 선택해주세요.")
            return
        
        item_id = selected[0]
        values = self.placement_info_tree.item(item_id, "values")
        
        # 필드 값 설정
        self.map_id_var.set(values[1])
        self.helper_name_var.set(values[2])
        self.hero_id_var.set(values[3])
        #self.hero_id_var.set(values[3])
        #self.hero_id_var.set(values[3])
        
        # 추가 정보 로드 (MapSpawn 파일에서)
        map_spawn_file = os.path.join(self.excel_path, "MapSpawn@_AutoGen.xlsx")
        if not os.path.exists(map_spawn_file):
            return
        
        try:
            # 엑셀 파일 로드
            workbook = openpyxl.load_workbook(map_spawn_file, read_only=True)
            
            # 시트 이름 설정
            sheet_name = self.get_sheet_name_from_filename(map_spawn_file)
            
            # 시트 선택
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                return
            
            # 헤더 확인
            header_row = list(sheet.rows)[0]
            header_values = [cell.value for cell in header_row]
            
            # 필요한 컬럼 인덱스 찾기
            map_id_idx = next((i for i, val in enumerate(header_values) if val == "MapID"), 1)
            hero_id_idx = next((i for i, val in enumerate(header_values) if val == "HeroID"), 4)
            map_spawn_group_idx = next((i for i, val in enumerate(header_values) if val == "MapSpawnGroupID"), 2)
            direction_id_idx = next((i for i, val in enumerate(header_values) if val == "DirectionID"), 5)
            show_condition_idx = next((i for i, val in enumerate(header_values) if val == "ShowConditionTID"), 6)
            hide_condition_idx = next((i for i, val in enumerate(header_values) if val == "HideConditionTID"), 7)
            
            # 데이터 찾기
            hero_id = values[3]
            for row in list(sheet.rows)[1:]:
                if str(row[hero_id_idx].value) == str(hero_id):
                    # 나머지 필드 값 설정
                    if map_spawn_group_idx < len(row):
                        self.map_spawn_group_id_var.set(str(row[map_spawn_group_idx].value or "0"))
                    
                    if direction_id_idx < len(row):
                        direction_val = row[direction_id_idx].value
                        if direction_val is not None:
                            self.direction_id_var.set(int(direction_val))
                    
                    if show_condition_idx < len(row):
                        self.show_condition_tid_var.set(str(row[show_condition_idx].value or "0"))
                    
                    if hide_condition_idx < len(row):
                        self.hide_condition_tid_var.set(str(row[hide_condition_idx].value or "0"))
                    
                    break
            
            # 워크북 닫기
            workbook.close()
            
        except Exception as e:
            print(f"배치 정보 상세 로드 오류: {str(e)}")
    
    #NPC 맵에 배치하기
    def place_npc(self):        
        """NPC 맵에 배치하기"""
        # 필수 필드 검사
        map_id = self.map_id_var.get().strip()
        map_spawn_group_id = self.map_spawn_group_id_var.get().strip()
        helper_name = self.helper_name_var.get().strip()
        hero_id = self.hero_id_var.get().strip()
        direction_id = self.direction_id_var.get()
        show_condition_tid = self.show_condition_tid_var.get().strip()
        hide_condition_tid = self.hide_condition_tid_var.get().strip()
        
        
        if not map_id or not hero_id:
            messagebox.showerror("오류", "MapID와 HeroID는 필수 입력 항목입니다.")
            return
            
        # 숫자 필드 검사
        for field_name, value in [("MapID", map_id), ("MapSpawnGroupID", map_spawn_group_id),
                                ("HeroID", hero_id), ("ShowConditionTID", show_condition_tid),
                                ("HideConditionTID", hide_condition_tid)]:
            if not value.isdigit():
                messagebox.showerror("오류", f"{field_name}는 숫자만 입력 가능합니다.")
                return
        
        # MapID 유효성 검사
        db_file = os.path.join(self.db_path, "MapTemplate.db")
        if os.path.exists(db_file):
            try:
                conn = sqlite3.connect(db_file)
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM MapTemplate WHERE UniqueID = ?", (int(map_id),))
                count = cursor.fetchone()[0]
                conn.close()
                
                if count == 0:
                    if not messagebox.askyesno("경고", f"MapID {map_id}가 MapTemplate.db에 존재하지 않습니다. 계속하시겠습니까?"):
                        return
            except Exception as e:
                messagebox.showwarning("경고", f"MapID 검증 중 오류: {str(e)}")
        
        # 조건 TID 유효성 검사
        if show_condition_tid != "0":
            self.validate_condition_tid("show")
        if hide_condition_tid != "0":
            self.validate_condition_tid("hide")
        
        # MapSpawn 파일 경로
        map_spawn_file = os.path.join(self.excel_path, "MapSpawn@_AutoGen.xlsx")
        
        # 파일 접근 가능 여부 확인
        if os.path.exists(map_spawn_file):
            try:
                # 파일이 열려있는지 테스트
                with open(map_spawn_file, 'a+b') as test_file:
                    pass
            except PermissionError:
                messagebox.showerror("파일 접근 오류", 
                    "MapSpawn@_AutoGen.xlsx 파일이 다른 프로그램에 의해 사용 중입니다.\n"
                    "엑셀이나 다른 프로그램에서 파일을 닫은 후 다시 시도해주세요.")
                return
        
        try:
            # 경고 무시
            import warnings
            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
            
            # 시트 이름
            sheet_name = self.get_sheet_name_from_filename(map_spawn_file)
            
            if os.path.exists(map_spawn_file):
                # 기존 파일 열기
                workbook = openpyxl.load_workbook(map_spawn_file)
                
                # 시트 확인 및 생성
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # 헤더 행 찾기
                    header_row = None
                    header_values = []
                    
                    for idx, row in enumerate(sheet.rows, 1):
                        row_values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
                        if "MapID" in row_values and "HeroID" in row_values:
                            header_row = idx
                            header_values = row_values
                            break
                    
                    if header_row is None:
                        # 헤더가 없으면 첫 번째 행에 생성
                        standard_headers = ["상태", "MapID", "MapSpawnGroupID", "HelperName", "HeroID", 
                                        "DirectionID", "ShowConditionTID", "HideConditionTID", 
                                        "ShowAlphaTime", "HideAlphaTime"]
                        for col_idx, header in enumerate(standard_headers, 1):
                            sheet.cell(row=1, column=col_idx, value=header)
                        header_row = 1
                        header_values = standard_headers
                else:
                    # 시트 생성 및 헤더 추가
                    sheet = workbook.create_sheet(sheet_name)
                    standard_headers = ["상태", "MapID", "MapSpawnGroupID", "HelperName", "HeroID", 
                                    "DirectionID", "ShowConditionTID", "HideConditionTID", 
                                    "ShowAlphaTime", "HideAlphaTime"]
                    for col_idx, header in enumerate(standard_headers, 1):
                        sheet.cell(row=1, column=col_idx, value=header)
                    header_row = 1
                    header_values = standard_headers
            else:
                # 새 엑셀 파일 생성
                workbook = openpyxl.Workbook()
                
                # 기본 시트 제거
                if "Sheet" in workbook.sheetnames:
                    del workbook["Sheet"]
                if "Sheet1" in workbook.sheetnames:
                    del workbook["Sheet1"]
                
                # 시트 생성 및 헤더 추가
                sheet = workbook.create_sheet(sheet_name)
                standard_headers = ["상태", "MapID", "MapSpawnGroupID", "HelperName", "HeroID", 
                                "DirectionID", "ShowConditionTID", "HideConditionTID", 
                                "ShowAlphaTime", "HideAlphaTime"]
                for col_idx, header in enumerate(standard_headers, 1):
                    sheet.cell(row=1, column=col_idx, value=header)
                header_row = 1
                header_values = standard_headers
            
            # 컬럼 인덱스 매핑 생성
            column_mapping = {}
            for idx, header in enumerate(header_values):
                if header and header != "상태":  # 비어있지 않고 "상태"가 아닌 헤더만 매핑
                    column_mapping[header] = idx + 1  # 실제 엑셀 컬럼은 1부터 시작
            
            # 새 행 위치
            new_row_idx = sheet.max_row + 1
            
            # 기본값으로 초기화 (A열 제외)
            for col_idx in range(2, len(header_values) + 1):  # 2부터 시작 (A열 제외)
                sheet.cell(row=new_row_idx, column=col_idx, value=0)
            
            # A열은 빈 값으로 설정 (상태 열)
            sheet.cell(row=new_row_idx, column=1, value="")
            
            # 사용자 입력 데이터 매핑
            data_to_insert = {
                "MapID": int(map_id),
                "MapSpawnGroupID": int(map_spawn_group_id),
                "HelperName": helper_name,
                "HeroID": int(hero_id),
                "DirectionID": int(direction_id),
                "ShowConditionTID": int(show_condition_tid),
                "HideConditionTID": int(hide_condition_tid)
            }
            
            # 특수 조건 처리
            show_alpha_time = 0
            hide_alpha_time = 0
            
            if int(show_condition_tid) != 0:
                show_alpha_time = 300
                
            if int(hide_condition_tid) != 0:
                hide_alpha_time = 300
                
            data_to_insert["ShowAlphaTime"] = show_alpha_time
            data_to_insert["HideAlphaTime"] = hide_alpha_time
            
            # 데이터 입력
            for header, value in data_to_insert.items():
                if header in column_mapping:
                    col_idx = column_mapping[header]
                    sheet.cell(row=new_row_idx, column=col_idx, value=value)
            
            # 파일 저장
            try:
                workbook.save(map_spawn_file)
            except PermissionError:
                messagebox.showerror("파일 저장 오류", 
                    "MapSpawn@_AutoGen.xlsx 파일을 저장할 수 없습니다.\n"
                    "엑셀이나 다른 프로그램에서 파일을 닫은 후 다시 시도해주세요.")
                return
            except Exception as e:
                messagebox.showerror("파일 저장 오류", f"파일 저장 중 오류 발생: {str(e)}")
                return
            
            # 배치 정보 목록 업데이트
            self.load_placement_info_list()
            
            messagebox.showinfo("완료", f"NPC (ID: {hero_id})가 맵 (ID: {map_id})에 성공적으로 배치되었습니다.")
            
        except Exception as e:
            import traceback
            traceback.print_exc()  # 디버깅용 상세 오류 출력
            messagebox.showerror("오류", f"NPC 배치 중 오류가 발생했습니다: {str(e)}")


    #선택한 NPC 배치 비활성화
    def remove_placement_info(self):    
        selected = self.placement_info_tree.selection()
        if not selected:
            messagebox.showwarning("선택 필요", "비활성화할 NPC 배치를 선택해주세요.")
            return
        
        map_spawn_file = os.path.join(self.excel_path, "MapSpawn@_AutoGen.xlsx")
        if not os.path.exists(map_spawn_file):
            messagebox.showerror("오류", "MapSpawn@_AutoGen.xlsx 파일이 존재하지 않습니다.")
            return
        
        # 파일 접근 가능 여부 확인
        try:
            with open(map_spawn_file, 'a+b') as test_file:
                pass
        except PermissionError:
            messagebox.showerror("파일 접근 오류", 
                "MapSpawn@_AutoGen.xlsx 파일이 다른 프로그램에 의해 사용 중입니다.\n"
                "엑셀이나 다른 프로그램에서 파일을 닫은 후 다시 시도해주세요.")
            return
        
        try:
            # 경고 무시
            import warnings
            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
            
            # 시트 이름
            sheet_name = self.get_sheet_name_from_filename(map_spawn_file)
            
            # 엑셀 파일 로드
            workbook = openpyxl.load_workbook(map_spawn_file)
            
            # 시트 선택
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                messagebox.showerror("오류", f"{sheet_name} 시트를 찾을 수 없습니다.")
                return
            
            # 헤더 확인
            header_row = [cell.value for cell in sheet[1]]
            hero_id_idx = next((i for i, val in enumerate(header_row) if val == "HeroID"), 4)
            
            # 실제 컬럼 인덱스는 1부터 시작
            hero_id_col = hero_id_idx + 1
            
            changed = False
            
            # 선택한 항목 처리
            for item in selected:
                values = self.placement_info_tree.item(item, "values")
                hero_id = values[3]
                
                # 해당 HeroID를 가진 행 찾기
                for row_idx in range(2, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row_idx, column=hero_id_col).value
                    if str(cell_value) == str(hero_id):
                        # A열에 # 추가
                        current_value = sheet.cell(row=row_idx, column=1).value
                        if current_value is None:
                            current_value = ""
                        
                        if not str(current_value).startswith("#"):
                            sheet.cell(row=row_idx, column=1).value = f"#{current_value}"
                            changed = True
                        break
            
            # 변경된 경우에만 저장
            if changed:
                try:
                    workbook.save(map_spawn_file)
                    # 목록 새로고침
                    self.load_placement_info_list()
                    messagebox.showinfo("완료", "선택한 NPC 배치가 비활성화되었습니다.")
                except Exception as e:
                    messagebox.showerror("파일 저장 오류", f"파일 저장 중 오류 발생: {str(e)}")
            else:
                messagebox.showinfo("알림", "변경된 항목이 없습니다.")
            
        except Exception as e:
            messagebox.showerror("오류", f"NPC 배치 비활성화 중 오류 발생: {str(e)}")


    #선택한 NPC 배치 활성화
    def activate_placement_info(self):        
        selected = self.placement_info_tree.selection()
        if not selected:
            messagebox.showwarning("선택 필요", "활성화할 NPC 배치를 선택해주세요.")
            return
        
        map_spawn_file = os.path.join(self.excel_path, "MapSpawn@_AutoGen.xlsx")
        if not os.path.exists(map_spawn_file):
            messagebox.showerror("오류", "MapSpawn@_AutoGen.xlsx 파일이 존재하지 않습니다.")
            return
        
        # 파일 접근 가능 여부 확인
        try:
            with open(map_spawn_file, 'a+b') as test_file:
                pass
        except PermissionError:
            messagebox.showerror("파일 접근 오류", 
                "MapSpawn@_AutoGen.xlsx 파일이 다른 프로그램에 의해 사용 중입니다.\n"
                "엑셀이나 다른 프로그램에서 파일을 닫은 후 다시 시도해주세요.")
            return
        
        try:
            # 경고 무시
            import warnings
            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
            
            # 시트 이름
            sheet_name = self.get_sheet_name_from_filename(map_spawn_file)
            
            # 엑셀 파일 로드
            workbook = openpyxl.load_workbook(map_spawn_file)
            
            # 시트 선택
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                messagebox.showerror("오류", f"{sheet_name} 시트를 찾을 수 없습니다.")
                return
            
            # 헤더 확인
            header_row = [cell.value for cell in sheet[1]]
            hero_id_idx = next((i for i, val in enumerate(header_row) if val == "HeroID"), 4)
            
            # 실제 컬럼 인덱스는 1부터 시작
            hero_id_col = hero_id_idx + 1
            
            changed = False
            
            # 선택한 항목 처리
            for item in selected:
                values = self.placement_info_tree.item(item, "values")
                hero_id = values[3]
                
                # 해당 HeroID를 가진 행 찾기
                for row_idx in range(2, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row_idx, column=hero_id_col).value
                    if str(cell_value) == str(hero_id):
                        # A열에서 # 제거
                        current_value = sheet.cell(row=row_idx, column=1).value
                        if current_value is not None and str(current_value).startswith("#"):
                            sheet.cell(row=row_idx, column=1).value = current_value.lstrip("#")
                            changed = True
                        break
            
            # 변경된 경우에만 저장
            if changed:
                try:
                    workbook.save(map_spawn_file)
                    # 목록 새로고침
                    self.load_placement_info_list()
                    messagebox.showinfo("완료", "선택한 NPC 배치가 활성화되었습니다.")
                except Exception as e:
                    messagebox.showerror("파일 저장 오류", f"파일 저장 중 오류 발생: {str(e)}")
            else:
                messagebox.showinfo("알림", "변경된 항목이 없습니다.")
            
        except Exception as e:
            messagebox.showerror("오류", f"NPC 배치 활성화 중 오류 발생: {str(e)}")

    
    def select_excel_file(self):
        """엑셀 파일 선택 대화상자"""
        file_path = filedialog.askopenfilename(
            title="HeroTemplate_AutoGen.xlsx 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialdir=self.excel_path
        )
        if file_path:
            self.excel_file_var.set(file_path)
            self.load_npc_list()

    def get_excel_structure(self, excel_file):
        """엑셀 파일의 구조 정보 가져오기"""
        try:
            # excel_cache에서 정보 확인
            if self.excel_cache and "HeroTemplate" in self.excel_cache:
                header_row = self.excel_cache["HeroTemplate"].get("header_row")
                if header_row is not None:
                    return {
                        "header_row": header_row,
                        "sample_row": header_row + 1  # 샘플 데이터는 헤더 바로 다음 행
                    }
            
            # 캐시에 정보가 없으면 엑셀 파일 분석
            df = pd.read_excel(excel_file, sheet_name="HeroTemplate", header=None)
            
            # 헤더 행 찾기 (UniqueID, Name 등 주요 컬럼이 있는 행)
            header_row = None
            for idx, row in df.iterrows():
                # 주요 컬럼명 포함된 행 찾기
                cols = [str(val).strip() for val in row.values if pd.notna(val)]
                if "UniqueID" in cols and "Name" in cols:
                    header_row = idx
                    break
            
            if header_row is None:
                # 헤더를 찾지 못하면 기본값 사용
                header_row = 2  # 3번째 행 (0-based)
                
            # 찾은 정보 캐시에 저장
            if not self.excel_cache:
                self.excel_cache = {}
            if "HeroTemplate" not in self.excel_cache:
                self.excel_cache["HeroTemplate"] = {}
            
            self.excel_cache["HeroTemplate"]["header_row"] = header_row
            
            return {
                "header_row": header_row,
                "sample_row": header_row + 1  # 샘플 데이터는 헤더 바로 다음 행
            }
        except Exception as e:
            print(f"엑셀 구조 분석 오류: {str(e)}")
            # 오류 발생 시 기본값 반환
            return {
                "header_row": 2,  # 기본값: 3번째 행 (0-based)
                "sample_row": 3   # 기본값: 4번째 행 (0-based)
            }


    def on_unique_id_change(self, *args):
        """UniqueID 변경 시 BaseGroupID와 EventGroupID 자동 업데이트"""
        unique_id = self.unique_id_var.get().strip()
        if unique_id and unique_id.isdigit() and int(unique_id) > 0:
            self.base_group_id_var.set(unique_id)
            self.event_group_id_var.set(unique_id)
    
    def validate_unique_id(self):
        """UniqueID의 유효성 검사 (고유값인지 확인)"""
        unique_id = self.unique_id_var.get().strip()
        
        # 기본 유효성 검사
        if not unique_id:
            messagebox.showerror("오류", "UniqueID를 입력해주세요.")
            return False
            
        if not unique_id.isdigit():
            messagebox.showerror("오류", "UniqueID는 숫자만 입력 가능합니다.")
            return False
            
        unique_id_int = int(unique_id)
        if unique_id_int <= 0:
            messagebox.showerror("오류", "UniqueID는 0보다 큰 숫자여야 합니다.")
            return False
        
        # DB에서 중복 확인
        db_file = os.path.join(self.db_path, "HeroTemplate.db")
        if not os.path.exists(db_file):
            messagebox.showerror("오류", "HeroTemplate.db 파일을 찾을 수 없습니다.")
            return False
            
        try:
            conn = sqlite3.connect(db_file)
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM HeroTemplate WHERE UniqueID = ?", (unique_id_int,))
            count = cursor.fetchone()[0]
            conn.close()
            
            if count > 0:
                messagebox.showerror("오류", f"UniqueID {unique_id}는 이미 사용 중입니다.")
                return False
            else:
                return True
                
        except Exception as e:
            messagebox.showerror("오류", f"데이터베이스 검증 오류: {str(e)}")
            return False
    
    def validate_model_id(self):
        """ModelHero ID 유효성 검사"""
        model_id = self.model_hero_id_var.get().strip()
        
        # 기본 유효성 검사
        if not model_id:
            messagebox.showerror("오류", "ModelHero ID를 입력해주세요.")
            return False
            
        if not model_id.isdigit():
            messagebox.showerror("오류", "ModelHero ID는 숫자만 입력 가능합니다.")
            return False
        
        # DB에서 확인
        db_file = os.path.join(self.db_path, "ModelHero.db")
        if not os.path.exists(db_file):
            messagebox.showerror("오류", "ModelHero.db 파일을 찾을 수 없습니다.")
            return False
            
        try:
            conn = sqlite3.connect(db_file)
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM ModelHero WHERE UniqueID = ?", (int(model_id),))
            count = cursor.fetchone()[0]
            conn.close()
            
            if count == 0:
                messagebox.showerror("오류", f"ModelHero ID {model_id}를 찾을 수 없습니다.")
                return False
            else:
                return True
                
        except Exception as e:
            messagebox.showerror("오류", f"데이터베이스 검증 오류: {str(e)}")
            return False
    
    def validate_condition_tid(self, condition_type):
        """조건 TID 유효성 검사"""
        if condition_type == "show":
            tid = self.show_condition_tid_var.get().strip()
        else:
            tid = self.hide_condition_tid_var.get().strip()
            
        # 기본값 0은 항상 유효함
        if tid == "0":
            return True
            
        # 기본 유효성 검사
        if not tid:
            if condition_type == "show":
                self.show_condition_tid_var.set("0")
            else:
                self.hide_condition_tid_var.set("0")
            return True
            
        if not tid.isdigit():
            messagebox.showerror("오류", "조건 TID는 숫자만 입력 가능합니다.")
            if condition_type == "show":
                self.show_condition_tid_var.set("0")
            else:
                self.hide_condition_tid_var.set("0")
            return False
        
        # DB에서 확인
        db_file = os.path.join(self.db_path, "ConditionTemplate.db")
        if not os.path.exists(db_file):
            messagebox.showerror("오류", "ConditionTemplate.db 파일을 찾을 수 없습니다.")
            if condition_type == "show":
                self.show_condition_tid_var.set("0")
            else:
                self.hide_condition_tid_var.set("0")
            return False
            
        try:
            conn = sqlite3.connect(db_file)
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM ConditionTemplate WHERE TemplateID = ?", (int(tid),))
            count = cursor.fetchone()[0]
            conn.close()
            
            if count == 0:
                messagebox.showerror("오류", f"조건 TID {tid}를 찾을 수 없습니다. 기본값 0으로 설정합니다.")
                if condition_type == "show":
                    self.show_condition_tid_var.set("0")
                else:
                    self.hide_condition_tid_var.set("0")
                return False
            else:
                return True
                
        except Exception as e:
            messagebox.showerror("오류", f"데이터베이스 검증 오류: {str(e)}")
            if condition_type == "show":
                self.show_condition_tid_var.set("0")
            else:
                self.hide_condition_tid_var.set("0")
            return False
    
    def clear_input_fields(self):
        """입력 필드 초기화"""
        self.unique_id_var.set("")
        self.base_group_id_var.set("")
        self.name_var.set("")
        self.nickname_var.set("")
        self.event_group_id_var.set("")
        self.model_hero_id_var.set("")
        self.size_x_var.set("3")
        self.size_y_var.set("3")
        
        # 카테고리 초기화
        for cat in self.category_types:
            self.category_vars[cat].set(False)
            
    def create_npc(self):
        """NPC 생성 및 엑셀에 데이터 입력"""
        # 모든 필드 유효성 검사
        unique_id = self.unique_id_var.get().strip()
        base_group_id = self.base_group_id_var.get().strip()
        event_group_id = self.event_group_id_var.get().strip()
        model_id = self.model_hero_id_var.get().strip()
        size_x = self.size_x_var.get().strip()
        size_y = self.size_y_var.get().strip()
        name = self.name_var.get().strip()
        
        # 필수 필드 검사
        if not unique_id or not base_group_id or not event_group_id or not model_id or not name:
            messagebox.showerror("오류", "필수 항목을 모두 입력해주세요.")
            return
        
        # 숫자 필드 검사
        for field_name, value in [("UniqueID", unique_id), ("BaseHeroID", base_group_id), 
                                ("EventGroupID", event_group_id), ("ModelHero ID", model_id),
                                ("SizeX", size_x), ("SizeY", size_y)]:
            if not value.isdigit():
                messagebox.showerror("오류", f"{field_name}는 숫자만 입력 가능합니다.")
                return
            if field_name != "ModelHero ID" and field_name != "SizeX" and field_name != "SizeY" and int(value) <= 0:
                messagebox.showerror("오류", f"{field_name}는 0보다 큰 숫자여야 합니다.")
                return
        
        # UniqueID 중복 확인
        if not self.validate_unique_id():
            return
            
        # ModelHero ID 확인
        if not self.validate_model_id():
            return
        
        # CategoryType 처리 - 확실히 정수값으로 변환
        category_value = int(self.category_var.get())
        
        try:
            # 엑셀 파일 경로 설정
            excel_file = self.excel_file_var.get()
            
            # 엑셀 파일 읽기
            try:
                # 엑셀 구조 정보 가져오기
                excel_structure = self.get_excel_structure(excel_file)
                header_row_idx = excel_structure["header_row"]
                sample_row_idx = excel_structure["sample_row"]
                
                # 엑셀 파일 읽기
                raw_df = pd.read_excel(excel_file, header=None, sheet_name="HeroTemplate")
                
                # 행 수 확인
                if len(raw_df) > sample_row_idx:
                    header_row = raw_df.iloc[header_row_idx]
                    sample_row = raw_df.iloc[sample_row_idx]
                    
                    # 컬럼 매핑 생성
                    columns = {}
                    for i, col_name in enumerate(header_row):
                        if pd.notna(col_name) and col_name:
                            columns[col_name] = i
                    
                    # HeroTemplate.DB의 컬럼 목록 가져오기
                    db_columns = self.get_db_columns("HeroTemplate")

                    # 새 NPC 데이터 생성 (빈 행으로 시작)
                    new_row = [None] * len(sample_row)

                    # A열(0번 인덱스)은 빈값으로 설정 (활성/삭제 관련 컬럼)
                    new_row[0] = ""

                    # 필요한 컬럼 확인 및 업데이트
                    for col_name, col_idx in columns.items():
                        # 사용자 입력 값 설정
                        if col_name == 'UniqueID':
                            new_row[col_idx] = int(unique_id)
                        elif col_name == 'BaseHeroID':
                            new_row[col_idx] = int(base_group_id)
                        elif col_name == 'Name':
                            new_row[col_idx] = name
                        elif col_name == 'NickName':
                            new_row[col_idx] = self.nickname_var.get()
                        elif col_name == 'CategoryType':
                            new_row[col_idx] = category_value
                        elif col_name == 'EventGroupID':
                            new_row[col_idx] = int(event_group_id)
                        elif col_name == 'ModelID':
                            new_row[col_idx] = int(model_id)
                        elif col_name == 'SizeX':
                            new_row[col_idx] = int(size_x)
                        elif col_name == 'SizeY':
                            new_row[col_idx] = int(size_y)
                        # 사용자 입력값이 아닌 경우 DB 컬럼에 포함된 항목만 샘플 데이터에서 복사
                        elif col_idx != 0 and col_name in db_columns:  # A열이 아니고 DB 컬럼에 포함된 경우만
                            new_row[col_idx] = sample_row[col_idx]
                        else:
                            # DB에 없는 컬럼은 빈값으로
                            new_row[col_idx] = ""
                    
                    # 기존 엑셀 파일 불러오기 및 서식 유지하면서 데이터 추가
                    import openpyxl
                    try:
                        # 기존 파일 열기
                        workbook = openpyxl.load_workbook(excel_file)
                        sheet = workbook["HeroTemplate"]
                        
                        # 새 행 데이터 추가
                        row_values = []
                        for i in range(len(new_row)):
                            if i < len(new_row):
                                row_values.append(new_row[i])
                            else:
                                row_values.append("")
                        
                        # 새 행 추가
                        sheet.append(row_values)
                        
                        # 파일 저장
                        workbook.save(excel_file)
                    except Exception as e:
                        # 오류 발생 시 pandas로 백업 저장
                        raw_df.loc[len(raw_df)] = new_row
                        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
                            raw_df.to_excel(writer, sheet_name="HeroTemplate", index=False, header=False)
                    
                    # 히스토리에 저장할 데이터 준비
                    history_data = {
                        "UniqueID": int(unique_id),
                        "BaseHeroID": int(base_group_id),
                        "Name": name,
                        "NickName": self.nickname_var.get(),
                        "CategoryType": category_value,
                        "EventGroupID": int(event_group_id),
                        "ModelID": int(model_id),
                        "SizeX": int(size_x),
                        "SizeY": int(size_y)
                    }
                    
                    # 히스토리에 저장
                    self.save_history(history_data)
                    
                    # NPC 목록 새로고침
                    self.load_npc_list()
                    
                    messagebox.showinfo("완료", f"NPC '{name}' (ID: {unique_id})가 성공적으로 생성되었습니다.")
                else:
                    raise ValueError("엑셀 파일의 형식이 올바르지 않습니다. 헤더와 샘플 데이터 행이 필요합니다.")
                    
            except Exception as e:
                # 파일이 없거나 형식이 올바르지 않으면 새로 생성
                messagebox.showinfo("정보", f"{os.path.basename(excel_file)} 파일이 없거나 형식이 올바르지 않아 새로 생성합니다.")
                
                # 기본 양식 생성
                basic_columns = ["UniqueID", "BaseHeroID", "Name", "NickName", "CategoryType", 
                                "EventGroupID", "ModelID", "SizeX", "SizeY"]
                
                # 기본 데이터 생성
                data = [
                    ["#제목행"] + [""] * (len(basic_columns)-1),  # 1행: 제목행
                    ["#설명행"] + [""] * (len(basic_columns)-1),  # 2행: 설명행
                    basic_columns,  # 3행: 컬럼명
                    [1001, 1001, "템플릿NPC", "별명", 30, 1001, 2001, 3, 3]  # 4행: 샘플 데이터
                ]
                
                # 새 데이터 추가
                new_row = [int(unique_id), int(base_group_id), name, self.nickname_var.get(), 
                        category_value, int(event_group_id), int(model_id), int(size_x), int(size_y)]
                data.append(new_row)
                
                # 데이터프레임 생성 및 저장 (HeroTemplate 시트 지정)
                with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
                    pd.DataFrame(data).to_excel(writer, sheet_name="HeroTemplate", index=False, header=False)
                
                # excel_cache 업데이트
                if not self.excel_cache:
                    self.excel_cache = {}
                if "HeroTemplate" not in self.excel_cache:
                    self.excel_cache["HeroTemplate"] = {}
                
                self.excel_cache["HeroTemplate"]["header_row"] = 2
                
                # 히스토리에 저장할 데이터 준비
                history_data = {
                    "UniqueID": int(unique_id),
                    "BaseHeroID": int(base_group_id),
                    "Name": name,
                    "NickName": self.nickname_var.get(),
                    "CategoryType": category_value,
                    "EventGroupID": int(event_group_id),
                    "ModelID": int(model_id),
                    "SizeX": int(size_x),
                    "SizeY": int(size_y)
                }
                
                # 히스토리에 저장
                self.save_history(history_data)
                
                # NPC 목록 새로고침
                self.load_npc_list()
                
                messagebox.showinfo("완료", f"NPC '{name}' (ID: {unique_id})가 성공적으로 생성되었습니다.")
                
        except Exception as e:
            messagebox.showerror("오류", f"NPC 생성 중 오류가 발생했습니다: {str(e)}")


    def update_history_list(self):
        """히스토리 목록 업데이트"""
        # 트리뷰 초기화
        self.history_tree.delete(*self.history_tree.get_children())
        
        # 히스토리 데이터 추가
        for entry in self.npc_history:
            self.history_tree.insert("", "end", values=(
                entry["UniqueID"],
                entry["Name"],
                entry.get("CategoryName", "Unknown"),  # 표시용 카테고리 이름
                entry["EventGroupID"]
            ), tags=(str(entry["UniqueID"]),))
    
    def load_from_history(self):
        """히스토리에서 NPC 정보 불러오기"""
        selected = self.history_tree.selection()
        if not selected:
            messagebox.showwarning("선택 필요", "히스토리에서 항목을 선택해주세요.")
            return
        
        item_id = selected[0]
        values = self.history_tree.item(item_id, "values")
        unique_id = values[0]
        
        # 히스토리에서 해당 항목 찾기
        for entry in self.npc_history:
            if str(entry["UniqueID"]) == unique_id:
                # 입력 필드에 값 설정
                self.unique_id_var.set(str(entry["UniqueID"]))
                self.base_group_id_var.set(str(entry["BaseHeroID"]))
                self.name_var.set(entry["Name"])
                self.nickname_var.set(entry["NickName"])
                self.event_group_id_var.set(str(entry["EventGroupID"]))
                self.model_hero_id_var.set(str(entry["ModelID"]))
                self.size_x_var.set(str(entry["SizeX"]))
                self.size_y_var.set(str(entry["SizeY"]))
                
                # 카테고리 설정
                category_value = entry["CategoryType"]
                try:
                    category_value = int(category_value)
                except (ValueError, TypeError):
                    # 문자열일 경우 기본값으로 NPC(30) 설정
                    category_value = 30
                    
                self.category_var.set(category_value)
                
                messagebox.showinfo("불러오기 완료", f"NPC '{entry['Name']}' (ID: {entry['UniqueID']})의 정보를 불러왔습니다.")
                break
    
    def load_npc_list(self):
        """NPC 목록 불러오기"""
        # 트리뷰 초기화
        self.npc_tree.delete(*self.npc_tree.get_children())
        
        excel_file = self.excel_file_var.get()
        if not os.path.exists(excel_file):
            return
            
        try:
            # 엑셀 구조 정보 가져오기
            excel_structure = self.get_excel_structure(excel_file)
            header_row_idx = excel_structure["header_row"]
            
            # 엑셀 파일을 raw로 읽어서 A열 확인 가능하게 함
            raw_df = pd.read_excel(excel_file, header=None, sheet_name="HeroTemplate")
            
            # 헤더 기준으로 데이터 읽기
            df = pd.read_excel(excel_file, sheet_name="HeroTemplate", header=header_row_idx)
            
            # 각 행을 트리뷰에 추가
            for idx, row in df.iterrows():
                # 첫 번째 행은 샘플이므로 제외
                if idx < 1:
                    continue
                
                # 실제 엑셀 파일에서의 행 인덱스 계산 (헤더 + 현재 idx + 1)
                excel_row_idx = header_row_idx + idx + 1
                
                try:
                    # A열의 값 확인 (인덱스가 범위를 벗어나지 않는지 체크)
                    if excel_row_idx < len(raw_df):
                        a_col_value = raw_df.iloc[excel_row_idx, 0]
                        # A열에 #이 있으면 비활성, 없으면 활성
                        if pd.notna(a_col_value) and str(a_col_value).startswith("#"):
                            status = "비활성"
                        else:
                            status = "활성"
                    else:
                        status = "활성"  # 기본값
                except:
                    status = "활성"  # 오류 발생 시 기본값
                
                try:
                    # CategoryType 처리
                    category_val = row.get("CategoryType", 30)
                    if pd.isna(category_val):
                        category_val = 30
                    try:
                        category_val = int(category_val)
                    except:
                        category_val = 30
                    
                    category_name = next((name for name, value in self.category_types if value == category_val), "NPC")
                    
                    # 트리뷰에 추가
                    self.npc_tree.insert("", "end", values=(
                        status,
                        str(row.get("UniqueID", "")),
                        str(row.get("Name", "")),
                        category_name,
                        str(row.get("EventGroupID", "")),
                        str(row.get("ModelID", ""))
                    ), tags=(str(row.get("UniqueID", "")),))
                    
                except Exception as e:
                    print(f"행 처리 오류: {str(e)}, 행: {row}")
                    continue
                
            # 배치 탭의 NPC 목록도 업데이트
            self.load_placement_npc_list()
                    
        except Exception as e:
            messagebox.showerror("오류", f"NPC 목록 로드 중 오류가 발생했습니다: {str(e)}")
            print(f"오류 상세: {str(e)}")

    #NPC 목록에서 정보 불러오기
    def load_from_npc_list(self):
        selected = self.npc_tree.selection()
        if not selected:
            messagebox.showwarning("선택 필요", "NPC 목록에서 항목을 선택해주세요.")
            return
        
        item_id = selected[0]
        values = self.npc_tree.item(item_id, "values")
        unique_id = values[1]
        
        print(f"[DEBUG] load_from_npc_list values: {unique_id}")  # 추가!!
        
        excel_file = self.excel_file_var.get()
        if not os.path.exists(excel_file):
            return
            
        try:
            # 엑셀 파일 읽기
            df = pd.read_excel(excel_file)
            
            # UniqueID로 행 찾기
            for idx, row in df.iterrows():
                if str(row["UniqueID"]) == unique_id:
                    # 입력 필드에 값 설정
                    self.unique_id_var.set(str(row["UniqueID"]))
                    self.base_group_id_var.set(str(row["BaseHeroID"]))
                    self.name_var.set(row["Name"].replace("#", "") if isinstance(row["Name"], str) else row["Name"])
                    self.nickname_var.set(row["NickName"] if "NickName" in row else "")
                    self.event_group_id_var.set(str(row["EventGroupID"]))
                    self.model_hero_id_var.set(str(row["ModelID"]))
                    self.size_x_var.set(str(row["SizeX"]))
                    self.size_y_var.set(str(row["SizeY"]))
                    
                    # 카테고리 설정
                    try:
                        category_value = int(row["CategoryType"])
                    except (ValueError, TypeError):
                        # 문자열이거나 변환 불가능한 경우 기본값 NPC(30)
                        category_value = 30
                    self.category_var.set(category_value)
                    
                    messagebox.showinfo("불러오기 완료", f"NPC '{row['Name']}' (ID: {row['UniqueID']})의 정보를 불러왔습니다.")
                    break
                
        except Exception as e:
            messagebox.showerror("오류", f"NPC 정보 불러오기 중 오류가 발생했습니다: {str(e)}")
    

    #배치 탭의 NPC 목록 불러오기
    def load_placement_npc_list(self):
        
        # 트리뷰 초기화
        self.placement_npc_tree.delete(*self.placement_npc_tree.get_children())
        
        excel_file = self.excel_file_var.get()
        if not os.path.exists(excel_file):
            return
            
        try:
            # 엑셀 구조 정보 가져오기
            excel_structure = self.get_excel_structure(excel_file)
            header_row_idx = excel_structure["header_row"]
            
            # 엑셀 파일을 raw로 읽어서 A열 확인 가능하게 함
            raw_df = pd.read_excel(excel_file, header=None, sheet_name="HeroTemplate")
            
            # 헤더 기준으로 데이터 읽기
            df = pd.read_excel(excel_file, sheet_name="HeroTemplate", header=header_row_idx)
            
            # 각 행을 트리뷰에 추가 (활성 NPC만)
            for idx, row in df.iterrows():
                # 첫 번째 행은 샘플이므로 제외
                if idx < 1:
                    continue
                    
                # 실제 엑셀 파일에서의 행 인덱스 계산 (헤더 + 현재 idx + 1)
                excel_row_idx = header_row_idx + idx + 1
                
                try:
                    # A열의 값 확인 (인덱스가 범위를 벗어나지 않는지 체크)
                    if excel_row_idx < len(raw_df):
                        a_col_value = raw_df.iloc[excel_row_idx, 0]
                        # A열에 #이 있으면 비활성 (표시하지 않음)
                        if pd.notna(a_col_value) and str(a_col_value).startswith("#"):
                            continue
                    
                    # CategoryType 처리
                    category_val = row.get("CategoryType", 30)
                    if pd.isna(category_val):
                        category_val = 30
                    try:
                        category_val = int(category_val)
                    except:
                        category_val = 30
                    
                    category_name = next((name for name, value in self.category_types if value == category_val), "NPC")
                    
                    # 트리뷰에 추가
                    self.placement_npc_tree.insert("", "end", values=(
                        str(row.get("UniqueID", "")),
                        str(row.get("Name", "")),
                        category_name
                    ), tags=(str(row.get("UniqueID", "")),))
                
                except Exception as e:
                    print(f"행 처리 오류: {str(e)}, 행: {row}")
                    continue
                    
        except Exception as e:
            messagebox.showerror("오류", f"NPC 목록 로드 중 오류가 발생했습니다: {str(e)}")
            print(f"오류 상세: {str(e)}")

    #배치 탭에서 NPC 선택 시 이벤트
    def on_placement_npc_selected(self, event):
        
        selected = self.placement_npc_tree.selection()
        if not selected:
            return
            
        item_id = selected[0]
        values = self.placement_npc_tree.item(item_id, "values")
        unique_id = values[0]
        
        # HeroID 필드에 선택한 NPC의 UniqueID 설정
        self.hero_id_var.set(str(unique_id))


    def remove_npc(self):
        """NPC 제거 (A열에 # 추가)"""
        selected = self.npc_tree.selection()
        if not selected:
            messagebox.showwarning("선택 필요", "제거할 NPC를 선택해주세요.")
            return
        
        item_id = selected[0]
        values = self.npc_tree.item(item_id, "values")
        status = values[0]
        unique_id = values[1]
        name = values[2]
        
        # 이미 제거된 NPC인지 확인
        if status == "제거됨":
            messagebox.showinfo("알림", f"이미 제거된 NPC입니다: {name}")
            return
        
        # 제거 확인
        if not messagebox.askyesno("확인", f"NPC '{name}' (ID: {unique_id})를 제거하시겠습니까?"):
            return
        
        excel_file = self.excel_file_var.get()
        if not os.path.exists(excel_file):
            return
            
        try:
            # excel 구조 정보 가져오기
            excel_structure = self.get_excel_structure(excel_file)
            header_row_idx = excel_structure["header_row"]
            
            # 엑셀 파일을 raw로 읽기
            raw_df = pd.read_excel(excel_file, header=None, sheet_name="HeroTemplate")
            
            # 헤더 행 가져오기
            header_row = raw_df.iloc[header_row_idx]
            
            # UniqueID 컬럼 인덱스 찾기
            unique_id_col = None
            for i, col_name in enumerate(header_row):
                if col_name == "UniqueID":
                    unique_id_col = i
                    break
            
            if unique_id_col is not None:
                # UniqueID로 행 찾기
                found_row = None
                for idx, row in raw_df.iterrows():
                    if idx > header_row_idx and str(row[unique_id_col]) == unique_id:
                        # A열(0번 컬럼)에 # 추가
                        current_value = raw_df.iloc[idx, 0]
                        if pd.isna(current_value):
                            current_value = ""
                        if not str(current_value).startswith("#"):
                            raw_df.iloc[idx, 0] = f"#{current_value}"
                        found_row = idx
                        break
                
                if found_row is not None:
                    # 서식 유지하며 저장
                    try:
                        import openpyxl
                        workbook = openpyxl.load_workbook(excel_file)
                        sheet = workbook["HeroTemplate"]
                        
                        # 찾은 행의 A열 셀 업데이트
                        cell_value = raw_df.iloc[found_row, 0]
                        sheet.cell(row=found_row+1, column=1).value = cell_value
                        
                        workbook.save(excel_file)
                    except Exception as e:
                        print(f"서식 유지 저장 오류: {str(e)}")
                        # 실패 시 일반 저장
                        raw_df.to_excel(excel_file, sheet_name="HeroTemplate", index=False, header=False)
                    
                    # NPC 목록 새로고침
                    self.load_npc_list()
                    
                    messagebox.showinfo("완료", f"NPC '{name}' (ID: {unique_id})가 제거되었습니다.")
                else:
                    messagebox.showwarning("경고", f"NPC ID {unique_id}를 엑셀 파일에서 찾을 수 없습니다.")
            else:
                messagebox.showwarning("경고", "엑셀 파일에 UniqueID 컬럼이 없습니다.")
                
        except Exception as e:
            messagebox.showerror("오류", f"NPC 제거 중 오류가 발생했습니다: {str(e)}")


    def activate_npc(self):
        """NPC 활성화 (A열의 # 제거)"""
        selected = self.npc_tree.selection()
        if not selected:
            messagebox.showwarning("선택 필요", "활성화할 NPC를 선택해주세요.")
            return
        
        item_id = selected[0]
        values = self.npc_tree.item(item_id, "values")
        status = values[0]
        unique_id = values[1]
        name = values[2]
        
        # 이미 활성화된 NPC인지 확인
        if status == "활성":
            messagebox.showinfo("알림", f"이미 활성화된 NPC입니다: {name}")
            return
        
        # 활성화 확인
        if not messagebox.askyesno("확인", f"NPC '{name}' (ID: {unique_id})를 활성화하시겠습니까?"):
            return
        
        excel_file = self.excel_file_var.get()
        if not os.path.exists(excel_file):
            return
            
        try:
            # excel 구조 정보 가져오기
            excel_structure = self.get_excel_structure(excel_file)
            header_row_idx = excel_structure["header_row"]
            
            # 엑셀 파일을 raw로 읽기
            raw_df = pd.read_excel(excel_file, header=None, sheet_name="HeroTemplate")
            
            # 헤더 행 가져오기
            header_row = raw_df.iloc[header_row_idx]
            
            # UniqueID 컬럼 인덱스 찾기
            unique_id_col = None
            for i, col_name in enumerate(header_row):
                if col_name == "UniqueID":
                    unique_id_col = i
                    break
            
            if unique_id_col is not None:
                # UniqueID로 행 찾기
                found_row = None
                for idx, row in raw_df.iterrows():
                    if idx > header_row_idx and str(row[unique_id_col]) == unique_id:
                        # A열(0번 컬럼)의 # 제거
                        current_value = raw_df.iloc[idx, 0]
                        if pd.notna(current_value) and str(current_value).startswith("#"):
                            # # 제거하여 A열 값 업데이트
                            new_value = str(current_value).replace("#", "", 1)
                            raw_df.iloc[idx, 0] = new_value if new_value else ""
                        found_row = idx
                        break
                
                if found_row is not None:
                    # 서식 유지하며 저장
                    try:
                        import openpyxl
                        workbook = openpyxl.load_workbook(excel_file)
                        sheet = workbook["HeroTemplate"]
                        
                        # 찾은 행의 A열 셀 업데이트
                        cell_value = raw_df.iloc[found_row, 0]
                        sheet.cell(row=found_row+1, column=1).value = cell_value
                        
                        workbook.save(excel_file)
                    except Exception as e:
                        print(f"서식 유지 저장 오류: {str(e)}")
                        # 실패 시 일반 저장
                        raw_df.to_excel(excel_file, sheet_name="HeroTemplate", index=False, header=False)
                    
                    # NPC 목록 새로고침
                    self.load_npc_list()
                    
                    messagebox.showinfo("완료", f"NPC '{name}' (ID: {unique_id})가 활성화되었습니다.")
                else:
                    messagebox.showwarning("경고", f"NPC ID {unique_id}를 엑셀 파일에서 찾을 수 없습니다.")
            else:
                messagebox.showwarning("경고", "엑셀 파일에 UniqueID 컬럼이 없습니다.")
                
        except Exception as e:
            messagebox.showerror("오류", f"NPC 활성화 중 오류가 발생했습니다: {str(e)}")


    def show_model_presets(self):
        """모델 프리셋 팝업 표시"""
        preset_popup = tk.Toplevel(self.popup)
        preset_popup.title("모델 프리셋")
        preset_popup.geometry("500x400")
        preset_popup.grab_set()
        
        # 프리셋 리스트
        preset_frame = ttk.Frame(preset_popup, padding=10)
        preset_frame.pack(fill=tk.BOTH, expand=True)
        
        # 프리셋 트리뷰
        columns = ("이름", "모델ID", "크기")
        preset_tree = ttk.Treeview(preset_frame, columns=columns, show="headings", height=15)
        
        for col in columns:
            preset_tree.heading(col, text=col)
            preset_tree.column(col, width=100)
        
        # 스크롤바
        preset_scroll = ttk.Scrollbar(preset_frame, orient="vertical", command=preset_tree.yview)
        preset_tree.configure(yscrollcommand=preset_scroll.set)
        
        preset_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        preset_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 버튼 프레임
        btn_frame = ttk.Frame(preset_popup, padding=10)
        btn_frame.pack(fill=tk.X)
        
        # 프리셋 입력 필드
        input_frame = ttk.Frame(btn_frame)
        input_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(input_frame, text="이름:").grid(row=0, column=0, padx=5)
        name_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=name_var, width=15).grid(row=0, column=1, padx=5)
        
        ttk.Label(input_frame, text="모델ID:").grid(row=0, column=2, padx=5)
        model_id_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=model_id_var, width=10).grid(row=0, column=3, padx=5)
        
        ttk.Label(input_frame, text="SizeX:").grid(row=0, column=4, padx=5)
        size_x_var = tk.StringVar(value="3")
        ttk.Entry(input_frame, textvariable=size_x_var, width=5).grid(row=0, column=5, padx=5)
        
        ttk.Label(input_frame, text="SizeY:").grid(row=0, column=6, padx=5)
        size_y_var = tk.StringVar(value="3")
        ttk.Entry(input_frame, textvariable=size_y_var, width=5).grid(row=0, column=7, padx=5)
        
        # 버튼들
        button_frame = ttk.Frame(btn_frame)
        button_frame.pack(pady=10)
        
        def add_preset():
            name = name_var.get().strip()
            model_id = model_id_var.get().strip()
            size_x = size_x_var.get().strip()
            size_y = size_y_var.get().strip()
            
            if not name or not model_id or not size_x or not size_y:
                messagebox.showwarning("입력 오류", "모든 필드를 입력해주세요.")
                return
                
            if not model_id.isdigit() or not size_x.isdigit() or not size_y.isdigit():
                messagebox.showwarning("입력 오류", "모델ID와 크기는 숫자만 입력 가능합니다.")
                return
            
            new_preset = {
                "name": name,
                "model_id": int(model_id),
                "size_x": int(size_x),
                "size_y": int(size_y)
            }
            
            self.model_presets.append(new_preset)
            self.save_presets()
            update_preset_list()
            
            # 입력 필드 초기화
            name_var.set("")
            model_id_var.set("")
            size_x_var.set("3")
            size_y_var.set("3")
        
        def remove_preset():
            selected = preset_tree.selection()
            if not selected:
                messagebox.showwarning("선택 필요", "삭제할 프리셋을 선택해주세요.")
                return
                
            item_id = selected[0]
            idx = preset_tree.index(item_id)
            
            if 0 <= idx < len(self.model_presets):
                del self.model_presets[idx]
                self.save_presets()
                update_preset_list()
        
        def apply_preset():
            selected = preset_tree.selection()
            if not selected:
                messagebox.showwarning("선택 필요", "적용할 프리셋을 선택해주세요.")
                return
                
            item_id = selected[0]
            idx = preset_tree.index(item_id)
            
            if 0 <= idx < len(self.model_presets):
                preset = self.model_presets[idx]
                self.model_hero_id_var.set(str(preset["model_id"]))
                self.size_x_var.set(str(preset["size_x"]))
                self.size_y_var.set(str(preset["size_y"]))
                preset_popup.destroy()
        
        def update_preset_list():
            preset_tree.delete(*preset_tree.get_children())
            for preset in self.model_presets:
                preset_tree.insert("", "end", values=(
                    preset["name"],
                    preset["model_id"],
                    f"{preset['size_x']} x {preset['size_y']}"
                ))
        
        ttk.Button(button_frame, text="추가", command=add_preset).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="삭제", command=remove_preset).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="적용", command=apply_preset).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="닫기", command=preset_popup.destroy).pack(side=tk.LEFT, padx=5)
        
        # 프리셋 목록 업데이트
        update_preset_list()
        
        # 더블 클릭 이벤트
        preset_tree.bind("<Double-1>", lambda e: apply_preset())
    
    
    def show_map_templates(self):
        """맵 템플릿 목록 팝업 표시"""
        # DB 파일 확인
        db_file = os.path.join(self.db_path, "MapTemplate.db")
        if not os.path.exists(db_file):
            messagebox.showerror("오류", "MapTemplate.db 파일을 찾을 수 없습니다.")
            return
            
        # 팝업 생성
        map_popup = tk.Toplevel(self.popup)
        map_popup.title("맵 템플릿 목록")
        map_popup.geometry("700x500")
        map_popup.grab_set()
        
        # 맵 목록 프레임
        map_frame = ttk.Frame(map_popup, padding=10)
        map_frame.pack(fill=tk.BOTH, expand=True)
        
        # 맵 검색
        search_frame = ttk.Frame(map_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(search_frame, text="맵 검색:").pack(side=tk.LEFT, padx=(0, 5))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # 맵 트리뷰
        columns = ("UniqueID", "Path")
        map_tree = ttk.Treeview(map_frame, columns=columns, show="headings", height=20)
        
        for col, width in zip(columns, [80, 400, 150]):
            map_tree.heading(col, text=col)
            map_tree.column(col, width=width)
        
        # 스크롤바
        map_scroll = ttk.Scrollbar(map_frame, orient="vertical", command=map_tree.yview)
        map_tree.configure(yscrollcommand=map_scroll.set)
        
        map_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        map_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 버튼 프레임
        btn_frame = ttk.Frame(map_popup, padding=10)
        btn_frame.pack(fill=tk.X)
        
        # 맵 목록 로드
        def load_map_templates():
            map_tree.delete(*map_tree.get_children())
            
            try:
                conn = sqlite3.connect(db_file)
                cursor = conn.cursor()
                
                search_text = search_var.get().strip().lower()
                if search_text:
                    cursor.execute(
                        "SELECT UniqueID, Path, Name FROM MapTemplate WHERE "
                        "lower(Path) LIKE ? OR lower(Name) LIKE ? OR UniqueID LIKE ?",
                        (f"%{search_text}%", f"%{search_text}%", f"%{search_text}%")
                    )
                else:
                    cursor.execute("SELECT UniqueID, Path FROM MapTemplate")
                    
                rows = cursor.fetchall()
                conn.close()
                
                for row in rows:
                    map_tree.insert("", "end", values=row)
                    
            except Exception as e:
                messagebox.showerror("오류", f"맵 템플릿 로드 중 오류가 발생했습니다: {str(e)}")
        
        def apply_map():
            selected = map_tree.selection()
            if not selected:
                messagebox.showwarning("선택 필요", "적용할 맵을 선택해주세요.")
                return
                
            item_id = selected[0]
            values = map_tree.item(item_id, "values")
            
            self.map_id_var.set(values[0])  # UniqueID 설정
            map_popup.destroy()
        
        # 검색 이벤트
        def on_search(*args):
            load_map_templates()
            
        search_var.trace_add("write", on_search)
        
        # 버튼
        ttk.Button(btn_frame, text="적용", command=apply_map).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="닫기", command=map_popup.destroy).pack(side=tk.LEFT, padx=5)
        
        # 초기 로드
        load_map_templates()
        
        # 더블 클릭 이벤트
        map_tree.bind("<Double-1>", lambda e: apply_map())


    def get_db_columns(self, db_name):
        """DB 파일에서 컬럼 목록 가져오기"""
        db_file = os.path.join(self.db_path, f"{db_name}.db")
        if not os.path.exists(db_file):
            return []
        
        try:
            conn = sqlite3.connect(db_file)
            cursor = conn.cursor()
            
            # 테이블 스키마 정보 가져오기
            cursor.execute(f"PRAGMA table_info({db_name})")
            columns = [row[1] for row in cursor.fetchall()]
            conn.close()
            
            return columns
        except Exception as e:
            print(f"DB 컬럼 로드 오류: {str(e)}")
            return []


    #NPC 검색기
    def show_hero_search_popup(self):
        """Hero 검색 팝업 표시"""
        # DB 파일 확인
        db_file = os.path.join(self.db_path, "HeroTemplate.db")
        if not os.path.exists(db_file):
            messagebox.showerror("오류", "HeroTemplate.db 파일을 찾을 수 없습니다.")
            return  
            
        # 팝업 생성
        search_popup = tk.Toplevel(self.popup)
        search_popup.title("Hero 검색")
        search_popup.geometry("800x500")
        search_popup.grab_set()
        
        # 검색 프레임
        search_frame = ttk.Frame(search_popup, padding=10)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 검색 필드
        ttk.Label(search_frame, text="검색 조건:").grid(row=0, column=0, padx=5, pady=5)
        
        search_type_var = tk.StringVar(value="UniqueID")
        search_type_combo = ttk.Combobox(search_frame, textvariable=search_type_var, 
                                        values=["UniqueID", "Name", "EventGroupID"], 
                                        state="readonly", width=15)
        search_type_combo.grid(row=0, column=1, padx=5, pady=5)
        
        search_value_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=search_value_var, width=30).grid(row=0, column=2, padx=5, pady=5)
        
        # 결과 프레임
        result_frame = ttk.Frame(search_popup, padding=10)
        result_frame.pack(fill=tk.BOTH, expand=True)
        
        # 결과 트리뷰
        columns = ("UniqueID", "Name", "NickName", "CategoryType", "EventGroupID")
        result_tree = ttk.Treeview(result_frame, columns=columns, show="headings", height=15)
        
        for col, width in zip(columns, [80, 150, 150, 100, 80]):
            result_tree.heading(col, text=col)
            result_tree.column(col, width=width)
        
        # 스크롤바
        tree_scroll = ttk.Scrollbar(result_frame, orient="vertical", command=result_tree.yview)
        result_tree.configure(yscrollcommand=tree_scroll.set)
        
        result_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 버튼 프레임
        btn_frame = ttk.Frame(search_popup, padding=10)
        btn_frame.pack(fill=tk.X)
        
        # 검색 함수
        def search_hero():
            # 트리뷰 초기화
            result_tree.delete(*result_tree.get_children())
            
            search_type = search_type_var.get()
            search_value = search_value_var.get().strip()
            
            if not search_value:
                messagebox.showwarning("입력 필요", "검색어를 입력해주세요.")
                return
            
            try:
                conn = sqlite3.connect(db_file)
                cursor = conn.cursor()
                
                if search_type == "UniqueID":
                    # 숫자 검증
                    if not search_value.isdigit():
                        messagebox.showwarning("입력 오류", "UniqueID는 숫자만 입력 가능합니다.")
                        return
                    
                    query = "SELECT UniqueID, Name, NickName, CategoryType, EventGroupID FROM HeroTemplate WHERE UniqueID = ?"
                    cursor.execute(query, (int(search_value),))
                
                elif search_type == "Name":
                    # 부분 일치 검색
                    query = "SELECT UniqueID, Name, NickName, CategoryType, EventGroupID FROM HeroTemplate WHERE Name LIKE ?"
                    cursor.execute(query, (f'%{search_value}%',))
                
                elif search_type == "EventGroupID":
                    # 숫자 검증
                    if not search_value.isdigit():
                        messagebox.showwarning("입력 오류", "EventGroupID는 숫자만 입력 가능합니다.")
                        return
                    
                    query = "SELECT UniqueID, Name, NickName, CategoryType, EventGroupID FROM HeroTemplate WHERE EventGroupID = ?"
                    cursor.execute(query, (int(search_value),))
                
                rows = cursor.fetchall()
                conn.close()
                
                if not rows:
                    messagebox.showinfo("검색 결과", "검색 결과가 없습니다.")
                    return
                    
                # 결과를 트리뷰에 추가
                for row in rows:
                    # CategoryType 이름 얻기
                    category_value = row[3]
                    category_name = next((name for name, value in self.category_types if value == category_value), "Unknown")
                    
                    # 트리뷰에 표시
                    result_tree.insert("", "end", values=(
                        row[0],  # UniqueID
                        row[1],  # Name
                        row[2],  # NickName
                        category_name,  # CategoryType
                        row[4]   # EventGroupID
                    ))
                
            except Exception as e:
                messagebox.showerror("검색 오류", f"검색 중 오류가 발생했습니다: {str(e)}")
        
        # 선택 함수
        def select_hero():
            selected = result_tree.selection()
            if not selected:
                messagebox.showwarning("선택 필요", "Hero를 선택해주세요.")
                return
                
            item_id = selected[0]
            values = result_tree.item(item_id, "values")
            
            # DB에서 전체 정보 가져오기
            try:
                conn = sqlite3.connect(db_file)
                cursor = conn.cursor()
                
                query = "SELECT * FROM HeroTemplate WHERE UniqueID = ?"
                cursor.execute(query, (values[0],))
                
                row = cursor.fetchone()
                if not row:
                    messagebox.showerror("오류", "선택한 Hero 정보를 찾을 수 없습니다.")
                    return
                    
                # 컬럼 이름 가져오기
                cursor.execute("PRAGMA table_info(HeroTemplate)")
                columns = [info[1] for info in cursor.fetchall()]
                
                conn.close()
                
                # Hero 정보를 딕셔너리로 변환
                hero_data = {}
                for i, col in enumerate(columns):
                    if i < len(row):
                        hero_data[col] = row[i]
                
                # 입력 필드에 데이터 설정
                if "UniqueID" in hero_data:
                    self.unique_id_var.set(str(hero_data["UniqueID"]))
                if "BaseHeroID" in hero_data:
                    self.base_group_id_var.set(str(hero_data["BaseHeroID"]))
                if "Name" in hero_data:
                    self.name_var.set(hero_data["Name"])
                if "NickName" in hero_data:
                    self.nickname_var.set(hero_data["NickName"])
                if "CategoryType" in hero_data:
                    self.category_var.set(int(hero_data["CategoryType"]))
                if "EventGroupID" in hero_data:
                    self.event_group_id_var.set(str(hero_data["EventGroupID"]))
                if "ModelID" in hero_data:
                    self.model_hero_id_var.set(str(hero_data["ModelID"]))
                if "SizeX" in hero_data:
                    self.size_x_var.set(str(hero_data["SizeX"]))
                if "SizeY" in hero_data:
                    self.size_y_var.set(str(hero_data["SizeY"]))
                
                # 팝업 닫기
                search_popup.destroy()
                
            except Exception as e:
                messagebox.showerror("데이터 로드 오류", f"Hero 정보 로드 중 오류가 발생했습니다: {str(e)}")
        
        # 버튼 추가
        ttk.Button(btn_frame, text="검색", command=search_hero, width=15).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="선택", command=select_hero, width=15).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="닫기", command=search_popup.destroy, width=15).pack(side=tk.LEFT, padx=10)
        
        # 더블클릭 이벤트
        result_tree.bind("<Double-1>", lambda e: select_hero())
        
        # 엔터키 이벤트
        search_popup.bind("<Return>", lambda e: search_hero())   

    #Condition 검색기
    def show_condition_search_popup(self, target="show"):
        """Condition 검색 팝업 표시 (Show/Hide 용)"""
        import json

        # DB 파일 확인
        db_file = os.path.join(self.db_path, "ConditionTemplate.db")
        enum_db_file = os.path.join(self.db_path, "enumName.db")
        if not os.path.exists(db_file) or not os.path.exists(enum_db_file):
            messagebox.showerror("오류", "ConditionTemplate.db 또는 enumName.db 파일을 찾을 수 없습니다.")
            return

        # 팝업 생성
        popup = tk.Toplevel(self.popup)
        popup.title("Condition 검색")
        popup.geometry("1200x600")
        popup.grab_set()

        # 검색 프레임
        search_frame = ttk.Frame(popup, padding=10)
        search_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(search_frame, text="검색 조건:").grid(row=0, column=0, padx=5, pady=5)

        search_type_var = tk.StringVar(value="TemplateID")
        search_type_combo = ttk.Combobox(search_frame, textvariable=search_type_var,
                                        values=["TemplateID", "TypeName", "Description", "ConditionType", "Condition1", "Condition2", "TutorialID"],
                                        state="readonly", width=15)
        search_type_combo.grid(row=0, column=1, padx=5, pady=5)

        search_value_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=search_value_var, width=30).grid(row=0, column=2, padx=5, pady=5)

        # 결과 프레임
        result_frame = ttk.Frame(popup, padding=10)
        result_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("TemplateID", "TypeName", "Description", "ConditionType", "Condition1", "Condition2", "TutorialID")
        tree = ttk.Treeview(result_frame, columns=columns, show="headings", height=20)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120)

        scroll = ttk.Scrollbar(result_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # enumName 매핑 가져오기
        conn_enum = sqlite3.connect(enum_db_file)
        enum_df = pd.read_sql_query("SELECT Key, Value, KR FROM enumName", conn_enum)
        conn_enum.close()

        econdition_map = {str(row["Value"]): row["KR"] for _, row in enum_df.iterrows() if row["Key"] == "EConditionType"}

        # 조건 불러오기
        def load_conditions():
            tree.delete(*tree.get_children())

            conn = sqlite3.connect(db_file)
            condition_df = pd.read_sql_query("SELECT * FROM ConditionTemplate", conn)
            conn.close()

            # TypeName 추가
            condition_df["TypeName"] = condition_df["ConditionType"].map(lambda x: econdition_map.get(str(x), "Unknown"))

            # Description 추가
            descriptions = []
            for _, row in condition_df.iterrows():
                cond_type = row["ConditionType"]
                cond1 = row["Condition1"]
                desc = "-"
                if cond_type in (1010, 1020, 1021):
                    if cond_type == 1010:
                        db_name = "StageTemplate.db"
                        id_column = "UniqueID"
                        name_column = "Name"
                    else:
                        db_name = "QuestTemplate.db"
                        id_column = "TemplateID"
                        name_column = "QuestName"
                    db_path_full = os.path.join(self.db_path, db_name)
                    if os.path.exists(db_path_full):
                        conn2 = sqlite3.connect(db_path_full)
                        try:
                            cursor = conn2.execute(f"SELECT {name_column} FROM {db_name.replace('.db','')} WHERE {id_column} = ?", (cond1,))
                            result = cursor.fetchone()
                            if result and result[0]:
                                desc = result[0]
                        except:
                            pass
                        conn2.close()
                descriptions.append(desc)

            condition_df["Description"] = descriptions

            # 검색
            search_col = search_type_var.get()
            search_value = search_value_var.get().strip().lower()

            if search_value:
                condition_df = condition_df[condition_df[search_col].astype(str).str.lower().str.contains(search_value)]

            for _, row in condition_df.iterrows():
                tree.insert("", "end", values=(
                    row["TemplateID"],
                    row["TypeName"],
                    row["Description"],
                    row["ConditionType"],
                    row["Condition1"],
                    row["Condition2"],
                    row["TutorialID"]
                ))

        # 선택 적용
        def apply_selection():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("선택 필요", "조건을 선택해주세요.")
                return

            item_id = selected[0]
            values = tree.item(item_id, "values")
            template_id = values[0]

            if target == "show":
                self.show_condition_tid_var.set(str(template_id))
            else:
                self.hide_condition_tid_var.set(str(template_id))

            popup.destroy()

        # 버튼 프레임
        btn_frame = ttk.Frame(popup, padding=10)
        btn_frame.pack(fill=tk.X)

        ttk.Button(btn_frame, text="적용", command=apply_selection).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="닫기", command=popup.destroy).pack(side=tk.LEFT, padx=5)

        # 검색 연동
        search_value_var.trace_add("write", lambda *args: load_conditions())

        load_conditions()

        tree.bind("<Double-1>", lambda e: apply_selection())


    def get_sheet_name_from_filename(self, filename):
        """파일 이름에서 시트 이름 추출 (@ 앞 부분)"""
        base_name = os.path.basename(filename)  # 파일 이름만 추출
        if '@' in base_name:
            sheet_name = base_name.split('@')[0]  # @ 앞부분 반환
            return sheet_name
        else:
            # @ 없으면 확장자 제외한 파일명 반환
            return os.path.splitext(base_name)[0]


# 메인 창에 NPC 생성기 버튼 추가하는 함수
def add_npc_creator_button(app):
    """ExcelSearchApp에 NPC 생성기 버튼 추가"""
    if not hasattr(app, "btn_frame"):
        return
    
    import tkinter as tk
    tk.Button(app.btn_frame, text="👤 NPC 생성기", command=lambda: open_npc_creator(app)).pack(side="left", padx=5)

def open_npc_creator(app):
    """NPC 생성기 팝업 열기"""
    NPCCreatorPopup(
        app.root,
        app.folder_path.get(),
        app.db_folder_path.get(),
        app.excel_cache
    )

# 테스트용 코드 (독립적으로 실행할 경우)
if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()  # 메인 창 숨기기
    
    excel_path = "."  # 현재 디렉토리
    db_path = "."     # 현재 디렉토리
    
    app = NPCCreatorPopup(root, excel_path, db_path)
    root.mainloop()