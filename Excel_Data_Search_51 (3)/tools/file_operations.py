import os
import platform
import logging
import traceback
from openpyxl import load_workbook
from utils.string_utils import find_excel_file, normalize_file_path

class ExcelFileManager:
    """엑셀 파일 처리와 관련된 기능을 담당하는 클래스"""
    
    def __init__(self, folder_path, progress_update=None):
        """
        ExcelFileManager 초기화
        
        Args:
            folder_path: 기본 폴더 경로
            progress_update: 진행 상태 업데이트 함수 (선택 사항)
        """
        self.folder_path = folder_path
        self.progress_update = progress_update
        self.files_with_external_links = set()  # 외부 링크가 있는 파일
        self.open_excel_files = set()  # 현재 열려 있는 엑셀 파일
    
    def set_progress_callback(self, callback):
        """진행 상태 업데이트 콜백 함수 설정"""
        self.progress_update = callback
    
    def _log_progress(self, message, success=True):
        """진행 상태 로깅 (콜백이 설정된 경우만)"""
        if self.progress_update:
            self.progress_update(message, success)
    
    def check_open_excel_files(self):
        """현재 시스템에서 열려 있는 엑셀 파일 목록을 반환 - 읽기 전용 모드 사용"""
        open_files = []
        
        try:
            import psutil
            
            # 프로세스 이름으로 Excel 실행 여부 확인
            excel_running = False
            for proc in psutil.process_iter(['name']):
                try:
                    if proc.info['name'].lower() in ('excel.exe', 'microsoft excel'):
                        excel_running = True
                        break
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                    pass
            
            if not excel_running:
                self._log_progress("Excel이 실행 중이지 않음, 열린 파일 없음", success=True)
                return []
            
            # 운영체제별 처리
            if platform.system() == 'Windows':
                try:
                    # openpyxl을 사용하여 열린 파일 확인 (read_only 모드로 안전하게)
                    import openpyxl
                    from openpyxl.utils.exceptions import InvalidFileException
                    
                    # Windows에서 열린 Excel 파일 목록 가져오기
                    for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
                        try:
                            if proc.info['name'].lower() == 'excel.exe':
                                # Excel 프로세스의 명령줄 인자에서 파일 경로 추출
                                if proc.info['cmdline'] and len(proc.info['cmdline']) > 1:
                                    for arg in proc.info['cmdline'][1:]:
                                        if arg.lower().endswith(('.xlsx', '.xls')):
                                            if os.path.exists(arg):
                                                open_files.append(os.path.normpath(arg))
                        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                            continue
                    
                    # 추가로 COM 인터페이스를 통해 열린 파일 확인
                    try:
                        import win32com.client
                        
                        # GetObject 사용하여 기존 Excel 인스턴스에 연결 시도
                        excel = win32com.client.GetObject(Class="Excel.Application")
                        
                        # 열린 워크북 확인
                        for wb in excel.Workbooks:
                            try:
                                file_path = wb.FullName
                                if os.path.exists(file_path):
                                    abs_path = os.path.abspath(file_path)
                                    if abs_path not in open_files:
                                        open_files.append(abs_path)
                                        self._log_progress(f"열린 파일 발견: {os.path.basename(file_path)}", success=True)
                            except Exception as wb_err:
                                print(f"워크북 정보 가져오기 오류: {wb_err}")
                                continue
                        
                        # 참조 해제
                        del excel
                        
                    except Exception as com_err:
                        print(f"COM 인터페이스 오류: {com_err}")
                        # COM 오류는 무시하고 계속 진행
                    
                except ImportError:
                    self._log_progress("win32com 모듈 없음, 열린 파일 확인 제한됨", success=False)
                    
            elif platform.system() == 'Darwin':  # macOS
                # macOS에서는 lsof 명령어 사용
                try:
                    import subprocess
                    
                    # Excel 프로세스 확인
                    excel_processes = subprocess.check_output(
                        ["ps", "-axco", "comm"], 
                        universal_newlines=True
                    )
                    
                    if "Microsoft Excel" in excel_processes:
                        # Excel이 실행 중인 경우, 열린 파일 확인
                        result = subprocess.check_output(
                            ["lsof", "-c", "Microsoft Excel"], 
                            universal_newlines=True
                        )
                        
                        for line in result.splitlines():
                            if ".xls" in line.lower():
                                parts = line.split()
                                if len(parts) > 8:
                                    file_path = parts[8]
                                    if os.path.exists(file_path) and file_path.lower().endswith((".xlsx", ".xls")):
                                        open_files.append(file_path)
                                        self._log_progress(f"열린 파일 발견: {os.path.basename(file_path)}", success=True)
                except Exception as e:
                    self._log_progress(f"macOS에서 열린 파일 확인 중 오류: {str(e)}", success=False)
                    
            else:  # Linux 등 기타 운영체제
                self._log_progress(f"현재 운영체제({platform.system()})에서는 열린 파일 확인이 제한됨", success=False)
            
            # 중복 제거
            open_files = list(set(open_files))
            
        except Exception as e:
            self._log_progress(f"열린 파일 확인 중 예외 발생: {str(e)}", success=False)
            import traceback
            traceback.print_exc()
        
        self._log_progress(f"총 {len(open_files)}개의 열린 엑셀 파일 확인됨", success=True)
        
        # 인스턴스 변수에 저장
        self.open_excel_files = set(open_files)
        
        return open_files

    def check_external_links(self, file_path):
        """파일에 외부 링크가 있는지 확인 - openpyxl의 read_only 모드 사용"""
        if not os.path.exists(file_path):
            print(f"⚠️ 파일이 존재하지 않음: {file_path}")
            return False

        try:
            # openpyxl을 사용하여 read_only 모드로 파일 열기
            import openpyxl
            print(f"openpyxl로 파일 열기 시도 (read_only): {file_path}")
            
            # 먼저 read_only 모드로 시도
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                
                # openpyxl의 _external_links 속성 확인
                has_links = hasattr(wb, '_external_links') and len(getattr(wb, '_external_links', [])) > 0
                
                if has_links:
                    link_count = len(wb._external_links)
                    print(f"🔗 외부 링크 발견 ({link_count}개): {file_path}")
                    if link_count <= 5 and hasattr(wb._external_links[0], 'Target'):
                        links = [link.Target for link in wb._external_links if hasattr(link, 'Target')]
                        print(f"  링크 목록: {links}")
                
                wb.close()
                return has_links
                
            except Exception as e:
                print(f"⚠️ read_only 모드로 파일 열기 실패: {str(e)}")
                
                # read_only 모드 실패 시 일반 모드로 재시도
                try:
                    wb = openpyxl.load_workbook(file_path, keep_links=True)
                    
                    # openpyxl의 _external_links 속성 확인
                    has_links = hasattr(wb, '_external_links') and len(getattr(wb, '_external_links', [])) > 0
                    
                    if has_links:
                        link_count = len(wb._external_links)
                        print(f"🔗 외부 링크 발견 ({link_count}개): {file_path}")
                    
                    wb.close()
                    return has_links
                    
                except Exception as e2:
                    print(f"⚠️ 일반 모드로도 파일 열기 실패: {str(e2)}")
                    
                    # 두 가지 모두 실패하면 xlwings로 마지막 시도
                    return self._check_external_links_with_xlwings(file_path)

        except ImportError:
            print("❌ openpyxl 모듈이 설치되어 있지 않습니다.")
            # openpyxl 없으면 xlwings로 대체
            return self._check_external_links_with_xlwings(file_path)
            
        except Exception as e:
            print(f"⚠️ 외부 링크 확인 중 예외 발생: {str(e)}")
        
        return False

    def _check_external_links_with_xlwings(self, file_path):
        """xlwings를 사용하여 외부 링크 확인 (대체 방법)"""
        try:
            import xlwings as xw
            print(f"xlwings로 파일 열기 시도: {file_path}")
            
            app = xw.App(visible=False)
            app.display_alerts = False
            
            try:
                wb = app.books.open(file_path, update_links=False)
                links = wb.api.LinkSources()
                has_links = links is not None and isinstance(links, (list, tuple)) and len(links) > 0
                
                if has_links:
                    print(f"🔗 xlwings: 외부 링크 발견 ({len(links)}개): {file_path}")
                
                wb.close()
                app.quit()
                return has_links
                
            except Exception as e:
                print(f"⚠️ xlwings로 파일 열기 실패: {str(e)}")
                if app:
                    app.quit()
                return False
                
        except ImportError:
            print("❌ xlwings 모듈이 설치되어 있지 않습니다.")
        except Exception as e:
            print(f"⚠️ xlwings 초기화 실패: {str(e)}")
        
        return False

    def open_excel_with_sheet(self, file_path, sheet_name):
        """엑셀 파일을 열고 특정 시트를 활성화합니다."""
        try:
            import platform
            
            if platform.system() == 'Windows':
                try:
                    # 먼저 파일 자체를 열기
                    import os
                    os.startfile(file_path)
                    
                    # 시간 지연 - 파일이 열릴 때까지 대기
                    import time
                    time.sleep(2)
                    
                    # COM 인터페이스를 통해 엑셀에 접근
                    try:
                        import win32com.client
                        
                        # 이미 실행 중인 Excel 인스턴스에 연결
                        excel = win32com.client.GetObject(Class="Excel.Application")
                        excel.Visible = True
                        
                        # 현재 열린 모든 워크북을 확인
                        found_wb = None
                        abs_file_path = os.path.abspath(file_path)
                        
                        for wb in excel.Workbooks:
                            if os.path.abspath(wb.FullName) == abs_file_path:
                                found_wb = wb
                                break
                        
                        # 워크북을 찾았으면 시트 활성화
                        if found_wb:
                            try:
                                # 시트 이름으로 시트 찾기
                                sheet = found_wb.Worksheets(sheet_name)
                                sheet.Activate()
                                print(f"시트 활성화 성공: {sheet_name}")
                            except Exception as e:
                                print(f"시트 활성화 실패: {e}")
                        else:
                            print("워크북을 찾을 수 없음")
                        
                        # 참조 해제
                        del excel
                        
                    except Exception as e:
                        print(f"COM 인터페이스 오류: {e}")
                        # COM 접근이 실패해도 파일은 이미 열림
                
                except Exception as e:
                    print(f"Excel 작업 오류: {e}")
                    # 일반적인 방법으로 파일만 열기
                    os.startfile(file_path)
                    
            else:  # 비 Windows 환경
                import subprocess
                subprocess.Popen(['xdg-open', file_path] if platform.system() == 'Linux' else ['open', file_path])
                print(f"파일 열기 성공 (비Windows): {file_path}")
            
            print(f"엑셀 파일 열기 성공: {file_path}")
            return True
            
        except Exception as e:
            print(f"백그라운드 Excel 작업 오류: {e}")
            # 오류 발생시 일반 방법으로 시도
            try:
                import os
                os.startfile(file_path) if platform.system() == 'Windows' else None
                return True
            except:
                return False

    def ensure_files_saved(self, file_paths):
        """여러 엑셀 파일이 완전히 저장되도록 함 (xlwings 사용)"""
        try:
            import xlwings as xw
            
            self._log_progress("xlwings로 파일 저장 확인 중...", success=True)
            
            # 단일 엑셀 인스턴스 사용
            app = xw.App(visible=False)
            app.display_alerts = False
            
            try:
                for idx, file_path in enumerate(file_paths):
                    try:
                        # 진행 상황 업데이트
                        self._log_progress(f"파일 저장 확인 중 ({idx+1}/{len(file_paths)}): {os.path.basename(file_path)}", success=True)
                        
                        # 워크북 열기
                        wb = app.books.open(file_path)
                        # 저장
                        wb.save()
                        # 닫기
                        wb.close()
                        
                    except Exception as e:
                        self._log_progress(f"파일 저장 확인 중 오류: {file_path} - {str(e)}", success=False)
            finally:
                # 항상 엑셀 앱 종료
                try:
                    app.quit()
                except:
                    pass
                
            self._log_progress("모든 파일 저장 확인 완료", success=True)
            
        except ImportError:
            self._log_progress("xlwings 모듈이 설치되어 있지 않아 파일 저장 확인을 건너뜁니다.", success=False)
        except Exception as e:
            self._log_progress(f"파일 저장 확인 중 오류 발생: {str(e)}", success=False)

    def modify_excel_file(self, file_path, sheet_name, tasks, header_row, col_map):
        """
        엑셀 파일의 특정 시트에서 작업 목록을 실행합니다.
        
        Args:
            file_path: 엑셀 파일 경로
            sheet_name: 시트 이름
            tasks: 작업 목록 (변경할 항목 정보)
            header_row: 헤더 행 번호
            col_map: 컬럼 이름과 인덱스 매핑
            
        Returns:
            (성공 여부, 변경된 행 목록)
        """
        # 파일 존재 확인
        if not os.path.exists(file_path):
            alt_path = find_excel_file(self.folder_path, os.path.basename(file_path))
            if alt_path:
                file_path = alt_path
            else:
                self._log_progress(f"파일을 찾을 수 없음: {os.path.basename(file_path)}", success=False)
                return False, []
        
        # 외부 링크 확인
        if self.check_external_links(file_path):
            self._log_progress(f"외부 링크 포함: '{os.path.basename(file_path)}'", success=False)
            self.files_with_external_links.add(file_path)
            return False, []
        
        # 열린 파일 확인
        if file_path in self.open_excel_files:
            self._log_progress(f"파일이 열려 있음: '{os.path.basename(file_path)}'", success=False)
            return False, []
        
        try:
            # 파일 열기
            workbook = load_workbook(file_path)
            
            # 시트 존재 확인
            if sheet_name not in workbook.sheetnames:
                self._log_progress(f"시트 없음: {sheet_name}", success=False)
                workbook.close()
                return False, []
                
            sheet = workbook[sheet_name]
            
            # STRING_ID 컬럼 확인
            string_id_col = col_map.get("STRING_ID")
            if not string_id_col:
                self._log_progress(f"STRING_ID 컬럼 없음", success=False)
                workbook.close()
                return False, []
            
            # 모든 STRING_ID와 행 번호 매핑을 미리 생성 (성능 향상)
            string_id_to_row = {}
            for row in range(header_row + 1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=string_id_col).value
                if cell_value:
                    string_id_to_row[str(cell_value).strip()] = row
            
            # 변경된 행 추적
            highlighted_rows = set()
            
            # 작업 처리
            self._log_progress(f"{len(tasks)}개 작업 처리 시작", success=True)
            
            # 언어 컬럼 리스트 정의
            LANG_COLUMNS = ["KR", "EN", "CN", "TW", "TH", "PT", "ES", "DE", "FR", "JP"]
            
            for task_idx, task in enumerate(tasks):
                mode, string_id, from_text, to_text, langs = task
                
                # 작업 진행 상황 업데이트 (10개 단위로)
                if task_idx % 10 == 0 or task_idx == 0:
                    self._log_progress(f"작업 진행 중: {task_idx+1}/{len(tasks)}", success=True)
                
                # 사전 생성된 매핑에서 행 찾기
                target_row = string_id_to_row.get(string_id)
                if not target_row:
                    continue
                
                # 변경 작업 수행 여부 추적
                changed = False
                
                if mode == "replace":
                    # 단어 치환 모드
                    for lang in langs if langs != ["ALL"] else LANG_COLUMNS:
                        # KR은 단어 치환 모드에서만 처리
                        if lang == "KR" and langs == ["ALL"]:
                            continue  # ALL이 선택되었을 때 KR은 건너뜀
                            
                        col_idx = col_map.get(lang)
                        if col_idx:
                            # 직접 셀 접근하여 값 변경
                            cell = sheet.cell(row=target_row, column=col_idx)
                            
                            if cell.value is not None:
                                # 문자열로 변환하여 처리
                                original_text = str(cell.value)
                                if from_text in original_text:
                                    # 변경 후 값을 직접 셀에 대입
                                    new_text = original_text.replace(from_text, to_text)
                                    cell.value = new_text
                                    changed = True
                                    
                elif mode == "bulk":
                    # 일괄 변경 모드
                    for lang in langs if langs != ["ALL"] else LANG_COLUMNS:
                        if lang == "KR":  # KR은 제외
                            continue
                            
                        col_idx = col_map.get(lang)
                        if col_idx:
                            # 직접 값 설정
                            sheet.cell(row=target_row, column=col_idx).value = to_text
                            changed = True
                            
                elif mode == "unique":
                    # 고유 텍스트 치환 모드
                    kr_text = from_text
                    lang_data = to_text
                    
                    for lang in langs if langs != ["ALL"] else LANG_COLUMNS:
                        if lang == "KR":  # KR은 이미 원본 텍스트
                            continue
                            
                        # 해당 언어의 치환 데이터가 있는지 확인
                        if lang in lang_data:
                            col_idx = col_map.get(lang)
                            if col_idx:
                                # 직접 값 설정
                                sheet.cell(row=target_row, column=col_idx).value = lang_data[lang]
                                changed = True
                
                # 변경된 행은 목록에 추가
                if changed:
                    highlighted_rows.add(target_row)
            
            # 변경된 행에 배경색 적용 (A~N열까지)
            if highlighted_rows:
                from openpyxl.styles import PatternFill
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                self._log_progress(f"{len(highlighted_rows)}개 행에 배경색 적용 중", success=True)
                for row in highlighted_rows:
                    for col in range(1, 15):  # A~N열 (1~14)
                        cell = sheet.cell(row=row, column=col)
                        cell.fill = yellow_fill
            
            # 파일 저장
            self._log_progress(f"파일 저장 중: {os.path.basename(file_path)}", success=True)
            
            try:
                # 파일 저장
                workbook.save(file_path)
                self._log_progress(f"✅ 저장 완료: {os.path.basename(file_path)}", success=True)
                return True, list(highlighted_rows)
            except Exception as save_error:
                self._log_progress(f"저장 실패: {str(save_error)}", success=False)
                return False, []
            finally:
                # 워크북 닫기
                workbook.close()
                
        except Exception as e:
            self._log_progress(f"파일 처리 중 오류 발생: {str(e)}", success=False)
            traceback.print_exc()
            return False, []