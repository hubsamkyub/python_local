#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 파일의 외부 링크 검사 도구 (GUI 포함)
"""

import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
import re
import threading

class ExcelExternalLinkChecker:
    def __init__(self):
        self.external_patterns = [
            r'\[.*\.xl.*\]',  # [파일명.xlsx] 패턴
            r"'.*:.*'!",      # '경로:파일명'! 패턴
            r'\\.*\.xl.*!',   # \경로\파일명.xlsx! 패턴
            r"'[A-Z]:[^']*'!"  # 'C:\경로'! 패턴
        ]
    
    def check_file(self, file_path, max_cells=1000, detailed=False):
        """
        엑셀 파일의 외부 링크 검사
        
        Args:
            file_path (str): 검사할 엑셀 파일 경로
            max_cells (int): 검사할 최대 셀 개수 (성능 최적화)
            detailed (bool): 상세 정보 반환 여부
            
        Returns:
            dict: 검사 결과
        """
        result = {
            'has_external_links': False,
            'external_links': [],
            'error': None,
            'file_name': os.path.basename(file_path)
        }
        
        try:
            # 파일 존재 확인
            if not os.path.exists(file_path):
                result['error'] = "파일을 찾을 수 없습니다"
                return result
            
            # 워크북 로드
            try:
                workbook = load_workbook(file_path, data_only=False)  # 공식도 가져오기
            except Exception as e:
                result['error'] = f"파일 열기 실패: {str(e)}"
                return result
            
            # 방법 1: 워크북 레벨 외부 링크 확인
            try:
                if hasattr(workbook, 'external_links') and workbook.external_links:
                    for link in workbook.external_links:
                        result['external_links'].append({
                            'type': 'workbook_link',
                            'target': str(link),
                            'location': 'workbook'
                        })
            except:
                pass
            
            # 방법 1.5: 명명된 범위(Named Ranges) 검사 - 중요!
            try:
                if hasattr(workbook, 'defined_names') and workbook.defined_names:
                    # 올바른 접근 방법: 딕셔너리 키로 접근
                    for name_key in workbook.defined_names.keys():
                        try:
                            defined_name = workbook.defined_names[name_key]
                            if hasattr(defined_name, 'value') and defined_name.value:
                                name_formula = str(defined_name.value)
                                
                                # #REF! 오류가 있는 명명된 범위 검사
                                ref_error_found = False
                                for ref_pattern in self.ref_error_patterns:
                                    if re.search(ref_pattern, name_formula):
                                        result['external_links'].append({
                                            'type': 'named_range_ref_error',
                                            'name': name_key,
                                            'formula': name_formula[:100] + ('...' if len(name_formula) > 100 else ''),
                                            'location': f"명명된_범위:{name_key}"
                                        })
                                        ref_error_found = True
                                        break
                                
                                # #REF! 오류가 없는 경우에만 외부 참조 패턴 검사
                                if not ref_error_found:
                                    for pattern in self.external_patterns:
                                        if re.search(pattern, name_formula):
                                            result['external_links'].append({
                                                'type': 'named_range_external_link',
                                                'name': name_key,
                                                'formula': name_formula[:100] + ('...' if len(name_formula) > 100 else ''),
                                                'location': f"명명된_범위:{name_key}"
                                            })
                                            break
                        except Exception as e:
                            # 개별 명명된 범위 처리 중 오류가 발생해도 계속 진행
                            pass
            except Exception as e:
                # 명명된 범위 검사 중 오류가 발생해도 계속 진행
                pass
            
            # 방법 2: 셀별 외부 참조 검사
            cell_count = 0
            for sheet_name in workbook.sheetnames:
                if cell_count >= max_cells:
                    break
                    
                worksheet = workbook[sheet_name]
                
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell_count += 1
                        if cell_count > max_cells:
                            break
                        
                        # 공식이 있는 셀 검사
                        if cell.data_type == 'f' and cell.value:
                            formula = str(cell.value)
                            
                            # 외부 참조 패턴 검사
                            for pattern in self.external_patterns:
                                if re.search(pattern, formula):
                                    link_info = {
                                        'type': 'formula_link',
                                        'location': f"{sheet_name}!{cell.coordinate}",
                                        'formula': formula[:100] + ('...' if len(formula) > 100 else '')
                                    }
                                    if detailed:
                                        link_info['full_formula'] = formula
                                    result['external_links'].append(link_info)
                                    break
                        
                        # #REF! 오류 검사
                        elif cell.value and str(cell.value).startswith('#REF!'):
                            result['external_links'].append({
                                'type': 'ref_error',
                                'location': f"{sheet_name}!{cell.coordinate}",
                                'value': str(cell.value)
                            })
                    
                    if cell_count > max_cells:
                        break
            
            workbook.close()
            
            # 외부 링크 발견 여부 설정
            result['has_external_links'] = len(result['external_links']) > 0
            
        except Exception as e:
            result['error'] = f"검사 중 오류 발생: {str(e)}"
        
        return result
    
    def check_folder(self, folder_path, pattern="*.xlsx", max_files=None, progress_callback=None):
        """
        폴더 내 엑셀 파일들의 외부 링크 검사
        
        Args:
            folder_path (str): 검사할 폴더 경로
            pattern (str): 파일 패턴 (기본: *.xlsx)
            max_files (int): 검사할 최대 파일 수
            progress_callback (func): 진행 상황 콜백 함수
            
        Returns:
            dict: 검사 결과 요약
        """
        import glob
        
        results = {
            'total_files': 0,
            'files_with_links': 0,
            'files_with_errors': 0,
            'problem_files': [],
            'summary': {}
        }
        
        if not os.path.exists(folder_path):
            results['error'] = "폴더를 찾을 수 없습니다"
            return results
        
        # 패턴에 맞는 파일 찾기
        search_pattern = os.path.join(folder_path, pattern)
        files = glob.glob(search_pattern)
        
        if max_files:
            files = files[:max_files]
        
        results['total_files'] = len(files)
        
        for i, file_path in enumerate(files):
            if progress_callback:
                progress_callback(f"검사 중: {os.path.basename(file_path)}", i + 1, len(files))
            
            file_result = self.check_file(file_path)
            
            if file_result['error']:
                results['files_with_errors'] += 1
                results['problem_files'].append({
                    'file': os.path.basename(file_path),
                    'status': 'error',
                    'message': file_result['error']
                })
            elif file_result['has_external_links']:
                results['files_with_links'] += 1
                results['problem_files'].append({
                    'file': os.path.basename(file_path),
                    'status': 'has_links',
                    'link_count': len(file_result['external_links']),
                    'links': file_result['external_links'][:3]  # 처음 3개만
                })
        
        # 요약 정보
        results['summary'] = {
            'clean_files': results['total_files'] - results['files_with_links'] - results['files_with_errors'],
            'problem_rate': round((results['files_with_links'] + results['files_with_errors']) / results['total_files'] * 100, 1) if results['total_files'] > 0 else 0
        }
        
        return results


class ExcelLinkCheckerGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("엑셀 외부 링크 검사 도구")
        self.root.geometry("800x600")
        
        self.checker = ExcelExternalLinkChecker()
        
        # 패턴들을 직접 정의 (참조 오류 방지)
        self.external_patterns = [
            r"'[^']*\.xl[sx]?[xm]?'!",  # '파일명.xlsx'! 또는 '경로\파일명.xlsx'!
            r'\[.*\.xl[sx]?[xm]?\]',    # [파일명.xlsx] 패턴
            r"'[A-Z]:[^']*\.xl[sx]?[xm]?'!", # 'C:\경로\파일명.xlsx'! 패턴  
            r'\\[^\\]*\.xl[sx]?[xm]?!', # \파일명.xlsx! 패턴
            r"=[^=]*'[A-Z]:[^']*'",     # =으로 시작하는 드라이브 경로
            r'\[\d+\]!',                # [숫자]! 패턴 (시트 참조)
        ]
        
        self.ref_error_patterns = [
            r'#REF!',                   # #REF! 오류
            r'OFFSET\(#REF!',          # OFFSET 함수에서 #REF! 오류
        ]
        
        self.setup_ui()
        
    def setup_ui(self):
        """UI 구성"""
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 파일/폴더 선택 영역
        input_frame = ttk.LabelFrame(main_frame, text="검사 대상 선택", padding="10")
        input_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 파일 선택
        file_frame = ttk.Frame(input_frame)
        file_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(file_frame, text="파일:").pack(side=tk.LEFT)
        self.file_path_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.file_path_var, width=50).pack(side=tk.LEFT, padx=(5, 5), fill=tk.X, expand=True)
        ttk.Button(file_frame, text="파일 선택", command=self.select_file).pack(side=tk.RIGHT)
        
        # 폴더 선택
        folder_frame = ttk.Frame(input_frame)
        folder_frame.pack(fill=tk.X, pady=(5, 0))
        
        ttk.Label(folder_frame, text="폴더:").pack(side=tk.LEFT)
        self.folder_path_var = tk.StringVar()
        ttk.Entry(folder_frame, textvariable=self.folder_path_var, width=50).pack(side=tk.LEFT, padx=(5, 5), fill=tk.X, expand=True)
        ttk.Button(folder_frame, text="폴더 선택", command=self.select_folder).pack(side=tk.RIGHT)
        
        # 옵션 영역
        option_frame = ttk.LabelFrame(main_frame, text="검사 옵션", padding="10")
        option_frame.pack(fill=tk.X, pady=(0, 10))
        
        options_grid = ttk.Frame(option_frame)
        options_grid.pack(fill=tk.X)
        
        ttk.Label(options_grid, text="최대 셀 검사 수:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.max_cells_var = tk.IntVar(value=1000)
        ttk.Spinbox(options_grid, from_=100, to=10000, increment=100, textvariable=self.max_cells_var, width=10).grid(row=0, column=1, sticky=tk.W, padx=(0, 20))
        
        self.detailed_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_grid, text="상세 정보 표시", variable=self.detailed_var).grid(row=0, column=2, sticky=tk.W)
        
        # 실행 버튼
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(button_frame, text="파일 검사", command=self.check_file).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="폴더 검사", command=self.check_folder).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="결과 지우기", command=self.clear_results).pack(side=tk.LEFT, padx=(5, 0))
        
        # 진행 상황 표시
        progress_frame = ttk.Frame(main_frame)
        progress_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.progress_label = ttk.Label(progress_frame, text="대기 중...")
        self.progress_label.pack(side=tk.LEFT)
        
        self.progress_bar = ttk.Progressbar(progress_frame, length=300, mode='determinate')
        self.progress_bar.pack(side=tk.RIGHT)
        
        # 결과 표시 영역
        result_frame = ttk.LabelFrame(main_frame, text="검사 결과", padding="10")
        result_frame.pack(fill=tk.BOTH, expand=True)
        
        # 텍스트 위젯과 스크롤바
        text_frame = ttk.Frame(result_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        self.result_text = tk.Text(text_frame, wrap=tk.WORD, font=("Consolas", 9))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.result_text.yview)
        self.result_text.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
    def select_file(self):
        """파일 선택 다이얼로그"""
        file_path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xls"), ("모든 파일", "*.*")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.folder_path_var.set("")  # 폴더 경로 지우기
    
    def select_folder(self):
        """폴더 선택 다이얼로그"""
        folder_path = filedialog.askdirectory(title="폴더 선택")
        if folder_path:
            self.folder_path_var.set(folder_path)
            self.file_path_var.set("")  # 파일 경로 지우기
    
    def check_file(self):
        """단일 파일 검사"""
        file_path = self.file_path_var.get().strip()
        if not file_path:
            messagebox.showwarning("경고", "검사할 파일을 선택하세요.")
            return
        
        self.progress_label.config(text="파일 검사 중...")
        self.progress_bar.config(value=0, maximum=1)
        
        def run_check():
            try:
                result = self.checker.check_file(
                    file_path, 
                    max_cells=self.max_cells_var.get(),
                    detailed=self.detailed_var.get()
                )
                
                self.root.after(0, lambda: self.display_file_result(result))
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("오류", f"검사 중 오류 발생: {str(e)}"))
            finally:
                self.root.after(0, lambda: [
                    self.progress_label.config(text="완료"),
                    self.progress_bar.config(value=1)
                ])
        
        threading.Thread(target=run_check, daemon=True).start()
    
    def check_folder(self):
        """폴더 검사"""
        folder_path = self.folder_path_var.get().strip()
        if not folder_path:
            messagebox.showwarning("경고", "검사할 폴더를 선택하세요.")
            return
        
        def progress_callback(message, current, total):
            self.root.after(0, lambda: [
                self.progress_label.config(text=message),
                self.progress_bar.config(value=current, maximum=total)
            ])
        
        def run_check():
            try:
                result = self.checker.check_folder(
                    folder_path,
                    pattern="*.xlsx",
                    progress_callback=progress_callback
                )
                
                self.root.after(0, lambda: self.display_folder_result(result))
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("오류", f"검사 중 오류 발생: {str(e)}"))
            finally:
                self.root.after(0, lambda: [
                    self.progress_label.config(text="완료"),
                    self.progress_bar.config(value=self.progress_bar.cget('maximum'))
                ])
        
        threading.Thread(target=run_check, daemon=True).start()
    
    def display_file_result(self, result):
        """파일 검사 결과 표시"""
        self.result_text.delete(1.0, tk.END)
        
        self.result_text.insert(tk.END, f"=== 파일 검사 결과: {result['file_name']} ===\n\n")
        
        if result['error']:
            self.result_text.insert(tk.END, f"❌ 오류: {result['error']}\n")
        elif result['has_external_links']:
            self.result_text.insert(tk.END, f"⚠️  외부 링크 발견: {len(result['external_links'])}개\n\n")
            
            # 명명된 범위 오류 먼저 표시 (중요도 높음)
            named_range_errors = [link for link in result['external_links'] if 'named_range' in link['type']]
            cell_errors = [link for link in result['external_links'] if 'named_range' not in link['type']]
            
            if named_range_errors:
                self.result_text.insert(tk.END, "🚨 명명된 범위 오류 (번역 적용 시 문제 발생 가능):\n")
                for i, link in enumerate(named_range_errors, 1):
                    self.result_text.insert(tk.END, f"{i}. 이름: {link.get('name', 'Unknown')}\n")
                    self.result_text.insert(tk.END, f"   타입: {link['type']}\n")
                    self.result_text.insert(tk.END, f"   공식: {link['formula']}\n\n")
            
            if cell_errors:
                self.result_text.insert(tk.END, "📋 셀 단위 외부 링크:\n")
                for i, link in enumerate(cell_errors, len(named_range_errors) + 1):
                    self.result_text.insert(tk.END, f"{i}. 타입: {link['type']}\n")
                    self.result_text.insert(tk.END, f"   위치: {link['location']}\n")
                    
                    if 'formula' in link:
                        self.result_text.insert(tk.END, f"   공식: {link['formula']}\n")
                    elif 'value' in link:
                        self.result_text.insert(tk.END, f"   값: {link['value']}\n")
                    elif 'target' in link:
                        self.result_text.insert(tk.END, f"   대상: {link['target']}\n")
                    
                    self.result_text.insert(tk.END, "\n")
        else:
            self.result_text.insert(tk.END, "✅ 외부 링크 없음 - 정상 파일입니다.\n")
    
    def display_folder_result(self, result):
        """폴더 검사 결과 표시"""
        self.result_text.delete(1.0, tk.END)
        
        if 'error' in result:
            self.result_text.insert(tk.END, f"❌ 오류: {result['error']}\n")
            return
        
        self.result_text.insert(tk.END, "=== 폴더 검사 결과 ===\n\n")
        self.result_text.insert(tk.END, f"📊 요약:\n")
        self.result_text.insert(tk.END, f"  • 총 파일 수: {result['total_files']}\n")
        self.result_text.insert(tk.END, f"  • 외부 링크 있는 파일: {result['files_with_links']}\n")
        self.result_text.insert(tk.END, f"  • 오류 발생 파일: {result['files_with_errors']}\n")
        self.result_text.insert(tk.END, f"  • 정상 파일: {result['summary']['clean_files']}\n")
        self.result_text.insert(tk.END, f"  • 문제 비율: {result['summary']['problem_rate']}%\n\n")
        
        if result['problem_files']:
            self.result_text.insert(tk.END, "🚨 문제 파일 목록:\n\n")
            
            for i, problem in enumerate(result['problem_files'], 1):
                if problem['status'] == 'has_links':
                    self.result_text.insert(tk.END, f"{i}. 🔗 {problem['file']}\n")
                    self.result_text.insert(tk.END, f"   외부 링크: {problem['link_count']}개\n")
                    
                    if problem['links']:
                        self.result_text.insert(tk.END, "   링크 예시:\n")
                        for link in problem['links']:
                            self.result_text.insert(tk.END, f"     - {link['location']}: {link['type']}\n")
                else:
                    self.result_text.insert(tk.END, f"{i}. ❌ {problem['file']}\n")
                    self.result_text.insert(tk.END, f"   오류: {problem['message']}\n")
                
                self.result_text.insert(tk.END, "\n")
        else:
            self.result_text.insert(tk.END, "✅ 모든 파일이 정상입니다!\n")
    
    def clear_results(self):
        """결과 지우기"""
        self.result_text.delete(1.0, tk.END)
        self.progress_label.config(text="대기 중...")
        self.progress_bar.config(value=0)
    
    def run(self):
        """GUI 실행"""
        self.root.mainloop()


def main():
    """메인 함수"""
    if len(sys.argv) > 1:
        # 명령행 모드
        target_path = sys.argv[1]
        checker = ExcelExternalLinkChecker()
        
        if os.path.isfile(target_path):
            # 단일 파일 검사
            print(f"파일 검사 중: {target_path}")
            result = checker.check_file(target_path, detailed=True)
            
            print(f"\n=== 검사 결과: {result['file_name']} ===")
            
            if result['error']:
                print(f"❌ 오류: {result['error']}")
            elif result['has_external_links']:
                print(f"⚠️  외부 링크 발견: {len(result['external_links'])}개")
                
                # 명명된 범위 오류와 셀 오류 구분
                named_range_errors = [link for link in result['external_links'] if 'named_range' in link['type']]
                cell_errors = [link for link in result['external_links'] if 'named_range' not in link['type']]
                
                if named_range_errors:
                    print(f"\n🚨 명명된 범위 오류 ({len(named_range_errors)}개) - 번역 적용 시 문제 발생 가능:")
                    for i, link in enumerate(named_range_errors, 1):
                        print(f"  {i}. {link.get('name', 'Unknown')} - {link['type']}")
                        print(f"     공식: {link['formula']}")
                
                if cell_errors:
                    print(f"\n📋 셀 단위 외부 링크 ({len(cell_errors)}개):")
                    for i, link in enumerate(cell_errors[:5], 1):  # 처음 5개만
                        print(f"  {i}. {link['type']} - {link['location']}")
                        if 'formula' in link:
                            print(f"     공식: {link['formula']}")
                        elif 'value' in link:
                            print(f"     값: {link['value']}")
                    if len(cell_errors) > 5:
                        print(f"  ... 외 {len(cell_errors) - 5}개")
            else:
                print("✅ 외부 링크 없음")
        
        elif os.path.isdir(target_path):
            # 폴더 검사
            print(f"폴더 검사 중: {target_path}")
            result = checker.check_folder(target_path)
            
            print(f"\n=== 폴더 검사 결과 ===")
            print(f"총 파일 수: {result['total_files']}")
            print(f"외부 링크 있는 파일: {result['files_with_links']}")
            print(f"오류 발생 파일: {result['files_with_errors']}")
            print(f"정상 파일: {result['summary']['clean_files']}")
            print(f"문제 비율: {result['summary']['problem_rate']}%")
            
            if result['problem_files']:
                print(f"\n=== 문제 파일 목록 ===")
                for problem in result['problem_files']:
                    if problem['status'] == 'has_links':
                        print(f"🔗 {problem['file']} - 외부 링크 {problem['link_count']}개")
                    else:
                        print(f"❌ {problem['file']} - {problem['message']}")
        
        else:
            print(f"❌ 경로를 찾을 수 없습니다: {target_path}")
    else:
        # GUI 모드
        app = ExcelLinkCheckerGUI()
        app.run()


if __name__ == "__main__":
    main()