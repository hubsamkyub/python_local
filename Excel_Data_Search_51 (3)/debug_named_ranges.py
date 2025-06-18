#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
명명된 범위 디버깅 도구 - 실제 데이터 확인용
"""

import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
import re

class NamedRangeDebugger:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("명명된 범위 디버깅 도구")
        self.root.geometry("900x700")
        
        # 개선된 외부 참조 패턴들
        self.external_patterns = [
            r"'[^']*\.xl[sx]?[xm]?'!",  # '파일명.xlsx'! 또는 '경로\파일명.xlsx'!
            r'\[.*\.xl[sx]?[xm]?\]',    # [파일명.xlsx] 패턴
            r"'[A-Z]:[^']*\.xl[sx]?[xm]?'!", # 'C:\경로\파일명.xlsx'! 패턴  
            r'\\[^\\]*\.xl[sx]?[xm]?!', # \파일명.xlsx! 패턴
            r"=[^=]*'[A-Z]:[^']*'",     # =으로 시작하는 드라이브 경로
        ]
        
        self.setup_ui()
        
    def setup_ui(self):
        """UI 구성"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 파일 선택
        file_frame = ttk.Frame(main_frame)
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(file_frame, text="엑셀 파일:").pack(side=tk.LEFT)
        self.file_path_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.file_path_var, width=60).pack(side=tk.LEFT, padx=(5, 5), fill=tk.X, expand=True)
        ttk.Button(file_frame, text="파일 선택", command=self.select_file).pack(side=tk.RIGHT)
        
        # 버튼들
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(button_frame, text="명명된 범위 분석", command=self.analyze_named_ranges).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="결과 지우기", command=self.clear_results).pack(side=tk.LEFT)
        
        # 결과 표시
        result_frame = ttk.LabelFrame(main_frame, text="분석 결과", padding="10")
        result_frame.pack(fill=tk.BOTH, expand=True)
        
        text_frame = ttk.Frame(result_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        self.result_text = tk.Text(text_frame, wrap=tk.WORD, font=("Consolas", 9))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.result_text.yview)
        self.result_text.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
    def select_file(self):
        """파일 선택"""
        file_path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xls"), ("모든 파일", "*.*")]
        )
        if file_path:
            self.file_path_var.set(file_path)
    
    def analyze_named_ranges(self):
        """명명된 범위 상세 분석"""
        file_path = self.file_path_var.get().strip()
        if not file_path:
            messagebox.showwarning("경고", "파일을 선택하세요.")
            return
        
        try:
            self.result_text.delete(1.0, tk.END)
            self.result_text.insert(tk.END, f"=== 명명된 범위 분석: {os.path.basename(file_path)} ===\n\n")
            
            # 워크북 로드
            workbook = load_workbook(file_path, data_only=False)
            
            # 방법 1: workbook.defined_names 확인
            self.result_text.insert(tk.END, "🔍 방법 1: workbook.defined_names\n")
            self.result_text.insert(tk.END, f"defined_names 존재: {hasattr(workbook, 'defined_names')}\n")
            
            if hasattr(workbook, 'defined_names'):
                defined_names = workbook.defined_names
                self.result_text.insert(tk.END, f"defined_names 타입: {type(defined_names)}\n")
                self.result_text.insert(tk.END, f"정의된 이름 개수: {len(defined_names) if defined_names else 0}\n")
                
                # DefinedNameDict 구조 자세히 분석
                self.result_text.insert(tk.END, f"defined_names 속성들: {[attr for attr in dir(defined_names) if not attr.startswith('_')]}\n\n")
                
                if defined_names:
                    external_found = 0
                    ref_error_found = 0
                    
                    # 방법 A: 직접 반복
                    self.result_text.insert(tk.END, "🔍 방법 A: 직접 반복\n")
                    for i, name in enumerate(defined_names):
                        self.result_text.insert(tk.END, f"--- 명명된 범위 {i+1} (직접 반복) ---\n")
                        self.result_text.insert(tk.END, f"객체 타입: {type(name)}\n")
                        self.result_text.insert(tk.END, f"객체 속성들: {[attr for attr in dir(name) if not attr.startswith('_')]}\n")
                        
                        # 다양한 방법으로 이름과 값 추출 시도
                        name_candidates = ['name', 'localSheetId', 'comment']
                        value_candidates = ['value', 'attr_text', 'text']
                        
                        for attr in name_candidates:
                            try:
                                val = getattr(name, attr, None)
                                if val:
                                    self.result_text.insert(tk.END, f"{attr}: {val}\n")
                            except:
                                pass
                        
                        for attr in value_candidates:
                            try:
                                val = getattr(name, attr, None)
                                if val:
                                    self.result_text.insert(tk.END, f"{attr}: {val}\n")
                                    
                                    # 이 값에 대해 패턴 검사
                                    val_str = str(val)
                                    if '#REF!' in val_str:
                                        ref_error_found += 1
                                        self.result_text.insert(tk.END, "🚨 #REF! 오류 발견!\n")
                                    
                                    for j, pattern in enumerate(self.external_patterns):
                                        if re.search(pattern, val_str):
                                            external_found += 1
                                            self.result_text.insert(tk.END, f"🔗 외부 참조 패턴 {j+1} 매칭: {pattern}\n")
                                            break
                            except:
                                pass
                        
                        self.result_text.insert(tk.END, "\n")
                    
                    # 방법 B: definedName 속성 확인
                    self.result_text.insert(tk.END, "🔍 방법 B: definedName 속성\n")
                    if hasattr(defined_names, 'definedName'):
                        def_names = defined_names.definedName
                        self.result_text.insert(tk.END, f"definedName 타입: {type(def_names)}\n")
                        if def_names:
                            for i, name in enumerate(def_names):
                                self.result_text.insert(tk.END, f"--- definedName {i+1} ---\n")
                                self.result_text.insert(tk.END, f"타입: {type(name)}\n")
                                self.result_text.insert(tk.END, f"속성들: {[attr for attr in dir(name) if not attr.startswith('_')]}\n")
                                
                                # 모든 속성 출력
                                for attr in ['name', 'value', 'attr_text', 'text', 'localSheetId']:
                                    try:
                                        val = getattr(name, attr, None)
                                        if val is not None:
                                            self.result_text.insert(tk.END, f"{attr}: {val}\n")
                                    except:
                                        pass
                                self.result_text.insert(tk.END, "\n")
                    
                    # 방법 C: 키-값 쌍으로 접근
                    self.result_text.insert(tk.END, "🔍 방법 C: 키-값 접근\n")
                    try:
                        items = list(defined_names.items()) if hasattr(defined_names, 'items') else []
                        self.result_text.insert(tk.END, f"items() 결과: {len(items)}개\n")
                        for key, value in items:
                            self.result_text.insert(tk.END, f"키: {key}, 값 타입: {type(value)}\n")
                            if hasattr(value, 'value'):
                                self.result_text.insert(tk.END, f"값: {value.value}\n")
                    except Exception as e:
                        self.result_text.insert(tk.END, f"items() 접근 실패: {e}\n")
                    
                    # 방법 D: 딕셔너리 키들 확인
                    self.result_text.insert(tk.END, "🔍 방법 D: 딕셔너리 키 확인\n")
                    try:
                        keys = list(defined_names.keys()) if hasattr(defined_names, 'keys') else []
                        self.result_text.insert(tk.END, f"키들: {keys}\n")
                        for key in keys:
                            try:
                                value = defined_names[key]
                                self.result_text.insert(tk.END, f"키 '{key}': {type(value)} - {value}\n")
                                if hasattr(value, 'value'):
                                    val_str = str(value.value)
                                    self.result_text.insert(tk.END, f"  실제 값: {val_str}\n")
                                    
                                    # 패턴 검사
                                    if '#REF!' in val_str:
                                        ref_error_found += 1
                                        self.result_text.insert(tk.END, "  🚨 #REF! 오류 발견!\n")
                                    
                                    for j, pattern in enumerate(self.external_patterns):
                                        if re.search(pattern, val_str):
                                            external_found += 1
                                            self.result_text.insert(tk.END, f"  🔗 외부 참조 패턴 {j+1} 매칭: {pattern}\n")
                                            break
                            except Exception as e:
                                self.result_text.insert(tk.END, f"키 '{key}' 접근 실패: {e}\n")
                    except Exception as e:
                        self.result_text.insert(tk.END, f"키 확인 실패: {e}\n")
                    
                    self.result_text.insert(tk.END, f"\n📊 요약:\n")
                    self.result_text.insert(tk.END, f"  • #REF! 오류: {ref_error_found}개\n")
                    self.result_text.insert(tk.END, f"  • 외부 참조: {external_found}개\n\n")
            
            # 방법 2: 워크북의 다른 속성들 확인
            self.result_text.insert(tk.END, "🔍 방법 2: 기타 워크북 속성\n")
            
            # external_links 확인
            if hasattr(workbook, 'external_links'):
                ext_links = workbook.external_links
                self.result_text.insert(tk.END, f"external_links: {ext_links}\n")
                if ext_links:
                    for link in ext_links:
                        self.result_text.insert(tk.END, f"  - {link}\n")
            
            # 워크북 속성들 나열
            self.result_text.insert(tk.END, f"\n워크북 속성들:\n")
            attrs = [attr for attr in dir(workbook) if not attr.startswith('_')]
            for attr in sorted(attrs):
                try:
                    value = getattr(workbook, attr)
                    if not callable(value):
                        self.result_text.insert(tk.END, f"  {attr}: {type(value)}\n")
                except:
                    pass
            
            # 방법 3: 시트별 이름 확인
            self.result_text.insert(tk.END, f"\n🔍 방법 3: 각 시트별 확인\n")
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                self.result_text.insert(tk.END, f"\n시트: {sheet_name}\n")
                
                # 시트 속성들 확인
                sheet_attrs = [attr for attr in dir(sheet) if 'name' in attr.lower() and not attr.startswith('_')]
                for attr in sheet_attrs:
                    try:
                        value = getattr(sheet, attr)
                        self.result_text.insert(tk.END, f"  {attr}: {value}\n")
                    except:
                        pass
            
            workbook.close()
            
        except Exception as e:
            self.result_text.insert(tk.END, f"❌ 오류 발생: {str(e)}\n")
            import traceback
            self.result_text.insert(tk.END, f"\n상세 오류:\n{traceback.format_exc()}\n")
    
    def clear_results(self):
        """결과 지우기"""
        self.result_text.delete(1.0, tk.END)
    
    def run(self):
        """실행"""
        self.root.mainloop()

def test_patterns():
    """패턴 테스트"""
    test_strings = [
        "='C:\\LegendOfHeroes\\Assets\\DB\\퀘스트.xlsx'!표1[#데이터]",
        "=OFFSET(#REF!#REF!,,3,COUNTA(#REF!#REF!))",
        "=[다른파일.xlsx]Sheet1!A1",
        "'C:\\temp\\파일.xlsx'!Sheet1!A1:B10",
        "#REF!",
        "=Sheet1!A1"
    ]
    
    patterns = [
        r"'[^']*\.xl[sx]?[xm]?'!",  # '파일명.xlsx'! 또는 '경로\파일명.xlsx'!
        r'\[.*\.xl[sx]?[xm]?\]',    # [파일명.xlsx] 패턴
        r"'[A-Z]:[^']*\.xl[sx]?[xm]?'!", # 'C:\경로\파일명.xlsx'! 패턴  
        r'\\[^\\]*\.xl[sx]?[xm]?!', # \파일명.xlsx! 패턴
        r"=[^=]*'[A-Z]:[^']*'",     # =으로 시작하는 드라이브 경로
    ]
    
    print("=== 패턴 테스트 ===")
    for test_str in test_strings:
        print(f"\n테스트 문자열: {test_str}")
        for i, pattern in enumerate(patterns):
            match = re.search(pattern, test_str)
            print(f"  패턴 {i+1}: {'✅ 매칭' if match else '❌ 미매칭'} - {pattern}")

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--test":
        test_patterns()
    else:
        app = NamedRangeDebugger()
        app.run()